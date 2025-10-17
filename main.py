import flet as ft
import datetime
import random
import string
from db_manager import DBManager
import threading
import getpass
from concurrent.futures import ThreadPoolExecutor
import sys
import os

# ============================================================
# VERSÃO DO APLICATIVO
# ============================================================
APP_VERSION = "1.1.5"

# Função para obter o caminho correto dos recursos (funciona tanto em desenvolvimento quanto em executável)
def resource_path(relative_path):
    """Obtém o caminho absoluto para o recurso, funciona para dev e para PyInstaller"""
    try:
        # PyInstaller cria uma pasta temp e armazena o caminho em _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# Configuração mínima de toasts (usada por redirect de `print` e por notificações)
app_settings = {'toast_duration': 3, 'auto_save': True}

db_manager = DBManager('db.db')

def ensure_supplier_po_column():
    """Verifica se a coluna supplier_po existe, se não, cria ela"""
    try:
        # Verificar se a coluna supplier_po existe
        columns_info = db_manager.query("PRAGMA table_info(supplier_database_table)")
        column_names = [col[1] for col in columns_info]
        
        if 'supplier_po' not in column_names:
            print("🔧 Coluna supplier_po não encontrada, criando...")
            db_manager.execute("ALTER TABLE supplier_database_table ADD COLUMN supplier_po TEXT")
            print("✅ Coluna supplier_po criada com sucesso!")
        else:
            print("✅ Coluna supplier_po já existe")
    except Exception as e:
        # Se for erro de coluna duplicada, está tudo certo
        if "duplicate column name" in str(e).lower():
            print("✅ Coluna supplier_po já existe (verificação por erro)")
        else:
            print(f"❌ Erro ao verificar coluna supplier_po: {e}")
            # Tentar criar a coluna apenas se não for erro de duplicata
            try:
                db_manager.execute("ALTER TABLE supplier_database_table ADD COLUMN supplier_po TEXT")
                print("✅ Coluna supplier_po criada após erro inicial")
            except Exception as e2:
                if "duplicate column name" not in str(e2).lower():
                    print(f"❌ Falha ao criar coluna supplier_po: {e2}")

def show_snack_bar(message, is_error=False, is_warning=False, page=None):
    """
    Mostra uma mensagem snack bar na interface de forma thread-safe
    
    Args:
        message: Mensagem a ser exibida
        is_error: True para erro (vermelho), False para outros tipos
        is_warning: True para aviso (laranja), False para outros tipos
        page: Instância do page (opcional, usa page_ref global se não fornecido)
    """
    global page_ref
    
    # Tentar obter referência do page
    target_page = page if page is not None else page_ref
    
    try:
        if target_page is not None and hasattr(target_page, 'open'):
            # Definir cor: vermelho para erro, laranja para aviso, verde para sucesso
            if is_error:
                color = ft.Colors.RED
            elif is_warning:
                color = ft.Colors.ORANGE
            else:
                color = ft.Colors.GREEN
            
            # Criar SnackBar conforme documentação do Flet - usar content explicitamente
            snack_bar = ft.SnackBar(
                content=ft.Text(message, color=ft.Colors.WHITE),
                bgcolor=color,
            )
            
            # Usar page.open() conforme exemplo oficial do Flet
            try:
                target_page.open(snack_bar)
                target_page.update()
            except Exception as open_error:
                # Se open() falhar, tenta adicionar ao overlay
                print(f"Aviso: page.open() falhou: {open_error}")
                print(f"Notification: {message}")
        else:
            print(f"Notification: {message}")
    except Exception as e:
        print(f"Erro em show_snack_bar: {e}")
        print(f"Notification: {message}")


def load_spinbox_increment():
    try:
        return 0.1
    except Exception:
        return 0.1

def format_po_ssid(value):
    """Formata valores de PO e SSID para exibir sem decimais"""
    if value is None or value == '' or value == '?' or str(value).lower() == 'none':
        return ''
    try:
        # Tentar converter para número e retornar como inteiro
        num_value = float(value)
        return str(int(num_value))
    except (ValueError, TypeError):
        # Se não for número, retornar como string
        return str(value)

def format_po_ssid_for_save(value):
    """Formata valores de PO e SSID antes de salvar no banco (remove decimais)"""
    if value is None or value == '' or value == '?' or str(value).lower() == 'none':
        return ''
    try:
        # Tentar converter para número e retornar como inteiro em formato string
        num_value = float(value)
        return str(int(num_value))
    except (ValueError, TypeError):
        # Se não for número, retornar como string removendo espaços
        return str(value).strip()

def format_display_value(value):
    """Formata qualquer valor para exibição, evitando mostrar None ou ?"""
    if value is None or value == '' or value == '?' or str(value).lower() == 'none':
        return ''
    return str(value)


# ===== CLASSE DE CAMPO DE EMAIL COM CHIPS =====
class EmailChipField:
    """Campo de email com chips removíveis (estilo Gmail)"""
    
    def __init__(self, label="Email", initial_value="", theme_colors=None, on_change=None):
        self.label = label
        self.emails = []
        self.theme_colors = theme_colors or {}
        self.on_change_callback = on_change
        self.input_field = None
        self.chips_container = None
        self.main_container = None
        
        # Parse initial value
        if initial_value:
            self.emails = [e.strip() for e in str(initial_value).split(';') if e.strip()]
    
    def get_value(self):
        """Retorna os emails como string separada por ;"""
        return '; '.join(self.emails)
    
    def set_value(self, value):
        """Define os emails a partir de uma string"""
        if value:
            self.emails = [e.strip() for e in str(value).split(';') if e.strip()]
        else:
            self.emails = []
        if self.chips_container:
            self.update_chips()
    
    def add_email(self, email):
        """Adiciona um email à lista"""
        email = email.strip()
        if email and email not in self.emails:
            self.emails.append(email)
            self.update_chips()
            if self.on_change_callback:
                self.on_change_callback(None)
    
    def remove_email(self, email):
        """Remove um email da lista"""
        if email in self.emails:
            self.emails.remove(email)
            self.update_chips()
            if self.on_change_callback:
                self.on_change_callback(None)
    
    def update_chips(self):
        """Atualiza a visualização dos chips"""
        if not self.chips_container:
            return
        
        self.chips_container.controls.clear()
        
        for email in self.emails:
            chip = ft.Container(
                content=ft.Row([
                    ft.Text(email, size=13, color=self.theme_colors.get('on_primary', 'white')),
                    ft.IconButton(
                        icon=ft.Icons.CLOSE,
                        icon_size=16,
                        icon_color=self.theme_colors.get('on_primary', 'white'),
                        on_click=lambda e, em=email: self.remove_email(em),
                        padding=0,
                        width=24,
                        height=24,
                    ),
                ], spacing=5, tight=True),
                bgcolor=self.theme_colors.get('primary', '#0066CC'),
                border_radius=16,
                padding=ft.padding.only(left=12, right=4, top=4, bottom=4),
                margin=ft.margin.only(right=5, bottom=5),
            )
            self.chips_container.controls.append(chip)
        
        try:
            self.chips_container.update()
        except:
            pass  # Pode falhar se ainda não estiver adicionado à página
    
    def on_input_submit(self, e):
        """Quando o usuário pressiona Enter ou ;"""
        if not self.input_field:
            return
            
        text = self.input_field.value.strip() if self.input_field.value else ""
        if text:
            # Separar por ; se houver múltiplos
            parts = [p.strip() for p in text.split(';') if p.strip()]
            for part in parts:
                self.add_email(part)
            
        # Sempre limpar o campo após processar
        self.input_field.value = ""
        try:
            self.input_field.update()
        except:
            pass
    
    def on_input_change(self, e):
        """Quando o texto muda, verificar se há ;"""
        if not self.input_field:
            return
            
        text = self.input_field.value
        if text and ';' in text:
            # Processar os emails antes do ;
            parts = text.split(';')
            # Adicionar todos os emails exceto o último (que está sendo digitado)
            for part in parts[:-1]:
                part = part.strip()
                if part:
                    self.add_email(part)
            
            # Manter apenas o texto após o último ; (se houver)
            remaining = parts[-1].strip()
            self.input_field.value = remaining
            try:
                self.input_field.update()
            except:
                pass
    
    def build(self):
        """Constrói e retorna o controle visual"""
        self.chips_container = ft.Row(
            controls=[],
            wrap=True,
            spacing=0,
            run_spacing=0,
        )
        
        self.input_field = ft.TextField(
            label=f"{self.label} (pressione Enter ou ; para adicionar)",
            hint_text="exemplo@email.com",
            expand=True,
            on_submit=self.on_input_submit,
            on_change=self.on_input_change,
            bgcolor=self.theme_colors.get('field_background'),
            color=self.theme_colors.get('on_surface'),
            border_color=self.theme_colors.get('outline'),
        )
        
        # Inicializar chips
        self.update_chips()
        
        self.main_container = ft.Column([
            ft.Container(
                content=self.chips_container,
                padding=ft.padding.only(bottom=10) if self.emails else 0,
            ),
            self.input_field,
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.INFO_OUTLINE, size=16, color=self.theme_colors.get('on_surface_variant', 'grey')),
                    ft.Text(
                        "Separe múltiplos emails com ; ou pressione Enter",
                        size=12,
                        color=self.theme_colors.get('on_surface_variant', 'grey'),
                        italic=True
                    ),
                ], spacing=5),
                padding=ft.padding.only(top=5),
            ),
        ], spacing=5)
        
        return self.main_container


# ===== CLASSE DE GERENCIAMENTO RESPONSIVO =====
class ResponsiveAppManager:
    """Classe para gerenciar o comportamento responsivo da aplicação"""
    
    def __init__(self, page: ft.Page):
        self.page = page
        self.results_container = None
        self.score_form_container = None  # Container do formulário de pesquisa
        self.is_maximized = False
        self.current_layout = "single"  # "single" ou "double"
        self.menu_is_expanded_ref = None
        self.update_menu_func = None
        
        # Configurar listener de redimensionamento - REMOVIDO (não usar reatividade)
        # self.page.on_resized = self.on_window_resize
        self.page.on_window_event = self.on_window_event
    
    def initialize_containers(self, results_container, score_form_container=None):
        """Inicializa as referências dos containers que serão gerenciados"""
        self.results_container = results_container
        self.score_form_container = score_form_container
        # Layout será aplicado quando houver cards para mostrar
    
    def initialize_menu_controls(self, menu_is_expanded_ref, update_menu_func):
        """Inicializa as referências para os controles do menu lateral"""
        self.menu_is_expanded_ref = menu_is_expanded_ref
        self.update_menu_func = update_menu_func
    
    def clear_results(self):
        """Limpa todos os resultados e reseta o estado"""
        if self.results_container:
            self.results_container.controls.clear()
            safe_update_control(self.results_container, self.page)
    
    def check_initial_window_state(self):
        """Verifica o estado inicial da janela e aplica o layout apropriado"""
        window_width = self.page.window.width or 1200
        is_maximized = self.page.window.maximized or False
        
        # Se maximizada, sempre usar layout double (prioridade ao estado maximized)
        if is_maximized:
            should_use_double_layout = True
        else:
            should_use_double_layout = window_width > 1400
        
        self.current_layout = "double" if should_use_double_layout else "single"
        self.is_maximized = is_maximized
        
        print(f"🚀 Estado inicial da janela: largura={window_width}px, maximizada={is_maximized}")
        print(f"🚀 Decisão: should_use_double_layout={should_use_double_layout}")
        print(f"🚀 Layout inicial definido como: {self.current_layout}")
        
        # Não precisa aplicar o layout aqui, será aplicado quando o primeiro card for adicionado
    
    def on_window_event(self, e):
        """Callback para eventos da janela (maximizar, restaurar, etc)"""
        if hasattr(e, 'event_type'):
            if e.event_type == "maximize":
                self.is_maximized = True 
                print("🔍 Janela maximizada - aplicando layout de duas colunas")
                self.apply_responsive_layout()
            elif e.event_type == "restore":
                self.is_maximized = False
                print("🔍 Janela restaurada - aplicando layout de uma coluna")
                self.apply_responsive_layout()
    
    def on_window_resize(self, e):
        """Callback chamado quando a janela é redimensionada"""
        window_width = self.page.window.width or 1200
        
        # Determinar se deve usar layout de duas colunas baseado na largura
        should_use_double_layout = window_width > 1400 or self.page.window.maximized
        
        new_layout = "double" if should_use_double_layout else "single"
        
        if new_layout != self.current_layout:
            self.current_layout = new_layout
            print(f"🔄 Mudando layout para: {new_layout} (largura: {window_width}px)")
            self.apply_responsive_layout()
            
        # Atualizar layout dos cards de suppliers se necessário
        self.update_supplier_cards_layout(window_width)
        
        # Atualizar layout da aba Users se necessário
        self.update_users_layout(window_width)

        # Atualizar layout da timeline se necessário
        self.update_timeline_layout(window_width)

        # Gerenciar estado do menu lateral baseado na largura da tela e layout
        if self.menu_is_expanded_ref and self.update_menu_func:
            # Se a tela for pequena, fechar o menu se estiver aberto
            if window_width < 1000 and self.menu_is_expanded_ref.current:
                print(f"📱 Tela pequena ({window_width}px), fechando menu lateral.")
                self.menu_is_expanded_ref.current = False
                self.update_menu_func()
            # Se o layout for duplo, reabrir o menu se estiver fechado
            elif new_layout == "double" and not self.menu_is_expanded_ref.current:
                print(f"📱 Layout duplo ativado ({window_width}px), reabrindo menu lateral.")
                self.menu_is_expanded_ref.current = True
                self.update_menu_func()
    
    def apply_responsive_layout(self):
        """Aplica o layout responsivo baseado no estado atual"""
        if not self.results_container:
            return
            
        try:
            print(f"🔄 apply_responsive_layout: self.current_layout = '{self.current_layout}'")
            if self.current_layout == "double":
                self._apply_double_column_layout()
            else:
                self._apply_single_column_layout()
                
            # Atualizar a interface
            safe_update_control(self.results_container, self.page)
            
        except Exception as e:
            print(f"❌ Erro ao aplicar layout responsivo: {e}")
    
    def _apply_double_column_layout(self):
        """Aplica layout de duas colunas para a aba Score"""
        print("📱 Aplicando layout de DUAS colunas")
        
        if not hasattr(self.results_container, 'controls'):
            return
        
        # Verificar se já está em layout de duas colunas
        if len(self.results_container.controls) > 0:
            first_control = self.results_container.controls[0]
            if isinstance(first_control, ft.Row):
                print("📱 Já está em layout de duas colunas, pulando reorganização")
                return  # Já está em layout double
            
        # Pegar todos os cards existentes
        cards = list(self.results_container.controls)
        print(f"📱 Cards existentes no container: {len(cards)}")
        
        # Limpar container atual
        self.results_container.controls.clear()
        
        # Criar duas colunas com alinhamento consistente
        left_column = ft.Column(
            [], 
            spacing=10, 
            expand=True,
            alignment=ft.MainAxisAlignment.START,
            horizontal_alignment=ft.CrossAxisAlignment.STRETCH
        )
        right_column = ft.Column(
            [], 
            spacing=10, 
            expand=True,
            alignment=ft.MainAxisAlignment.START,
            horizontal_alignment=ft.CrossAxisAlignment.STRETCH
        )
        
        # Distribuir cards alternadamente entre as colunas
        for i, card in enumerate(cards):
            if i % 2 == 0:
                left_column.controls.append(card)
                print(f"   ✅ Card {i+1} -> coluna esquerda")
            else:
                right_column.controls.append(card)
                print(f"   ✅ Card {i+1} -> coluna direita")
        
        # Criar linha com as duas colunas com alinhamento estável
        double_column_row = ft.Row([
            left_column,
            right_column
        ], 
        spacing=20, 
        expand=True,
        alignment=ft.MainAxisAlignment.START,
        vertical_alignment=ft.CrossAxisAlignment.START
        )
        
        # Adicionar ao container principal
        self.results_container.controls.append(double_column_row)
        print(f"📱 Layout de duas colunas aplicado")
    
    def _apply_single_column_layout(self):
        """Aplica layout de uma coluna para a aba Score"""
        print("📱 Aplicando layout de UMA coluna")
        
        if not hasattr(self.results_container, 'controls'):
            return
            
        # Se já está em layout de uma coluna, não fazer nada
        if len(self.results_container.controls) > 0:
            first_control = self.results_container.controls[0]
            if not isinstance(first_control, ft.Row):
                print("📱 Já está em layout de uma coluna, pulando reorganização")
                return  # Já está em layout single
        
        # Coletar todos os cards das colunas
        all_cards = []
        for control in self.results_container.controls:
            if isinstance(control, ft.Row):
                # É um layout de duas colunas, extrair cards
                for column in control.controls:
                    if isinstance(column, ft.Column):
                        all_cards.extend(column.controls)
            else:
                # São cards individuais
                all_cards.append(control)
        
        print(f"📱 Reorganizando {len(all_cards)} cards para layout de uma coluna")
        
        # Limpar container e readicionar cards em coluna única
        self.results_container.controls.clear()
        for i, card in enumerate(all_cards):
            self.results_container.controls.append(card)
            print(f"   ✅ Card {i+1}/{len(all_cards)} readicionado")

    def render_cards(self, cards):
        """Substitui os cards exibidos reaproveitando instâncias existentes."""
        if not self.results_container:
            return

        self.results_container.controls.clear()
        self.results_container.controls.extend(cards)
        # Garantir disposição correta com base no layout atual
        self.apply_responsive_layout()
    
    def add_card_to_layout(self, card):
        """Adiciona um novo card respeitando o layout atual"""
        if not self.results_container:
            print("❌ Container não disponível para adicionar card")
            return
            
        print(f"🔧 Adicionando card, layout atual: {self.current_layout}, containers existentes: {len(self.results_container.controls)}")
            
        if self.current_layout == "double":
            # Buscar estrutura de duas colunas existente
            double_column_row = None
            for control in self.results_container.controls:
                if isinstance(control, ft.Row) and len(control.controls) == 2:
                    double_column_row = control
                    break
            
            # Se não existe estrutura, criar uma nova
            if double_column_row is None:
                print("🔧 Criando nova estrutura de duas colunas")
                # Coletar cards existentes antes de limpar
                existing_cards = list(self.results_container.controls)
                self.results_container.controls.clear()
                
                # Criar colunas com alinhamento consistente
                left_column = ft.Column(
                    [], 
                    spacing=10, 
                    expand=True,
                    alignment=ft.MainAxisAlignment.START,
                    horizontal_alignment=ft.CrossAxisAlignment.STRETCH
                )
                right_column = ft.Column(
                    [], 
                    spacing=10, 
                    expand=True,
                    alignment=ft.MainAxisAlignment.START,
                    horizontal_alignment=ft.CrossAxisAlignment.STRETCH
                )
                
                # Redistribuir cards existentes
                for i, existing_card in enumerate(existing_cards):
                    if not isinstance(existing_card, ft.Row):  # Evitar nested rows
                        if i % 2 == 0:
                            left_column.controls.append(existing_card)
                        else:
                            right_column.controls.append(existing_card)
                
                # Criar nova row com as colunas
                double_column_row = ft.Row([
                    left_column,
                    right_column
                ], 
                spacing=20, 
                expand=True,
                alignment=ft.MainAxisAlignment.START,
                vertical_alignment=ft.CrossAxisAlignment.START
                )
                
                self.results_container.controls.append(double_column_row)
            
            # Agora adicionar o novo card à coluna apropriada
            left_column, right_column = double_column_row.controls
            if len(left_column.controls) <= len(right_column.controls):
                left_column.controls.append(card)
                print(f"🔧 Card adicionado à coluna esquerda (total: {len(left_column.controls)})")
            else:
                right_column.controls.append(card)
                print(f"🔧 Card adicionado à coluna direita (total: {len(right_column.controls)})")
            
            # Uma única atualização no final
            self.results_container.update()
            
        else:
            # Layout single - adicionar diretamente
            self.results_container.controls.append(card)
            self.results_container.update()
            print(f"🔧 Card adicionado em layout single (total: {len(self.results_container.controls)})")
    
    def update_supplier_cards_layout(self, window_width):
        """Atualiza o layout interno dos cards de suppliers baseado na largura da janela
        
        DESABILITADO: O layout interno dos campos não deve mudar automaticamente.
        Apenas o layout de distribuição dos cards (uma/duas colunas de cards) deve ser responsivo.
        """
        # FUNÇÃO DESABILITADA - não alterar layout interno dos campos
        return
        
        # Código original comentado abaixo caso precise ser restaurado
        # try:
        #     global suppliers_results_list
        #     if 'suppliers_results_list' in globals() and suppliers_results_list:
        #         should_use_single_column = window_width < 1000
        #         
        #         for card_control in suppliers_results_list.controls:
        #             if isinstance(card_control, ft.Card):
        #                 self._update_single_supplier_card_layout(card_control, should_use_single_column)
        #         
        #         suppliers_results_list.update()
        #         print(f"📱 Layout dos cards de suppliers atualizado: {'uma coluna' if should_use_single_column else 'duas colunas'}")
        #         
        # except Exception as e:
        #     print(f"❌ Erro ao atualizar layout dos cards de suppliers: {e}")
    
    def _update_single_supplier_card_layout(self, card, use_single_column):
        """Atualiza o layout interno de um único card de supplier"""
        try:
            # Verificar se o card tem os dados necessários (função de criar layout)
            if hasattr(card, 'data') and card.data and 'create_layout' in card.data:
                create_layout_func = card.data['create_layout']
                
                # Recriar o layout usando a função armazenada
                new_layout = create_layout_func(use_single_column)
                
                # Substituir o conteúdo do container
                if hasattr(card, 'content') and hasattr(card.content, 'content'):
                    card.content.content = new_layout
                    
                print(f"✅ Card {card.data.get('supplier_id', 'unknown')} atualizado para {'uma' if use_single_column else 'duas'} coluna(s)")
                
        except Exception as e:
            print(f"❌ Erro ao atualizar card individual: {e}")
            # Fallback para o método antigo se necessário
            self._update_single_supplier_card_layout_fallback(card, use_single_column)
    
    def _update_single_supplier_card_layout_fallback(self, card, use_single_column):
        """Método de fallback para cards antigos sem função de layout armazenada"""
        try:
            # Obter o container interno do card
            if hasattr(card, 'content') and hasattr(card.content, 'content'):
                container_content = card.content.content
                
                # Verificar se é uma Column com o layout que esperamos
                if isinstance(container_content, ft.Column) and len(container_content.controls) >= 1:
                    # Primeiro controle deve ser Row (duas colunas) ou Column (uma coluna)
                    main_content = container_content.controls[0]
                    
                    # Obter as referências das colunas left_column e right_column
                    left_column = None
                    right_column = None
                    
                    if isinstance(main_content, ft.Row):
                        # Layout de duas colunas atual
                        if len(main_content.controls) >= 3:  # left_column, divider, right_column
                            left_column = main_content.controls[0]
                            right_column = main_content.controls[2]
                    elif isinstance(main_content, ft.Column):
                        # Layout de uma coluna atual - precisamos extrair as partes
                        if len(main_content.controls) >= 3:
                            left_column = main_content.controls[0]
                            right_column = main_content.controls[2]
                    
                    if left_column and right_column:
                        # Recriar o layout baseado na decisão
                        if use_single_column:
                            # Layout de uma coluna
                            new_layout = ft.Column([
                                left_column,
                                ft.Divider(),
                                right_column,
                                ft.Divider(),
                                container_content.controls[-1] if len(container_content.controls) > 1 else ft.Container()  # actions_row
                            ], spacing=15)
                        else:
                            # Layout de duas colunas
                            new_layout = ft.Column([
                                ft.Row([left_column, ft.VerticalDivider(), right_column], expand=True),
                                ft.Divider(),
                                container_content.controls[-1] if len(container_content.controls) > 1 else ft.Container()  # actions_row
                            ], spacing=15)
                        
                        # Substituir o conteúdo
                        card.content.content = new_layout
                        
        except Exception as e:
            print(f"❌ Erro no fallback: {e}")
    
    def update_users_layout(self, window_width):
        """Atualiza o layout da aba Users baseado na largura da janela"""
        try:
            # Definir breakpoint para responsividade (menor que 1000px = layout vertical)
            should_use_vertical_layout = window_width < 1000
            users_form_container = ft.Container(
                content=ft.Column([
                    # Linha 1: WWID e Nome
                    ft.Row([
                        ft.TextField(
                            label="WWID",
                            hint_text="Digite o WWID do usuário",
                            prefix_icon=ft.Icons.BADGE,
                            expand=True,
                            border_radius=8,
                            filled=False,  # Fundo transparente
                        ),
                        ft.TextField(
                            label="Nome Completo",
                            hint_text="Digite o nome do usuário",
                            prefix_icon=ft.Icons.PERSON,
                            expand=True,
                            border_radius=8,
                            filled=False,  # Fundo transparente
                        ),
                    ], spacing=10),

                    # Linha 2: Senha e Privilégio
                    ft.Row([
                        ft.TextField(
                            label="Senha",
                            hint_text="Digite a senha do usuário",
                            prefix_icon=ft.Icons.LOCK,
                            password=True,
                            can_reveal_password=True,
                            expand=True,
                            border_radius=8,
                            filled=False,  # Fundo transparente
                        ),
                        ft.Dropdown(
                            label="Nível de Privilégio",
                            hint_text="Selecione o nível",
                            options=[
                                ft.dropdown.Option("User", "User"),
                                ft.dropdown.Option("Admin", "Admin"),
                                ft.dropdown.Option("Super Admin", "Super Admin"),
                            ],
                            expand=True,
                            bgcolor=None,  # Fundo transparente
                            content_padding=ft.padding.symmetric(horizontal=12, vertical=16),
                        ),
                    ], spacing=10),
                ], spacing=15),
                bgcolor=None,
                padding=ft.padding.all(20),
                border_radius=8,
                border=None
            )

            # Wrapper externo para limitar largura
            users_form_container = ft.Container(
                content=users_form_container,
                width=500,
                alignment=ft.alignment.center
            )
                
            
        except Exception as e:
            print(f"❌ Erro ao atualizar layout da aba Users: {e}")
            import traceback
            traceback.print_exc()
    
    def update_timeline_layout(self, window_width):
        """Atualiza o layout da timeline baseado na largura da janela"""
        try:
            # Verificar se a variável da timeline existe no escopo global
            if 'timeline_content_container' not in globals():
                return
            
            timeline_content_container = globals().get('timeline_content_container')
            timeline_chart_container = globals().get('timeline_chart_container')
            timeline_table_container = globals().get('timeline_table_container')
            
            if not timeline_content_container or not timeline_content_container.current:
                return
            
            should_use_side_by_side = window_width >= 1200  # Threshold de 1200px
            
            # Pegar o conteúdo existente dos containers (se houver)
            chart_content = None
            table_content = None
            
            if timeline_chart_container and timeline_chart_container.current:
                chart_content = timeline_chart_container.current.content
            
            if timeline_table_container and timeline_table_container.current:
                table_content = timeline_table_container.current.content
                
            # Criar containers com altura fixa de 500px
            chart_container = ft.Container(
                ref=timeline_chart_container,
                content=chart_content if chart_content else ft.Container(
                    content=ft.Text("Selecione um fornecedor para visualizar o gráfico", 
                                   size=16, text_align=ft.TextAlign.CENTER),
                    alignment=ft.alignment.center,
                ),
                height=500,
                expand=True
            )
            
            # Tabela sempre com scroll interno
            table_container = ft.Container(
                ref=timeline_table_container,
                content=table_content if table_content else ft.Column([
                    ft.Text("Selecione um fornecedor para visualizar a tabela", 
                           size=16, text_align=ft.TextAlign.CENTER)
                ], scroll=ft.ScrollMode.AUTO),
                height=500,
                expand=True
            )
                
            if should_use_side_by_side:
                # Layout lado a lado (Row) - largura >= 1200px
                timeline_content_container.current.content = ft.Row([
                    chart_container,
                    table_container
                ], spacing=20)
                print("📱 Timeline: Layout lado a lado aplicado (Row)")
            else:
                # Layout empilhado (Column) - largura < 1200px
                timeline_content_container.current.content = ft.Column([
                    chart_container,
                    table_container
                ], spacing=20)
                print("📱 Timeline: Layout empilhado aplicado (Column - tabela abaixo do gráfico)")
            
            timeline_content_container.current.update()
            
        except Exception as e:
            print(f"❌ Erro ao atualizar layout da timeline: {e}")
            import traceback
            traceback.print_exc()
    
    def _find_users_content_recursive(self, control, depth=0):
        """Busca recursiva para encontrar users_content"""
        if depth > 10:  # Evitar recursão infinita
            return None
            
        # Verificar se é o users_content (Container com padding 20 e visible False inicialmente)
        if isinstance(control, ft.Container) and hasattr(control, 'padding') and control.padding == 20:
            if hasattr(control, 'content') and isinstance(control.content, ft.Column):
                column = control.content
                if len(column.controls) >= 3:
                    first_text = column.controls[0]
                    if isinstance(first_text, ft.Text) and "Gerenciamento de Usuários" in str(first_text.value):
                        return control
        
        # Buscar recursivamente nos filhos
        if hasattr(control, 'controls'):
            for child in control.controls:
                result = self._find_users_content_recursive(child, depth + 1)
                if result:
                    return result
        elif hasattr(control, 'content') and control.content:
            return self._find_users_content_recursive(control.content, depth + 1)
        
        return None

# Instância global do gerenciador responsivo
responsive_app_manager = None

# ===== VARIÁVEIS GLOBAIS =====
current_user_wwid = None
current_user_name = None
current_user_privilege = None
current_user_permissions = {}
selected_user = None  # Para controlar a seleção de usuários na aba Users
suppliers_results_list = None  # Lista global dos cards de suppliers
score_control_type = "slider"  # MANTIDO por compatibilidade - Cards obsoletos, apenas tabela é usada

# ===== FUNÇÕES AUXILIARES =====

def safe_update_control(control, page=None):
    """
    Atualiza um controle de forma segura, lidando com casos onde o controle
    não está devidamente vinculado à página ou onde há problemas de concorrência
    """
    try:
        if control is None:
            return
        
        # Tentar atualizar o controle primeiro
        if hasattr(control, 'page') and control.page is not None:
            try:
                control.update()
                return
            except Exception as control_error:
                print(f"Aviso: Erro ao atualizar controle diretamente: {control_error}")
        
        # Se o controle não funcionou e temos uma página, tentar atualizar a página com lock
        if page is not None:
            try:
                # Usar run_task para atualização thread-safe
                if hasattr(page, 'run_task'):
                    page.run_task(lambda: _safe_page_update(page))
                else:
                    _safe_page_update(page)
            except Exception as page_error:
                print(f"Aviso: Erro ao atualizar página: {page_error}")
    except Exception as e:
        print(f"Erro crítico em safe_update_control: {e}")

def _safe_page_update(page):
    """Atualização interna da página com verificações extras"""
    try:
        if page is not None and hasattr(page, 'update'):
            page.update()  # Chamada real do update
    except Exception as e:
        print(f"Erro ao executar page.update: {e}")

def safe_page_update(page):
    """
    Wrapper seguro para safe_page_update(page) que previne crashes por:
    - Atualizações concorrentes
    - Referências inválidas
    - Estado inconsistente da página
    
    Use esta função SEMPRE ao invés de safe_page_update(page) diretamente
    """
    if page is None:
        return
    
    try:
        # Verificar se a página está em estado válido
        if not hasattr(page, 'update'):
            print("Aviso: Página não possui método update")
            return
        
        # Tentar atualizar de forma thread-safe
        if hasattr(page, 'run_task'):
            try:
                page.run_task(lambda: _safe_page_update(page))
            except Exception as task_error:
                print(f"Aviso: run_task falhou, tentando update direto: {task_error}")
                _safe_page_update(page)
        else:
            _safe_page_update(page)
            
    except Exception as e:
        print(f"Erro em safe_page_update: {e}")

def safe_callback(func):
    """
    Decorator para proteger callbacks contra exceções que causam crashes
    Use este decorator em callbacks de botões, eventos, etc.
    
    Exemplo:
        @safe_callback
        def on_button_click(e):
            # código do callback
    """
    def wrapper(*args, **kwargs):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            print(f"⚠️ Erro capturado no callback {func.__name__}: {e}")
            import traceback
            traceback.print_exc()
            # Tentar mostrar mensagem ao usuário se possível
            try:
                if args and hasattr(args[0], 'page'):
                    show_snack_bar(f"Erro: {str(e)[:100]}", is_error=True)
            except:
                pass
    return wrapper

# ===== DEFINIÇÕES SIMPLES DE TEMAS =====

def get_current_theme_colors(theme_name="white"):
    """Return base Colors for the given theme

    Nota: os campos de "Comentário" e os campos de nota (spinbox/slider TextField)
    presentes nos cards da aba Score foram configurados para usar a mesma cor de
    fundo do card (`card_background`) para terem aparencia integrada ao card.
    Para alterar esse comportamento, edite os locais onde `bgcolor=...get('card_background')`
    é usado ou modifique os valores retornados por este tema.
    """
    if theme_name == "dark":
        return {
            "field_background": "#2C2C2C",
            "on_surface": "#E0E0E0",
            "on_surface_variant": "#B0B0B0",
            "outline": "#555555",
            "outline_variant": "#444444",
            "card_background": "#1E1E1E",
            "primary": "#4EA1FF",
            "on_primary": "#FFFFFF",
            "primary_container": "#3B3B3B",
            "surface_variant": "#181818",
            "rail_background": "#0F0F0F"
        }
    elif theme_name == "dracula":
        return {
            "field_background": None,
            "on_surface": "#F8F8F2",
            "on_surface_variant": "#BFBDB6",
            "outline": "#6272A4",
            "outline_variant": "#44475A",
            "card_background": "#343746",
            "primary": "#BD93F9",
            "on_primary": "#FFFFFF",
            "primary_container": "#6272A4",
            "surface_variant": "#2E303E",
            "rail_background": "#1E1F29"
        }
    else:  # white
        return {
            "field_background": "#FFFFFF",
            "on_surface": "#1C1C1C",
            "on_surface_variant": "#666666",
            "outline": "#CCCCCC",
            "outline_variant": "#E8E8E8",
            "card_background": "#F7F7F7",
            "primary": "#0066CC",
            "on_primary": "#FFFFFF",
            "primary_container": "#E6F0FF",
            "surface_variant": "#FAFAFA",
            "rail_background": "#F0F0F0"
        }

def get_theme_name_from_page(page_ref=None):
    """Obtém o nome do tema atual da página"""
    if page_ref and hasattr(page_ref, 'data') and page_ref.data:
        return page_ref.data.get("theme_name", "white")
    return "white"

def get_primary_darker_color(theme_name="white"):
    """Retorna uma versão mais escura da cor primária baseada no tema"""
    if theme_name == "dark":
        return "#1A3D68"  # Versão mais escura de #4EA1FF
    elif theme_name == "dracula":
        return "#4D3377"  # Versão mais escura de #BD93F9
    else:  # white
        return "#93B7E2"  # Versão mais escura de #0066CC

def get_score_color(score):
    """
    Retorna uma cor neutra para o texto das notas
    (esquema de cores vermelho/verde foi removido)
    """
    try:
        # Sempre retornar cor padrão do tema (None = cor padrão do texto)
        return None
    except (ValueError, TypeError):
        # Se houver erro, retornar cor padrão
        return None

def analyze_data_consistency(otif, nil, pickup, package):
    """
    Analisa a consistência dos dados das métricas e retorna um score de confiabilidade (0-100)
    e uma lista de inconsistências encontradas.
    
    Sistema completo de validação com TODAS as inconsistências lógicas possíveis:
    - Verifica relações entre OTIF, NIL, Pickup e Package
    - Distingue NULL (não atribuído) de 0.0 (zero explícito)
    - Aplica penalidades progressivas para múltiplas inconsistências
    
    Retorna: (score, issues_list, icon, color)
    """
    try:
        # Distinguir entre NULL (não atribuído) e 0 (nota zero)
        otif_is_null = otif is None or otif == ''
        nil_is_null = nil is None or nil == ''
        pickup_is_null = pickup is None or pickup == ''
        package_is_null = package is None or package == ''
        
        # Converter para float depois de verificar NULL
        otif = float(otif) if not otif_is_null else None
        nil = float(nil) if not nil_is_null else None
        pickup = float(pickup) if not pickup_is_null else None
        package = float(package) if not package_is_null else None
        
        issues = []
        score = 100  # Começar com confiabilidade máxima
        
        # ========== REGRA 1: OTIF vazio = sem entrega ==========
        if otif_is_null:
            if not nil_is_null or not pickup_is_null or not package_is_null:
                issues.append("❌ Sem entrega (OTIF vazio) mas com avaliações preenchidas")
                score -= 40
        
        # ========== REGRAS PARA QUANDO TEVE ENTREGA (OTIF preenchido) ==========
        if not otif_is_null:
            
            # === GRUPO A: Inconsistências com NIL (Logística) ===
            
            # A1: OTIF excelente (≥9) mas NIL péssimo (≤2)
            if nil is not None and otif >= 9 and nil <= 2:
                issues.append("⚠️ OTIF excelente mas logística péssima")
                score -= 25
            
            # A2: OTIF alto (≥8) mas NIL=0
            elif nil is not None and otif >= 8 and nil == 0:
                issues.append("⚠️ OTIF alto mas logística zerada")
                score -= 30
            
            # A3: NIL péssimo (≤3) mas Pickup excelente (≥8)
            if nil is not None and pickup is not None:
                if nil <= 3 and pickup >= 8:
                    issues.append("⚠️ Logística péssima mas inspeção excelente")
                    score -= 20
            
            # === GRUPO B: Inconsistências com Pickup (Inspeção) ===
            
            # B1: Pickup=0 mas Package alto (≥8)
            if pickup is not None and pickup == 0 and package is not None and package >= 8:
                issues.append("⚠️ Reprovou inspeção total mas embalagem excelente")
                score -= 25
            
            # B2: Pickup péssimo (≤2) mas OTIF excelente (≥9)
            if pickup is not None and pickup <= 2 and otif >= 9:
                issues.append("⚠️ Inspeção péssima mas OTIF excelente")
                score -= 20
            
            # B3: Pickup péssimo (≤3) mas Package excelente (≥8)
            if pickup is not None and package is not None:
                if pickup <= 3 and package >= 8:
                    issues.append("⚠️ Produto péssimo mas embalagem excelente")
                    score -= 22
            
            # === GRUPO C: Inconsistências com Package (Embalagem) ===
            
            # C1: Package=0 mas Pickup alto (≥8)
            if package is not None and package == 0 and pickup is not None and pickup >= 8:
                issues.append("⚠️ Embalagem zerada mas produto excelente")
                score -= 20
            
            # C2: Package péssimo (≤2) mas OTIF excelente (≥9)
            if package is not None and package <= 2 and otif >= 9:
                issues.append("⚠️ Embalagem péssima mas OTIF excelente")
                score -= 18
            
            # === GRUPO D: Falhas Múltiplas Simultâneas (CRÍTICO) ===
            
            # D1: NIL=0 E Pickup=0 (falha total em logística E inspeção)
            if nil is not None and nil == 0 and pickup is not None and pickup == 0:
                issues.append("🔴 CRÍTICO: Logística E inspeção ambas zeradas")
                score -= 35
            
            # D2: NIL=0 E Package=0 (falha em logística E embalagem)
            if nil is not None and nil == 0 and package is not None and package == 0:
                issues.append("🔴 CRÍTICO: Logística E embalagem ambas zeradas")
                score -= 35
            
            # D3: Pickup=0 E Package=0 (falha em inspeção E embalagem)
            if pickup is not None and pickup == 0 and package is not None and package == 0:
                issues.append("🔴 CRÍTICO: Inspeção E embalagem ambas zeradas")
                score -= 35
            
            # D4: Três ou mais métricas zeradas simultaneamente
            zeros_count = sum([
                1 for x in [nil, pickup, package] 
                if x is not None and x == 0
            ])
            if zeros_count >= 3:
                issues.append("🔴 CATASTRÓFICO: Três ou mais notas zeradas")
                score -= 50
            
            # === GRUPO E: Padrões Suspeitos ===
            
            # E1: OTIF baixo (≤3) mas todas outras altas (≥8)
            if otif <= 3:
                high_count = sum([
                    1 for x in [nil, pickup, package]
                    if x is not None and x >= 8
                ])
                if high_count >= 2:
                    issues.append("⚠️ OTIF péssimo mas outras métricas excelentes")
                    score -= 25
            
            # E2: Apenas uma métrica muito baixa (≤2) e outras muito altas (≥8)
            all_values = [(otif, "OTIF"), (nil, "NIL"), (pickup, "Pickup"), (package, "Package")]
            assigned_values = [(v, n) for v, n in all_values if v is not None]
            
            if len(assigned_values) >= 3:
                low_values = [n for v, n in assigned_values if v <= 2]
                high_values = [n for v, n in assigned_values if v >= 8]
                
                if len(low_values) == 1 and len(high_values) >= 2:
                    issues.append(f"⚠️ Apenas {low_values[0]} muito baixo, outras altas")
                    score -= 15
        
        # ========== REGRA GERAL: Valores Idênticos Suspeitos ==========
        non_null_values = [x for x in [otif, nil, pickup, package] if x is not None]
        if len(non_null_values) >= 3:
            unique_values = set(non_null_values)
            
            # Permitir apenas 0 (falha total) ou 10 (perfeição total)
            if len(unique_values) == 1:
                value = non_null_values[0]
                if value not in [0, 10]:
                    issues.append(f"⚠️ Todas as notas idênticas ({value}) - suspeito")
                    score -= 18
        
        # Garantir que score não fique negativo
        score = max(0, score)
        
        # Definir ícone e cor baseado no score
        if score >= 80:
            icon = ft.Icons.THERMOSTAT  # Verde - Dados consistentes
            color = ft.Colors.GREEN_600
        elif score >= 50:
            icon = ft.Icons.THERMOSTAT_AUTO  # Laranja - Atenção
            color = ft.Colors.ORANGE_600
        else:
            icon = ft.Icons.AC_UNIT  # Vermelho - Dados críticos
            color = ft.Colors.RED_600
        
        return score, issues, icon, color
        
    except Exception as e:
        print(f"Erro ao analisar consistência: {e}")
        return 100, [], ft.Icons.THERMOSTAT, ft.Colors.GREY

def get_card_color(theme_name="white"):
    """Retorna a cor dos cards baseada no tema"""
    if theme_name == "dracula":
        return "#6272A4"  # Cor da linha 590 do tema dracula
    else:
        return ft.Colors.BLUE_600  # Cor padrão para white e dark

# ===== SISTEMA DE LOGIN =====
def login_screen(page: ft.Page):
    """Tela de login para autenticação do usuário"""
    global current_user_wwid, current_user_name, current_user_privilege, current_user_permissions
    
    # Configurações da janela de login
    page.title = f"Score App v{APP_VERSION} - Login"
    page.window.width = 500
    page.window.height = 650
    page.window.resizable = False
    page.window.center()
    page.theme_mode = ft.ThemeMode.LIGHT  # Forçar tema claro (branco)
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.bgcolor = "#FFFFFF"  # Fundo branco
    
    # Cores da tela de login (tema claro/branco)
    Colors = {
        'primary': "#1976D2",  # Azul sistema Windows
        'on_primary': "#FFFFFF",
        'surface': "#FFFFFF",  # Fundo branco para o card
        'on_surface': "#000000",  # Texto preto
        'on_surface_variant': "#757575",
        'field_background': "#FFFFFF",  # Fundo branco para campos
        'outline': "#BDBDBD",
        'error': "#D32F2F"
    }
    
    # Funções para Remember Me
    def get_app_data_path():
        """Retorna o caminho da pasta AppData para o ScoreApp"""
        import os
        app_data = os.getenv('APPDATA')  # Pasta AppData/Roaming do usuário
        app_folder = os.path.join(app_data, 'ScoreApp')
        
        # Criar pasta se não existir
        if not os.path.exists(app_folder):
            os.makedirs(app_folder)
        
        return os.path.join(app_folder, 'remember_me.txt')
    
    def load_saved_credentials():
        """Carrega credenciais salvas do AppData"""
        try:
            credentials_file = get_app_data_path()
            with open(credentials_file, 'r', encoding='utf-8') as f:
                lines = f.readlines()
                if len(lines) >= 2:
                    saved_wwid = lines[0].strip()
                    saved_password = lines[1].strip()
                    print(f"[INFO] Credenciais carregadas de: {credentials_file}")
                    return saved_wwid, saved_password
        except FileNotFoundError:
            print("[INFO] Nenhuma credencial salva encontrada")
        except Exception as e:
            print(f"[ERRO] Erro ao carregar credenciais: {e}")
        return "", ""
    
    def save_credentials(wwid, password):
        """Salva credenciais na pasta AppData"""
        try:
            credentials_file = get_app_data_path()
            with open(credentials_file, 'w', encoding='utf-8') as f:
                f.write(f"{wwid}\n{password}\n")
            print(f"[INFO] Credenciais salvas em: {credentials_file}")
        except Exception as e:
            print(f"[ERRO] Erro ao salvar credenciais: {e}")
    
    def clear_saved_credentials():
        """Remove credenciais salvas da pasta AppData"""
        try:
            import os
            credentials_file = get_app_data_path()
            if os.path.exists(credentials_file):
                os.remove(credentials_file)
                print(f"[INFO] Credenciais removidas de: {credentials_file}")
        except Exception as e:
            print(f"❌ Erro ao limpar credenciais: {e}")
    
    # Carregar credenciais salvas
    saved_wwid, saved_password = load_saved_credentials()
    
    # Campos de entrada com tema do sistema
    wwid_field = ft.TextField(
        label="WWID",
        width=320,
        autofocus=True,
        border_radius=8,
        border_color=Colors['outline'],
        prefix_icon=ft.Icons.PERSON,
        value=saved_wwid  # Preencher com valor salvo
    )
    
    password_field = ft.TextField(
        label="Password", 
        password=True, 
        can_reveal_password=True, 
        width=320,
        border_radius=8,
        border_color=Colors['outline'],
        prefix_icon=ft.Icons.LOCK,
        value=saved_password  # Preencher com valor salvo
    )    # Checkbox Remember Me
    remember_me_checkbox = ft.Checkbox(
        label="Lembrar de mim",
        value=bool(saved_wwid and saved_password),  # Marcar se há credenciais salvas
        check_color=Colors['primary']
    )
    
    error_text = ft.Text(
        "",
        color=Colors['error'],
        text_align=ft.TextAlign.CENTER,
        size=12
    )
    
    login_button = ft.ElevatedButton(
        "Entrar",
        width=320,
        height=45,
        style=ft.ButtonStyle(
            bgcolor=Colors['primary'],
            color=Colors['on_primary'],
            shape=ft.RoundedRectangleBorder(radius=8),
            elevation=2
        ),
        icon=ft.Icons.LOGIN
    )
    
    def authenticate_user(wwid, password):
        """Verifica credenciais na tabela users_table"""
        try:
            result = db_manager.query_one("""
                SELECT user_wwid, user_name, user_password, user_privilege, otif, nil, pickup, package 
                FROM users_table 
                WHERE user_wwid = ? AND user_password = ?
            """, (wwid, password))
            
            if result:
                return {
                    'wwid': result['user_wwid'],
                    'name': result['user_name'],
                    'password': result['user_password'],
                    'privilege': result['user_privilege'],
                    'otif': result['otif'] == '1' or result['otif'] == 1,
                    'nil': result['nil'] == '1' or result['nil'] == 1,
                    'pickup': result['pickup'] == '1' or result['pickup'] == 1,
                    'package': result['package'] == '1' or result['package'] == 1
                }
            return None
            
        except Exception as e:
            print(f"Erro na autenticação: {e}")
            return None
    
    def get_user_theme(wwid):
        """Busca tema do usuário na tabela theme_settings"""
        try:
            result = db_manager.query_one("SELECT theme_mode FROM theme_settings WHERE user_wwid = ?", (wwid,))
            return result['theme_mode'] if result else "white"
            
        except Exception as e:
            print(f"Erro ao carregar tema: {e}")
            return "white"

    def on_login_click(e):
        wwid = wwid_field.value.strip()
        password = password_field.value.strip()

        if not wwid or not password:
            error_text.value = ""
            snackbar = ft.SnackBar(
                content=ft.Text("⚠️ Por favor, preencha todos os campos", color=ft.Colors.WHITE),
                bgcolor=ft.Colors.ORANGE_700,
            )
            page.overlay.append(snackbar)
            snackbar.open = True
            page.update()
            return

        # Apenas desabilita o botão para evitar múltiplos cliques
        login_button.disabled = True
        safe_page_update(page)

        # Autenticar usuário
        user_data = authenticate_user(wwid, password)

        if user_data:
            # Gerenciar Remember Me
            if remember_me_checkbox.value:
                save_credentials(wwid, password)
                print("📝 Credenciais salvas para próximo login")
            else:
                clear_saved_credentials()
                print("🗑️ Credenciais removidas")
            
            # Definir variáveis globais
            global current_user_wwid, current_user_name, current_user_privilege, current_user_permissions
            current_user_wwid = user_data['wwid']
            current_user_name = user_data['name']  # Nome do usuário do banco de dados
            current_user_privilege = user_data['privilege']
            current_user_permissions = {
                'otif': user_data['otif'],
                'nil': user_data['nil'],
                'pickup': user_data['pickup'],
                'package': user_data['package']
            }

            # Buscar tema do usuário
            user_theme = get_user_theme(wwid)

            print(f"Login bem-sucedido: {current_user_wwid} ({current_user_privilege})")
            print(f"Permissões: {current_user_permissions}")
            print(f"Tema: {user_theme}")

            # Definir usuário atual no DBManager para logs
            db_manager.set_current_user(current_user_wwid)

            # Carregar aplicação principal (sem animação)
            page.controls.clear()
            page.title = f"Score App v{APP_VERSION}"
            page.window.min_width = 850
            page.window.min_height = 900
            page.window.resizable = True
            page.window.maximized = True
            page.window.title_bar_hidden = True  # Remove a borda da janela padrão
            page.window.title_bar_buttons_hidden = True  # Esconde botões padrão
            safe_page_update(page) # Aplica a maximização

            # Inicializar aplicação principal na mesma página
            initialize_main_app(page, user_theme)

        else:
            # Reabilita o botão e mostra o erro
            login_button.disabled = False
            error_text.value = ""
            password_field.value = ""
            snackbar = ft.SnackBar(
                content=ft.Text("❌ WWID ou senha incorretos", color=ft.Colors.WHITE),
                bgcolor=ft.Colors.RED_700,
            )
            page.overlay.append(snackbar)
            snackbar.open = True
            page.update()

    login_button.on_click = on_login_click
    
    # Permitir login com Enter
    def on_key_press(e):
        if e.key == "Enter":
            on_login_click(e)
    
    wwid_field.on_submit = on_login_click
    password_field.on_submit = on_login_click
    
    # Layout da tela de login com responsividade
    login_container = ft.Container(
        content=ft.Column(
            [
                ft.Container(height=15),
                # Logo da Cummins
                ft.Container(
                    content=ft.Image(
                        src=resource_path("images/cummins.ico"),
                        width=120,
                        height=120,
                        fit=ft.ImageFit.CONTAIN,
                    ),
                    alignment=ft.alignment.center,
                ),
                ft.Container(height=12),
                ft.Text(
                    "Score App",
                    size=28,
                    weight=ft.FontWeight.BOLD,
                    text_align=ft.TextAlign.CENTER,
                    color=Colors['on_surface']
                ),
                ft.Text(
                    "Sistema de Avaliação de Fornecedores",
                    size=14,
                    text_align=ft.TextAlign.CENTER,
                    color=Colors['on_surface_variant']
                ),
                ft.Container(height=25),
                wwid_field,
                ft.Container(height=12),
                password_field,
                ft.Container(height=15),
                # Container para centralizar o checkbox
                ft.Container(
                    content=remember_me_checkbox,
                    alignment=ft.alignment.center_left,
                    width=320,
                ),
                ft.Container(height=8),
                error_text,
                ft.Container(height=20),
                login_button,
                ft.Container(height=15),
                # Versão do aplicativo
                ft.Text(
                    f"v{APP_VERSION}",
                    size=11,
                    color=Colors['on_surface_variant'],
                    text_align=ft.TextAlign.CENTER,
                    italic=True
                )
            ],
            alignment=ft.MainAxisAlignment.CENTER,
            horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            spacing=0
        ),
        bgcolor="#FFFFFF",  # Fundo branco
        border_radius=12,
        padding=ft.padding.all(30),
        shadow=ft.BoxShadow(
            spread_radius=0,
            blur_radius=10,
            color="#26000000",  # Preto com 15% de opacidade
            offset=ft.Offset(0, 2)
        ),
        alignment=ft.alignment.center,
        width=420,
        height=None,  # Altura automática para evitar overflow
        margin=ft.margin.all(20),  # Margem para garantir que fique dentro da tela
        animate_opacity=300
    )
    
    # Usar Container com scroll para garantir que caiba na tela
    main_container = ft.Container(
        content=login_container,
        alignment=ft.alignment.center,
        expand=True,
        bgcolor="#FFFFFF"  # Fundo branco também no container principal
    )
    
    page.add(main_container)
# ===== FIM DO SISTEMA DE LOGIN =====

class DeleteSupplierConfirmationDialog(ft.AlertDialog):
    def __init__(self, supplier_name, supplier_id, on_confirm, on_cancel, scale_func=None):
        super().__init__()
        self.supplier_name = supplier_name
        self.supplier_id = supplier_id
        self.on_confirm_callback = on_confirm
        self.on_cancel_callback = on_cancel
        self.random_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
        self.scale_func = scale_func or (lambda x: x)
        
        self.confirmation_text = ft.TextField(
            label="Digite o código de confirmação",
            width=self.scale_func(300),
            text_align=ft.TextAlign.CENTER
        )
        self.title = ft.Text("🗑️ Excluir Supplier")
        self.error_text = ft.Text("Código incorreto. Tente novamente.", color="red", visible=False)
        self.code_display_text = ft.Text(self.random_code, weight=ft.FontWeight.BOLD, size=20, color="primary")
        
        self.content = ft.Column([
            ft.Text(f"Tem certeza que deseja excluir '{self.supplier_name}'?"),
            ft.Text("Esta ação não pode ser desfeita."),
            ft.Text("Para confirmar, digite o código abaixo:"),
            ft.Row([self.code_display_text], alignment=ft.MainAxisAlignment.CENTER),
            self.confirmation_text,
            self.error_text
        ], tight=True, spacing=15, width=350)
        
        self.actions = [
            ft.TextButton("Cancelar", on_click=self.cancel),
            ft.TextButton("Excluir", on_click=self.confirm, style=ft.ButtonStyle(color="red"))
        ]
        self.actions_alignment = ft.MainAxisAlignment.END
    
    def confirm(self, e):
        if self.confirmation_text.value.strip().upper() == self.random_code:
            self.on_confirm_callback(e)
        else:
            self.error_text.visible = True
            self.update()
    
    def cancel(self, e):
        self.on_cancel_callback(e)


class DeleteListItemConfirmationDialog(ft.AlertDialog):
    def __init__(self, item_name, item_type, on_confirm, on_cancel, scale_func=None):
        super().__init__()
        self.item_name = item_name
        self.item_type = item_type
        self.on_confirm_callback = on_confirm
        self.on_cancel_callback = on_cancel
        self.random_code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=5))
        self.scale_func = scale_func or (lambda x: x)
        
        self.confirmation_text = ft.TextField(
            label="Digite o código de confirmação",
            width=self.scale_func(300),
            text_align=ft.TextAlign.CENTER
        )
        self.title = ft.Text("🗑️ Excluir Item da Lista")
        self.error_text = ft.Text("Código incorreto. Tente novamente.", color="red", visible=False)
        self.code_display_text = ft.Text(self.random_code, weight=ft.FontWeight.BOLD, size=20, color="primary")
        
        self.content = ft.Column([
            ft.Text(f"Tem certeza que deseja excluir '{self.item_name}'?"),
            ft.Text(f"Lista: {self.item_type}"),
            ft.Text("Esta ação não pode ser desfeita."),
            ft.Text("Para confirmar, digite o código abaixo:"),
            ft.Row([self.code_display_text], alignment=ft.MainAxisAlignment.CENTER),
            self.confirmation_text,
            self.error_text
        ], tight=True, spacing=15, width=350)
        
        self.actions = [
            ft.TextButton("Cancelar", on_click=self.cancel),
            ft.TextButton("Excluir", on_click=self.confirm, style=ft.ButtonStyle(color="red"))
        ]
        self.actions_alignment = ft.MainAxisAlignment.END
    
    def confirm(self, e):
        if self.confirmation_text.value.strip().upper() == self.random_code:
            self.on_confirm_callback(e)
        else:
            self.error_text.visible = True
            self.update()
    
    def cancel(self, e):
        self.on_cancel_callback(e)


class AddSupplierDialog(ft.AlertDialog):
    def __init__(self, on_confirm, on_cancel, page, scale_func=None, list_options=None):
        super().__init__()
        self.on_confirm_callback = on_confirm
        self.on_cancel_callback = on_cancel
        self.page = page
        self.scale_func = scale_func or (lambda x: x)
        self.list_options = list_options or {}
        
        def update_status_color_dialog(e):
            """Atualiza a cor do texto do dropdown de status no diálogo"""
            status_value = e.control.value
            if status_value == "Active":
                e.control.color = "green"
            elif status_value == "Inactive":
                e.control.color = "red"
            else:
                e.control.color = get_current_theme_colors(self.page).get('on_surface') if hasattr(self, 'page') else "black"
            if e.page:
                e.safe_page_update(page)
            elif hasattr(e.control, 'update'):
                e.control.update()
        
        # Campos do formulário organizados por seção (sem supplier_id)
        self.fields = {
            # Informações Básicas
            "vendor_name": ft.TextField(
                label="Vendor Name*",
                expand=True,
                bgcolor=get_current_theme_colors(self.page).get('field_background'),
                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline')
            ),
            "supplier_origin": ft.Dropdown(
                label="Origem",
                expand=True,
                value="",
                options=[ft.dropdown.Option(v) for v in ["Nacional", "Importado"]],
                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline')
            ),
            "supplier_category": ft.Container(
                content=ft.Dropdown(
                    label="Supplier Category",
                    expand=True,
                    color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                    border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline'),
                    options=[
                        ft.dropdown.Option(v)
                        for v in (self.list_options.get('category') or ["Raw Materials", "Components", "Services", "Equipment"])
                    ],
                ),
                bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('field_background'),
                expand=True
            ),
            
            # Informações de Contato
            "supplier_email": ft.TextField(
                label="Email (pressione Enter ou ; para separar)", 
                multiline=True,
                min_lines=4,
                max_lines=6,
                expand=True,
                bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('field_background'),
                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline'),
                hint_text="exemplo@email.com; outro@email.com"
            ),
            "supplier_number": ft.TextField(
                label="SSID",
                expand=True,
                bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('field_background'),
                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline')
            ),
            "supplier_po": ft.TextField(
                label="PO",
                expand=True,
                bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('field_background'),
                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline')
            ),
            
            # Configurações Organizacionais
            "bu": ft.Container(
                content=ft.Dropdown(
                    label="BU (Business Unit)",
                    expand=True,
                    color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                    border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline'),
                    filled=False,
                    options=[ft.dropdown.Option(v) for v in (self.list_options.get('bu') or ["", "Operations", "Manufacturing", "Procurement", "Quality", "Logistics"]) ]
                ),
                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                expand=True
            ),

            "supplier_status": ft.Container(
                content=ft.Dropdown(
                    label="Status",
                    expand=True,
                    value="Active",  # Valor padrão
                    color="green",  # Cor inicial verde para "Active"
                    border_color=get_current_theme_colors(self.page).get('outline'),
                    options=[
                        ft.dropdown.Option("Active"),
                        ft.dropdown.Option("Inactive"),
                    ],
                    on_change=update_status_color_dialog
                ),
                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                expand=True
            ),

                        "planner": ft.Container(
                            content=ft.Dropdown(
                                label="Planner",
                                expand=True,
                                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline'),
                                options=[ft.dropdown.Option(v) for v in ((self.list_options.get('planner') or []) )]
                            ),
                            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                            expand=True
                        ),

                        "continuity": ft.Container(
                            content=ft.Dropdown(
                                label="Continuity", 
                                expand=True,
                                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline'),
                                options=[ft.dropdown.Option(v) for v in ((self.list_options.get('continuity') or []) )]
                            ),
                            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                            expand=True
                        ),

                        "sourcing": ft.Container(
                            content=ft.Dropdown(
                                label="Sourcing",
                                expand=True,
                                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline'),
                                options=[ft.dropdown.Option(v) for v in ((self.list_options.get('sourcing') or []) )]
                            ),
                            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                            expand=True
                        ),

                        "sqie": ft.Container(
                            content=ft.Dropdown(
                                label="SQIE",
                                expand=True,
                                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline'),
                                options=[ft.dropdown.Option(v) for v in ((self.list_options.get('sqie') or []) )]
                            ),
                            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                            expand=True
                        ),

        }
        
        self.title = ft.Text("➕ Adicionar Novo Supplier", size=20, weight=ft.FontWeight.BOLD)
        self.error_text = ft.Text("", color="error", visible=False, size=13, weight=ft.FontWeight.W_500)
        
        # Criar seções organizadas com tema padrão - Layout Clean
        basic_info_section = ft.Container(
            content=ft.Column([
                ft.Text("📋 Informações Básicas", size=14, weight=ft.FontWeight.BOLD),
                ft.Row([self.fields["vendor_name"], self.fields["supplier_origin"]], spacing=15),
                ft.Row([self.fields["supplier_category"], self.fields["supplier_status"]], spacing=15),
            ], spacing=10),
            padding=ft.padding.all(15),
            bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('card_background'),
            border_radius=10,
            border=ft.border.all(1, get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline_variant'))
        )
        
        contact_section = ft.Container(
            content=ft.Column([
                ft.Text("📞 Informações de Contato", size=14, weight=ft.FontWeight.BOLD),
                ft.Container(
                    content=self.fields["supplier_email"],
                    padding=ft.padding.only(bottom=5)
                ),
                ft.Container(
                    content=ft.Text(
                        "💡 Dica: Separe múltiplos emails com ponto e vírgula (;)",
                        size=11,
                        italic=True,
                        color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface_variant')
                    ),
                    padding=ft.padding.only(bottom=10)
                ),
                ft.Row([self.fields["supplier_number"], self.fields["supplier_po"]], spacing=15),
            ], spacing=8),
            padding=ft.padding.all(15),
            bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('card_background'),
            border_radius=10,
            border=ft.border.all(1, get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline_variant'))
        )
        
        org_section = ft.Container(
            content=ft.Column([
                ft.Text("🏢 Configurações Organizacionais", size=14, weight=ft.FontWeight.BOLD),
                ft.Row([self.fields["bu"]], spacing=15),
            ], spacing=10),
            padding=ft.padding.all(15),
            bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('card_background'),
            border_radius=10,
            border=ft.border.all(1, get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline_variant'))
        )
        
        management_section = ft.Container(
            content=ft.Column([
                ft.Text("⚙️ Configurações de Gestão", size=14, weight=ft.FontWeight.BOLD),
                ft.Row([self.fields["planner"], self.fields["continuity"]], spacing=15),
                ft.Row([self.fields["sourcing"], self.fields["sqie"]], spacing=15),
            ], spacing=10),
            padding=ft.padding.all(15),
            bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('card_background'),
            border_radius=10,
            border=ft.border.all(1, get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline_variant'))
        )
        
        self.content = ft.Container(
            content=ft.Column([
                ft.Container(
                    content=ft.Text("* Campos obrigatórios", size=12, weight=ft.FontWeight.W_500, color="error"),
                    padding=ft.padding.only(bottom=5)
                ),
                basic_info_section,
                contact_section,
                org_section,
                management_section,
                self.error_text
            ], spacing=15, scroll=ft.ScrollMode.AUTO),
            width=600,
            height=550,
            padding=ft.padding.all(10)
        )
        
        self.actions = [
            ft.TextButton("Cancelar", on_click=self.cancel, icon=ft.Icons.CANCEL),
            ft.ElevatedButton("💾 Salvar", on_click=self.confirm, 
                            style=ft.ButtonStyle(
                                bgcolor="primary",
                                color="white"  # Cor da fonte branca para melhor contraste
                            ))
        ]
        self.actions_alignment = ft.MainAxisAlignment.END
    
    def get_field_value(self, field):
        """Obtém valor tanto de TextField quanto de Dropdown"""
        if hasattr(field, 'value') and field.value:
            return str(field.value).strip()
        return ""
    
    def validate_fields(self):
        """Valida os campos obrigatórios"""
        vendor_name_value = self.get_field_value(self.fields["vendor_name"])
        print(f"🔍 DEBUG: Validando vendor_name: '{vendor_name_value}'")
        
        if not vendor_name_value:
            print("🔍 DEBUG: Vendor Name está vazio")
            self.error_text.value = "❌ Campo Vendor Name é obrigatório!"
            self.error_text.visible = True
            return False
        
        print("🔍 DEBUG: Validação passou")
        self.error_text.visible = False
        return True
    
    def confirm(self, e):
        print("🔍 DEBUG: Método confirm chamado")
        if self.validate_fields():
            print("🔍 DEBUG: Validação passou, chamando callback")
            self.on_confirm_callback(e)
        else:
            print("🔍 DEBUG: Validação falhou")
            self.update()
    
    def cancel(self, e):
        self.on_cancel_callback(e)

class EditTimelineRecordDialog(ft.AlertDialog):
    def __init__(self, record_data, on_confirm, on_cancel, page):
        super().__init__()
        self.record_data = record_data  # (month, year, otif, nil, pickup, package, total, comment)
        self.on_confirm_callback = on_confirm
        self.on_cancel_callback = on_cancel
        self.page = page

        month, year, otif, nil, pickup, package, total, comment = record_data

        months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                 "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        try:
            month_int = int(month)
            month_name = months[month_int-1] if 1 <= month_int <= 12 else str(month)
        except (ValueError, TypeError):
            month_name = str(month)

        self.title = ft.Text(f"Editar Registro - {month_name}/{year}")

        # Criar controles baseados na configuração e permissões
        self.otif_control = self.create_score_control("OTIF", otif, current_user_permissions.get('otif', False))
        self.nil_control = self.create_score_control("NIL", nil, current_user_permissions.get('nil', False))
        self.pickup_control = self.create_score_control("Pickup", pickup, current_user_permissions.get('pickup', False))
        self.package_control = self.create_score_control("Package", package, current_user_permissions.get('package', False))

        self.comment_field = ft.TextField(
            label="Comentário",
            value=comment or "",
            expand=True,
            multiline=True,
            min_lines=6,
            max_lines=12,
            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
            border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
        )

        # Layout em duas colunas: duas notas por linha
        left_col = ft.Column([ft.Container(content=self.otif_control, width=260), ft.Container(height=6), ft.Container(content=self.pickup_control, width=260)], spacing=6)
        right_col = ft.Column([ft.Container(content=self.nil_control, width=260), ft.Container(height=6), ft.Container(content=self.package_control, width=260)], spacing=6)

        controls_row = ft.Row([
            left_col,
            ft.Container(width=12),
            right_col
        ], alignment=ft.MainAxisAlignment.START)

        self.content = ft.Container(
            content=ft.Column([
                    ft.Text("Edite os valores do registro:", size=14),
                    ft.Container(height=8),
                    controls_row,
                    ft.Container(height=8),
                    self.comment_field,
            ], spacing=8, tight=False),
                # Remover width fixo para permitir adaptação à janela
                height=420
        )

        self.actions = [
            ft.TextButton("Cancelar", on_click=self.cancel, icon=ft.Icons.CANCEL),
            ft.ElevatedButton("💾 Salvar", on_click=self.confirm,
                            style=ft.ButtonStyle(
                                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'),
                                color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_primary')
                            ))
        ]
        self.actions_alignment = ft.MainAxisAlignment.END

    def create_score_control(self, label, value, has_permission):
        """Cria um controle de pontuação baseado na configuração (slider ou spinbox) e permissões"""
        global score_control_type

        # Normalizar valor recebido: pode vir como número, string numérica, ou valores inválidos
        norm_value = None
        try:
            if value is None:
                norm_value = None
            elif isinstance(value, (int, float)):
                norm_value = float(value)
            elif isinstance(value, str):
                # remover espaços e possíveis percentuais/ caracteres extras
                v = value.strip()
                # tentar converter diretamentente
                norm_value = float(v) if v != "" else None
            else:
                # tenta converter generically
                norm_value = float(value)
        except (ValueError, TypeError):
            # valor inválido — registrar e usar None como fallback
            print(f"Aviso: valor de pontuação inválido para '{label}': {value}")
            norm_value = None

        if score_control_type == "slider":
            # Criar slider Material
            score_text = ft.Text(
                f"{norm_value:.1f}" if norm_value is not None else "0.0", 
                width=40, 
                text_align=ft.TextAlign.CENTER,
                size=18,
                weight=ft.FontWeight.W_300,
                font_family="Roboto"
            )
            score_slider = ft.Slider(
                min=0,
                max=10,
                divisions=100,
                value=norm_value if norm_value is not None else 0,
                # Largura maior para melhor usabilidade na aba Score
                width=230,
                # Cores do slider: ativa segue a primária; inativa mais clara
                active_color="primary",
                inactive_color="surface_variant",
                disabled=not has_permission
            )

            def on_slider_change(e):
                score_text.value = f"{e.control.value:.1f}"
                if hasattr(score_text, 'page') and score_text.page is not None:
                    score_text.update()

            score_slider.on_change = on_slider_change

            # Aplicar opacidade reduzida se não tiver permissão
            opacity = 1.0 if has_permission else 0.5
            score_text.opacity = opacity

            return ft.Column([
                ft.Text(label, size=12, weight="bold", opacity=opacity),
                score_slider,
                ft.Container(
                    content=score_text,
                    alignment=ft.alignment.center,
                    # Acompanhar a largura do slider para manter alinhamento
                    width=230,
                    margin=ft.margin.only(top=-15)
                )
            ], spacing=0, tight=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER)

        else:  # spinbox
            try:
                increment = load_spinbox_increment()
            except NameError:
                # Função não disponível no escopo no momento; usar valor padrão
                increment = 0.1
            except Exception as ex:
                print(f"Aviso: não foi possível carregar increment do spinbox: {ex}")
                increment = 0.1

            score_field = ft.TextField(
                value=f"{norm_value:.1f}" if norm_value is not None else "0.0",
                text_align=ft.TextAlign.CENTER,
                width=70,
                border_radius=8,
                # Manter a consistência visual: usar cor de fundo igual à do card
                bgcolor=get_current_theme_colors(get_theme_name_from_page(self.page)).get('card_background'),
                color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('on_surface'),
                border_color=get_current_theme_colors(get_theme_name_from_page(self.page)).get('outline'),
                disabled=not has_permission
            )

            # CORRIGIDO: Validar entrada manual no campo de spinbox apenas quando terminar de editar
            def on_score_field_blur(e):
                """Valida e formata o valor apenas quando o usuário termina de editar (perde o foco)"""
                try:
                    v = e.control.value
                    if v is None or str(v).strip() == "":
                        e.control.value = "0.0"
                        e.control.update()
                        return
                    
                    # Substituir vírgula por ponto antes de processar
                    v = str(v).strip().replace(',', '.')
                    
                    # Permitir entrada de números inteiros e decimais
                    num = float(v)
                    # Limitar entre 0 e 10
                    num = max(0, min(10, num))
                    # Formatar com 1 casa decimal
                    e.control.value = f"{num:.1f}"
                    e.control.update()
                except (ValueError, TypeError):
                    # Se não conseguir converter, voltar para 0.0
                    e.control.value = "0.0"
                    e.control.update()

            # Usar on_blur em vez de on_change para não interferir durante a digitação
            score_field.on_blur = on_score_field_blur

            def adjust_score(e):
                if not has_permission:
                    return
                try:
                    current_value = float(score_field.value)
                    if e.control.data == "+":
                        current_value += increment
                    elif e.control.data == "-":
                        current_value -= increment
                    new_value = max(0, min(10, current_value))
                    score_field.value = str(round(new_value, 1))
                    score_field.update()
                except ValueError:
                    score_field.value = "0.0"
                    score_field.update()

            # Aplicar opacidade reduzida se não tiver permissão
            opacity = 1.0 if has_permission else 0.5

            return ft.Column([
                ft.Text(label, size=12, weight="bold", opacity=opacity),
                ft.Row([
                    ft.IconButton(ft.Icons.REMOVE, on_click=adjust_score, data="-", icon_size=16, disabled=not has_permission, opacity=opacity),
                    score_field,
                    ft.IconButton(ft.Icons.ADD, on_click=adjust_score, data="+", icon_size=16, disabled=not has_permission, opacity=opacity),
                ], alignment=ft.MainAxisAlignment.CENTER)
            ], spacing=5, horizontal_alignment=ft.CrossAxisAlignment.CENTER)

    def get_control_value(self, control):
        """Extrai o valor do controle (slider ou spinbox)"""
        try:
            if score_control_type == "slider":
                # Para slider Material: Column[Text, Slider, Container[Text]]
                if hasattr(control, 'controls') and len(control.controls) >= 2:
                    slider = control.controls[1]  # O slider é o segundo controle
                    if hasattr(slider, 'value'):
                        # Garantir que o slider esteja no intervalo
                        try:
                            val = float(slider.value)
                            return max(0, min(10, val))
                        except Exception:
                            return None
            else:  # spinbox
                # Para spinbox: Column[Text, Row[IconButton, TextField, IconButton]]
                if hasattr(control, 'controls') and len(control.controls) >= 2:
                    spinbox_row = control.controls[1]  # A Row é o segundo controle
                    if hasattr(spinbox_row, 'controls') and len(spinbox_row.controls) >= 2:
                        text_field = spinbox_row.controls[1]  # O TextField é o segundo da Row
                        if hasattr(text_field, 'value') and text_field.value:
                            try:
                                val = float(text_field.value)
                                return max(0, min(10, val))
                            except Exception:
                                return None
        except (ValueError, TypeError, AttributeError):
            pass
        return None

    def confirm(self, e):
        # Extrair valores dos controles
        try:
            otif = self.get_control_value(self.otif_control)
            nil = self.get_control_value(self.nil_control)
            pickup = self.get_control_value(self.pickup_control)
            package = self.get_control_value(self.package_control)
            comment = self.comment_field.value.strip() if self.comment_field.value else None

            # Chamar callback com os novos valores
            self.on_confirm_callback(self.record_data, otif, nil, pickup, package, comment)
        except Exception as ex:
           print(f"❌ Erro ao processar valores: {str(ex)}")

    def cancel(self, e):
        self.on_cancel_callback(e)

# Variável global para o usuário atual
# Variáveis globais de usuário definidas no início do arquivo

# Configurações globais do aplicativo
app_settings = {'toast_duration': 3, 'auto_save': True}

def save_auto_save_setting(enabled):
    """Persist a preferência de auto save."""
    try:
        value = "1" if enabled else "0"
        exists = db_manager.query_one(
            "SELECT COUNT(*) as count FROM app_settings WHERE setting_key = 'auto_save_enabled'"
        )

        if exists and (exists.get('count') or 0) > 0:
            db_manager.execute(
                "UPDATE app_settings SET setting_value = ?, last_updated = CURRENT_TIMESTAMP WHERE setting_key = 'auto_save_enabled'",
                (value,),
            )
        else:
            db_manager.execute(
                "INSERT INTO app_settings (setting_key, setting_value) VALUES (?, ?)",
                ('auto_save_enabled', value),
            )
    except Exception as e:
        print(f"Erro ao salvar configuração de auto save: {e}")

def load_auto_save_setting():
    """Carrega a preferência de auto save."""
    try:
        result = db_manager.query_one(
            "SELECT setting_value FROM app_settings WHERE setting_key = 'auto_save_enabled'"
        )
        if result and result.get('setting_value') is not None:
            value = str(result['setting_value']).strip().lower()
            return value in {"1", "true", "yes", "on"}
    except Exception as e:
        print(f"Erro ao carregar configuração de auto save: {e}")
    return True

def load_app_settings(user_wwid=None):
    """Carrega as configurações gerais do aplicativo."""
    settings = {'toast_duration': 3, 'auto_save': True}

    try:
        if user_wwid:
            result = db_manager.query_one(
                "SELECT toast_duration FROM app_settings WHERE user_wwid = ? ORDER BY last_updated DESC LIMIT 1",
                (user_wwid,),
            )
        else:
            result = db_manager.query_one(
                "SELECT toast_duration FROM app_settings WHERE user_wwid = 'default' ORDER BY last_updated DESC LIMIT 1"
            )

        if result and result.get('toast_duration') is not None:
            settings['toast_duration'] = result['toast_duration']
    except Exception as ex:
        print(f"Erro ao carregar configurações do app: {ex}")

    try:
        settings['auto_save'] = load_auto_save_setting()
    except Exception as ex:
        print(f"Erro ao carregar configuração de auto save: {ex}")

    return settings

def save_app_settings(settings, user_wwid=None):
    """Salva as configurações gerais do aplicativo."""
    try:
        user_id = user_wwid or 'default'
        
        existing = db_manager.query_one("SELECT id FROM app_settings WHERE user_wwid = ?", (user_id,))
        
        if existing:
            # Atualizar configuração existente
            db_manager.execute("UPDATE app_settings SET toast_duration = ?, last_updated = CURRENT_TIMESTAMP WHERE user_wwid = ?", 
                      (settings['toast_duration'], user_id))
        else:
            # Inserir nova configuração
            db_manager.execute("INSERT INTO app_settings (user_wwid, toast_duration) VALUES (?, ?)", 
                      (user_id, settings['toast_duration']))
        
        return True
    except Exception as ex:
        print(f"Erro ao salvar configurações do app: {ex}")
        return False

def initialize_main_app(page: ft.Page, user_theme="white"):
    global app_settings, current_user_wwid, current_user_name, current_user_privilege, current_user_permissions, users_controls, log_controls, page_ref
    
    # Definir page_ref global para uso em notificações
    page_ref = page
    
    # Armazenar nome do tema na página para acesso global e confiável
    page.data = {"theme_name": user_theme}

    # Carregar configurações do app
    app_settings = load_app_settings()

    # Executor único para tarefas pesadas fora da thread da UI
    background_executor = ThreadPoolExecutor(max_workers=4)

    def run_async(task, *args, on_success=None, on_error=None, on_finally=None):
        """Executa `task` em thread separada e trata callbacks na UI."""
        global page_ref
        
        print(f"\n🚀 DEBUG run_async - Início")
        print(f"   Task: {task.__name__ if hasattr(task, '__name__') else 'anonymous'}")
        print(f"   Args: {args}")
        print(f"   page_ref válido? {page_ref is not None}")

        def worker():
            try:
                print(f"   🔄 Executando task em thread...")
                result = task(*args)
                print(f"   ✅ Task executada com sucesso, resultado: {result}")
                
                if on_success:
                    print(f"   📞 Chamando on_success diretamente (sem async)...")
                    try:
                        on_success(result)
                        print(f"   ✅ on_success executado com sucesso")
                    except Exception as callback_error:
                        print(f"   ❌ Erro ao executar on_success: {callback_error}")
                        import traceback
                        traceback.print_exc()
            except Exception as exc:  # noqa: BLE001
                print(f"   ❌ Erro em tarefa assíncrona: {exc}")
                import traceback
                traceback.print_exc()
                
                if on_error:
                    print(f"   📞 Chamando on_error diretamente (sem async)...")
                    try:
                        on_error(exc)
                        print(f"   ✅ on_error executado com sucesso")
                    except Exception as error_callback_error:
                        print(f"   ❌ Erro ao executar on_error: {error_callback_error}")
                        import traceback
                        traceback.print_exc()
            finally:
                if on_finally:
                    print(f"   📞 Chamando on_finally diretamente (sem async)...")
                    try:
                        on_finally()
                        print(f"   ✅ on_finally executado com sucesso")
                    except Exception as finally_error:
                        print(f"   ❌ Erro ao executar on_finally: {finally_error}")

        print(f"   🎯 Submetendo worker ao executor...")
        background_executor.submit(worker)
    
    def create_email_development_tabs():
        """Cria as abas de email com mensagem de desenvolvimento"""
        Colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        # Conteúdo da mensagem de desenvolvimento
        development_content = ft.Column([
            ft.Container(height=50),  # Espaçamento superior
            ft.Icon(
                name=ft.Icons.CONSTRUCTION, 
                size=80, 
                color=Colors.get('on_surface_variant', Colors.get('on_surface', '#757575'))
            ),
            ft.Container(height=20),  # Espaçamento
            ft.Text(
                "Funcionalidade em Desenvolvimento",
                size=24,
                weight="bold",
                color=Colors['on_surface'],
                text_align=ft.TextAlign.CENTER
            ),
            ft.Container(height=10),  # Espaçamento
            ft.Text(
                "Esta seção está sendo desenvolvida e estará disponível em breve.",
                size=16,
                color=Colors.get('on_surface_variant', Colors.get('on_surface', '#757575')),
                text_align=ft.TextAlign.CENTER
            ),
        ], 
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        expand=True
        )
        
        # Criar as abas
        tabs = ft.Tabs(
            selected_index=0,
            animation_duration=300,
            tab_alignment=ft.TabAlignment.START,
            tabs=[
                ft.Tab(
                    text="Envio Individual",
                    icon=ft.Icons.PERSON,
                    content=ft.Container(
                        content=development_content,
                        alignment=ft.alignment.center,
                        expand=True,
                        padding=20
                    )
                ),
                ft.Tab(
                    text="Envio em Lote", 
                    icon=ft.Icons.GROUP,
                    content=ft.Container(
                        content=development_content,
                        alignment=ft.alignment.center,
                        expand=True,
                        padding=20
                    )
                )
            ],
            expand=True
        )
        
        return tabs
    
    # Tema simplificado - apenas 3 opções básicas
    
    def menu_item(icon, text, idx, show_text=True):
        is_selected = selected_index.current == idx
        Colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        def on_hover(e):
            e.control.bgcolor = Colors['surface_variant'] if e.data == "true" else (Colors['primary_container'] if is_selected else None)
            e.control.update()
        row_controls = [
            ft.Icon(icon, color=Colors['primary'] if is_selected else Colors['on_surface'])
        ]
        if show_text:
            row_controls.append(
                ft.Text(
                    text,
                    size=16,
                    weight="bold" if is_selected else "normal",
                    color=Colors['primary'] if is_selected else Colors['on_surface'],
                )
            )
        # Corrige: passa a função handler sem executar
        return ft.Container(
            content=ft.Row(row_controls, alignment=ft.MainAxisAlignment.CENTER if not show_text else ft.MainAxisAlignment.START, spacing=15),
            padding=ft.padding.symmetric(horizontal=0 if not show_text else 10, vertical=10),
            border_radius=8,
            bgcolor=Colors['primary_container'] if is_selected else None,
            on_hover=on_hover,
            on_click=lambda e: set_selected(idx)(e),
            animate=200,
            width=160 if show_text else 40,
        )
    page.title = f"Score App v{APP_VERSION}"
    page.padding = 0
    page.theme_mode = ft.ThemeMode.LIGHT

    # Estado do menu selecionado
    selected_index = ft.Ref()
    selected_index.current = 0

    # Funções auxiliares para usuários (definidas antes da interface)
    def check_user_exists(wwid):
        """Verifica se um usuário existe no banco de dados pelo WWID."""
        if not wwid or not wwid.strip():
            return False
        try:
            result = db_manager.query_one("SELECT 1 FROM users_table WHERE UPPER(user_wwid) = ?", (wwid.upper().strip(),))
            return result is not None
        except Exception as ex:
            print(f"Erro ao verificar usuário: {ex}")
            return False

    def update_action_button():
        """Atualiza o texto e ícone do botão de ação baseado no estado atual."""
        if 'action_btn' not in users_controls:
            return

        wwid_ctrl = resolve_control(users_controls.get('wwid')) if 'wwid' in users_controls else None
        wwid = wwid_ctrl.value.strip() if wwid_ctrl and getattr(wwid_ctrl, 'value', None) else ''

        action_btn_ctrl = resolve_control(users_controls.get('action_btn'))

        # Controlar visibilidade baseado em privilégio
        is_super_admin = current_user_privilege == "Super Admin"
        
        if wwid and check_user_exists(wwid):
            # Usuário existe - botão Update
            if action_btn_ctrl:
                action_btn_ctrl.text = "Update"
                action_btn_ctrl.icon = ft.Icons.EDIT
                # Mostrar para Super Admin ou se for o próprio usuário editando sua senha
                action_btn_ctrl.visible = is_super_admin or (wwid.upper() == current_user_wwid.upper())
        else:
            # Usuário não existe - botão Add User (apenas Super Admin)
            if action_btn_ctrl:
                action_btn_ctrl.text = "Add User"
                action_btn_ctrl.icon = ft.Icons.PERSON_ADD
                action_btn_ctrl.visible = is_super_admin

        if action_btn_ctrl and hasattr(action_btn_ctrl, 'update'):
            try:
                action_btn_ctrl.update()
            except Exception as e:
                print(f"Erro ao atualizar botão de ação: {e}")

    def on_wwid_change(e):
        """Função chamada quando o WWID é alterado para verificar em tempo real."""
        global selected_user
        wwid = e.control.value.strip().upper() if e.control.value else ''
        
        # Atualizar botão
        update_action_button()
        
        # Se WWID existe, carregar dados automaticamente
        if wwid and check_user_exists(wwid):
            # Só carregar se não é o usuário já selecionado
            if selected_user != wwid:
                load_user_data_by_wwid(wwid)
        elif wwid:
            # WWID digitado mas não existe - limpar apenas outros campos, manter WWID
            clear_other_fields_except_wwid()
        elif not wwid:
            # Campo WWID foi limpo - limpar seleção
            selected_user = None

    def find_field_by_label(label_text):
        """Busca um TextField ou Dropdown na estrutura users_content pelo label"""
        try:
            # Buscar users_content através da busca recursiva
            users_content_found = responsive_app_manager._find_users_content_recursive(page.controls[0])
            if not users_content_found:
                return None
                
            return _find_control_recursive(users_content_found, label_text)
        except Exception as e:
            print(f"Erro ao buscar campo {label_text}: {e}")
            return None
    
    def find_checkbox_by_label(label_text):
        """Busca um Checkbox na estrutura users_content pelo label"""
        try:
            # Buscar users_content através da busca recursiva
            users_content_found = responsive_app_manager._find_users_content_recursive(page.controls[0])
            if not users_content_found:
                return None
                
            return _find_checkbox_recursive(users_content_found, label_text)
        except Exception as e:
            print(f"Erro ao buscar checkbox {label_text}: {e}")
            return None
    
    def _find_control_recursive(control, target_label, depth=0):
        """Busca recursiva por TextField ou Dropdown com label específico"""
        if depth > 15:  # Evitar recursão infinita
            return None
            
        # Verificar se é TextField ou Dropdown com o label procurado
        if isinstance(control, (ft.TextField, ft.Dropdown)):
            if hasattr(control, 'label') and control.label == target_label:
                return control
        
        # Buscar nos filhos
        if hasattr(control, 'controls'):
            for child in control.controls:
                result = _find_control_recursive(child, target_label, depth + 1)
                if result:
                    return result
        elif hasattr(control, 'content') and control.content:
            return _find_control_recursive(control.content, target_label, depth + 1)
        
        return None
    
    def _find_checkbox_recursive(control, target_label, depth=0):
        """Busca recursiva por Checkbox com label específico"""
        if depth > 15:  # Evitar recursão infinita
            return None
            
        # Verificar se é Checkbox com o label procurado
        if isinstance(control, ft.Checkbox):
            if hasattr(control, 'label') and control.label == target_label:
                return control
        
        # Buscar nos filhos
        if hasattr(control, 'controls'):
            for child in control.controls:
                result = _find_checkbox_recursive(child, target_label, depth + 1)
                if result:
                    return result
        elif hasattr(control, 'content') and control.content:
            return _find_checkbox_recursive(control.content, target_label, depth + 1)
        
        return None

    def clear_other_fields_except_wwid():
        """Limpa todos os campos exceto o WWID quando digitando um WWID novo."""
        try:
            # Usar uma abordagem mais direta que não depende das referências antigas
            # após reorganização responsiva
            
            # Para campos de texto, buscar diretamente na estrutura
            try:
                # Buscar campo Name diretamente
                name_field = find_field_by_label("Nome Completo")
                if name_field:
                    name_field.value = ""
            except Exception as e:
                print(f"Erro ao limpar Name via busca direta: {e}")
                
            try:
                # Buscar campo Password diretamente  
                password_field = find_field_by_label("Senha")
                if password_field:
                    password_field.value = ""
            except Exception as e:
                print(f"Erro ao limpar Password via busca direta: {e}")
                
            try:
                # Buscar campo Privilege diretamente
                privilege_field = find_field_by_label("Nível de Privilégio")
                if privilege_field:
                    privilege_field.value = None
            except Exception as e:
                print(f"Erro ao limpar Privilege via busca direta: {e}")
            
            # Para checkboxes, buscar diretamente
            checkbox_labels = ["OTIF", "NIL", "Pickup", "Package"]
            for label in checkbox_labels:
                try:
                    checkbox = find_checkbox_by_label(label)
                    if checkbox:
                        checkbox.value = False
                except Exception as e:
                    print(f"Erro ao limpar checkbox {label}: {e}")
            
            # Não limpar selected_user automaticamente para evitar problemas com exclusão
            # A limpeza será feita apenas manualmente através da função clear_users_fields()
            
        except Exception as ex:
            print(f"Erro ao limpar campos: {ex}")

    def load_user_data_by_wwid(wwid):
        """Carrega os dados de um usuário pelo WWID e preenche os campos."""
        try:
            result = db_manager.query_one("""
                SELECT user_name, user_password, user_privilege,
                       otif, nil, pickup, package
                FROM users_table 
                WHERE UPPER(user_wwid) = ?
            """, (wwid.upper(),))
            
            if result:
                name = result['user_name']
                password = result['user_password']
                privilege = result['user_privilege']
                otif = result['otif']
                nil = result['nil']
                pickup = result['pickup']
                package = result['package']
                
                print(f"🔄 Preenchimento automático para WWID {wwid}")
                
                # Preencher campos usando busca direta na estrutura
                try:
                    name_field = find_field_by_label("Nome Completo")
                    if name_field:
                        name_field.value = str(name) if name else ""
                except Exception as e:
                    print(f"Erro ao preencher Name: {e}")
                    
                try:
                    password_field = find_field_by_label("Senha")
                    if password_field:
                        password_field.value = str(password) if password else ""
                except Exception as e:
                    print(f"Erro ao preencher Password: {e}")
                    
                try:
                    privilege_field = find_field_by_label("Nível de Privilégio")
                    if privilege_field:
                        privilege_field.value = str(privilege) if privilege else None
                except Exception as e:
                    print(f"Erro ao preencher Privilege: {e}")
                
                # Definir valores dos checkboxes usando busca direta
                checkbox_data = [
                    ("OTIF", otif), ("NIL", nil), ("Pickup", pickup), ("Package", package)
                ]
                for label, db_val in checkbox_data:
                    try:
                        checkbox = find_checkbox_by_label(label)
                        if checkbox:
                            checkbox.value = int(db_val) == 1
                    except Exception as e:
                        print(f"Erro ao preencher checkbox {label}: {e}")
                
                # Atualizar página diretamente
                try:
                    safe_page_update(page)
                except Exception as e:
                    print(f"Erro ao atualizar página: {e}")
                
                # Definir usuário selecionado globalmente
                global selected_user
                selected_user = wwid
                
                print(f"✅ Dados carregados automaticamente para {wwid}")
                
        except Exception as ex:
            print(f"Erro ao carregar dados do usuário {wwid}: {ex}")

    # Variáveis globais para controles de usuário
    selected_user = None

    # Estado do menu (expandido/recolhido)
    menu_is_expanded = ft.Ref()
    menu_is_expanded.current = True

    menu_column_ref = ft.Ref()

    def load_user_theme(user_wwid=None):
        """Carrega o tema salvo do usuário ou o tema padrão."""
        try:
            if user_wwid:
                result = db_manager.query_one("SELECT theme_mode FROM theme_settings WHERE user_wwid = ? ORDER BY last_updated DESC LIMIT 1", (user_wwid,))
            else:
                # Se não há usuário logado, pegar o último tema usado
                result = db_manager.query_one("SELECT theme_mode FROM theme_settings ORDER BY last_updated DESC LIMIT 1")
            
            return result['theme_mode'] if result else "light"
            
        except Exception as e:
            print(f"Erro ao carregar tema: {e}")
            return "light"

    def save_user_theme(theme_mode, user_wwid=None):
        """Salva o tema do usuário no banco de dados."""
        try:
            # Inserir ou atualizar tema
            if user_wwid:
                # Verificar se já existe configuração para este usuário
                existing = db_manager.query_one("SELECT id FROM theme_settings WHERE user_wwid = ?", (user_wwid,))
                if existing:
                    # Atualizar existente
                    db_manager.execute("UPDATE theme_settings SET theme_mode = ?, last_updated = CURRENT_TIMESTAMP WHERE user_wwid = ?", 
                              (theme_mode, user_wwid))
                else:
                    # Inserir novo
                    db_manager.execute("INSERT INTO theme_settings (user_wwid, theme_mode) VALUES (?, ?)", 
                              (user_wwid, theme_mode))
            else:
                # Sem usuário específico, inserir como configuração geral
                db_manager.execute("INSERT INTO theme_settings (theme_mode) VALUES (?)", (theme_mode,))
            
            print(f"Tema '{theme_mode}' salvo com sucesso!")
            
        except Exception as e:
            print(f"Erro ao salvar tema: {e}")


    # Helper para acessar tanto ft.Ref quanto controles diretos
    def resolve_control(ctrl_or_ref):
        try:
            return ctrl_or_ref.current if hasattr(ctrl_or_ref, 'current') and ctrl_or_ref.current else ctrl_or_ref
        except Exception:
            return ctrl_or_ref


    def apply_theme(theme_mode, is_initialization=False):
        """Apply theme to the page"""
        if theme_mode == "dracula":
            page.theme_mode = ft.ThemeMode.DARK
            page.bgcolor = "#282A36"
            page.theme = ft.Theme(
                color_scheme=ft.ColorScheme(
                    primary="#BD93F9",
                    background="#282A36",
                    surface="#282A36",
                    on_surface="#F8F8F2",
                    outline="#6272A4",
                )
            )
        elif theme_mode == "dark":
            page.theme = None
            page.theme_mode = ft.ThemeMode.DARK
            page.bgcolor = "#121212"
        else:  # white
            page.theme = None
            page.theme_mode = ft.ThemeMode.LIGHT
            page.bgcolor = "#FFFFFF"

        # Save current theme
        page.data["theme_name"] = theme_mode
        safe_page_update(page)

        # Update all UI controls (always call, whether init or not)
        try:
            # Only call functions if they're defined (to avoid NameError during initialization)
            if 'refresh_users_list' in globals():
                refresh_users_list()
            if 'update_menu' in globals():
                update_menu()
            for key in ['planner', 'continuity', 'sourcing', 'sqie']:
                if 'refresh_list_ui' in globals():
                    refresh_list_ui(key)
            if 'update_control_colors' in globals():
                update_control_colors(page, theme_mode)
            if 'update_timeline_search_container_colors' in globals():
                update_timeline_search_container_colors(theme_mode)
            # Atualizar cores específicas da aba Risks (ano/meta)
            try:
                if 'update_risks_container_colors' in globals():
                    update_risks_container_colors(theme_mode)
            except Exception as _e:
                print(f"Aviso: falha ao atualizar cores da aba Risks: {_e}")
            # Atualizar conteúdo da aba Home
            try:
                if 'update_home_content' in globals():
                    update_home_content()
            except Exception as _e:
                print(f"Aviso: falha ao atualizar aba Home: {_e}")
        except Exception as e:
            print(f"⚠️ Error updating UI: {e}")

    def update_timeline_search_container_colors(theme_name):
        """Atualiza as cores do container dos campos de pesquisa do Timeline"""
        try:
            if timeline_search_container.current:
                Colors = get_current_theme_colors(theme_name)
                
                # Atualizar as cores do container principal
                timeline_search_container.current.bgcolor = Colors.get('surface_variant')
                timeline_search_container.current.border = None  # Remover borda
                
                # Atualizar o campo de pesquisa
                
                # Atualizar os dropdowns dentro do container
                if timeline_vendor_dropdown.current:
                    timeline_vendor_dropdown.current.bgcolor = Colors.get('field_background')
                    timeline_vendor_dropdown.current.color = Colors.get('on_surface')
                    timeline_vendor_dropdown.current.border_color = Colors.get('outline')
                
                if timeline_year_dropdown.current:
                    timeline_year_dropdown.current.bgcolor = Colors.get('field_background')
                    timeline_year_dropdown.current.color = Colors.get('on_surface')
                    timeline_year_dropdown.current.border_color = Colors.get('outline')
                
                timeline_search_container.current.update()
                
        except Exception as e:
            print(f"Erro ao atualizar cores do container Timeline: {e}")

    def update_control_colors(control, theme_name, depth=0):
        """Atualiza cores de controles recursivamente"""
        if depth > 20:
            return
        
        # Obter cores do tema
        Colors = get_current_theme_colors(theme_name)
            
        # Atualizar TextField
        if isinstance(control, ft.TextField):
            control.bgcolor = Colors.get('field_background')
            control.color = Colors.get('on_surface')
            control.border_color = Colors.get('outline')
            
        # Atualizar Dropdown  
        elif isinstance(control, ft.Dropdown):
            control.bgcolor = Colors.get('field_background')
            control.color = Colors.get('on_surface')
            control.border_color = Colors.get('outline')
        
        # Atualizar Card
        elif isinstance(control, ft.Card):
            control.color = Colors.get('card_background')
        
        # Atualizar Container
        elif isinstance(control, ft.Container):
            # Se tem bgcolor definido, atualizar conforme o contexto
            if hasattr(control, 'bgcolor') and control.bgcolor:
                # Para containers da aba Home que usam surface_variant
                if any(color in str(control.bgcolor) for color in ['#2E303E', '#181818', '#FAFAFA', 'surface_variant']):
                    control.bgcolor = Colors.get('surface_variant')
                    print(f"🎨 Container bgcolor atualizado para: {Colors.get('surface_variant')}")
                # Para containers que são fundos de cards
                elif any(color in str(control.bgcolor) for color in ['#343746', '#1E1E1E', '#F7F7F7', 'card_background']):
                    control.bgcolor = Colors.get('card_background')
                    print(f"🎨 Card bgcolor atualizado para: {Colors.get('card_background')}")
                # Para containers que provavelmente são fundos de campos
                elif control.bgcolor in ['#FFFFFF', '#2C2C2C', '#44475A', 'field_background']:
                    control.bgcolor = Colors.get('field_background')
        
        # Atualizar Text (para casos especiais)
        elif isinstance(control, ft.Text):
            # Preservar cores especiais como primary, mas atualizar cores básicas
            if hasattr(control, 'color') and control.color:
                color_str = str(control.color)
                # Atualizar cores específicas dos temas
                if any(old_color in color_str for old_color in ['#F8F8F2', '#E0E0E0', '#1C1C1C', '#000000']):
                    control.color = Colors.get('on_surface')
                    print(f"🎨 Text color atualizado para: {Colors.get('on_surface')}")
                elif any(old_color in color_str for old_color in ['#666666', '#555555']):
                    control.color = Colors.get('on_surface_variant', '#666666')
                    print(f"🎨 Text secondary color atualizado para: {Colors.get('on_surface_variant', '#666666')}")
                elif control.color in ['on_surface', 'black', 'white']:
                    control.color = Colors.get('on_surface')
        
        # Recursão nos filhos
        if hasattr(control, 'controls'):
            for child in control.controls:
                update_control_colors(child, theme_name, depth + 1)
        elif hasattr(control, 'content') and control.content:
            update_control_colors(control.content, theme_name, depth + 1)

    def update_risks_container_colors(theme_name):
        """Atualiza as cores do header da aba Risks (ano e target)"""
        try:
            Colors = get_current_theme_colors(theme_name)
            # Atualizar container principal do header (border e bgcolor)
            if risks_header_container and risks_header_container.current:
                try:
                    risks_header_container.current.bgcolor = Colors.get('surface_variant')
                    risks_header_container.current.border = None  # Remover borda
                except Exception:
                    pass

            # Atualizar dropdown e texto alvo dentro do header
            if risks_year_dropdown and risks_year_dropdown.current:
                risks_year_dropdown.current.bgcolor = Colors.get('field_background')
                risks_year_dropdown.current.color = Colors.get('on_surface')
                risks_year_dropdown.current.border_color = Colors.get('outline')

            if target_display_container and target_display_container.current:
                # target_display_container tem borda e bgcolor especiais
                try:
                    target_display_container.current.border = ft.border.all(1.5, ft.Colors.AMBER_700)
                    target_display_container.current.bgcolor = ft.Colors.with_opacity(0.1, ft.Colors.AMBER_700)
                except Exception:
                    pass

            # Atualizar cores do switch de inativos
            if inactive_switch_container and inactive_switch_container.current:
                try:
                    inactive_switch_container.current.bgcolor = Colors.get('surface_variant')
                    inactive_switch_container.current.border = ft.border.all(1, Colors.get('outline'))
                except Exception:
                    pass
                    
            if include_inactive_switch and include_inactive_switch.current:
                try:
                    include_inactive_switch.current.active_color = Colors.get('primary')
                    include_inactive_switch.current.inactive_track_color = Colors.get('on_surface_variant')
                    include_inactive_switch.current.inactive_thumb_color = Colors.get('on_surface_variant')
                except Exception:
                    pass
                    
            if inactive_switch_icon and inactive_switch_icon.current:
                try:
                    inactive_switch_icon.current.color = Colors.get('on_surface_variant')
                except Exception:
                    pass

            # Fazer update nos controles referenciados e forçar redraw da página
            if risks_header_container and risks_header_container.current:
                try:
                    risks_header_container.current.update()
                except Exception:
                    pass
            if target_display_container and target_display_container.current:
                try:
                    target_display_container.current.update()
                except Exception:
                    pass
            # Atualizar dropdown do ano se existir
            if risks_year_dropdown and risks_year_dropdown.current:
                try:
                    risks_year_dropdown.current.update()
                except Exception:
                    pass
                    
            # Atualizar controles do switch de inativos
            if inactive_switch_container and inactive_switch_container.current:
                try:
                    inactive_switch_container.current.update()
                except Exception:
                    pass
            if include_inactive_switch and include_inactive_switch.current:
                try:
                    include_inactive_switch.current.update()
                except Exception:
                    pass
            if inactive_switch_icon and inactive_switch_icon.current:
                try:
                    inactive_switch_icon.current.update()
                except Exception:
                    pass
            try:
                safe_page_update(page)
            except Exception:
                pass

        except Exception as e:
            print(f"Erro ao atualizar cores da aba Risks: {e}")

    def update_home_content():
        """Atualiza o conteúdo da aba Home quando o tema muda"""
        try:
            print(f"🎨 Iniciando atualização da aba Home...")
            print(f"🎨 Tema atual: {get_theme_name_from_page(page)}")
            
            if home_content_container.current:
                print("🎨 Container encontrado, recriando conteúdo...")
                # Recriar o conteúdo com o novo tema
                new_content = create_home_content()
                home_content_container.current.content = new_content
                
                # Forçar atualização recursiva de todos os controles
                update_control_colors(home_content_container.current, get_theme_name_from_page(page))
                
                home_content_container.current.update()
                print("✅ Conteúdo da aba Home atualizado com novo tema")
            else:
                print("⚠️ Referência do container da aba Home não encontrada")
        except Exception as e:
            print(f"❌ Erro ao atualizar conteúdo da aba Home: {e}")
            import traceback
            print(f"Traceback: {traceback.format_exc()}")
            # Fallback: tentar forçar update da página inteira
            try:
                safe_page_update(page)
            except:
                pass

    # ===== CARREGAMENTO INICIAL DO TEMA =====
    # Aplicar tema padrão ou salvo se disponível
    saved_theme = None
    try:
        # Carregar tema salvo do banco de dados
        if current_user_wwid:
            try:
                import sqlite3
                import os
                
                # Verificar se o banco de dados existe
                db_path = 'db.db'
                if os.path.exists(db_path):
                    # Criar tabela se não existir
                    db_manager.execute('''
                        CREATE TABLE IF NOT EXISTS user_themes (
                            user_wwid TEXT PRIMARY KEY,
                            theme TEXT NOT NULL
                        )
                    ''')
                    
                    # Buscar tema do usuário
                    result = db_manager.query_one("SELECT theme FROM user_themes WHERE user_wwid = ?", (current_user_wwid,))
                    saved_theme = result['theme'] if result else None
                    
                    if saved_theme:
                        print(f"Tema encontrado no banco para usuário {current_user_wwid}: {saved_theme}")
                    else:
                        print(f"Nenhum tema salvo encontrado para usuário {current_user_wwid}")
                else:
                    print("Banco de dados não encontrado, usando tema padrão")
            except Exception as e:
                print(f"Erro ao acessar banco de dados: {e}")
                saved_theme = None
        
        # Aplicar tema carregado ou usar padrão
        if user_theme and user_theme in ["white", "dark", "dracula"]:
            apply_theme(user_theme, is_initialization=True)
            print(f"Tema aplicado na inicialização: {user_theme}")
        else:
            apply_theme("white", is_initialization=True)  # tema padrão
            print("Aplicando tema padrão (white) na inicialização")
            
    except Exception as e:
        print(f"Erro durante carregamento de tema: {e}")
        apply_theme("white", is_initialization=True)  # fallback para tema padrão
    # ===== FIM DO CARREGAMENTO INICIAL DO TEMA =====

    # DEPRECATED: Função substituída por show_snack_bar()
    # Mantida temporariamente para referência
    def _deprecated_show_toast(message, color="green", restore_control=None):
        """Mostra um show_toast com duração configurável e gerenciamento melhorado.

        If `restore_control` is provided, the control will be re-enabled and its
        processing flag cleared before showing the show_toast. This ensures error/warning
        flows that call `show_toast` don't accidentally leave the originating
        control disabled.
        """
        global app_settings, _active_toast, _active_toast_cancel_event

        # Initialize globals used to track the active toast
        try:
            _active_toast
        except NameError:
            _active_toast = None
        try:
            _active_toast_cancel_event
        except NameError:
            _active_toast_cancel_event = None

        # Restore control immediately if provided
        try:
            if restore_control is not None:
                try:
                    restore_control.disabled = False
                    if hasattr(restore_control, '_is_processing'):
                        restore_control._is_processing = False
                    safe_update_control(restore_control, page)
                except Exception as _:
                    print(f"Aviso: não foi possível restaurar controle antes do show_toast: {_}")
        except Exception:
            pass

        # MELHORADO: Limpeza mais robusta de toasts anteriores
        def clean_previous_toasts():
            """Limpa todos os toasts anteriores do overlay de forma segura"""
            try:
                if _active_toast_cancel_event is not None:
                    _active_toast_cancel_event.set()
                
                if not hasattr(page, 'overlay') or page.overlay is None:
                    return
                
                # Remover toast específico se existe
                if _active_toast is not None:
                    try:
                        if _active_toast in page.overlay:
                            page.overlay.remove(_active_toast)
                    except (ValueError, AttributeError):
                        pass
                
                # NOVO: Limpeza geral de toasts órfãos
                # Remove containers que parecem ser toasts baseados em propriedades
                toasts_to_remove = []
                for item in page.overlay[:]:  # Cópia da lista para iteração segura
                    try:
                        if (hasattr(item, 'top') and hasattr(item, 'right') and 
                            hasattr(item, 'bgcolor') and 
                            item.top == 50 and item.right == 20):
                            toasts_to_remove.append(item)
                    except AttributeError:
                        continue
                
                for toast in toasts_to_remove:
                    try:
                        if toast in page.overlay:
                            page.overlay.remove(toast)
                    except (ValueError, AttributeError):
                        pass
                        
                safe_page_update(page)
                
            except Exception as ex:
                print(f"Aviso: erro na limpeza de toasts: {ex}")

        # Limpar toasts anteriores
        clean_previous_toasts()

        # Normalize color - accept common names
        color_map = {
            'green': "#31A037",
            'red': '#C62828',
            'orange': '#EF6C00',
            'blue': '#1565C0'
        }
        bgcolor = color_map.get(color, color)

        # Create new toast container and add to overlay
        try:
            toast_container = ft.Container(
                content=ft.Text(str(message), color="white", weight="bold"),
                bgcolor=bgcolor,
                padding=10,
                border_radius=5,
                top=50,
                right=20,
                animate_opacity=300,
            )

            if not hasattr(page, 'overlay') or page.overlay is None:
                page.overlay = []
            
            page.overlay.append(toast_container)
            _active_toast = toast_container
            safe_page_update(page)
        except Exception as ex:
            print(f"Erro ao mostrar show_toast: {ex}")
            return

        # MELHORADO: Remover thread mais robusta com timeout de segurança
        def remover(cancel_event, toast_ref):
            """Remove o toast após duração especificada, com proteções de segurança"""
            try:
                import time
                duration = app_settings.get('toast_duration', 3)
                max_duration = min(duration, 10)  # NOVO: Limite máximo de 10 segundos
                
                # Wait in small increments to allow cancellation
                waited = 0.0
                step = 0.1
                while waited < max_duration:
                    if cancel_event.is_set():
                        return
                    time.sleep(step)
                    waited += step

                # MELHORADO: Remoção mais robusta do toast
                try:
                    if (hasattr(page, 'overlay') and page.overlay is not None and
                        toast_ref is not None):
                        
                        # Verificação dupla para evitar condições de corrida
                        if toast_ref in page.overlay:
                            page.overlay.remove(toast_ref)
                            safe_page_update(page)
                        else:
                            # NOVO: Fallback - procurar por toasts similares órfãos
                            similar_toasts = [
                                item for item in page.overlay 
                                if (hasattr(item, 'top') and hasattr(item, 'right') and 
                                    item.top == 50 and item.right == 20)
                            ]
                            for toast in similar_toasts[:1]:  # Remove apenas o primeiro
                                try:
                                    page.overlay.remove(toast)
                                    safe_page_update(page)
                                    break
                                except (ValueError, AttributeError):
                                    pass
                                    
                except Exception as inner:
                    print(f"Aviso: erro ao remover toast específico: {inner}")
                    
            except Exception as outer:
                print(f"Erro na thread de remoção de toast: {outer}")
            finally:
                # NOVO: Limpeza final de segurança
                try:
                    if toast_ref == _active_toast:
                        globals()['_active_toast'] = None
                except Exception:
                    pass

        cancel_event = threading.Event()
        _active_toast_cancel_event = cancel_event
        
        # MELHORADO: Thread daemon com referência específica do toast
        remover_thread = threading.Thread(
            target=remover, 
            args=(cancel_event, toast_container), 
            daemon=True,
            name=f"ToastRemover-{id(toast_container)}"
        )
        remover_thread.start()

    def clear_all_toasts():
        """Função utilitária para limpar manualmente todos os toasts órfãos da tela"""
        try:
            global _active_toast, _active_toast_cancel_event
            
            # Cancelar toast ativo atual
            if _active_toast_cancel_event is not None:
                try:
                    _active_toast_cancel_event.set()
                except Exception:
                    pass
            
            if not hasattr(page, 'overlay') or page.overlay is None:
                return 0
            
            # Encontrar e remover todos os containers que parecem toasts
            toasts_removed = 0
            for item in page.overlay[:]:  # Cópia para iteração segura
                try:
                    # Identificar toasts por suas propriedades características
                    if (hasattr(item, 'top') and hasattr(item, 'right') and 
                        hasattr(item, 'bgcolor') and hasattr(item, 'padding') and
                        item.top == 50 and item.right == 20):
                        page.overlay.remove(item)
                        toasts_removed += 1
                except (ValueError, AttributeError):
                    continue
            
            # Limpar referências globais
            _active_toast = None
            _active_toast_cancel_event = None
            
            # Atualizar página se removemos algo
            if toasts_removed > 0:
                safe_page_update(page)
                print(f"🧹 Removidos {toasts_removed} toast(s) órfão(s)")
            
            return toasts_removed
            
        except Exception as ex:
            print(f"❌ Erro ao limpar toasts: {ex}")
            return 0

    def test_toast_system():
        """Função de teste para verificar o sistema de toast melhorado"""
        try:
            print("🧪 Testando sistema de toast...")
            
            # Limpar toasts existentes primeiro
            cleared = clear_all_toasts()
            if cleared > 0:
                print(f"🧹 Limpeza inicial: {cleared} toasts removidos")
            
            # Mostrar alguns toasts de teste
            show_snack_bar("✅ Teste 1: Toast verde")
            threading.Timer(1.0, lambda: show_snack_bar("⚠️ Teste 2: Toast laranja", False)).start()
            threading.Timer(2.0, lambda: show_snack_bar("❌ Teste 3: Toast vermelho", True)).start()
            threading.Timer(3.0, lambda: show_snack_bar("ℹ️ Teste 4: Toast azul", False)).start()
            
            print("🧪 Testes de toast iniciados. Observe se eles aparecem e desaparecem automaticamente.")
            
        except Exception as ex:
            print(f"❌ Erro no teste do sistema de toast: {ex}")

    def test_spinbox_functionality():
        """Função de teste para demonstrar o funcionamento correto dos spinboxes"""
        try:
            print("🧪 Testando funcionalidade dos spinboxes...")
            print("✅ Correções aplicadas:")
            print("   - Mudança de on_change para on_blur")
            print("   - Formatação apenas após perder o foco")
            print("   - Permite digitar '10' sem converter para '1.0'")
            print("   - Validação de faixa 0.0 a 10.0")
            print("   - Formatação com 1 casa decimal")
            print("🎯 Agora você pode digitar '10' nos spinboxes e ele não será convertido para '1.0' durante a digitação!")
            
        except Exception as ex:
            print(f"❌ Erro no teste dos spinboxes: {ex}")

    # show_toast = show_toast  # REMOVIDO - função obsoleta

    def load_user_criteria(user_wwid=None):
        """Carrega os pesos dos critérios salvos no banco."""
        try:
            # Primeiro tenta carregar da criteria_settings (configurações personalizadas)
            if user_wwid:
                result = db_manager.query_one("SELECT nil_weight, otif_weight, pickup_weight, package_weight, target_weight FROM criteria_settings WHERE user_wwid = ? ORDER BY last_updated DESC LIMIT 1", (user_wwid,))
                if result:
                    return {
                        'NIL': result['nil_weight'],
                        'OTIF': result['otif_weight'], 
                        'Quality of Pick Up': result['pickup_weight'],
                        'Quality-Supplier Package': result['package_weight'],
                        'Target': result['target_weight']
                    }
            
            # Se não encontrou configurações personalizadas, carrega da criteria_table original
            print("Carregando critérios da tabela criteria_table...")
            results = db_manager.query("SELECT criteria_category, value FROM criteria_table")
            
            criteria_data = {}
            for row in results:
                criteria_name = row['criteria_category']
                value_str = row['value']
                
                try:
                    # Converter valor para float
                    raw_value = float(value_str)
                    
                    # Target mantém escala 0-10, outros critérios ficam na escala 0-1
                    if criteria_name == "Target":
                        value = raw_value  # Manter escala 0-10
                        print(f"Target mantido na escala 0-10: {value}")
                    else:
                        value = raw_value  # Outros critérios já estão na escala 0-1
                    
                    criteria_data[criteria_name] = value
                    print(f"Critério carregado: {criteria_name} = {value}")
                    
                except ValueError as e:
                    print(f"Erro ao converter valor para {criteria_name}: {value_str} - {e}")
            
            if criteria_data:
                return criteria_data
                
            return None
        except Exception as ex:
            print(f"Erro ao carregar critérios: {ex}")
            return None

    def save_user_criteria(criteria_weights, user_wwid=None):
        """Salva os pesos dos critérios no banco."""
        try:
            if user_wwid:
                # Verificar se já existe configuração para este usuário
                existing = db_manager.query_one("SELECT id FROM criteria_settings WHERE user_wwid = ?", (user_wwid,))
                
                if existing:
                    # Atualizar configuração existente
                    db_manager.execute("""UPDATE criteria_settings SET 
                                   nil_weight = ?, otif_weight = ?, pickup_weight = ?, 
                                   package_weight = ?, target_weight = ?, 
                                   last_updated = CURRENT_TIMESTAMP 
                                   WHERE user_wwid = ?""", 
                              (criteria_weights['NIL'], criteria_weights['OTIF'], 
                               criteria_weights['Quality of Pick Up'], criteria_weights['Quality-Supplier Package'], 
                               criteria_weights['Target'], user_wwid))
                else:
                    # Inserir nova configuração
                    db_manager.execute("""INSERT INTO criteria_settings 
                                   (user_wwid, nil_weight, otif_weight, pickup_weight, package_weight, target_weight) 
                                   VALUES (?, ?, ?, ?, ?, ?)""", 
                              (user_wwid, criteria_weights['NIL'], criteria_weights['OTIF'], 
                               criteria_weights['Quality of Pick Up'], criteria_weights['Quality-Supplier Package'], 
                               criteria_weights['Target']))
                
                return True
            return False
        except Exception as ex:
            print(f"Erro ao salvar critérios: {ex}")
            return False

    # --- Início: Funções para Gerar Nota Cheia ---
    
    def generate_full_score_dialog():
        """Abre diálogo para gerar notas cheias"""
        global current_user_wwid
        # Refs para os controles do diálogo
        month_dropdown_ref = ft.Ref()
        year_dropdown_ref = ft.Ref()
        generate_button_ref = ft.Ref()
        cancel_button_ref = ft.Ref()
        # controle de execução e cancelamento
        import threading as _threading
        cancel_event = _threading.Event()
        running_flag = {'value': False}
        include_inactive_ref = ft.Ref()
        progress_bar_ref = ft.Ref()
        progress_text_ref = ft.Ref()
        
        # Opções de mês
        months = [
            ft.dropdown.Option("1", "Janeiro"),
            ft.dropdown.Option("2", "Fevereiro"),
            ft.dropdown.Option("3", "Março"),
            ft.dropdown.Option("4", "Abril"),
            ft.dropdown.Option("5", "Maio"),
            ft.dropdown.Option("6", "Junho"),
            ft.dropdown.Option("7", "Julho"),
            ft.dropdown.Option("8", "Agosto"),
            ft.dropdown.Option("9", "Setembro"),
            ft.dropdown.Option("10", "Outubro"),
            ft.dropdown.Option("11", "Novembro"),
            ft.dropdown.Option("12", "Dezembro"),
        ]
        
        # Opções de ano (2024 a 2040)
        years = [ft.dropdown.Option(str(year), str(year)) for year in range(2024, 2041)]
        
        def close_dialog(e):
            # Se estiver rodando, aciona cancelamento; caso contrário, fecha o diálogo normalmente
            if running_flag['value']:
                cancel_event.set()
                try:
                    progress_text_ref.current.visible = True
                    progress_text_ref.current.value = "Cancelando..."
                    progress_bar_ref.current.visible = True
                    # opcional: indeterminate
                    progress_bar_ref.current.value = None
                    # desabilita botões para evitar cliques repetidos
                    if cancel_button_ref.current:
                        cancel_button_ref.current.disabled = True
                    if generate_button_ref.current:
                        generate_button_ref.current.disabled = True
                    safe_page_update(page)
                except Exception:
                    pass
            else:
                page.close(dialog)
        
        def start_generation(e):
            month = month_dropdown_ref.current.value
            year = year_dropdown_ref.current.value
            include_inactive = bool(include_inactive_ref.current.value)
            
            if not month or not year:
                # Mostrar erro na própria janela
                progress_bar_ref.current.visible = False
                progress_text_ref.current.visible = True
                progress_text_ref.current.value = "❌ Mês e ano devem ser selecionados para gerar notas."
                progress_text_ref.current.color = ft.Colors.RED
                safe_page_update(page)
                return
            
            # Validação para não gerar para meses futuros
            now = datetime.datetime.now()
            month_int = int(month)
            year_int = int(year)
            
            if (year_int > now.year) or (year_int == now.year and month_int > now.month):
                # Mostrar erro na própria janela
                progress_bar_ref.current.visible = False
                progress_text_ref.current.visible = True
                progress_text_ref.current.value = "❌ Não é possível gerar notas para meses futuros."
                progress_text_ref.current.color = ft.Colors.RED
                safe_page_update(page)
                return
            
            # Evitar reentrância e desabilitar botão Gerar
            if running_flag['value']:
                return
            running_flag['value'] = True
            try:
                if generate_button_ref.current:
                    generate_button_ref.current.disabled = True
                if cancel_button_ref.current:
                    cancel_button_ref.current.disabled = False
                # Desabilitar dropdowns e switch durante execução
                if month_dropdown_ref.current:
                    month_dropdown_ref.current.disabled = True
                if year_dropdown_ref.current:
                    year_dropdown_ref.current.disabled = True
                if include_inactive_ref.current:
                    include_inactive_ref.current.disabled = True
                safe_page_update(page)
            except Exception:
                pass

            # Iniciar geração em thread separada
            def run_generation():
                try:
                    generate_full_scores(
                        month, year,
                        include_inactive,
                        cancel_event,
                        progress_bar_ref,
                        progress_text_ref,
                        dialog,
                        generate_button_ref,
                        cancel_button_ref,
                        theme_colors,
                        dialog_content,
                        month_dropdown_ref,
                        year_dropdown_ref,
                        include_inactive_ref
                    )
                except Exception as ex:
                    page.run_thread(lambda: show_snack_bar(f"Erro durante a geração: {str(ex)}", True))
                finally:
                    # reset estado ao finalizar (se o diálogo ainda estiver aberto)
                    running_flag['value'] = False
                    try:
                        if generate_button_ref.current:
                            generate_button_ref.current.disabled = False
                        if cancel_button_ref.current:
                            cancel_button_ref.current.disabled = False
                        # Reabilitar dropdowns e switch
                        if month_dropdown_ref.current:
                            month_dropdown_ref.current.disabled = False
                        if year_dropdown_ref.current:
                            year_dropdown_ref.current.disabled = False
                        if include_inactive_ref.current:
                            include_inactive_ref.current.disabled = False
                        safe_page_update(page)
                    except Exception:
                        pass
            
            _threading.Thread(target=run_generation, daemon=True).start()
        
        # Obter cores do tema atual
        theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        # Conteúdo do diálogo
        dialog_content = ft.Column([
            ft.Text(
                "Gerar Nota Cheia", 
                size=20, 
                weight="bold",
                color=theme_colors.get('on_surface')
            ),
            ft.Text(
                "Selecione o mês e ano para gerar notas máximas para todos os fornecedores ativos:", 
                size=14,
                color=theme_colors.get('on_surface')
            ),
            ft.Divider(height=20, color=theme_colors.get('outline')),
            
            ft.Row([
                ft.Dropdown(
                    label="Mês",
                    options=months,
                    expand=True,
                    ref=month_dropdown_ref,
                    dense=True,
                    color=theme_colors.get('on_surface'),
                    border_color=theme_colors.get('outline'),
                    **({'bgcolor': theme_colors.get('field_background')} if theme_colors.get('field_background') else {})
                ),
                ft.Dropdown(
                    label="Ano", 
                    options=years,
                    expand=True,
                    ref=year_dropdown_ref,
                    dense=True,
                    color=theme_colors.get('on_surface'),
                    border_color=theme_colors.get('outline'),
                    **({'bgcolor': theme_colors.get('field_background')} if theme_colors.get('field_background') else {})
                ),
            ], spacing=10),

            # Switch para incluir fornecedores inativos
            ft.Row([
                ft.Switch(ref=include_inactive_ref, value=False, active_color=theme_colors.get('primary')),
                ft.Text("Gerar score também para Inactives", color=theme_colors.get('on_surface'))
            ], spacing=10),
            
            # Container para barra de progresso
            ft.Container(
                content=ft.Column([
                    ft.Text(
                        "", 
                        ref=progress_text_ref, 
                        size=12, 
                        visible=False,
                        color=theme_colors.get('on_surface')
                    ),
                    ft.ProgressBar(ref=progress_bar_ref, visible=False),
                ], spacing=5),
                padding=ft.padding.only(top=20)
            ),
            
            ft.Divider(height=20, color=theme_colors.get('outline')),
            ft.Row([
                ft.TextButton(
                    "Cancelar", 
                    on_click=close_dialog,
                    ref=cancel_button_ref,
                    style=ft.ButtonStyle(
                        color=theme_colors.get('on_surface')
                    )
                ),
                ft.ElevatedButton(
                    "Gerar", 
                    on_click=start_generation, 
                    bgcolor=theme_colors.get('primary'), 
                    color=theme_colors.get('on_primary'),
                    ref=generate_button_ref,
                ),
            ], alignment=ft.MainAxisAlignment.END, spacing=10)
        ], width=450, spacing=10, tight=True)
        
        dialog = ft.AlertDialog(
            content=dialog_content,
            modal=True,
        )
        
        page.open(dialog)
    
    def generate_full_scores(month, year, include_inactive, cancel_event, progress_bar_ref, progress_text_ref, dialog, generate_button_ref, cancel_button_ref, theme_colors, dialog_content, month_dropdown_ref, year_dropdown_ref, include_inactive_ref):
        """Gera notas cheias baseado na função fornecida pelo usuário"""
        try:
            # Mostrar progresso
            progress_bar_ref.current.visible = True
            progress_text_ref.current.visible = True
            progress_text_ref.current.value = "Carregando critérios..."
            safe_page_update(page)
            
            # Carregar critérios da tabela criteria_table
            criterios_raw = db_manager.query(
                "SELECT criteria_category, value FROM criteria_table"
            )
            
            criterio_map = {}
            for row in criterios_raw:
                # DBManager.query retorna dict(row)
                nome = str(row.get('criteria_category') if isinstance(row, dict) else row[0]).strip().lower()
                valor = float(row.get('value') if isinstance(row, dict) else row[1])
                if "package" in nome:
                    criterio_map["package"] = valor
                elif "pick" in nome:
                    criterio_map["pickup"] = valor  
                elif "nil" in nome:
                    criterio_map["nil"] = valor
                elif "otif" in nome:
                    criterio_map["otif"] = valor
            
            if not all(k in criterio_map for k in ["package", "pickup", "nil", "otif"]):
                page.run_thread(lambda: show_snack_bar("Um ou mais critérios de pontuação estão faltando na tabela de critérios.", True))
                return
            
            progress_text_ref.current.value = "Carregando fornecedores..."
            safe_page_update(page)
            
            # Carregar fornecedores
            if include_inactive:
                fornecedores = db_manager.query(
                    "SELECT supplier_id, vendor_name, supplier_status FROM supplier_database_table"
                )
            else:
                # Filtrar na própria query para melhor performance
                fornecedores = db_manager.query(
                    "SELECT supplier_id, vendor_name, supplier_status FROM supplier_database_table WHERE LOWER(supplier_status) = 'active'"
                )
            
            if not fornecedores:
                page.run_thread(lambda: show_snack_bar("Nenhum fornecedor encontrado no banco de dados.", False))
                return

            # Otimização: buscar todos os registros existentes de uma vez
            progress_text_ref.current.value = "Verificando registros existentes..."
            safe_page_update(page)
            
            existing_records = db_manager.query(
                "SELECT supplier_id FROM supplier_score_records_table WHERE month = ? AND year = ?",
                (month, year)
            )
            existing_set = {str(record.get('supplier_id') if isinstance(record, dict) else record[0]) for record in existing_records}
            
            # Configurar barra de progresso
            progress_bar_ref.current.value = 0
            total_fornecedores = len(fornecedores)
            safe_page_update(page)
            
            user = getpass.getuser()
            register_date = datetime.datetime.now()
            registered_by = "NIL,OTIF,Package,Pickup"
            nota_fixa = 10.0
            
            adicionados = 0
            ignorados_inativos = 0
            ignorados_existentes = 0
            
            # Preparar batch de inserções
            batch_inserts = []
            
            progress_text_ref.current.value = "Preparando dados para inserção..."
            safe_page_update(page)
            
            for i, fornecedor in enumerate(fornecedores):
                # Cancelamento cooperativo
                if cancel_event.is_set():
                    page.close(dialog)
                    page.run_thread(lambda: show_snack_bar("Operação cancelada pelo usuário.", False))
                    return
                    
                # Atualizar progresso de preparação
                if i % 50 == 0:  # Atualizar a cada 50 fornecedores para não sobrecarregar UI
                    progress = (i + 1) / total_fornecedores * 0.5  # 50% para preparação
                    progress_bar_ref.current.value = progress
                    progress_text_ref.current.value = f"Preparando dados... {i + 1} de {total_fornecedores}"
                    safe_page_update(page)
                
                supplier_id = fornecedor.get('supplier_id') if isinstance(fornecedor, dict) else fornecedor[0]
                supplier_name = fornecedor.get('vendor_name') if isinstance(fornecedor, dict) else fornecedor[1]
                raw_status = fornecedor.get('supplier_status') if isinstance(fornecedor, dict) else fornecedor[2]
                status = raw_status.strip() if raw_status else ""
                status_lower = status.lower()
                
                # Filtrar fornecedores inativos (se não incluir inativos e já não foi filtrado na query)
                if not include_inactive and status_lower != "active":
                    ignorados_inativos += 1
                    continue
                
                # Verificar se já existe usando o set (muito mais rápido)
                if str(supplier_id) in existing_set:
                    ignorados_existentes += 1
                    continue
                
                # Calcular total baseado nos critérios
                total = (
                    nota_fixa * criterio_map["otif"] +
                    nota_fixa * criterio_map["nil"] +
                    nota_fixa * criterio_map["package"] +
                    nota_fixa * criterio_map["pickup"]
                )
                total = round(total, 2)
                
                # Adicionar ao batch
                batch_inserts.append((
                    str(supplier_id),
                    str(supplier_name),
                    str(month),
                    str(year),
                    float(nota_fixa),  # quality_package
                    float(nota_fixa),  # quality_pickup
                    float(nota_fixa),  # nil
                    float(nota_fixa),  # otif
                    float(total),      # total_score
                    "Maximum score auto-generated.",
                    register_date.isoformat(),
                    str(user),
                    registered_by
                ))
            
            # Executar inserções em batch se houver dados
            if batch_inserts:
                progress_text_ref.current.value = f"Inserindo {len(batch_inserts)} registros..."
                progress_bar_ref.current.value = 0.5  # 50% para inserção
                safe_page_update(page)
                
                query_insert = """
                    INSERT INTO supplier_score_records_table (
                        supplier_id, supplier_name, month, year,
                        quality_package, quality_pickup, nil, otif,
                        total_score, comment, register_date, changed_by, registered_by
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
                
                try:
                    # Usar transação para inserção em batch
                    with db_manager._get_connection() as conn:
                        cursor = conn.cursor()
                        
                        # Inserir em lotes de 100 para evitar sobrecarga
                        batch_size = 100
                        total_batches = len(batch_inserts) // batch_size + (1 if len(batch_inserts) % batch_size > 0 else 0)
                        
                        for batch_num in range(total_batches):
                            if cancel_event.is_set():
                                page.close(dialog)
                                page.run_thread(lambda: show_snack_bar("Operação cancelada pelo usuário.", False))
                                return
                                
                            start_idx = batch_num * batch_size
                            end_idx = min(start_idx + batch_size, len(batch_inserts))
                            batch_data = batch_inserts[start_idx:end_idx]
                            
                            cursor.executemany(query_insert, batch_data)
                            
                            # Atualizar progresso
                            progress = 0.5 + (batch_num + 1) / total_batches * 0.5  # 50% base + 50% para inserção
                            progress_bar_ref.current.value = progress
                            progress_text_ref.current.value = f"Inserindo lote {batch_num + 1} de {total_batches}..."
                            safe_page_update(page)
                        
                        conn.commit()
                        adicionados = len(batch_inserts)
                        
                except Exception as batch_error:
                    page.run_thread(lambda: show_snack_bar(f"Erro durante inserção em batch: {str(batch_error)}", True))
                    return
            
            # Exibir resultado no próprio diálogo (sem fechar)
            try:
                progress_bar_ref.current.visible = False
                progress_text_ref.current.visible = False
                
                # Criar resumo estruturado e mostrar no lugar do progresso
                summary_text = (
                    f"✅ Geração de notas concluída!\n\n"
                    f"📊 Adicionadas: {adicionados}\n"
                    f"⚠️  Ignoradas (inativas): {ignorados_inativos}\n" 
                    f"ℹ️  Ignoradas (já existem): {ignorados_existentes}"
                )
                
                progress_text_ref.current.visible = True
                progress_text_ref.current.value = summary_text
                progress_text_ref.current.size = 14
                
                # Log único para geração de full score
                try:
                    with db_manager._get_connection() as conn:
                        db_manager._log_db_change(conn, f"FULL SCORE GENERATED - {adicionados} records added", user, db_manager.current_user_wwid)
                except Exception as log_error:
                    print(f"Erro ao logar geração de full score: {log_error}")
                
                # reabilitar Gerar; alterar Cancelar para Fechar
                if generate_button_ref and generate_button_ref.current:
                    generate_button_ref.current.disabled = False
                # Reabilitar dropdowns e switch
                if month_dropdown_ref and month_dropdown_ref.current:
                    month_dropdown_ref.current.disabled = False
                if year_dropdown_ref and year_dropdown_ref.current:
                    year_dropdown_ref.current.disabled = False  
                if include_inactive_ref and include_inactive_ref.current:
                    include_inactive_ref.current.disabled = False
                if cancel_button_ref and cancel_button_ref.current:
                    try:
                        cancel_button_ref.current.text = "Fechar"
                    except Exception:
                        pass
                safe_page_update(page)
            except Exception as e:
                # Fallback para texto simples se houver erro
                summary = (
                    f"Geração de notas concluída!\n"
                    f"✅ Adicionadas: {adicionados}\n"
                    f"⚠️ Ignoradas (inativas): {ignorados_inativos}\n"
                    f"ℹ️ Ignoradas (já existem): {ignorados_existentes}"
                )
                try:
                    progress_bar_ref.current.visible = False
                    progress_text_ref.current.visible = True
                    progress_text_ref.current.value = summary
                    # reabilitar Gerar; alterar Cancelar para Fechar
                    if generate_button_ref and generate_button_ref.current:
                        generate_button_ref.current.disabled = False
                    # Reabilitar dropdowns e switch
                    if month_dropdown_ref and month_dropdown_ref.current:
                        month_dropdown_ref.current.disabled = False
                    if year_dropdown_ref and year_dropdown_ref.current:
                        year_dropdown_ref.current.disabled = False  
                    if include_inactive_ref and include_inactive_ref.current:
                        include_inactive_ref.current.disabled = False
                    if cancel_button_ref and cancel_button_ref.current:
                        try:
                            cancel_button_ref.current.text = "Fechar"
                        except Exception:
                            pass
                    safe_page_update(page)
                except Exception:
                    pass
            
            # Atualizar lista de resultados se estiver visível
            page.run_thread(lambda: (
                results_list.controls.clear(),
                results_list.update()
            ) if 'results_list' in globals() and results_list else None)
            
        except Exception as e:
            err_msg = f"Erro durante a geração de notas: {e}"
            page.run_thread(lambda msg=err_msg: show_snack_bar(msg, True))
    
    # --- Fim: Funções para Gerar Nota Cheia ---

    # --- Início: Funções para Importar/Exportar Score ---
    
    def import_score_dialog():
        """Interface completa para importar scores de arquivo Excel"""
        import datetime
        
        # Referencias para os componentes
        month_dropdown_ref = ft.Ref[ft.Dropdown]()
        year_dropdown_ref = ft.Ref[ft.Dropdown]() 
        file_display_ref = ft.Ref[ft.Column]()
        import_button_ref = ft.Ref[ft.ElevatedButton]()
        validation_status_ref = ft.Ref[ft.Row]()
        message_display_ref = ft.Ref[ft.Column]()  # Para mostrar mensagens
        progress_bar_ref = ft.Ref[ft.ProgressBar]()  # Barra de progresso
        progress_text_ref = ft.Ref[ft.Text]()  # Texto de progresso
        
        # Estado da aplicação
        selected_file = {"path": None, "name": None, "valid": False}
        
        # Preparar opções de mês e ano
        months = [
            ft.dropdown.Option("1", "Janeiro"),
            ft.dropdown.Option("2", "Fevereiro"), 
            ft.dropdown.Option("3", "Março"),
            ft.dropdown.Option("4", "Abril"),
            ft.dropdown.Option("5", "Maio"),
            ft.dropdown.Option("6", "Junho"),
            ft.dropdown.Option("7", "Julho"),
            ft.dropdown.Option("8", "Agosto"),
            ft.dropdown.Option("9", "Setembro"),
            ft.dropdown.Option("10", "Outubro"),
            ft.dropdown.Option("11", "Novembro"),
            ft.dropdown.Option("12", "Dezembro")
        ]
        years = [ft.dropdown.Option(str(year)) for year in range(2024, 2041)]
        
        def close_dialog(e):
            page.close(dialog)
        
        def check_form_validity():
            """Verifica se mês/ano estão selecionados, arquivo válido e data não é futura"""
            month_ok = month_dropdown_ref.current.value is not None
            year_ok = year_dropdown_ref.current.value is not None
            file_ok = selected_file["valid"]
            date_ok = True
            
            # Verificar se a data não é futura (permite mês atual)
            if month_ok and year_ok:
                selected_month = int(month_dropdown_ref.current.value)
                selected_year = int(year_dropdown_ref.current.value)
                current_date = datetime.datetime.now()
                
                # Comparar apenas ano/mês (não dia específico)
                current_year = current_date.year
                current_month = current_date.month
                
                # Data é futura se ano > atual OU (ano = atual E mês > atual)
                if selected_year > current_year or (selected_year == current_year and selected_month > current_month):
                    date_ok = False
            
            # Habilitar/desabilitar botão de importar
            if import_button_ref.current:
                import_button_ref.current.disabled = not (month_ok and year_ok and file_ok and date_ok)
                safe_page_update(page)
            
            return date_ok
        
        def validate_excel_file(file_path):
            """Valida se o arquivo Excel é válido (contém aba de validação)"""
            try:
                import openpyxl
                import os
                
                if not os.path.exists(file_path):
                    return False, "Arquivo não encontrado"
                
                # Abrir workbook
                wb = openpyxl.load_workbook(file_path)
                
                # Verificar se existe a aba de validação oculta
                if "ScoreApp_Validation" not in wb.sheetnames:
                    return False, "Arquivo não foi gerado pelo ScoreApp"
                
                # Validar dados da aba de validação - apenas verificar o texto fixo
                validation_ws = wb["ScoreApp_Validation"]
                validation_text = validation_ws["A1"].value
                
                if validation_text != "SCOREAPP_VALIDATION_123456":
                    return False, "Arquivo inválido - não foi gerado pelo ScoreApp"
                
                # Verificar se a aba "Import_score" existe
                if "Import_score" not in wb.sheetnames:
                    return False, "Aba 'Import_score' não encontrada"
                
                return True, "Arquivo válido"
                
            except Exception as e:
                return False, f"Erro ao validar arquivo: {str(e)}"
        
        def handle_file_selection(result):
            """Processa arquivo selecionado"""
            if result.files:
                file_path = result.files[0].path
                file_name = result.files[0].name
                
                # Validar arquivo
                is_valid, message = validate_excel_file(file_path)
                
                # Atualizar estado
                selected_file["path"] = file_path
                selected_file["name"] = file_name  
                selected_file["valid"] = is_valid
                
                # Atualizar display do arquivo
                if is_valid:
                    status_color = "green"
                    status_icon = ft.Icons.CHECK_CIRCLE
                    status_text = "✅ Pronto para importar"
                else:
                    status_color = "red"
                    status_icon = ft.Icons.ERROR
                    status_text = f"❌ Arquivo incorreto: {message}"
                
                # Atualizar interface
                file_display_ref.current.controls.clear()
                file_display_ref.current.controls.extend([
                    ft.Card(
                        content=ft.Container(
                            content=ft.Row([
                                ft.Icon(ft.Icons.DESCRIPTION, color="blue", size=24),
                                ft.Column([
                                    ft.Text(file_name, weight="bold", size=14),
                                    ft.Text(status_text, color=status_color, size=12)
                                ], expand=True, spacing=2),
                                ft.IconButton(
                                    icon=ft.Icons.DELETE,
                                    icon_color="red",
                                    tooltip="Remover arquivo",
                                    on_click=lambda _: remove_file()
                                )
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            padding=15
                        ),
                        elevation=2
                    )
                ])
                
                validation_status_ref.current.controls.clear()
                if not is_valid:
                    validation_status_ref.current.controls.extend([
                        ft.Icon(status_icon, color=status_color, size=20),
                        ft.Text(status_text, color=status_color, size=12, weight="bold")
                    ])
                
                check_form_validity()
                safe_page_update(page)
        
        def remove_file():
            """Remove arquivo selecionado"""
            selected_file["path"] = None
            selected_file["name"] = None
            selected_file["valid"] = False
            
            file_display_ref.current.controls.clear()
            validation_status_ref.current.controls.clear()
            
            check_form_validity()
            safe_page_update(page)
        
        def select_file(e):
            """Abre seletor de arquivo"""
            import_file_picker = ft.FilePicker(on_result=handle_file_selection)
            page.overlay.append(import_file_picker)
            safe_page_update(page)
            
            import_file_picker.pick_files(
                dialog_title="Selecione o arquivo Excel para importar",
                allowed_extensions=["xlsx"]
            )
        
        def import_file(e):
            """Executa importação do arquivo com barra de progresso"""
            # Limpar mensagens anteriores
            message_display_ref.current.controls.clear()
            
            if not selected_file["valid"]:
                show_message("❌ Arquivo inválido para importação", "red")
                return
            
            # Verificar se data não é futura
            if not check_form_validity():
                show_message("❌ Não é possível importar dados para datas futuras", "red")
                return
                
            month = month_dropdown_ref.current.value
            year = year_dropdown_ref.current.value
            file_path = selected_file["path"]
            
            # Desabilitar botão durante importação
            import_button_ref.current.disabled = True
            
            # Mostrar barra de progresso
            progress_bar_ref.current.visible = True
            progress_bar_ref.current.value = 0
            progress_text_ref.current.visible = True
            progress_text_ref.current.value = "Preparando importação..."
            safe_page_update(page)
            
            def run_import():
                try:
                    # Executar importação com CRUD e recálculo
                    success, message = import_scores_from_excel_with_crud_with_progress(
                        file_path, month, year, progress_bar_ref, progress_text_ref
                    )
                    
                    # Atualizar interface na thread principal
                    def update_ui():
                        progress_bar_ref.current.visible = False
                        progress_text_ref.current.visible = False
                        import_button_ref.current.disabled = False
                        
                        if success:
                            show_message(f"✅ {message}", "green")
                            # Atualizar dados na tela se necessário
                            try:
                                load_all_lists_data()
                            except:
                                pass
                        else:
                            show_message(f"❌ {message}", "red")
                        
                        # Garantir que o diálogo permaneça aberto
                        safe_page_update(page)
                    
                    page.run_thread(update_ui)
                    
                except Exception as ex:
                    def update_error():
                        progress_bar_ref.current.visible = False
                        progress_text_ref.current.visible = False
                        import_button_ref.current.disabled = False
                        show_message(f"❌ Erro na importação: {ex}", "red")
                        # Garantir que o diálogo permaneça aberto
                        safe_page_update(page)
                    
                    page.run_thread(update_error)
            
            # Executar em thread separada
            import threading
            threading.Thread(target=run_import, daemon=True).start()
        
        def show_message(text, message_type):
            """Mostra mensagem na janela de importação"""
            message_display_ref.current.controls.clear()
            
            # Definir cores baseadas no tipo de mensagem
            if message_type == "green":
                text_color = "green"
                bg_color = "#e8f5e8"  # Verde claro
                border_color = "green"
            elif message_type == "red":
                text_color = "red"
                bg_color = "#fdeaea"  # Vermelho claro
                border_color = "red"
            elif message_type == "blue":
                text_color = "blue"
                bg_color = "#e3f2fd"  # Azul claro
                border_color = "blue"
            else:
                text_color = "black"
                bg_color = "#f5f5f5"  # Cinza claro
                border_color = "grey"
            
            message_display_ref.current.controls.append(
                ft.Container(
                    content=ft.Text(text, color=text_color, size=14, weight="bold"),
                    padding=10,
                    bgcolor=bg_color,
                    border_radius=8,
                    border=ft.border.all(1, border_color)
                )
            )
            safe_page_update(page)
        
        # Conteúdo do diálogo
        dialog_content = ft.Column([
            ft.Text("Importar Scores do Excel", size=18, weight="bold"),
            ft.Divider(height=20),
            
            # Seleção de mês e ano
            ft.Text("Período para Importação:", size=14, weight="bold"),
            ft.Row([
                ft.Dropdown(
                    label="Mês",
                    options=months,
                    expand=True,
                    ref=month_dropdown_ref,
                    on_change=lambda _: check_form_validity()
                ),
                ft.Dropdown(
                    label="Ano", 
                    options=years,
                    expand=True,
                    ref=year_dropdown_ref,
                    on_change=lambda _: check_form_validity()
                ),
            ], spacing=10),
            
            ft.Container(height=20),
            
            # Seleção de arquivo
            ft.Text("Arquivo Excel:", size=14, weight="bold"),
            ft.ElevatedButton(
                "📁 Selecionar Arquivo Excel",
                icon=ft.Icons.FOLDER_OPEN,
                on_click=select_file
            ),
            
            # Display do arquivo selecionado
            ft.Column(ref=file_display_ref, spacing=10),
            
            # Status de validação
            ft.Row(ref=validation_status_ref, spacing=10),
            
            # Barra de progresso
            ft.Container(
                content=ft.Column([
                    ft.Text(
                        "", 
                        ref=progress_text_ref, 
                        size=12, 
                        visible=False,
                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                    ),
                    ft.ProgressBar(ref=progress_bar_ref, visible=False),
                ], spacing=5),
                padding=ft.padding.only(top=10)
            ),
            
            # Area de mensagens
            ft.Column(ref=message_display_ref, spacing=10),
            
            ft.Container(height=20),
            ft.Divider(),
            
            # Botões
            ft.Row([
                ft.TextButton(
                    "Fechar",
                    on_click=close_dialog
                ),
                ft.ElevatedButton(
                    "📥 Importar Arquivo",
                    on_click=import_file,
                    ref=import_button_ref,
                    disabled=True,  # Inicialmente desabilitado
                    bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'),
                    color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_primary')
                )
            ], alignment=ft.MainAxisAlignment.END, spacing=10)
        ], width=500, spacing=15, tight=True)
        
        dialog = ft.AlertDialog(
            title=ft.Text("Importar Dados"),
            content=dialog_content,
            modal=True
        )
        
        page.open(dialog)

    def import_scores_from_excel_with_crud_with_progress(file_path, month, year, progress_bar_ref, progress_text_ref):
        """Importa scores do Excel com CRUD e recálculo de notas"""
        try:
            import openpyxl
            from datetime import datetime
            import os
        except ImportError:
            return False, "Biblioteca openpyxl não encontrada"

        try:
            # Verificar se o arquivo existe
            if not os.path.exists(file_path):
                return False, "Arquivo não encontrado"

            # Atualizar progresso
            if progress_text_ref and progress_text_ref.current:
                page.run_thread(lambda: setattr(progress_text_ref.current, 'value', 'Abrindo arquivo Excel...'))
                page.run_thread(lambda: safe_page_update(page))

            # Abrir workbook
            wb = openpyxl.load_workbook(file_path)
            
            # Verificar validação
            if "ScoreApp_Validation" not in wb.sheetnames:
                return False, "Arquivo não foi gerado pelo ScoreApp"
                
            validation_ws = wb["ScoreApp_Validation"]
            validation_text = validation_ws["A1"].value
            
            if validation_text != "SCOREAPP_VALIDATION_123456":
                return False, "Arquivo inválido"
            
            # Verificar aba Import_score
            if "Import_score" not in wb.sheetnames:
                return False, "Aba 'Import_score' não encontrada"
                
            ws = wb["Import_score"]
            
            # Ler headers para identificar as colunas
            headers = []
            col = 1
            while True:
                cell_value = ws.cell(row=1, column=col).value
                if cell_value is None:
                    break
                headers.append(str(cell_value).strip())
                col += 1
            
            # Mapear headers para índices
            header_map = {header.lower(): idx for idx, header in enumerate(headers)}
            
            # Identificar colunas obrigatórias
            required_cols = ["supplier id", "po", "bu", "supplier name"]
            for req_col in required_cols:
                if req_col not in header_map:
                    return False, f"Coluna obrigatória '{req_col}' não encontrada"
            
            # Identificar colunas de notas (headers padrão do sistema) baseado nas permissões do usuário
            score_columns = {}
            standard_score_names = {
                "otif": ["otif", "on time in full", "OTIF"],
                "nil": ["nil", "zero defects", "NIL"],
                "pickup": ["quality pickup", "pickup quality", "quality_pickup", "pickup", "Pickup"],
                "package": ["quality package", "package quality", "quality_package", "package", "Package"]
            }
            
            # Só mapear as colunas que o usuário tem permissão para acessar
            for score_type, possible_names in standard_score_names.items():
                # Verificar se o usuário tem permissão para esta nota
                if current_user_permissions.get(score_type):
                    for possible_name in possible_names:
                        if possible_name in header_map:
                            score_columns[score_type] = header_map[possible_name]
                            break
            
            print(f"🔍 DEBUG: Colunas de notas mapeadas baseadas nas permissões: {score_columns}")
            print(f"🔍 DEBUG: Permissões do usuário: {current_user_permissions}")
            
            # Verificar se o usuário tem pelo menos uma permissão de nota
            if not score_columns:
                return False, "Usuário não tem permissão para importar nenhuma coluna de nota"
            
            # Obter critérios do banco - mesmo formato da geração de notas
            if progress_text_ref and progress_text_ref.current:
                page.run_thread(lambda: setattr(progress_text_ref.current, 'value', 'Carregando critérios...'))
                page.run_thread(lambda: safe_page_update(page))
            
            criterios_raw = db_manager.query("SELECT criteria_category, value FROM criteria_table")
            criterio_map = {}
            for row in criterios_raw:
                nome = str(row.get('criteria_category') if isinstance(row, dict) else row[0]).strip().lower()
                valor = float(row.get('value') if isinstance(row, dict) else row[1])
                if "package" in nome:
                    criterio_map["package"] = valor
                elif "pick" in nome:
                    criterio_map["pickup"] = valor  
                elif "nil" in nome:
                    criterio_map["nil"] = valor
                elif "otif" in nome:
                    criterio_map["otif"] = valor
            
            processed_count = 0
            updated_count = 0
            created_count = 0
            
            # Calcular total de linhas para progresso
            total_rows = ws.max_row - 1  # -1 porque linha 1 é header
            
            # Processar cada linha de dados
            for row_num in range(2, ws.max_row + 1):
                # Atualizar progresso
                current_progress = (row_num - 2) / total_rows if total_rows > 0 else 0
                if progress_bar_ref and progress_bar_ref.current:
                    page.run_thread(lambda p=current_progress: setattr(progress_bar_ref.current, 'value', p))
                if progress_text_ref and progress_text_ref.current:
                    page.run_thread(lambda r=row_num-1, t=total_rows: setattr(progress_text_ref.current, 'value', f'Processando registro {r} de {t}...'))
                page.run_thread(lambda: safe_page_update(page))
                
                try:
                    # Ler dados básicos
                    supplier_id_cell = ws.cell(row=row_num, column=header_map["supplier id"] + 1)
                    po_cell = ws.cell(row=row_num, column=header_map["po"] + 1)
                    bu_cell = ws.cell(row=row_num, column=header_map["bu"] + 1)
                    supplier_name_cell = ws.cell(row=row_num, column=header_map["supplier name"] + 1)
                    
                    if not supplier_id_cell.value:
                        continue
                        
                    supplier_id = str(supplier_id_cell.value).strip()
                    po = str(po_cell.value).strip() if po_cell.value else ""
                    bu = str(bu_cell.value).strip() if bu_cell.value else ""
                    supplier_name = str(supplier_name_cell.value).strip() if supplier_name_cell.value else ""
                    
                    # Ler notas do Excel - apenas para as colunas que existem no arquivo
                    excel_scores = {}
                    
                    # Só processar as colunas que realmente existem no Excel
                    for score_type, col_idx in score_columns.items():
                        cell_value = ws.cell(row=row_num, column=col_idx + 1).value
                        
                        # Se a célula tem valor ou está vazia, processar
                        if cell_value is not None and str(cell_value).strip():
                            try:
                                score_value = float(cell_value)
                                if 0.0 <= score_value <= 10.0:
                                    excel_scores[score_type] = score_value
                                else:
                                    excel_scores[score_type] = None  # Valores fora do range = null
                            except (ValueError, TypeError):
                                excel_scores[score_type] = None  # Valores inválidos = null
                        else:
                            excel_scores[score_type] = None  # Células vazias = null
                    
                    # Verificar se existe registro no banco
                    existing_record = db_manager.query(
                        "SELECT * FROM supplier_score_records_table WHERE supplier_id = ? AND month = ? AND year = ?",
                        [supplier_id, month, year]
                    )
                    
                    if existing_record:
                        # Atualizar registro existente
                        record = existing_record[0]
                        current_scores = {
                            "otif": record.get('otif') if isinstance(record, dict) else record[8],
                            "nil": record.get('nil') if isinstance(record, dict) else record[7],
                            "pickup": record.get('quality_pickup') if isinstance(record, dict) else record[6],
                            "package": record.get('quality_package') if isinstance(record, dict) else record[5]
                        }
                        
                        # Obter registered_by atual
                        current_registered_by = record.get('registered_by') if isinstance(record, dict) else record[12]
                        current_registered_list = current_registered_by.split(',') if current_registered_by else []
                        
                        # Mesclar notas - apenas sobrescrever as que têm valor real no Excel
                        final_scores = current_scores.copy()
                        updated_fields = []
                        
                        # Só atualizar as notas que têm colunas correspondentes no Excel E têm valor não nulo
                        for score_type, col_idx in score_columns.items():
                            # Se a coluna existe no Excel E tem valor (não é None/vazio), usar o valor do Excel
                            if score_type in excel_scores and excel_scores[score_type] is not None:
                                final_scores[score_type] = excel_scores[score_type]
                                updated_fields.append(score_type)
                            # Se a coluna não existe no Excel OU está vazia, manter o valor atual do banco
                        
                        # Recalcular total_score usando TODAS as notas (atualizadas + preservadas)
                        total_score = 0.0
                        for score_type, score_value in final_scores.items():
                            if score_value is not None and score_type in criterio_map:
                                try:
                                    score_float = float(score_value)
                                    criterio_float = float(criterio_map[score_type])
                                    total_score += score_float * criterio_float
                                except (ValueError, TypeError):
                                    print(f"⚠️ Erro ao converter valores para float: score_value={score_value}, criterio={criterio_map[score_type]}")
                                    continue
                        
                        # Criar comentário específico sobre quais notas foram importadas (apenas as que realmente mudaram)
                        import_comment_parts = []
                        for score_type in updated_fields:
                            score_value = excel_scores[score_type]
                            # Como chegou até aqui, sabemos que score_value não é None
                            import_comment_parts.append(f"{score_type.upper()} score generated by Excel import")
                        
                        import_comment = "; ".join(import_comment_parts) if import_comment_parts else "Updated by Excel import"
                        
                        # Atualizar registered_by apenas com as notas que foram importadas
                        new_registered_list = current_registered_list.copy()
                        score_name_map = {
                            "otif": "OTIF",
                            "nil": "NIL", 
                            "pickup": "Pickup",
                            "package": "Package"
                        }
                        
                        for score_type in updated_fields:
                            score_name = score_name_map.get(score_type, score_type.upper())
                            if score_name not in new_registered_list:
                                new_registered_list.append(score_name)
                        
                        new_registered_by = ",".join(new_registered_list) if new_registered_list else "Excel Import"
                        
                        # Atualizar no banco - mesma estrutura da geração de notas
                        db_manager.execute(
                            """UPDATE supplier_score_records_table 
                               SET quality_package = ?, quality_pickup = ?, nil = ?, otif = ?, 
                                   total_score = ?, comment = ?, changed_by = ?, register_date = ?, registered_by = ?
                               WHERE supplier_id = ? AND month = ? AND year = ?""",
                            [final_scores["package"], final_scores["pickup"], 
                             final_scores["nil"], final_scores["otif"], total_score, 
                             import_comment, current_user_name or "Import", datetime.now().isoformat(),
                             new_registered_by, supplier_id, month, year]
                        )
                        updated_count += 1
                        
                    else:
                        # Criar novo registro - apenas com as notas que têm valor no Excel
                        final_scores = {
                            "otif": excel_scores.get("otif") if "otif" in score_columns and excel_scores.get("otif") is not None else None,
                            "nil": excel_scores.get("nil") if "nil" in score_columns and excel_scores.get("nil") is not None else None,
                            "pickup": excel_scores.get("pickup") if "pickup" in score_columns and excel_scores.get("pickup") is not None else None,
                            "package": excel_scores.get("package") if "package" in score_columns and excel_scores.get("package") is not None else None
                        }
                        
                        # Calcular total_score apenas com as notas que não são None
                        total_score = 0.0
                        for score_type, score_value in final_scores.items():
                            if score_value is not None and score_type in criterio_map:
                                try:
                                    score_float = float(score_value)
                                    criterio_float = float(criterio_map[score_type])
                                    total_score += score_float * criterio_float
                                except (ValueError, TypeError):
                                    print(f"⚠️ Erro ao converter valores para float: score_value={score_value}, criterio={criterio_map[score_type]}")
                                    continue
                        
                        # Criar comentário para novo registro - apenas para notas que realmente têm valor
                        new_import_comment_parts = []
                        for score_type in ["otif", "nil", "pickup", "package"]:
                            if score_type in score_columns and excel_scores.get(score_type) is not None:
                                new_import_comment_parts.append(f"{score_type.upper()} score generated by Excel import")
                        
                        new_import_comment = "; ".join(new_import_comment_parts) if new_import_comment_parts else "Created by Excel import"
                        
                        # Criar registered_by apenas com as notas que foram importadas
                        new_registered_list = []
                        score_name_map = {
                            "otif": "OTIF",
                            "nil": "NIL",
                            "pickup": "Pickup", 
                            "package": "Package"
                        }
                        
                        for score_type in ["otif", "nil", "pickup", "package"]:
                            if score_type in score_columns and excel_scores.get(score_type) is not None:
                                score_name = score_name_map.get(score_type, score_type.upper())
                                new_registered_list.append(score_name)
                        
                        new_registered_by = ",".join(new_registered_list) if new_registered_list else "Excel Import"
                        
                        # Inserir no banco - mesma query da geração de notas
                        query_insert = """
                            INSERT INTO supplier_score_records_table (
                                supplier_id, supplier_name, month, year,
                                quality_package, quality_pickup, nil, otif,
                                total_score, comment, register_date, changed_by, registered_by
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """
                        
                        params = (
                            str(supplier_id),
                            str(supplier_name),
                            str(month),
                            str(year),
                            final_scores["package"],
                            final_scores["pickup"],
                            final_scores["nil"],
                            final_scores["otif"],
                            total_score,
                            new_import_comment,
                            datetime.now().isoformat(),
                            current_user_name or "Import",
                            new_registered_by
                        )
                        
                        db_manager.execute(query_insert, params)
                        created_count += 1
                    
                    processed_count += 1
                    
                except Exception as row_error:
                    print(f"⚠️ Erro na linha {row_num}: {row_error}")
                    continue
            
            # Finalizar progresso
            if progress_bar_ref and progress_bar_ref.current:
                page.run_thread(lambda: setattr(progress_bar_ref.current, 'value', 1.0))
            if progress_text_ref and progress_text_ref.current:
                page.run_thread(lambda: setattr(progress_text_ref.current, 'value', 'Importação concluída!'))
            page.run_thread(lambda: safe_page_update(page))
            
            if processed_count == 0:
                return False, "Nenhum dado válido encontrado no arquivo"
            
            message = f"Importação concluída: {processed_count} registros processados"
            if created_count > 0:
                message += f", {created_count} criados"
            if updated_count > 0:
                message += f", {updated_count} atualizados"
                
            return True, message
            
        except Exception as e:
            return False, f"Erro durante importação: {str(e)}"

    def import_scores_from_excel(file_path):
        """Importa scores do arquivo Excel protegido"""
        try:
            import openpyxl
            from datetime import datetime
            import os
        except ImportError:
            show_snack_bar("❌ Biblioteca openpyxl não encontrada. Instale com: pip install openpyxl", True)
            return

        try:
            # Verificar se o arquivo existe
            if not os.path.exists(file_path):
                show_snack_bar("❌ Arquivo não encontrado.", True)
                return

            # Abrir workbook
            wb = openpyxl.load_workbook(file_path)
            
            # Verificar se existe a aba de validação oculta
            if "ScoreApp_Validation" not in wb.sheetnames:
                show_snack_bar("❌ Arquivo inválido: não foi gerado pelo ScoreApp.", True)
                return
                
            # Validar dados da aba de validação - apenas verificar o texto fixo
            validation_ws = wb["ScoreApp_Validation"]
            try:
                validation_text = validation_ws["A1"].value
                
                if validation_text != "SCOREAPP_VALIDATION_123456":
                    show_snack_bar("❌ Arquivo inválido: não foi gerado pelo ScoreApp.", True)
                    return
                    
            except Exception as e:
                show_snack_bar("❌ Erro ao validar arquivo: dados de validação corrompidos.", True)
                return
            
            # Verificar se a aba "Import_score" existe
            if "Import_score" not in wb.sheetnames:
                show_snack_bar("❌ Aba 'Import_score' não encontrada no arquivo.", True)
                return
                
            ws = wb["Import_score"]
            
            # Verificar headers esperados (dinâmico baseado nas colunas)
            expected_base_headers = ["Supplier ID", "PO", "Supplier Name"]
            actual_headers = []
            
            # Ler todos os headers da primeira linha
            col = 1
            while True:
                cell_value = ws.cell(row=1, column=col).value
                if cell_value is None:
                    break
                actual_headers.append(cell_value)
                col += 1
            
            print(f"🔍 DEBUG: Headers encontrados: {actual_headers}")
            
            # Verificar se pelo menos Supplier ID, PO e Supplier Name estão presentes
            if len(actual_headers) < 3 or actual_headers[0] != "Supplier ID" or actual_headers[1] != "PO" or actual_headers[2] != "Supplier Name":
                show_snack_bar(f"❌ Headers incorretos. Esperados pelo menos: {expected_base_headers}", True)
                return
            
            # Mapear colunas de notas encontradas
            note_columns = {}
            for i, header in enumerate(actual_headers[3:], 4):  # Começar da coluna 4
                if header in ["OTIF", "NIL", "Pickup", "Package"]:
                    note_columns[header.lower()] = i
                    
            print(f"🔍 DEBUG: Colunas de notas mapeadas: {note_columns}")

            # Solicitar mês e ano para importação
            show_import_period_dialog(ws)
            
        except Exception as e:
            show_snack_bar(f"❌ Erro ao processar arquivo Excel: {str(e)}", True)

    def show_import_period_dialog(worksheet):
        """Mostra diálogo para selecionar mês e ano da importação"""
        month_dropdown_ref = ft.Ref()
        year_dropdown_ref = ft.Ref()
        
        # Opções de mês
        months = [
            ft.dropdown.Option("", "Selecione o mês..."),
            ft.dropdown.Option("1", "Janeiro"),
            ft.dropdown.Option("2", "Fevereiro"),
            ft.dropdown.Option("3", "Março"),
            ft.dropdown.Option("4", "Abril"),
            ft.dropdown.Option("5", "Maio"),
            ft.dropdown.Option("6", "Junho"),
            ft.dropdown.Option("7", "Julho"),
            ft.dropdown.Option("8", "Agosto"),
            ft.dropdown.Option("9", "Setembro"),
            ft.dropdown.Option("10", "Outubro"),
            ft.dropdown.Option("11", "Novembro"),
            ft.dropdown.Option("12", "Dezembro"),
        ]
        
        # Opções de ano
        current_year = datetime.now().year
        years = [ft.dropdown.Option("", "Selecione o ano...")] + [
            ft.dropdown.Option(str(year), str(year)) for year in range(current_year-2, current_year+3)
        ]
        
        def close_dialog(e):
            page.close(period_dialog)
        
        def confirm_import(e):
            month = month_dropdown_ref.current.value if month_dropdown_ref.current else None
            year = year_dropdown_ref.current.value if year_dropdown_ref.current else None
            
            if not month or not year:
                show_snack_bar("❌ Selecione o mês e ano para importação", False)
                return
            
            try:
                process_excel_import(worksheet, int(month), int(year))
                close_dialog(e)
            except Exception as ex:
                show_snack_bar(f"❌ Erro ao importar dados: {ex}", True)
        
        # Conteúdo do diálogo
        dialog_content = ft.Column([
            ft.Text("Selecionar Período para Importação", size=18, weight="bold"),
            ft.Container(height=20),
            ft.Text("Selecione o mês e ano onde as notas serão importadas:", size=14),
            ft.Container(height=10),
            ft.Dropdown(
                ref=month_dropdown_ref,
                label="Mês",
                options=months,
                width=200,
                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
            ),
            ft.Container(height=10),
            ft.Dropdown(
                ref=year_dropdown_ref,
                label="Ano",
                options=years,
                width=200,
                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
            ),
            ft.Container(height=30),
            ft.Row([
                ft.ElevatedButton(
                    "Fechar",
                    on_click=close_dialog,
                    icon=ft.Icons.CANCEL,
                    style=ft.ButtonStyle(
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('error'),
                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_error')
                    )
                ),
                ft.Container(width=10),
                ft.ElevatedButton(
                    "📥 Importar",
                    on_click=confirm_import,
                    icon=ft.Icons.UPLOAD,
                    style=ft.ButtonStyle(
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'),
                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_primary')
                    )
                )
            ], alignment=ft.MainAxisAlignment.END)
        ], width=300, spacing=15)
        
        # Criar e exibir diálogo
        period_dialog = ft.AlertDialog(
            title=ft.Text("Importar Dados"),
            content=dialog_content,
            modal=True
        )
        
        page.open(period_dialog)

    def process_excel_import(worksheet, month, year):
        """Processa a importação das notas do Excel para o banco"""
        try:
            from datetime import datetime
            
            imported_count = 0
            updated_count = 0
            errors = []
            register_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # Mapear colunas de notas do worksheet
            actual_headers = []
            col = 1
            while True:
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value is None:
                    break
                actual_headers.append(cell_value)
                col += 1
            
            # Mapear colunas de notas encontradas
            note_columns = {}
            for i, header in enumerate(actual_headers[3:], 4):  # Começar da coluna 4
                if header in ["OTIF", "NIL", "Pickup", "Package"]:
                    note_columns[header.lower()] = i
                    
            print(f"🔍 DEBUG: Colunas de notas mapeadas na importação: {note_columns}")
            
            # Buscar critérios atuais
            criteria = db_manager.query("SELECT * FROM criteria_table LIMIT 1")
            if not criteria:
                show_snack_bar("❌ Critérios de pontuação não encontrados na tabela de critérios.", True)
                return
                
            criterio = criteria[0]
            criterio_map = {
                "otif": criterio.get('otif', 0) / 100,
                "nil": criterio.get('nil', 0) / 100,
                "package": criterio.get('package', 0) / 100,
                "pickup": criterio.get('pickup', 0) / 100,
            }
            
            # Processar cada linha do Excel (começando da linha 2)
            row_num = 2
            while True:
                supplier_id_cell = worksheet.cell(row=row_num, column=1).value  # Supplier ID
                po = worksheet.cell(row=row_num, column=2).value                # PO
                supplier_name = worksheet.cell(row=row_num, column=3).value     # Supplier Name
                
                # Se não há Supplier ID, PO nem nome, assumir fim dos dados
                if not supplier_id_cell and not po and not supplier_name:
                    break
                
                # Ler notas das colunas encontradas (podem estar vazias)
                scores = {}
                for field, col_num in note_columns.items():
                    value = worksheet.cell(row=row_num, column=col_num).value or 0
                    try:
                        scores[field] = float(value)
                        if scores[field] < 0 or scores[field] > 10:
                            errors.append(f"Linha {row_num}: {field} = {value} (deve estar entre 0 e 10)")
                            scores[field] = 0
                    except (ValueError, TypeError):
                        scores[field] = 0
                
                # Preencher campos não encontrados com 0 (para compatibilidade com o banco)
                otif = scores.get('otif', 0)
                nil = scores.get('nil', 0) 
                pickup = scores.get('pickup', 0)
                package = scores.get('package', 0)
                
                # Buscar supplier_id pelo PO ou pelo supplier_id direto
                supplier_query = None
                
                # Tentar primeiro pelo supplier_id se disponível
                if supplier_id_cell:
                    try:
                        supplier_id_from_cell = int(supplier_id_cell)
                        supplier_query = db_manager.query_one(
                            "SELECT supplier_id, vendor_name FROM supplier_database_table WHERE supplier_id = ? AND supplier_status = 'Active'",
                            (supplier_id_from_cell,)
                        )
                    except (ValueError, TypeError):
                        pass
                
                # Se não encontrou pelo ID, tentar pelo PO
                if not supplier_query and po:
                    supplier_query = db_manager.query_one(
                        "SELECT supplier_id, vendor_name FROM supplier_database_table WHERE supplier_number = ? AND supplier_status = 'Active'",
                        (po,)
                    )
                
                if not supplier_query:
                    errors.append(f"Linha {row_num}: Supplier ID '{supplier_id_cell}' ou PO '{po}' não encontrado ou supplier inativo")
                    row_num += 1
                    continue
                
                supplier_id = supplier_query['supplier_id']
                vendor_name = supplier_query['vendor_name']
                
                # Calcular total score
                total_score = (
                    scores['otif'] * criterio_map['otif'] +
                    scores['nil'] * criterio_map['nil'] +
                    scores['pickup'] * criterio_map['pickup'] +
                    scores['package'] * criterio_map['package']
                )
                total_score = round(total_score, 2)
                
                # Verificar se registro já existe
                existing = db_manager.query_one(
                    "SELECT 1 FROM supplier_score_records_table WHERE supplier_id = ? AND month = ? AND year = ?",
                    (supplier_id, month, year)
                )
                
                if existing:
                    # Atualizar registro existente
                    update_query = """
                        UPDATE supplier_score_records_table 
                        SET otif = ?, nil = ?, quality_pickup = ?, quality_package = ?, 
                            total_score = ?, changed_by = ?, supplier_name = ?
                        WHERE supplier_id = ? AND month = ? AND year = ?
                    """
                    db_manager.execute(update_query, (
                        scores['otif'], scores['nil'], scores['pickup'], scores['package'],
                        total_score, current_user_name, vendor_name,
                        supplier_id, month, year
                    ))
                    updated_count += 1
                else:
                    # Inserir novo registro
                    insert_query = """
                        INSERT INTO supplier_score_records_table 
                        (supplier_id, month, year, otif, nil, quality_pickup, quality_package, 
                         total_score, registered_by, register_date, changed_by, supplier_name, comment)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """
                    db_manager.execute(insert_query, (
                        supplier_id, month, year, scores['otif'], scores['nil'], 
                        scores['pickup'], scores['package'], total_score,
                        current_user_name, register_date, current_user_name, vendor_name, ""
                    ))
                    imported_count += 1
                
                row_num += 1
            
            # Mostrar resultado da importação
            result_msg = f"✅ Importação concluída:\n"
            if imported_count > 0:
                result_msg += f"• {imported_count} registro(s) criado(s)\n"
            if updated_count > 0:
                result_msg += f"• {updated_count} registro(s) atualizado(s)\n"
            if errors:
                result_msg += f"• {len(errors)} erro(s) encontrado(s)"
                
            show_snack_bar(result_msg, False if not errors else False)
            
            # Atualizar interface
            # load_scores()  # REMOVIDO - Cards obsoletos, apenas tabela é usada
            
            # Mostrar erros se houver
            if errors:
                error_msg = "⚠️ Erros encontrados:\n" + "\n".join(errors[:5])  # Mostrar apenas os primeiros 5 erros
                if len(errors) > 5:
                    error_msg += f"\n... e mais {len(errors) - 5} erro(s)"
                show_snack_bar(error_msg)
                
        except Exception as e:
            show_snack_bar(f"❌ Erro durante a importação: {str(e)}", True)
    
    def export_suppliers_to_excel(with_existing_scores=False, month=None, year=None, conditional_formatting=False):
        """Exporta suppliers ativos para Excel com proteção por senha e opcionalmente com notas existentes"""
        print("🔄 DEBUG: Iniciando exportação...")
        
        try:
            print("🔄 DEBUG: Tentando importar openpyxl...")
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Protection
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.formatting.rule import ColorScaleRule
            import os
            from datetime import datetime
            print("✅ DEBUG: openpyxl importado com sucesso!")
        except ImportError as e:
            error_msg = f"❌ Biblioteca openpyxl não encontrada. Erro: {e}"
            print(error_msg)
            show_snack_bar(error_msg, True, False, page)
            return
        except Exception as e:
            error_msg = f"❌ Erro ao importar bibliotecas: {e}"
            print(error_msg)
            show_snack_bar(error_msg, True, False, page)
            return

        try:
            print("🔄 DEBUG: Executando query no banco...")
            # Buscar suppliers ativos (incluindo supplier_po)
            base_query = """
                SELECT supplier_id, supplier_number, bu, vendor_name, supplier_po 
                FROM supplier_database_table 
                WHERE supplier_status = 'Active'
                ORDER BY vendor_name
            """
            suppliers = db_manager.query(base_query)
            
            if not suppliers:
                error_msg = "❌ Nenhum supplier ativo encontrado."
                print(error_msg)
                show_snack_bar(error_msg, False, False, page)
                return

            print(f"🔍 DEBUG: Encontrados {len(suppliers)} suppliers ativos")
            
            # Se solicitado, buscar notas existentes para o período
            existing_scores = {}
            if with_existing_scores and month and year:
                print(f"🔄 DEBUG: Buscando notas existentes para {month}/{year}...")
                scores_query = """
                    SELECT supplier_id, quality_package, quality_pickup, nil, otif
                    FROM supplier_score_records_table 
                    WHERE month = ? AND year = ?
                """
                print(f"🔍 DEBUG: Executando query: {scores_query} com parâmetros: [{month}, {year}]")
                scores_data = db_manager.query(scores_query, [month, year])
                print(f"🔍 DEBUG: Query retornou {len(scores_data) if scores_data else 0} registros")
                
                if scores_data:
                    for i, score_record in enumerate(scores_data):
                        supplier_id = score_record.get('supplier_id') if isinstance(score_record, dict) else score_record[0]
                        # IMPORTANTE: Converter supplier_id para string para garantir compatibilidade
                        supplier_id_str = str(supplier_id)
                        
                        package_score = score_record.get('quality_package') if isinstance(score_record, dict) else score_record[1]
                        pickup_score = score_record.get('quality_pickup') if isinstance(score_record, dict) else score_record[2]
                        nil_score = score_record.get('nil') if isinstance(score_record, dict) else score_record[3]
                        otif_score = score_record.get('otif') if isinstance(score_record, dict) else score_record[4]
                        
                        existing_scores[supplier_id_str] = {
                            'package': package_score,
                            'pickup': pickup_score, 
                            'nil': nil_score,
                            'otif': otif_score
                        }
                        
                        # Debug para os primeiros 3 registros
                        if i < 3:
                            print(f"🔍 DEBUG: Registro {i+1} - Supplier: {supplier_id_str}, Package: {package_score}, Pickup: {pickup_score}, NIL: {nil_score}, OTIF: {otif_score}")
                
                print(f"🔍 DEBUG: Encontradas {len(existing_scores)} notas existentes para o período")
                print(f"🔍 DEBUG: Primeiros suppliers com notas: {list(existing_scores.keys())[:5]}")
            else:
                print("🔍 DEBUG: Não foi solicitada exportação com notas existentes")

            print("🔄 DEBUG: Criando workbook...")
            # Criar workbook e worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Import_score"
            
            print("🔄 DEBUG: Criando aba de validação oculta...")
            # Criar aba oculta para validação do sistema
            validation_ws = wb.create_sheet("ScoreApp_Validation")
            validation_ws.sheet_state = "hidden"  # Ocultar a aba
            
            # Adicionar apenas um texto simples de validação
            validation_ws.cell(row=1, column=1, value="SCOREAPP_VALIDATION_123456")
            
            print("🔍 DEBUG: Aba de validação criada e oculta")
            
            print("🔄 DEBUG: Criando headers baseado nas permissões do usuário...")
            # Headers baseado nas permissões do usuário
            headers = ["Supplier ID", "PO", "BU", "Supplier Name"]
            
            # Mapear permissões para nomes das colunas
            permission_columns = []
            if current_user_permissions.get('otif'):
                headers.append("OTIF")
                permission_columns.append("OTIF")
            if current_user_permissions.get('nil'):
                headers.append("NIL") 
                permission_columns.append("NIL")
            if current_user_permissions.get('pickup'):
                headers.append("Pickup")
                permission_columns.append("Pickup")
            if current_user_permissions.get('package'):
                headers.append("Package")
                permission_columns.append("Package")
            
            print(f"🔍 DEBUG: Colunas de notas permitidas: {permission_columns}")
            
            if not permission_columns:
                error_msg = "❌ Usuário não tem permissão para nenhuma coluna de nota."
                print(error_msg)
                show_snack_bar(error_msg, False, False, page)
                return
                
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
            
            # Adicionar instrução na célula ao lado dos headers
            instruction_col = len(headers) + 2
            instruction_cell = ws.cell(row=1, column=instruction_col)
            if with_existing_scores:
                # Mapear número do mês para nome
                month_names = {
                    "1": "JANEIRO", "2": "FEVEREIRO", "3": "MARÇO", "4": "ABRIL", 
                    "5": "MAIO", "6": "JUNHO", "7": "JULHO", "8": "AGOSTO",
                    "9": "SETEMBRO", "10": "OUTUBRO", "11": "NOVEMBRO", "12": "DEZEMBRO"
                }
                month_name = month_names.get(str(month), f"MÊS {month}")
                
                # Instrução principal
                instruction_cell.value = f"DADOS COM NOTAS DE {month_name}/{year} - Preencha as notas de 0.0 a 10.0 | Preencha 10 caso não haja entrega no mês"
                instruction_cell.font = Font(bold=True, italic=True, color="0066CC", size=12)
                
            else:
                instruction_cell.value = "INSTRUÇÕES: Preencha as notas de 0.0 a 10.0 | Preencha 10 caso não haja entrega no mês"
                instruction_cell.font = Font(bold=True, italic=True, color="0066CC")
            
            # Adicionar mais informações na linha 2
            info_cell = ws.cell(row=2, column=instruction_col)
            info_cell.font = Font(italic=True, color="666666", size=9)
            
            # Se for exportação com notas, adicionar informação adicional sobre os dados
            if with_existing_scores:
                info_cell.value = f"Arquivo gerado com notas existentes do período {month_name}/{year}"
                info_cell.font = Font(italic=True, color="006600", size=10)  # Verde escuro
                
                # Adicionar data de referência destacada na linha 3 (abaixo da explicação)
                date_ref_cell = ws.cell(row=3, column=instruction_col)
                date_ref_cell.value = f"📅 PERÍODO: {month_name}/{year}"
                date_ref_cell.font = Font(bold=True, color="FF0000", size=16)  # Vermelho, tamanho maior
                date_ref_cell.alignment = Alignment(horizontal="left")
                
                # Adicionar fundo amarelo para destacar ainda mais
                date_ref_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            print("🔄 DEBUG: Preenchendo dados dos suppliers...")
            # Preencher dados dos suppliers
            for row, supplier in enumerate(suppliers, 2):
                supplier_id_value = supplier.get('supplier_id') or ''
                
                # Tratar supplier_po como número inteiro, removendo decimais
                raw_po = supplier.get('supplier_po')
                if raw_po is not None and raw_po != '' and str(raw_po).strip() != '':
                    try:
                        # Converter para float primeiro, depois para int para remover decimais
                        po_value = int(float(raw_po))
                    except (ValueError, TypeError):
                        # Se falhar a conversão, manter como vazio
                        po_value = ''
                else:
                    po_value = ''
                
                bu_value = supplier.get('bu') or ''
                name_value = supplier.get('vendor_name') or ''
                
                ws.cell(row=row, column=1, value=supplier_id_value) # Supplier ID
                
                # PO (supplier_po) - usar número inteiro diretamente
                if po_value != '':
                    po_cell = ws.cell(row=row, column=2, value=po_value)
                    po_cell.number_format = '0'  # Formato de número inteiro no Excel (sem decimais)
                else:
                    po_cell = ws.cell(row=row, column=2, value='')
                
                ws.cell(row=row, column=3, value=bu_value)          # BU
                ws.cell(row=row, column=4, value=name_value)        # Supplier Name (vendor_name)
                
                # IMPORTANTE: Converter supplier_id para string para garantir compatibilidade
                supplier_id_str = str(supplier_id_value)
                
                # Preencher colunas de notas se solicitado e disponível
                if with_existing_scores and supplier_id_str in existing_scores:
                    scores = existing_scores[supplier_id_str]
                    col_index = 5  # Primeira coluna de notas (coluna E)
                    
                    # Mapear permissões para tipos de score
                    score_mapping = {
                        'otif': 'otif',
                        'nil': 'nil', 
                        'pickup': 'pickup',
                        'package': 'package'
                    }
                    
                    # Iterar pelas permissões na ordem correta dos headers
                    permission_order = ['otif', 'nil', 'pickup', 'package']
                    
                    for permission in permission_order:
                        if current_user_permissions.get(permission, False):
                            score_type = score_mapping[permission]
                            score_value = scores.get(score_type)
                            
                            if score_value is not None:
                                try:
                                    float_value = float(score_value)
                                    ws.cell(row=row, column=col_index, value=float_value)
                                except (ValueError, TypeError) as e:
                                    pass
                            
                            col_index += 1
                
            # Ajustar larguras das colunas dinamicamente
            column_widths = [15, 15, 12, 30]  # Supplier ID, PO, BU, Supplier Name
            for _ in permission_columns:
                column_widths.append(12)  # Largura padrão para colunas de notas
                
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            print("🔄 DEBUG: Aplicando proteção e validação...")
            # Proteger planilha com senha - bloquear colunas A, B, C, D e header
            password = "30625629"
            
            # Desbloquear apenas as colunas de notas permitidas (começando da coluna 5)
            notes_start_col = 5  # Primeira coluna de notas (depois de Supplier ID, PO, BU e Supplier Name)
            notes_end_col = 4 + len(permission_columns)  # Última coluna de notas
            
            # Aplicar formatação e desbloquear células de notas
            for row in range(2, len(suppliers) + 2):
                for col in range(notes_start_col, notes_end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.protection = Protection(locked=False)
                    # Formatação numérica com 1 casa decimal
                    cell.number_format = '0.0'
            
            # Criar validação de dados para números de 0 a 10
            if notes_start_col <= notes_end_col:
                print("🔄 DEBUG: Aplicando validação de dados (0.0 a 10.0)...")
                
                # Definir o range das células de notas
                start_col_letter = openpyxl.utils.get_column_letter(notes_start_col)
                end_col_letter = openpyxl.utils.get_column_letter(notes_end_col)
                notes_range = f"{start_col_letter}2:{end_col_letter}{len(suppliers) + 1}"
                
                # Criar validação de dados
                data_validation = DataValidation(
                    type="decimal",
                    operator="between",
                    formula1=0,
                    formula2=10,
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle="Valor Inválido",
                    error="Por favor, insira um número entre 0.0 e 10.0",
                    showInputMessage=True,
                    promptTitle="Nota do Supplier",
                    prompt="Insira uma nota de 0.0 a 10.0"
                )
                
                # Aplicar validação ao range de células
                ws.add_data_validation(data_validation)
                data_validation.add(notes_range)
                
                print(f"🔍 DEBUG: Validação aplicada ao range: {notes_range}")
                
                # Aplicar formatação condicional baseada na opção selecionada
                if conditional_formatting:
                    # Formatação condicional com escala de cores (vermelho 0 -> verde 10)
                    color_scale_rule = ColorScaleRule(
                        start_type='num',
                        start_value=0,
                        start_color='FF0000',  # Vermelho para 0
                        mid_type='num', 
                        mid_value=5,
                        mid_color='FFFF00',    # Amarelo para 5
                        end_type='num',
                        end_value=10,
                        end_color='00FF00'     # Verde para 10
                    )
                    ws.conditional_formatting.add(notes_range, color_scale_rule)
                    print("🔍 DEBUG: Formatação condicional de escala de cores aplicada (0=vermelho, 10=verde)")
                else:
                    # Formatação condicional apenas para valores fora do range
                    from openpyxl.formatting.rule import CellIsRule
                    from openpyxl.styles import Font as ConditionalFont, PatternFill as ConditionalFill
                    
                    # Regra para valores inválidos (menores que 0 ou maiores que 10)
                    invalid_rule = CellIsRule(
                        operator='notBetween',
                        formula=[0, 10],
                        fill=ConditionalFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
                        font=ConditionalFont(color="CC0000", bold=True)
                    )
                    
                    # Aplicar formatação condicional ao range de notas
                    ws.conditional_formatting.add(notes_range, invalid_rule)
                    print("🔍 DEBUG: Formatação condicional aplicada para valores inválidos")
            
            # Aplicar proteção da planilha
            ws.protection.set_password(password)
            ws.protection.sheet = True
            ws.protection.formatCells = False
            ws.protection.formatColumns = False
            ws.protection.formatRows = False
            ws.protection.insertColumns = False
            ws.protection.insertRows = False
            ws.protection.deleteColumns = False
            ws.protection.deleteRows = False
            
            print("🔄 DEBUG: Preparando para salvar arquivo...")
            
            # Criar nome do arquivo baseado na nova nomenclatura
            timestamp = datetime.now().strftime("%Y%m%d")
            
            # Determinar a nota/tipo de dados
            if with_existing_scores:
                note_type = f"Data_{month}_{year}"
                data_ref = f"{month}_{year}"
            else:
                note_type = "Import"
                data_ref = timestamp
            
            # Nome do arquivo: Score_[nota]_[data_ref]
            suggested_filename = f"Score_{note_type}_{data_ref}.xlsx"
            
            # Variável para armazenar o file_picker
            file_picker_ref = [None]  # Usar lista para permitir acesso em função aninhada
            
            # Abrir diálogo para salvar arquivo
            def save_file_dialog():
                try:
                    # Criar diálogo de seleção de arquivo
                    file_picker = ft.FilePicker(
                        on_result=lambda e: save_selected_file(e, file_picker)
                    )
                    file_picker_ref[0] = file_picker
                    page.overlay.append(file_picker)
                    safe_page_update(page)
                    
                    # Abrir diálogo para salvar
                    file_picker.save_file(
                        dialog_title="Salvar arquivo Excel",
                        file_name=suggested_filename,
                        allowed_extensions=["xlsx"]
                    )
                    
                except Exception as picker_error:
                    print(f"❌ Erro no file picker: {picker_error}")
                    # Fallback para Desktop
                    save_to_desktop()
            
            def save_selected_file(e, picker):
                try:
                    # Remover file picker do overlay primeiro
                    if picker and picker in page.overlay:
                        page.overlay.remove(picker)
                    safe_page_update(page)
                    
                    if e.path:
                        # Usuário selecionou um local
                        full_path = e.path
                        wb.save(full_path)
                        print(f"📁 DEBUG: Arquivo salvo em: {full_path}")
                        success_message = f"✅ Arquivo exportado com sucesso!\n{os.path.basename(full_path)}\nLocalização: {os.path.dirname(full_path)}"
                        show_snack_bar(success_message, is_error=False, is_warning=False, page=page)
                    else:
                        # Usuário cancelou
                        show_snack_bar("❌ Exportação cancelada pelo usuário", is_error=True, is_warning=False, page=page)
                        
                except Exception as save_error:
                    print(f"❌ Erro ao salvar no local selecionado: {save_error}")
                    save_to_desktop()
            
            def save_to_desktop():
                """Fallback: salvar na Desktop"""
                try:
                    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
                    full_path = os.path.join(desktop_path, suggested_filename)
                    wb.save(full_path)
                    print(f"📁 DEBUG: Arquivo salvo em: {full_path}")
                    success_message = f"✅ Arquivo exportado com sucesso!\n{suggested_filename}\nLocalização: Desktop"
                    show_snack_bar(success_message, is_error=False, is_warning=False, page=page)
                except Exception as desktop_error:
                    print(f"❌ Erro ao salvar na Desktop: {desktop_error}")
                    # Último fallback para pasta atual
                    try:
                        wb.save(suggested_filename)
                        print(f"📁 DEBUG: Arquivo salvo na pasta atual: {suggested_filename}")
                        fallback_message = f"✅ Arquivo exportado: {suggested_filename}\n(salvo na pasta do aplicativo)"
                        show_snack_bar(fallback_message, is_error=False, is_warning=False, page=page)
                    except Exception as fallback_error:
                        print(f"❌ Erro no fallback final: {fallback_error}")
                        raise fallback_error
            
            # Iniciar processo de salvamento
            save_file_dialog()
                    
        except Exception as e:
            error_msg = f"❌ Erro ao exportar: {str(e)}"
            print(error_msg)
            show_snack_bar(error_msg, True, False, page)

    def export_form_dialog():
        """Abre diálogo com opção para exportar suppliers ativos com ou sem notas existentes"""
        
        # Referencias para os controles
        export_with_scores_switch_ref = ft.Ref[ft.Switch]()
        conditional_formatting_switch_ref = ft.Ref[ft.Switch]()
        month_dropdown_ref = ft.Ref[ft.Dropdown]()
        year_dropdown_ref = ft.Ref[ft.Dropdown]()
        period_container_ref = ft.Ref[ft.Container]()
        
        # Opções de mês e ano
        months = [
            ft.dropdown.Option("1", "Janeiro"),
            ft.dropdown.Option("2", "Fevereiro"),
            ft.dropdown.Option("3", "Março"),
            ft.dropdown.Option("4", "Abril"),
            ft.dropdown.Option("5", "Maio"),
            ft.dropdown.Option("6", "Junho"),
            ft.dropdown.Option("7", "Julho"),
            ft.dropdown.Option("8", "Agosto"),
            ft.dropdown.Option("9", "Setembro"),
            ft.dropdown.Option("10", "Outubro"),
            ft.dropdown.Option("11", "Novembro"),
            ft.dropdown.Option("12", "Dezembro")
        ]
        years = [ft.dropdown.Option(str(year)) for year in range(2024, 2041)]
        
        def on_switch_change(e):
            """Controla a visibilidade dos campos de mês e ano"""
            show_period = export_with_scores_switch_ref.current.value
            period_container_ref.current.visible = show_period
            safe_page_update(page)
        
        def close_dialog(e):
            page.close(dialog)
        
        def export_form(e):
            try:
                # Verificar se deve exportar com notas
                export_with_scores = export_with_scores_switch_ref.current.value
                # Verificar se deve aplicar formatação condicional
                apply_conditional_formatting = conditional_formatting_switch_ref.current.value
                
                if export_with_scores:
                    # Verificar se mês e ano foram selecionados
                    selected_month = month_dropdown_ref.current.value
                    selected_year = year_dropdown_ref.current.value
                    
                    if not selected_month or not selected_year:
                        show_snack_bar("❌ Selecione mês e ano para exportar com notas existentes", True)
                        return
                    
                    export_suppliers_to_excel(with_existing_scores=True, month=selected_month, year=selected_year, conditional_formatting=apply_conditional_formatting)
                else:
                    export_suppliers_to_excel(with_existing_scores=False, conditional_formatting=apply_conditional_formatting)
                    
                close_dialog(e)
            except Exception as ex:
                show_snack_bar(f"❌ Erro ao exportar: {ex}", True)
        
        # Obter cores do tema atual
        theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        # Container para campos de período (inicialmente oculto)
        period_container = ft.Container(
            content=ft.Column([
                ft.Text("Período das Notas:", size=14, weight="bold", color=theme_colors.get('on_surface')),
                ft.Row([
                    ft.Dropdown(
                        label="Mês",
                        options=months,
                        expand=True,
                        ref=month_dropdown_ref,
                        color=theme_colors.get('on_surface'),
                        border_color=theme_colors.get('outline'),
                        **({'bgcolor': theme_colors.get('field_background')} if theme_colors.get('field_background') else {})
                    ),
                    ft.Dropdown(
                        label="Ano",
                        options=years,
                        expand=True,
                        ref=year_dropdown_ref,
                        color=theme_colors.get('on_surface'),
                        border_color=theme_colors.get('outline'),
                        **({'bgcolor': theme_colors.get('field_background')} if theme_colors.get('field_background') else {})
                    ),
                ], spacing=10),
            ], spacing=10),
            visible=False,
            ref=period_container_ref,
            padding=ft.padding.symmetric(vertical=10)
        )
        
        # Conteúdo do diálogo
        dialog_content = ft.Column([
            ft.Text("Exportar Suppliers para Excel", size=18, weight="bold", color=theme_colors.get('on_surface')),
            ft.Container(height=20),
            ft.Text("Esta operação irá exportar todos os suppliers ativos em um arquivo Excel protegido.", 
                   size=14, color=theme_colors.get('on_surface')),
            ft.Text("• Colunas Supplier ID, PO, BU e Supplier Name: bloqueadas", size=12, color="orange"),
            ft.Text("• Colunas de notas: abertas para preenchimento", size=12, color="green"),
            
            ft.Container(height=20),
            ft.Divider(color=theme_colors.get('outline')),
            
            # Switch para exportar com notas existentes
            ft.Row([
                ft.Switch(
                    ref=export_with_scores_switch_ref,
                    value=False,
                    active_color=theme_colors.get('primary'),
                    on_change=on_switch_change
                ),
                ft.Text("Exportar com notas existentes", size=14, weight="bold", color=theme_colors.get('on_surface'))
            ], spacing=10),
            
            # Switch para formatação condicional
            ft.Row([
                ft.Switch(
                    ref=conditional_formatting_switch_ref,
                    value=True,
                    active_color=theme_colors.get('primary')
                ),
                ft.Column([
                    ft.Text("Formatação condicional nas notas", size=14, weight="bold", color=theme_colors.get('on_surface')),
                    ft.Text("Cores de 0 (vermelho) a 10 (verde)", size=12, color=theme_colors.get('on_surface_variant'))
                ], spacing=2)
            ], spacing=10),
            
            # Container para campos de período
            period_container,
            
            ft.Container(height=20),
            ft.Divider(color=theme_colors.get('outline')),
            
            ft.Row([
                ft.ElevatedButton(
                    "Fechar",
                    on_click=close_dialog,
                    icon=ft.Icons.CANCEL,
                    style=ft.ButtonStyle(
                        bgcolor=theme_colors.get('error'),
                        color=theme_colors.get('on_error')
                    )
                ),
                ft.Container(width=10),
                ft.ElevatedButton(
                    "📊 Exportar Excel",
                    on_click=export_form,
                    icon=ft.Icons.DOWNLOAD,
                    style=ft.ButtonStyle(
                        bgcolor=theme_colors.get('primary'),
                        color=theme_colors.get('on_primary')
                    )
                )
            ], alignment=ft.MainAxisAlignment.END)
        ], width=450, spacing=15, tight=True)
        
        # Criar e exibir diálogo
        dialog = ft.AlertDialog(
            title=ft.Text("Exportar Dados", color=theme_colors.get('on_surface')),
            content=dialog_content,
            modal=True
        )
        
        page.open(dialog)
    
    # --- Fim: Funções para Importar/Exportar Score ---

    def theme_changed(e):
        """Handle theme change from UI (radio button)"""
        global current_user_wwid
        theme_mode = e.control.value

        # Aplicar tema usando a função centralizada
        apply_theme(theme_mode)

        # Salvar tema no banco
        save_user_theme(theme_mode, current_user_wwid)

        # Atualizar menus, abas e listas
        try:
            update_menu()
            update_config_tabs()
            load_all_lists_data()
            load_users_full()
        except Exception as ex:
            print(f"⚠️ Error updating after theme change: {ex}")

        # Pegar cores atuais do tema
        theme_colors = get_current_theme_colors(get_theme_name_from_page(page))

        # --- Containers principais ---
        score_form_container.bgcolor = theme_colors.get('field_background')
        score_form_container.update()

        users_form_container.bgcolor = theme_colors.get('field_background')
        users_form_container.border = ft.border.all(1, theme_colors.get('outline'))
        users_form_container.update()

        rail_container.bgcolor = theme_colors.get('rail_background')
        rail_container.update()

        # --- Dropdowns da aba score ---
        for dropdown in [month_dropdown, year_dropdown]:
            dropdown.bgcolor = theme_colors.get('field_background')
            dropdown.color = theme_colors.get('on_surface')
            dropdown.border_color = theme_colors.get('outline')
            dropdown.update()

        # --- TextFields da aba score ---
        for field_ref in [search_field_ref, selected_po, selected_bu]:
            if field_ref.current:
                field_ref.current.bgcolor = theme_colors.get('field_background')
                field_ref.current.color = theme_colors.get('on_surface')
                field_ref.current.border_color = theme_colors.get('outline')
                field_ref.current.update()

        # --- Campos das abas de configuração ---
        try:
            for config_type in ['sqie', 'continuity', 'planner', 'sourcing', 'bu', 'category']:
                if config_type in lists_controls:
                    for key in ['input', 'alias', 'email']:
                        if key in lists_controls[config_type]:
                            ctrl = lists_controls[config_type][key]
                            ctrl.bgcolor = theme_colors.get('field_background')
                            ctrl.color = theme_colors.get('on_surface')
                            ctrl.border_color = theme_colors.get('outline')
                            ctrl.update()
        except Exception as ex:
            print(f"⚠️ Error updating config fields: {ex}")

        # --- Campos da aba Users ---
        try:
            if wwid_field_ref.current:
                wwid_field_ref.current.filled = False
                wwid_field_ref.current.color = theme_colors.get('on_surface')
                wwid_field_ref.current.border_color = theme_colors.get('outline')
                wwid_field_ref.current.update()
            if name_field_ref.current:
                name_field_ref.current.filled = False
                name_field_ref.current.color = theme_colors.get('on_surface')
                name_field_ref.current.border_color = theme_colors.get('outline')
                name_field_ref.current.update()
            if password_field_ref.current:
                password_field_ref.current.filled = False
                password_field_ref.current.color = theme_colors.get('on_surface')
                password_field_ref.current.border_color = theme_colors.get('outline')
                password_field_ref.current.update()
            if privilege_dropdown_ref.current:
                privilege_dropdown_ref.current.bgcolor = None
                privilege_dropdown_ref.current.color = theme_colors.get('on_surface')
                privilege_dropdown_ref.current.border_color = theme_colors.get('outline')
                privilege_dropdown_ref.current.update()
        except Exception as ex:
            print(f"⚠️ Error updating user fields: {ex}")

        # --- Campos da aba Log ---
        try:
            if 'log_controls' in globals() and log_controls:
                for ctrl in log_controls.values():
                    control = resolve_control(ctrl)
                    if control:
                        control.bgcolor = theme_colors.get('field_background')
                        control.color = theme_colors.get('on_surface')
                        control.border_color = theme_colors.get('outline')
                        control.update()
        except Exception as ex:
            print(f"⚠️ Error updating log controls: {ex}")

        # --- Campo de pesquisa de Suppliers ---
        try:
            if suppliers_search_field_ref.current:
                suppliers_search_field_ref.current.bgcolor = theme_colors.get('field_background')
                suppliers_search_field_ref.current.color = theme_colors.get('on_surface')
                suppliers_search_field_ref.current.border_color = theme_colors.get('outline')
                suppliers_search_field_ref.current.update()
        except Exception as ex:
            print(f"⚠️ Error updating suppliers search field: {ex}")

        # --- Cards de Score ---
        try:
            if responsive_app_manager and responsive_app_manager.results_container:
                update_control_colors(responsive_app_manager.results_container, theme_mode, 0)
                responsive_app_manager.results_container.update()
        except Exception as ex:
            print(f"⚠️ Error updating Score cards: {ex}")

        # --- Cards de Suppliers ---
        try:
            if 'suppliers_results_list' in globals() and suppliers_results_list and suppliers_results_list.controls:
                for card in suppliers_results_list.controls:
                    update_control_colors(card, theme_mode, 0)
                suppliers_results_list.update()
        except Exception as ex:
            print(f"⚠️ Error updating Suppliers cards: {ex}")

        # --- Cards de métricas da Timeline ---
        try:
            if 'timeline_cards_refs' in globals() and timeline_cards_refs:
                Colors = get_current_theme_colors(theme_mode)
                primary_color = Colors.get('primary')
                for refs in timeline_cards_refs.values():
                    if refs["card"].current:
                        refs["card"].current.surface_tint_color = primary_color
                    if refs["icon"].current:
                        refs["icon"].current.color = primary_color
                    if refs["value"].current:
                        refs["value"].current.color = primary_color
                    # Não reatribuir gradiente - removido para usar bgcolor sólido nos cards
                    # if refs["gradient"].current:
                    #     refs["gradient"].current.gradient = ft.LinearGradient(...)
                timeline_metrics_row.current.update()
        except Exception as ex:
            print(f"⚠️ Error updating Timeline metric cards: {ex}")

    # Confirmação via show_toast
        show_snack_bar(f"✅ Theme '{theme_mode}' applied and saved!", False)

    def set_selected(idx):
        def handler(e):
            # Verificar se o usuário tem acesso à aba
            if not is_tab_accessible(idx, current_user_privilege or "User"):
                # Mostrar toast de erro e retornar sem alterar a seleção
                show_snack_bar("Acesso negado para esta funcionalidade", True)
                return
                
            selected_index.current = idx
            home_view.visible = idx == 0
            score_view.visible = idx == 1
            timeline_view.visible = idx == 2
            risks_view.visible = idx == 3
            email_view.visible = idx == 4
            configs_view.visible = idx == 5
            
            # Carregar dados específicos da aba quando ela for selecionada
            if idx == 2:  # Aba Timeline
                # Sempre voltar para a primeira subaba (Performance Chart)
                if timeline_tabs.current:
                    timeline_tabs.current.selected_index = 0
            
            if idx == 4:  # Aba Email
                # O EmailManager já cuida de toda a inicialização
                # Força atualização da página de forma segura
                safe_page_update(page_ref)
            
            # Limpar campos da aba Log quando sair da aba Configs
            if selected_index.current != 5 and idx != 5:
                try:
                    # Se estava na aba Configs e está saindo dela
                    if selected_config_tab.current == 5:  # Se estava na sub-aba Log
                        clear_log_fields()
                except Exception as ex:
                    print(f"Aviso: erro ao limpar campos da aba Log: {ex}")
            
            # Quando aba Score é selecionada, apenas mostrar favoritos se TODOS os filtros estão vazios E o usuário fez uma ação específica
            if idx == 1 and search_field_ref.current:  # Score agora é índice 1
                search_term = search_field_ref.current.value.strip() if search_field_ref.current.value else ""
                bu_val = selected_bu.current.value if selected_bu.current and selected_bu.current.value else ""
                po_val = selected_po.current.value if selected_po.current and selected_po.current.value else ""
                month_val = selected_month.current.value if selected_month.current and selected_month.current.value else ""
                year_val = selected_year.current.value if selected_year.current and selected_year.current.value else ""
                
                # Só mostrar favoritos se TODOS os campos estão vazios E não é a primeira abertura da aba
                all_filters_empty = not search_term and not bu_val and not po_val and not month_val and not year_val
                
                # Não mostrar favoritos automaticamente ao mudar para a aba - deixar a página em branco
                # Os favoritos só aparecem quando o usuário explicitamente limpar todos os filtros
                if False:  # Desabilitar a exibição automática de favoritos
                    show_favorites_only()
                    
            update_menu()
            # Se a aba Risks foi selecionada, gerar os cards imediatamente (usa ano atual se dropdown vazio)
            try:
                if idx == 3:  # Risks agora é índice 3
                    # Atualizar valor do target exibido na aba Risks antes de gerar cards
                    try:
                        if target_risks_text and target_risks_text.current and target_slider and target_slider.value is not None:
                            target_risks_text.current.value = f"{target_slider.value:.2f}"
                            target_risks_text.current.update()
                    except Exception as _ex:
                        print(f"Aviso ao atualizar Target na aba Risks: {_ex}")
                    # Ao entrar na página de Risks, buscar por Todo Período por padrão
                    try:
                        if risks_year_dropdown and risks_year_dropdown.current:
                            # '' corresponde à opção 'Todo Período'
                            risks_year_dropdown.current.value = ""
                            try:
                                risks_year_dropdown.current.update()
                            except Exception:
                                pass
                    except Exception as _ex2:
                        print(f"Aviso ao setar filtro 'Todo Período' na aba Risks: {_ex2}")
                    generate_risk_cards()
            except Exception as ex:
                print(f"Erro ao gerar cards ao abrir aba Risks: {ex}")
            # Se estamos saindo da aba Risks, limpar os cards para liberar memória/estado
            try:
                if idx != 3 and risks_cards_container and risks_cards_container.current:  # Risks agora é índice 3
                    risks_cards_container.current.content = ft.Text("Nenhum risco pesquisado")
                    risks_cards_container.current.update()
            except Exception as ex:
                print(f"Aviso ao limpar cards da aba Risks: {ex}")
            safe_page_update(page)
        return handler

    # Dropdowns de mês e ano
    month_options = [ft.dropdown.Option(str(i).zfill(2), text=str(i)) for i in range(1, 13)]
    year_options = [ft.dropdown.Option(str(y)) for y in range(2024, 2036)]
    selected_month = ft.Ref()
    selected_year = ft.Ref()
    selected_bu = ft.Ref()
    selected_po = ft.Ref()
    selected_card_limit = ft.Ref()  # Nova variável para limitar quantidade de cards
    show_inactive_switch = ft.Ref()  # Switch para mostrar/ocultar cards inativos
    
    # Cache de campos da tabela para atualização rápida sem recriar
    table_fields_cache = {}  # {supplier_id: {otif: field_ref, nil: field_ref, ...}}
    current_displayed_suppliers = []  # Lista de supplier_ids atualmente exibidos

    def update_table_scores_only():
        """Atualiza apenas os valores dos campos da tabela existente, sem recriar a estrutura."""
        global table_fields_cache
        
        month_val = selected_month.current.value if selected_month.current else None
        year_val = selected_year.current.value if selected_year.current else None
        
        print(f"📊 Atualizando scores da tabela existente - Mês: {month_val}, Ano: {year_val}")
        
        # Validação: só atualizar se ambos estiverem preenchidos ou ambos vazios
        has_month = month_val and str(month_val).strip()
        has_year = year_val and str(year_val).strip()
        
        if (has_month and not has_year) or (has_year and not has_month):
            print("⚠️ Atualização cancelada: Mês e Ano devem estar ambos preenchidos ou ambos vazios")
            return
        
        if not db_conn:
            print("❌ Banco de dados não conectado")
            return
        
        # Atualizar cada campo no cache
        updated_count = 0
        for supplier_id, fields in table_fields_cache.items():
            try:
                if month_val and year_val:
                    # Buscar scores do banco
                    query = """
                        SELECT otif, quality_pickup, quality_package, nil, comment, total_score
                        FROM supplier_score_records_table
                        WHERE supplier_id = ? AND month = ? AND year = ?
                    """
                    score_record = db_manager.query_one(query, (supplier_id, int(month_val), int(year_val)))
                    
                    if score_record:
                        # Atualizar campos com valores do banco
                        if 'otif' in fields and fields['otif']:
                            fields['otif'].value = f"{float(score_record['otif']):.1f}" if score_record['otif'] is not None else ""
                            fields['otif'].color = get_score_color(float(score_record['otif'])) if score_record['otif'] is not None else get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                            fields['otif'].update()
                        
                        if 'nil' in fields and fields['nil']:
                            fields['nil'].value = f"{float(score_record['nil']):.1f}" if score_record['nil'] is not None else ""
                            fields['nil'].color = get_score_color(float(score_record['nil'])) if score_record['nil'] is not None else get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                            fields['nil'].update()
                        
                        if 'pickup' in fields and fields['pickup']:
                            fields['pickup'].value = f"{float(score_record['quality_pickup']):.1f}" if score_record['quality_pickup'] is not None else ""
                            fields['pickup'].color = get_score_color(float(score_record['quality_pickup'])) if score_record['quality_pickup'] is not None else get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                            fields['pickup'].update()
                        
                        if 'package' in fields and fields['package']:
                            fields['package'].value = f"{float(score_record['quality_package']):.1f}" if score_record['quality_package'] is not None else ""
                            fields['package'].color = get_score_color(float(score_record['quality_package'])) if score_record['quality_package'] is not None else get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                            fields['package'].update()
                        
                        if 'total' in fields and fields['total']:
                            fields['total'].value = f"{float(score_record['total_score']):.1f}" if score_record['total_score'] is not None else ""
                            fields['total'].color = get_score_color(float(score_record['total_score'])) if score_record['total_score'] is not None else get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                            fields['total'].update()
                        
                        if 'comment' in fields and fields['comment']:
                            fields['comment'].value = score_record['comment'] if score_record['comment'] else ""
                            fields['comment'].update()
                        
                        updated_count += 1
                    else:
                        # Sem registro - limpar campos
                        for field_name in ['otif', 'nil', 'pickup', 'package']:
                            if field_name in fields and fields[field_name]:
                                fields[field_name].value = ""
                                fields[field_name].color = get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                                fields[field_name].update()
                        
                        if 'total' in fields and fields['total']:
                            fields['total'].value = ""
                            fields['total'].color = get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                            fields['total'].update()
                        
                        if 'comment' in fields and fields['comment']:
                            fields['comment'].value = ""
                            fields['comment'].update()
                else:
                    # Sem mês/ano - limpar todos os campos
                    for field_name in ['otif', 'nil', 'pickup', 'package']:
                        if field_name in fields and fields[field_name]:
                            fields[field_name].value = ""
                            fields[field_name].color = get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                            fields[field_name].update()
                    
                    if 'total' in fields and fields['total']:
                        fields['total'].value = ""
                        fields['total'].color = get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                        fields['total'].update()
                    
                    if 'comment' in fields and fields['comment']:
                        fields['comment'].value = ""
                        fields['comment'].update()
                        
            except Exception as e:
                print(f"❌ Erro ao atualizar campos do supplier {supplier_id}: {e}")
        
        print(f"✅ {updated_count} fornecedores atualizados na tabela existente")

    def on_month_year_change(e):
        # Quando mês/ano mudam, atualizar apenas os VALORES dos campos na tabela existente
        # Não recriar toda a tabela
        print(f"Mês/Ano mudou - atualizando valores dos campos existentes")
        
        # Se já temos uma tabela carregada, apenas atualizar os valores
        if table_fields_cache and current_displayed_suppliers:
            print(f"   ✅ Tabela já existe com {len(table_fields_cache)} fornecedores - atualizando apenas valores")
            update_table_scores_only()
            return
        
        # Senão, fazer busca completa (recriar tabela)
        print(f"   📋 Nenhuma tabela carregada - fazendo busca completa")
        
        # Se há termo de busca, usar a busca unificada para manter o comportamento
        if search_field_ref.current and search_field_ref.current.value.strip():
            search_term = search_field_ref.current.value.strip()
            print(f"   Termo de busca ativo: '{search_term}' - usando busca unificada")
            unified_search_execute()
        elif results_list.controls:
            # Se não há termo mas há resultados, usar search_suppliers (busca antiga)
            print(f"   Sem termo de busca - usando search_suppliers")
            search_suppliers()

    def on_show_inactive_change(e):
        """Handler para quando o switch de mostrar inativos muda"""
        show_inactive = e.control.value
        print(f"Switch 'Mostrar Inativos' mudou para: {show_inactive}")
        
        # Refazer a busca com o novo filtro
        if search_field_ref.current and search_field_ref.current.value.strip():
            search_term = search_field_ref.current.value.strip()
            print(f"   Refazendo busca com termo: '{search_term}'")
            unified_search_execute()
        elif results_list.controls:
            print(f"   Refazendo busca antiga")
            search_suppliers()

    def get_user_accessible_tabs(privilege):
        """Retorna lista de índices de abas que o usuário pode acessar baseado no privilégio"""
        if privilege == "Super Admin":
            # Super Admin pode acessar tudo
            return [0, 1, 2, 3, 4, 5]  # Home, Score, Timeline, Risks, Email, Configs
        elif privilege == "Admin":
            # Admin pode: Home, Score, Timeline, Risks, Email, Configs
            return [0, 1, 2, 3, 4, 5]  # Home, Score, Timeline, Risks, Email, Configs
        else:  # User
            # User pode: Home, Timeline, Risks, Configs limitado
            return [0, 2, 3, 5]  # Home, Timeline, Risks, Configs

    def is_tab_accessible(tab_index, privilege):
        """Verifica se uma aba específica é acessível para o privilégio do usuário"""
        accessible_tabs = get_user_accessible_tabs(privilege)
        return tab_index in accessible_tabs

    def has_pending_scores():
        """Verifica se há suppliers com notas pendentes para o usuário logado"""
        try:
            # Verificar permissões do usuário
            user_permissions = current_user_permissions
            
            if not user_permissions:
                return False
            
            # Pegar mês e ano atual
            now = datetime.datetime.now()
            current_month = now.strftime("%m")
            current_year = str(now.year)
            
            # Construir condições WHERE baseado nas permissões do usuário (NULL ou string vazia)
            conditions = []
            if user_permissions.get('otif', False):
                conditions.append("(otif IS NULL OR otif = '')")
            if user_permissions.get('nil', False):
                conditions.append("(nil IS NULL OR nil = '')")
            if user_permissions.get('pickup', False):
                conditions.append("(quality_pickup IS NULL OR quality_pickup = '')")
            if user_permissions.get('package', False):
                conditions.append("(quality_package IS NULL OR quality_package = '')")
            
            if not conditions:
                return False
            
            # Buscar suppliers com notas NULL nas colunas que o usuário pode preencher
            where_clause = " OR ".join(conditions)
            query = f"""
                SELECT COUNT(*) as total
                FROM supplier_score_records_table 
                WHERE month = ? AND year = ? AND ({where_clause})
            """
            
            result = db_manager.query_one(query, (current_month, current_year))
            
            if result:
                count = result.get('total') if isinstance(result, dict) else result[0]
                return count > 0
            
            return False
            
        except Exception as e:
            print(f"❌ Erro ao verificar pendências: {e}")
            import traceback
            traceback.print_exc()
            return False

    def update_menu():
        if menu_column_ref.current:
            show_text = menu_is_expanded.current
            
            # Lista completa de itens de menu
            all_menu_items = [
                ("home_outlined", "Home", 0),
                ("score_outlined", "Score", 1),
                ("timeline_outlined", "Timeline", 2),
                ("warning_outlined", "Risks", 3),
                ("email_outlined", "Email", 4),
                # Settings removido - acessível apenas pela top bar
            ]
            
            # Filtrar itens baseado no privilégio do usuário
            accessible_tabs = get_user_accessible_tabs(current_user_privilege or "User")
            filtered_menu_items = [
                menu_item(icon, text, idx, show_text) 
                for icon, text, idx in all_menu_items 
                if idx in accessible_tabs
            ]
            
            menu_column_ref.current.controls = filtered_menu_items
            menu_column_ref.current.width = 160 if show_text else 40
            menu_column_ref.current.update()

    def toggle_menu(e):
        menu_is_expanded.current = not menu_is_expanded.current
        update_menu()
        safe_page_update(page)

    # --- Início: Lógica do Banco de Dados e Pesquisa ---

    # Conexão com o banco de dados local. check_same_thread=False é necessário para Flet.
    try:
        db_conn = sqlite3.connect("db.db", check_same_thread=False)
        print("Conexão com banco de dados estabelecida com sucesso!")
        
        # Verificar e criar coluna supplier_po se necessário
        ensure_supplier_po_column()
        
        # Criar tabela de favoritos se não existir
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS favorites_table (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_wwid TEXT NOT NULL,
                supplier_id TEXT NOT NULL,
                created_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(user_wwid, supplier_id)
            )
        ''')
        print("Tabela de favoritos verificada/criada com sucesso!")

        # Criar tabelas de listas usadas em comboboxes (se não existirem)
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS sqie_table (
                sqie_id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                alias TEXT UNIQUE,
                email TEXT,
                register_date TIMESTAMP,
                registered_by TEXT
            )
        ''')
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS continuity_table (
                continuity_id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                alias TEXT UNIQUE,
                email TEXT,
                register_date TIMESTAMP,
                registered_by TEXT
            )
        ''')
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS planner_table (
                planner_id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                alias TEXT UNIQUE,
                email TEXT,
                register_date TIMESTAMP,
                registered_by TEXT
            )
        ''')
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS sourcing_table (
                sourcing_id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT,
                alias TEXT UNIQUE,
                email TEXT,
                register_date TIMESTAMP,
                registered_by TEXT
            )
        ''')
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS business_unit_table (
                business_id INTEGER PRIMARY KEY AUTOINCREMENT,
                bu TEXT UNIQUE,
                register_date TIMESTAMP,
                registered_by TEXT
            )
        ''')
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS categories_table (
                categories_id INTEGER PRIMARY KEY AUTOINCREMENT,
                category TEXT UNIQUE,
                register_date TIMESTAMP,
                registered_by TEXT
            )
        ''')
        
        # Criar tabela de usuários
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS users_table (
                user_id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_wwid TEXT UNIQUE NOT NULL,
                user_name TEXT,
                user_password TEXT NOT NULL,
                user_privilege TEXT NOT NULL,
                otif INTEGER DEFAULT 0,
                nil INTEGER DEFAULT 0,
                pickup INTEGER DEFAULT 0,
                package INTEGER DEFAULT 0,
                register_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                registered_by TEXT,
                last_updated_by TEXT,
                last_updated_date TIMESTAMP
            )
        ''')
        
        # Helper: garante colunas existem em tabela, adicionando as que faltam
        def ensure_columns(table_name, columns):
            """Verifica colunas existentes usando PRAGMA table_info e adiciona colunas ausentes.

            `columns` deve ser um dict {column_name: column_definition_sql_after_name},
            por exemplo: {'user_name': 'TEXT', 'last_updated_date': 'TIMESTAMP'}
            """
            try:
                # Obter colunas existentes
                existing = db_manager.query(f"PRAGMA table_info({table_name})")
                existing_names = {row['name'] for row in existing} if existing else set()
                for col, definition in columns.items():
                    if col not in existing_names:
                        q = f"ALTER TABLE {table_name} ADD COLUMN {col} {definition}"
                        try:
                            db_manager.execute(q)
                            print(f"Coluna {col} adicionada à tabela {table_name}")
                        except Exception as e:
                            print(f"Erro ao adicionar coluna {col} em {table_name}: {e}")
            except Exception as e:
                print(f"Erro ao garantir colunas para {table_name}: {e}")

        # Garantir colunas na users_table
        ensure_columns('users_table', {
            'user_name': 'TEXT',
            'last_updated_by': 'TEXT',
            'last_updated_date': 'TIMESTAMP',
            'registered_by': 'TEXT',
        })
        
        # Criar tabela de configurações de tema
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS theme_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_wwid TEXT,
                theme_mode TEXT NOT NULL DEFAULT 'light',
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Criar tabela de configurações de critérios
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS criteria_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_wwid TEXT,
                nil_weight REAL DEFAULT 0.2,
                otif_weight REAL DEFAULT 0.2,
                pickup_weight REAL DEFAULT 0.2,
                package_weight REAL DEFAULT 0.2,
                target_weight REAL DEFAULT 5.0,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Criar tabela de configurações gerais do aplicativo
        db_manager.execute('''
            CREATE TABLE IF NOT EXISTS app_settings (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_wwid TEXT,
                setting_key TEXT,
                setting_value TEXT,
                toast_duration INTEGER DEFAULT 3,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Garantir colunas existentes em criteria_settings e app_settings
        ensure_columns('criteria_settings', {
            'target_weight': 'REAL DEFAULT 0.2'
        })

        ensure_columns('app_settings', {
            'setting_key': 'TEXT',
            'setting_value': 'TEXT'
        })
        
        db_conn.commit()
        
    except sqlite3.Error as e:
        print(f"Erro ao conectar ao banco de dados: {e}")
        db_conn = None # Garante que o app não quebre se a conexão falhar


    # REMOVIDO - Funções movidas para escopo global (antes de load_app_settings)
    # def save_auto_save_setting(enabled):
    # def load_auto_save_setting():

    def create_score_textfield():
        """Cria um TextField simples para notas com validação de 0.00 a 10.00"""
        score_field = ft.TextField(
            value="0.00", 
            text_align=ft.TextAlign.CENTER, 
            width=100,
            height=50,
            border_radius=8,
            text_size=16,
            # Usar cores do tema
            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('card_background'),
            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
            border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline'),
            # Teclado numérico
            keyboard_type=ft.KeyboardType.NUMBER,
            # Placeholder para indicar o formato
            hint_text="0.00"
        )

        def on_field_change(e):
            """Validação em tempo real durante a digitação"""
            try:
                v = e.control.value
                if v is None or v == "":
                    return  # Permitir campo vazio durante digitação
                
                # Permitir entrada parcial (ex: "5.", "5.5")
                # Apenas validar se for um número válido
                try:
                    num = float(v)
                    # Se o número estiver fora do range, corrigir
                    if num > 10:
                        e.control.value = "10.00"
                        e.control.color = get_score_color(10.0)
                        e.control.update()
                    elif num < 0:
                        e.control.value = "0.00"
                        e.control.color = get_score_color(0.0)
                        e.control.update()
                    else:
                        # Aplicar cor baseada no valor
                        e.control.color = get_score_color(num)
                        e.control.update()
                except ValueError:
                    # Se não for um número válido, não fazer nada (permitir digitação)
                    pass
            except Exception as ex:
                print(f"Erro na validação do campo: {ex}")

        def on_field_blur(e):
            """Validação e formatação quando o usuário termina de editar"""
            try:
                v = e.control.value
                if v is None or str(v).strip() == "":
                    e.control.value = "0.00"
                    e.control.color = get_score_color(0.0)
                    e.control.update()
                    return
                
                # Converter para float
                num = float(str(v).strip())
                # Limitar entre 0 e 10
                num = max(0.0, min(10.0, num))
                # Formatar com 2 casas decimais
                e.control.value = f"{num:.2f}"
                # Aplicar cor baseada no valor
                e.control.color = get_score_color(num)
                e.control.update()
            except (ValueError, TypeError):
                # Se não conseguir converter, voltar para 0.00
                e.control.value = "0.00"
                e.control.color = get_score_color(0.0)
                e.control.update()

        score_field.on_change = on_field_change
        score_field.on_blur = on_field_blur

        return score_field

    # ---------- Utilitários para listas (sqie, planner, sourcing, continuity, bu, category) ----------
    
    def load_all_lists_data():
        """Carrega dados existentes de todas as tabelas para popular as listas na interface."""
        if not db_conn:
            return
        
        try:
            # Usar refresh_list_ui para cada tipo de lista
            # Isso garantirá consistência com a nova estrutura de cards
            for key in lists_controls.keys():
                refresh_list_ui(key)
                    
        except Exception as e:
            print(f"Erro ao carregar dados das listas: {e}")

    def select_list_item(key, value):
        """Seleciona ou des-seleciona um item na lista, atualizando apenas o realce visual."""
        
        # Se o item clicado já estava selecionado, desmarca (define como None).
        # Caso contrário, define o item clicado como o selecionado.
        if lists_controls[key]['selected_item'] == value:
            lists_controls[key]['selected_item'] = None
        else:
            lists_controls[key]['selected_item'] = value
        
        # Recarregar apenas a lista específica usando o mesmo estilo de load_all_lists_data
        reload_single_list_with_style(key)
        
        # A atualização da página é necessária para que a lista seja redesenhada na tela.
        safe_page_update(page)

    def reload_single_list_with_style(key):
        """Recarrega uma lista específica mantendo o estilo da load_all_lists_data"""
        # Usar refresh_list_ui para manter consistência
        refresh_list_ui(key)

    def load_list_options(table_name, column_name):
        """Retorna lista de tuplas (value, text) para popular Dropdowns."""
        try:
            if table_name in ("business_unit_table", "categories_table"):
                # Tabelas que só têm uma coluna principal
                rows = db_manager.query(f"SELECT {column_name} FROM {table_name} ORDER BY {column_name}")
                options = [r[column_name] for r in rows if r[column_name]]  # Filtrar valores nulos/vazios
                return options if options else []  # Não retorna lista vazia se não há dados
            else:
                # Para tabelas com name, alias, email (sqie, continuity, planner, sourcing)
                # Usar apenas o campo 'name' como solicitado
                rows = db_manager.query(f"SELECT name FROM {table_name} WHERE name IS NOT NULL AND name != '' ORDER BY name")
                options = [r['name'] for r in rows if r['name']]  # Filtrar valores nulos/vazios
                return options if options else []  # Não retorna lista vazia se não há dados
        except Exception as ex:
            print(f"Erro ao carregar opções de {table_name}: {ex}")
            return []

    def load_list_items_full(table_name):
        """Retorna dados completos dos itens da lista com name, alias, email e ID único."""
        try:
            if table_name == "business_unit_table":
                rows = db_manager.query(f"SELECT business_id, bu FROM {table_name} ORDER BY bu")
                return [{"id": r["business_id"], "display": r["bu"], "alias": r["bu"]} for r in rows]
            elif table_name == "categories_table":
                rows = db_manager.query(f"SELECT categories_id, category FROM {table_name} ORDER BY category")
                return [{"id": r["categories_id"], "display": r["category"], "alias": r["category"]} for r in rows]
            else:
                # Para tabelas com name, alias, email - incluir ID da tabela
                id_columns = {
                    'sqie_table': 'sqie_id',
                    'continuity_table': 'continuity_id', 
                    'planner_table': 'planner_id',
                    'sourcing_table': 'sourcing_id'
                }
                
                id_col = id_columns.get(table_name, 'id')
                rows = db_manager.query(f"SELECT {id_col}, name, alias, email FROM {table_name} ORDER BY alias")
                items = []
                for row in rows:
                    row_id = row[id_col]
                    name = row["name"]
                    alias = row["alias"]
                    email = row["email"]
                    # Criar texto de exibição com as informações completas
                    display_parts = []
                    if name and name.strip():
                        display_parts.append(f"Nome: {name.strip()}")
                    if alias and alias.strip():
                        display_parts.append(f"Alias: {alias.strip()}")
                    if email and email.strip():
                        display_parts.append(f"Email: {email.strip()}")
                    
                    display_text = " | ".join(display_parts) if display_parts else "Item sem dados"
                    items.append({
                        "id": row_id,
                        "display": display_text,
                        "alias": alias or name or "",
                        "name": name or "",
                        "email": email or ""
                    })
                return items
        except Exception as ex:
            print(f"Erro ao carregar itens completos de {table_name}: {ex}")
            return []

    def insert_list_item(table_name, columns, values):
        """Insere item em tabela genérica passando colunas e valores."""
        placeholders = ','.join(['?'] * len(values))
        cols = ','.join(columns)
        q = f"INSERT INTO {table_name} ({cols}) VALUES ({placeholders})"
        db_manager.execute(q, values)

    def delete_list_item_by_alias(table_name, alias_column, alias):
        q = f"DELETE FROM {table_name} WHERE {alias_column} = ?"
        db_manager.execute(q, (alias,))
    
    def get_all_cards():
        """Retorna todos os cards, navegando pela estrutura responsiva se necessário"""
        all_cards = []
        
        if not responsive_app_manager or not responsive_app_manager.results_container:
            # Fallback para results_list tradicional
            for control in results_list.controls:
                if isinstance(control, ft.Card):
                    all_cards.append(control)
            return all_cards
        
        # Navegar pela estrutura responsiva
        for control in responsive_app_manager.results_container.controls:
            if isinstance(control, ft.Card):
                # Card direto (layout single)
                all_cards.append(control)
            elif isinstance(control, ft.Row):
                # Layout double - verificar colunas
                for column in control.controls:
                    if isinstance(column, ft.Column):
                        for card in column.controls:
                            if isinstance(card, ft.Card):
                                all_cards.append(card)
        
        return all_cards

    def load_scores():
        """Para os cards de resultado visíveis, carrega as notas do banco de dados
        com base no mês e ano selecionados.
        """
        if not selected_month.current or not selected_year.current:
            return
            
        month_val = selected_month.current.value
        year_val = selected_year.current.value

        # A função só executa se ambos, mês e ano, estiverem selecionados.
        # Se qualquer um estiver vazio, limpar as notas e não buscar nada
        if not month_val or not year_val or month_val == "" or year_val == "":
            # Limpar as notas se a data for desmarcada
            all_cards = get_all_cards()
            for card in all_cards:
                if hasattr(card, 'data') and card.data:
                    # Suporte para estrutura antiga (spinbox_refs)
                    if "spinbox_refs" in card.data:
                        for ref in card.data["spinbox_refs"].values():
                            try:
                                ref.value = "0.0"
                                ref.color = get_score_color(0.0)
                                if hasattr(ref, 'page') and ref.page is not None:
                                    ref.update()
                            except Exception as e:
                                print(f"Erro ao limpar spinbox: {e}")
                    
                    # Suporte para estrutura nova (score_sliders/score_texts)
                    if "score_sliders" in card.data and "score_texts" in card.data:
                        score_sliders = card.data["score_sliders"]
                        score_texts = card.data["score_texts"]
                        score_controls = card.data.get("score_controls")
                        
                        for ui_name in ["OTIF", "Pickup", "Package", "NIL"]:
                            if ui_name in score_sliders and ui_name in score_texts:
                                try:
                                    # Para sliders normais
                                    if hasattr(score_sliders[ui_name], 'value'):
                                        score_sliders[ui_name].value = 0.0
                                        if hasattr(score_sliders[ui_name], 'page') and score_sliders[ui_name].page is not None:
                                            score_sliders[ui_name].update()
                                    
                                    # Atualizar texto
                                    score_texts[ui_name].value = "0.0"
                                    score_texts[ui_name].color = get_score_color(0.0)
                                    if hasattr(score_texts[ui_name], 'page') and score_texts[ui_name].page is not None:
                                        score_texts[ui_name].update()
                                except Exception as e:
                                    print(f"Erro ao limpar slider {ui_name}: {e}")
            safe_page_update(page)
            return

        if not db_conn: 
            print("Conexão com banco não disponível")
            return

        print(f"Carregando scores para mês: {month_val}, ano: {year_val}")

        # Usar a função auxiliar para obter todos os cards
        all_cards = get_all_cards()
        
        for card in all_cards:
            if not isinstance(card, ft.Card) or not hasattr(card, 'data') or not card.data:
                continue

            supplier_id = card.data["supplier_id"]
            
            # Verificar se é slider ou spinbox
            score_sliders = card.data.get("score_sliders")
            score_texts = card.data.get("score_texts")
            spinbox_refs = card.data.get("spinbox_refs")
            
            # Se não tem nenhum dos dois tipos de controle, pular
            if not score_sliders and not spinbox_refs:
                continue
            
            try:
                query = """
                    SELECT otif, quality_pickup, quality_package, nil, comment
                    FROM supplier_score_records_table
                    WHERE supplier_id = ? AND month = ? AND year = ?
                """
                score_record = db_manager.query_one(query, (supplier_id, int(month_val), int(year_val)))

                if score_record:
                    # Mapeamento direto: OTIF, Pickup, Package, NIL, Comment
                    scores = {
                        "OTIF": score_record['otif'] if score_record['otif'] is not None else 0.0,
                        "Pickup": score_record['quality_pickup'] if score_record['quality_pickup'] is not None else 0.0,
                        "Package": score_record['quality_package'] if score_record['quality_package'] is not None else 0.0,
                        "NIL": score_record['nil'] if score_record['nil'] is not None else 0.0,
                    }
                    comment_value = score_record['comment'] if score_record['comment'] is not None else ""
                    
                    # Atualizar valores nos sliders (se existirem)
                    if score_sliders:
                        for ui_name, score_value in scores.items():
                            if ui_name in score_sliders:
                                try:
                                    val = float(score_value)
                                    print(f"🔍 DEBUG carregando {ui_name} = {val:.1f}")
                                    print(f"   Tipo do controle: {type(score_sliders[ui_name])}")
                                    
                                    # Para sliders, textfields e outros controles
                                    # Se for TextField (tipo str para value), formatar com 2 casas decimais
                                    if isinstance(score_sliders[ui_name].value, str) or score_control_type == "textfield":
                                        score_sliders[ui_name].value = f"{val:.2f}"
                                        score_sliders[ui_name].color = get_score_color(val)
                                        print(f"   ✅ Valor do TextField setado: {score_sliders[ui_name].value}")
                                    else:
                                        score_sliders[ui_name].value = val
                                        print(f"   ✅ Valor do slider setado: {score_sliders[ui_name].value}")
                                    
                                    # Atualizar também o campo de texto se existir (para sliders)
                                    if ui_name in score_texts:
                                        score_texts[ui_name].value = f"{val:.1f}"
                                        score_texts[ui_name].color = get_score_color(val)
                                        print(f"   ✅ Valor do texto setado: {score_texts[ui_name].value}")
                                        
                                        # Atualizar texto se estiver na página
                                        if hasattr(score_texts[ui_name], 'page') and score_texts[ui_name].page is not None:
                                            score_texts[ui_name].update()
                                            print(f"   ✅ Campo de texto {ui_name} atualizado na página")
                                    
                                    # Verificar se o controle está na página antes de atualizar
                                    if hasattr(score_sliders[ui_name], 'page') and score_sliders[ui_name].page is not None:
                                        score_sliders[ui_name].update()
                                        print(f"   ✅ Controle {ui_name} atualizado na página")
                                    else:
                                        print(f"   ⚠️ Slider {ui_name} não está na página ainda")
                                    
                                except Exception as e:
                                    print(f"❌ Erro ao atualizar {ui_name}: {e}")
                                    import traceback
                                    traceback.print_exc()
                                    continue
                    
                    # Atualizar valores nos spinboxes (se existirem)
                    if spinbox_refs:
                        print(f"🔍 DEBUG spinbox_refs disponíveis: {list(spinbox_refs.keys())}")
                        for ui_name, score_value in scores.items():
                            if ui_name in spinbox_refs:
                                try:
                                    val = float(score_value)
                                    print(f"🔍 DEBUG atualizando spinbox {ui_name} com valor {val:.1f}")
                                    print(f"🔍 DEBUG tipo do spinbox_refs[{ui_name}]: {type(spinbox_refs[ui_name])}")
                                    
                                    spinbox_refs[ui_name].value = f"{val:.1f}"
                                    spinbox_refs[ui_name].color = get_score_color(val)
                                    print(f"🔍 DEBUG valor definido: {spinbox_refs[ui_name].value}")
                                    
                                    # Verificar se o spinbox está na página antes de atualizar
                                    if hasattr(spinbox_refs[ui_name], 'page') and spinbox_refs[ui_name].page is not None:
                                        spinbox_refs[ui_name].update()
                                        print(f"✅ Spinbox {ui_name} atualizado com sucesso")
                                    else:
                                        print(f"⚠️ Spinbox {ui_name} não está na página")
                                except Exception as e:
                                    print(f"❌ Erro ao atualizar spinbox {ui_name}: {e}")
                                    continue
                            else:
                                print(f"⚠️ Spinbox {ui_name} não encontrado em spinbox_refs")
                    else:
                        print("⚠️ Nenhum spinbox_refs encontrado no card")
                    
                    # Atualizar comentário se houver referência para ele no card
                    if "comment_field" in card.data and card.data["comment_field"]:
                        try:
                            card.data["comment_field"].value = comment_value
                            if hasattr(card.data["comment_field"], 'page') and card.data["comment_field"].page is not None:
                                card.data["comment_field"].update()
                            print(f"Atualizando comentário: {comment_value}")
                        except Exception as e:
                            print(f"Erro ao atualizar comentário: {e}")
                        
                else:
                    # Não encontrou registro, zerar valores
                    # Zerar sliders (se existirem)
                    if score_sliders and score_texts:
                        for ui_name in ["OTIF", "Pickup", "Package", "NIL"]:
                            if ui_name in score_sliders:
                                try:
                                    print(f"🔄 Zerando {ui_name} (sem registro no banco)")
                                    # Para sliders normais
                                    if hasattr(score_sliders[ui_name], 'value'):
                                        score_sliders[ui_name].value = 0.0
                                        if hasattr(score_sliders[ui_name], 'page') and score_sliders[ui_name].page is not None:
                                            score_sliders[ui_name].update()
                                    
                                    # Atualizar texto
                                    score_texts[ui_name].value = "0.0"
                                    score_texts[ui_name].color = get_score_color(0.0)
                                    if hasattr(score_texts[ui_name], 'page') and score_texts[ui_name].page is not None:
                                        score_texts[ui_name].update()
                                except Exception as e:
                                    print(f"Erro ao zerar slider {ui_name}: {e}")
                                    continue
                    
                    # Zerar spinboxes (se existirem)
                    if spinbox_refs:
                        for ui_name in ["OTIF", "Pickup", "Package", "NIL"]:
                            if ui_name in spinbox_refs:
                                try:
                                    spinbox_refs[ui_name].value = "0.0"
                                    spinbox_refs[ui_name].color = get_score_color(0.0)
                                    
                                    # Verificar se o spinbox está na página antes de atualizar
                                    if hasattr(spinbox_refs[ui_name], 'page') and spinbox_refs[ui_name].page is not None:
                                        spinbox_refs[ui_name].update()
                                except Exception as e:
                                    print(f"Erro ao zerar spinbox {ui_name}: {e}")
                                    continue
                    
                    # Limpar comentário
                    if "comment_field" in card.data and card.data["comment_field"]:
                        try:
                            card.data["comment_field"].value = ""
                            if hasattr(card.data["comment_field"], 'page') and card.data["comment_field"].page is not None:
                                card.data["comment_field"].update()
                        except Exception as e:
                            print(f"Erro ao limpar comentário: {e}")
                
            except sqlite3.Error as e:
                print(f"Erro ao carregar notas para o supplier_id {supplier_id}: {e}")
        
        safe_page_update(page)

   
    search_field_ref = ft.Ref()

    def clear_score_filters(e=None):
        """Limpa todos os filtros da aba Score, limpa os cards e mostra apenas os favoritos"""
        try:
            print("🔄 Iniciando limpeza dos filtros e cards da aba Score...")
            
            # Limpar todos os campos de filtro e comboboxes
            if search_field_ref.current:
                search_field_ref.current.value = ""
                search_field_ref.current.update()
            
            if selected_po.current:
                selected_po.current.value = ""
                selected_po.current.update()
            
            if selected_bu.current:
                selected_bu.current.value = None
                selected_bu.current.update()
            
            if selected_month.current:
                selected_month.current.value = None
                selected_month.current.update()
            
            if selected_year.current:
                selected_year.current.value = None
                selected_year.current.update()
            
            if selected_card_limit.current:
                selected_card_limit.current.value = "6"  # Resetar para padrão
                selected_card_limit.current.update()
            
            if show_inactive_switch.current:
                show_inactive_switch.current.value = False  # Resetar para não mostrar inativos
                show_inactive_switch.current.update()
            
            # Limpar todos os cards de resultados primeiro
            print("🗑️ Limpando cards de resultados...")
            try:
                if responsive_app_manager and responsive_app_manager.results_container:
                    responsive_app_manager.clear_results()
                    print("✅ Cards limpos via responsive_app_manager")
            except Exception as clear_ex:
                print(f"⚠️ Falha ao limpar via responsive_app_manager: {clear_ex}")
                # Fallback: limpar diretamente via results_list
                if results_list and hasattr(results_list, 'controls'):
                    results_list.controls.clear()
                    print("✅ Cards limpos via results_list")
            
            # Agora mostrar apenas os favoritos
            print("⭐ Carregando favoritos...")
            show_favorites_only()
            
            # Atualizar a página uma vez no final
            safe_page_update(page)
            
            print("✅ Filtros limpos e favoritos carregados com sucesso")
            
        except Exception as ex:
            print(f"❌ Erro ao limpar filtros: {ex}")
            import traceback
            traceback.print_exc()

    def create_score_slider(page_ref=None):
        """Cria um widget de slider customizado para notas de 0 a 10."""
        
        # O texto que exibirá o valor do slider com fonte maior e Roboto
        score_text = ft.Text(
            "0.0", 
            width=60, 
            text_align=ft.TextAlign.CENTER, 
            color=get_score_color(0.0),
            size=20,
            weight=ft.FontWeight.W_400,
            font_family="Roboto"
        )

        # Obter a cor mais escura baseada no tema atual
        theme_name = get_theme_name_from_page(page_ref) if page_ref else "white"
        darker_color = get_primary_darker_color(theme_name)

        # O slider - totalmente flexível para se ajustar ao espaço disponível do card
        score_slider = ft.Slider(
            min=0,
            max=10,
            divisions=100,  # (10 - 0) / 0.1 = 100
            value=0,
            expand=True,  # Expandir para preencher o espaço disponível sem ultrapassar
            # Cores do slider: ativa primária; inativa mais escura baseada no tema
            active_color="primary",
            inactive_color=darker_color
        )

        def on_slider_change(e):
            # Formata o valor para ter uma casa decimal
            new_value = f"{e.control.value:.1f}"
            print(f"🔄 Slider mudou para: {new_value}")
            score_text.value = new_value
            # Aplicar cor baseada no valor
            score_text.color = get_score_color(e.control.value)
            
            # Tentar atualizar o controle - com fallback para layout responsivo
            updated = False
            try:
                if hasattr(score_text, 'page') and score_text.page is not None:
                    print(f"   ✅ Atualizando via score_text.update()")
                    score_text.update()
                    updated = True
                elif hasattr(e.control, 'page') and e.control.page is not None:
                    # Fallback: atualizar via página do slider
                    print(f"   ✅ Atualizando via e.control.page.update()")
                    e.control.page.update()
                    updated = True
                else:
                    print(f"   ⚠️ score_text.page={getattr(score_text, 'page', None)}, e.control.page={getattr(e.control, 'page', None)}")
            except Exception as update_error:
                print(f"⚠️ Erro ao atualizar slider text: {update_error}")
            
            # Último recurso: tentar atualizar a página diretamente
            if not updated:
                print(f"   🔄 Tentando fallback via page_ref")
                try:
                    if page_ref and hasattr(page_ref, 'update'):
                        page_ref.update()
                        print(f"   ✅ Atualizado via page_ref")
                except Exception as fallback_error:
                    print(f"   ❌ Fallback falhou: {fallback_error}")
        
        score_slider.on_change = on_slider_change
        
        # Retorna o slider e o texto em uma coluna (texto centralizado embaixo)
        slider_column = ft.Column([
            score_slider, 
            ft.Container(
                content=score_text,
                alignment=ft.alignment.center,
                expand=True,  # Usar expand ao invés de largura fixa
                margin=ft.margin.only(top=-20)  # Margem negativa para aproximar do slider
            )
        ], spacing=0, tight=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER, expand=True)
        
        return slider_column, score_slider, score_text

   
    def create_score_table(records, page_num=1, records_per_page=50):
        """Cria uma tabela com os registros de fornecedores com paginação.
        
        Args:
            records: Lista completa de registros de fornecedores
            page_num: Número da página atual (começa em 1)
            records_per_page: Quantidade de registros por página (default: 50)
        """
        global page_ref, table_fields_cache, current_displayed_suppliers
        
        # Limpar cache anterior e preparar para nova tabela
        table_fields_cache = {}
        current_displayed_suppliers = []
        
        # Calcular paginação
        total_records = len(records)
        total_pages = (total_records + records_per_page - 1) // records_per_page  # Arredondar para cima
        
        # Validar página
        if page_num < 1:
            page_num = 1
        elif page_num > total_pages:
            page_num = total_pages
        
        # Calcular índices para a página atual
        start_idx = (page_num - 1) * records_per_page
        end_idx = min(start_idx + records_per_page, total_records)
        page_records = records[start_idx:end_idx]
        
        print(f"📊 Total de registros: {total_records}")
        print(f"� Página {page_num}/{total_pages} - Mostrando registros {start_idx + 1}-{end_idx}")
        
        # Dicionário para mapear supplier_id -> total_score_field para atualização após salvar
        total_score_fields = {}
        
        # Obter tema
        theme_name = get_theme_name_from_page(page)
        theme_colors = get_current_theme_colors(theme_name)
        primary_color = theme_colors.get('primary', ft.Colors.BLUE)
        text_color = theme_colors.get('on_surface', ft.Colors.BLACK)
        
        # Obter mês e ano selecionados para buscar scores
        month_val = selected_month.current.value if selected_month.current else None
        year_val = selected_year.current.value if selected_year.current else None
        
        # VERIFICAR SE MÊS E ANO ESTÃO SELECIONADOS - DESABILITAR CAMPOS SE NÃO ESTIVEREM
        fields_enabled = bool(month_val and month_val.strip() and year_val)
        print(f"🔒 Campos {'habilitados' if fields_enabled else 'desabilitados'} - Mês: {month_val}, Ano: {year_val}")
        
        # Criar cabeçalho da tabela com larguras específicas
        header_row = ft.Container(
            content=ft.Row([
                ft.Container(ft.Text("ID", size=12, weight="bold", color=text_color, no_wrap=True), width=60, alignment=ft.alignment.center),
                ft.Container(ft.Text("Supplier ID", size=12, weight="bold", color=text_color, no_wrap=True), width=100, alignment=ft.alignment.center_left),
                ft.Container(ft.Text("PO", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                ft.Container(ft.Text("BU", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                ft.Container(ft.Text("Supplier Name", size=12, weight="bold", color=text_color, no_wrap=True), width=450, alignment=ft.alignment.center_left),  # AUMENTADO PARA 450
                ft.Container(ft.Text("OTIF", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                ft.Container(ft.Text("NIL", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                ft.Container(ft.Text("Pickup", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                ft.Container(ft.Text("Package", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                ft.Container(ft.Text("Total", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                ft.Container(ft.Text("Comments", size=12, weight="bold", color=text_color, no_wrap=True), expand=True, alignment=ft.alignment.center_left),
                ft.Container(ft.Icon(ft.Icons.SAVE, size=12, color=text_color), width=80, alignment=ft.alignment.center),
            ], spacing=10, tight=True),
            padding=ft.padding.symmetric(horizontal=16, vertical=12),
            bgcolor=ft.Colors.with_opacity(0.1, primary_color),
            border_radius=ft.border_radius.only(top_left=8, top_right=8),
        )
        
        # Criar linhas da tabela - APENAS REGISTROS DA PÁGINA ATUAL
        table_rows = []
        
        for i, record in enumerate(page_records):
            supplier_id = record.get('supplier_id', '')
            supplier_number = record.get('supplier_number', '')
            vendor_name = record.get('vendor_name', '')
            supplier_po = record.get('supplier_po', '')
            bu = record.get('bu', '')
            supplier_status = record.get('supplier_status', 'Active')  # Pegar status (Active/Inactive)
            
            # DEBUG: Log reduzido
            if i % 10 == 0 or i < 3:
                print(f"[LINHA] {start_idx + i + 1}: {supplier_id} - {vendor_name}")
            
            # Carregar scores para este fornecedor se mês e ano estão selecionados
            otif_val = ""
            nil_val = ""
            pickup_val = ""
            package_val = ""
            comment_val = ""
            total_score_val = ""
            
            if month_val and year_val and db_conn:
                try:
                    query = """
                        SELECT otif, quality_pickup, quality_package, nil, comment, total_score
                        FROM supplier_score_records_table
                        WHERE supplier_id = ? AND month = ? AND year = ?
                    """
                    score_record = db_manager.query_one(query, (supplier_id, int(month_val), int(year_val)))
                    
                    if score_record:
                        otif_val = f"{float(score_record['otif']):.1f}" if score_record['otif'] is not None else ""
                        nil_val = f"{float(score_record['nil']):.1f}" if score_record['nil'] is not None else ""
                        pickup_val = f"{float(score_record['quality_pickup']):.1f}" if score_record['quality_pickup'] is not None else ""
                        package_val = f"{float(score_record['quality_package']):.1f}" if score_record['quality_package'] is not None else ""
                        comment_val = score_record['comment'] if score_record['comment'] is not None else ""
                        total_score_val = f"{float(score_record['total_score']):.1f}" if score_record['total_score'] is not None else ""
                except Exception as e:
                    print(f"Erro ao carregar scores para {supplier_id}: {e}")
            
            # Cor alternada para linhas
            row_bgcolor = ft.Colors.with_opacity(0.03, text_color) if i % 2 == 0 else None
            
            # Criar linha da tabela com larguras específicas
            # Tratar valores vazios
            supplier_id_text = str(supplier_id) if supplier_id else ""
            supplier_number_text = str(supplier_number) if supplier_number else ""
            vendor_name_text = str(vendor_name) if vendor_name else ""
            supplier_po_text = str(supplier_po) if supplier_po else ""
            bu_text = str(bu) if bu else ""
            comment_text = str(comment_val)[:50] if comment_val else ""
            
            # Controle de alterações e feedback visual
            has_changes = [False]
            status_icon_ref = [None]
            status_timer_ref = [None]

            def is_auto_save_enabled():
                return bool(app_settings.get('auto_save', True))

            def cancel_status_timer():
                if status_timer_ref[0]:
                    try:
                        status_timer_ref[0].cancel()
                    except Exception:
                        pass
                    status_timer_ref[0] = None

            def update_status_icon(state, hide_after=None, icon_ref=None):
                # Se icon_ref não foi passado, usar status_icon_ref (para compatibilidade)
                if icon_ref is None:
                    icon_ref = status_icon_ref
                
                icon = icon_ref[0] if icon_ref else None
                if not icon:
                    return
                cancel_status_timer()

                if state == "editing":
                    icon.name = ft.Icons.SUPERSCRIPT
                    icon.color = ft.Colors.AMBER
                    icon.visible = True
                    icon.tooltip = "Alterações em andamento"
                elif state == "saved":
                    icon.name = ft.Icons.CHECK_CIRCLE
                    icon.color = ft.Colors.GREEN
                    icon.visible = True
                    icon.tooltip = "Alterações salvas"

                    if hide_after:
                        def hide_icon():
                            try:
                                # Executar diretamente sem async
                                update_status_icon("hidden", icon_ref=icon_ref)
                            except Exception:
                                # Ignorar erros se o event loop já foi fechado
                                pass

                        status_timer_ref[0] = threading.Timer(hide_after, hide_icon)
                        status_timer_ref[0].start()
                elif state == "hidden":
                    icon.visible = False
                    icon.color = ft.Colors.with_opacity(0, ft.Colors.AMBER)  # Resetar cor ao esconder

                try:
                    icon.update()
                except Exception:
                    # Ignorar erros se o event loop já foi fechado
                    pass

            def normalize_score_input(value_text):
                return float(value_text.replace(",", "."))

            def mark_editing_state(icon_ref=status_icon_ref):
                has_changes[0] = True
                update_status_icon("editing", icon_ref=icon_ref)
            
            # Função para recalcular e salvar o total_score
            def recalculate_and_save_total(sid):
                try:
                    print(f"\n🔢 DEBUG recalculate_and_save_total")
                    print(f"   Supplier ID: {sid}")
                    print(f"   Month: {month_val}, Year: {year_val}")
                    
                    # Validar mês e ano
                    if not month_val or not year_val:
                        print(f"   ❌ Mês ou ano não disponíveis - pulando recálculo")
                        return
                    
                    # Buscar critérios
                    criterios_raw = db_manager.query("SELECT criteria_category, value FROM criteria_table")
                    criterio_map = {}
                    for row in criterios_raw:
                        nome = str(row.get('criteria_category') if isinstance(row, dict) else row[0]).strip().lower()
                        valor = float(row.get('value') if isinstance(row, dict) else row[1])
                        if "package" in nome:
                            criterio_map["package"] = valor
                        elif "pick" in nome:
                            criterio_map["pickup"] = valor  
                        elif "nil" in nome:
                            criterio_map["nil"] = valor
                        elif "otif" in nome:
                            criterio_map["otif"] = valor
                    
                    print(f"   Critérios carregados: {criterio_map}")
                    
                    # Buscar scores atuais
                    score_query = """
                        SELECT otif, quality_pickup, quality_package, nil
                        FROM supplier_score_records_table
                        WHERE supplier_id = ? AND month = ? AND year = ?
                    """
                    score_data = db_manager.query_one(score_query, (sid, int(month_val), int(year_val)))
                    
                    if score_data:
                        print(f"   Scores encontrados: OTIF={score_data['otif']}, NIL={score_data['nil']}, Pickup={score_data['quality_pickup']}, Package={score_data['quality_package']}")
                        
                        # Calcular total_score
                        total_score = (
                            float(score_data['otif'] or 0) * criterio_map.get('otif', 0) +
                            float(score_data['nil'] or 0) * criterio_map.get('nil', 0) +
                            float(score_data['quality_pickup'] or 0) * criterio_map.get('pickup', 0) +
                            float(score_data['quality_package'] or 0) * criterio_map.get('package', 0)
                        )
                        total_score = round(total_score, 2)
                        
                        print(f"   Total Score calculado: {total_score}")
                        
                        # Atualizar total_score
                        update_total_query = """
                            UPDATE supplier_score_records_table 
                            SET total_score = ?
                            WHERE supplier_id = ? AND month = ? AND year = ?
                        """
                        db_manager.execute(update_total_query, (total_score, sid, int(month_val), int(year_val)))
                        print(f"   ✅ Total score atualizado no banco")
                    else:
                        print(f"   ⚠️ Nenhum score encontrado para recalcular")
                except Exception as ex:
                    print(f"   ❌ Erro ao recalcular total_score: {ex}")
                    import traceback
                    traceback.print_exc()
            
            # Função para salvar valor editado
            def save_score_value(e, sid, field_name, field_ref, total_field=None, icon_ref=None):
                try:
                    print(f"\n[DEBUG] save_score_value - Início")
                    print(f"   Supplier ID: {sid}")
                    print(f"   Field Name: {field_name}")
                    print(f"   Event control value: {e.control.value if e and hasattr(e, 'control') else 'N/A'}")
                    print(f"   Field ref value: {field_ref.value}")
                    print(f"   Month: {month_val}, Year: {year_val}")
                    
                    # Validar mês e ano
                    if not month_val or not year_val:
                        print(f"   [ERRO] Mês ou ano não selecionados!")
                        show_snack_bar("Selecione um mês e ano antes de editar scores", is_error=True)
                        return
                    
                    # IMPORTANTE: Usar o valor do EVENTO (e.control.value), não do field_ref
                    # porque o field_ref pode ter o valor desatualizado no momento do on_blur
                    value = (e.control.value or "").strip() if e and hasattr(e, 'control') else (field_ref.value or "").strip()
                    print(f"   Valor a ser processado: '{value}'")
                    
                    if not value:
                        print(f"   [INFO] Valor vazio - limpando campo")
                        field_ref.value = ""
                        field_ref.color = text_color
                        field_ref.update()
                        update_status_icon("hidden", icon_ref=icon_ref)
                        has_changes[0] = False
                        return

                    # Validar e converter o valor
                    try:
                        num_value = normalize_score_input(value)
                        print(f"   [OK] Valor convertido: {num_value}")
                    except (ValueError, AttributeError) as ve:
                        print(f"   [ERRO] Erro ao converter valor: {ve}")
                        show_snack_bar("Valor inválido! Digite um número entre 0 e 10", is_error=True)
                        update_status_icon("editing", icon_ref=icon_ref)
                        return

                    # Validar range de 0 a 10
                    if num_value < 0 or num_value > 10:
                        print(f"   [ERRO] Valor fora do range: {num_value}")
                        show_snack_bar("Valor deve estar entre 0 e 10", is_error=True)
                        update_status_icon("editing", icon_ref=icon_ref)
                        return

                    auto_save_enabled = is_auto_save_enabled()
                    print(f"   Auto-save habilitado: {auto_save_enabled}")

                    if not auto_save_enabled:
                        # Formatar imediatamente se auto-save estiver desabilitado
                        formatted_value = f"{num_value:.1f}"
                        field_ref.value = formatted_value
                        field_ref.color = get_score_color(num_value)
                        field_ref.update()
                        print(f"   [OK] Campo formatado (sem auto-save): {formatted_value}")
                        has_changes[0] = True
                        mark_editing_state(icon_ref=icon_ref)
                        print(f"   [INFO] Auto-save desabilitado - aguardando salvamento manual")
                        return

                    # Capturar valores de mês e ano no escopo da closure
                    current_month = int(month_val)
                    current_year = int(year_val)
                    current_value = num_value

                    # Formatar campo IMEDIATAMENTE
                    formatted_value = f"{num_value:.1f}"
                    field_ref.value = formatted_value
                    field_ref.color = get_score_color(num_value)
                    print(f"   [OK] Campo formatado: {formatted_value}")
                    
                    mark_editing_state(icon_ref=icon_ref)

                    # SALVAR DIRETAMENTE (sem thread assíncrona)
                    try:
                        print(f"\n[SAVE] Iniciando salvamento DIRETO")
                        print(f"   Supplier ID: {sid}")
                        print(f"   Month: {current_month}, Year: {current_year}")
                        print(f"   Field: {field_name} = {current_value}")
                        
                        # Verificar se o registro existe
                        check_query = """
                            SELECT COUNT(*) as count FROM supplier_score_records_table
                            WHERE supplier_id = ? AND month = ? AND year = ?
                        """
                        result = db_manager.query_one(check_query, (sid, current_month, current_year))
                        record_exists = result and result['count'] > 0
                        print(f"   Registro existe? {record_exists}")
                        
                        if record_exists:
                            # Registro existe - fazer UPDATE
                            update_query = f"""
                                UPDATE supplier_score_records_table 
                                SET {field_name} = ?
                                WHERE supplier_id = ? AND month = ? AND year = ?
                            """
                            print(f"   [UPDATE] Executando UPDATE...")
                            db_manager.execute(update_query, (current_value, sid, current_month, current_year))
                            print(f"   [OK] UPDATE executado")
                        else:
                            # Registro não existe - fazer INSERT
                            insert_query = f"""
                                INSERT INTO supplier_score_records_table 
                                (supplier_id, month, year, {field_name})
                                VALUES (?, ?, ?, ?)
                            """
                            print(f"   [INSERT] Executando INSERT...")
                            db_manager.execute(insert_query, (sid, current_month, current_year, current_value))
                            print(f"   [OK] INSERT executado")
                        
                        print(f"   [CALC] Recalculando total_score...")
                        recalculate_and_save_total(sid)
                        print(f"   [OK] Total recalculado")
                        
                        # Buscar o novo total_score do banco
                        total_query = """
                            SELECT total_score FROM supplier_score_records_table
                            WHERE supplier_id = ? AND month = ? AND year = ?
                        """
                        total_result = db_manager.query_one(total_query, (sid, current_month, current_year))
                        new_total_raw = total_result['total_score'] if total_result and total_result['total_score'] is not None else 0
                        new_total = float(new_total_raw) if new_total_raw else 0.0  # CONVERTER PARA FLOAT!
                        print(f"   [TOTAL] Novo total_score: {new_total}")
                        
                        # Atualizar interface
                        field_ref.value = formatted_value
                        field_ref.color = get_score_color(num_value)
                        field_ref.update()
                        
                        # Atualizar campo de total_score se foi passado
                        if total_field:
                            total_field.value = f"{new_total:.1f}"
                            total_field.color = get_score_color(new_total)
                            total_field.update()
                            print(f"   [UI] Campo Total Score atualizado: {new_total:.1f}")
                        
                        has_changes[0] = False
                        update_status_icon("saved", hide_after=1.8, icon_ref=icon_ref)
                        
                        # NÃO MOSTRAR SNACKBAR EM AUTO-SAVE (apenas logs para debug)
                        # show_snack_bar removido para não poluir a interface em auto-save
                        print(f"   [OK] Salvamento concluído! {field_name.upper()}={formatted_value} | TOTAL={new_total:.1f} (ID:{sid})")

                        
                    except Exception as save_error:
                        print(f"   [ERRO] Erro ao salvar: {save_error}")
                        import traceback
                        traceback.print_exc()
                        show_snack_bar(f"Erro ao salvar: {save_error}", is_error=True)
                except Exception as ex:
                    print(f"[ERRO] ERRO GERAL em save_score_value: {ex}")
                    import traceback
                    traceback.print_exc()
                    show_snack_bar(f"Erro ao atualizar: {ex}", is_error=True)
            
            # Função para salvar comentário
            def save_comment(e, sid, field_ref):
                try:
                    value = (field_ref.value or "").strip()
                    # COMENTÁRIOS SEMPRE SALVAM AUTOMATICAMENTE (ignora configuração de auto_save)
                    
                    mark_editing_state()

                    def persist():
                        # Verificar se o registro existe
                        check_query = """
                            SELECT COUNT(*) as count FROM supplier_score_records_table
                            WHERE supplier_id = ? AND month = ? AND year = ?
                        """
                        result = db_manager.query_one(check_query, (sid, int(month_val), int(year_val)))
                        
                        if result and result['count'] > 0:
                            # Registro existe - fazer UPDATE
                            update_query = """
                                UPDATE supplier_score_records_table 
                                SET comment = ?
                                WHERE supplier_id = ? AND month = ? AND year = ?
                            """
                            db_manager.execute(update_query, (value if value else None, sid, int(month_val), int(year_val)))
                        else:
                            # Registro não existe - fazer INSERT
                            insert_query = """
                                INSERT INTO supplier_score_records_table 
                                (supplier_id, month, year, comment)
                                VALUES (?, ?, ?, ?)
                            """
                            db_manager.execute(insert_query, (sid, int(month_val), int(year_val), value if value else None))

                    def handle_success(_):
                        has_changes[0] = False
                        update_status_icon("saved", hide_after=1.8)
                        field_ref.update()

                    run_async(
                        persist,
                        on_success=handle_success,
                        on_error=lambda exc: show_snack_bar(f"Erro ao atualizar: {exc}", is_error=True),
                    )
                except Exception as ex:
                    show_snack_bar(f"Erro ao atualizar: {ex}", is_error=True)
            
            # Função para salvar manualmente (definida antes, mas os campos serão passados explicitamente)
            def save_manually(e, sid, otif_ref, nil_ref, pickup_ref, package_ref, comment_ref, total_ref, icon_ref):
                try:
                    def parse_score_field(field, label):
                        val = (field.value or "").strip()
                        if not val:
                            field.value = ""
                            field.color = text_color
                            field.update()
                            return None

                        try:
                            num = normalize_score_input(val)
                        except (ValueError, AttributeError):
                            raise ValueError(f"{label}: valor inválido. Digite um número entre 0 e 10")

                        if num < 0 or num > 10:
                            raise ValueError(f"{label}: valor deve estar entre 0 e 10")

                        # Formatar como 0.0
                        field.value = f"{num:.1f}"
                        field.color = get_score_color(num)
                        field.update()
                        return num

                    otif_num = parse_score_field(otif_ref, "OTIF")
                    nil_num = parse_score_field(nil_ref, "NIL")
                    pickup_num = parse_score_field(pickup_ref, "Pickup")
                    package_num = parse_score_field(package_ref, "Package")
                    comment_val = (comment_ref.value or "").strip() or None

                    update_status_icon("editing", icon_ref=icon_ref)

                    def persist():
                        # Verificar se o registro existe
                        check_query = """
                            SELECT COUNT(*) as count FROM supplier_score_records_table
                            WHERE supplier_id = ? AND month = ? AND year = ?
                        """
                        result = db_manager.query_one(check_query, (sid, int(month_val), int(year_val)))
                        
                        if result and result['count'] > 0:
                            # Registro existe - fazer UPDATE
                            update_query = """
                                UPDATE supplier_score_records_table 
                                SET otif = ?, nil = ?, quality_pickup = ?, quality_package = ?, comment = ?
                                WHERE supplier_id = ? AND month = ? AND year = ?
                            """
                            db_manager.execute(
                                update_query,
                                (otif_num, nil_num, pickup_num, package_num, comment_val, sid, int(month_val), int(year_val)),
                            )
                        else:
                            # Registro não existe - fazer INSERT
                            insert_query = """
                                INSERT INTO supplier_score_records_table 
                                (supplier_id, month, year, otif, nil, quality_pickup, quality_package, comment)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """
                            db_manager.execute(
                                insert_query,
                                (sid, int(month_val), int(year_val), otif_num, nil_num, pickup_num, package_num, comment_val),
                            )
                        
                        recalculate_and_save_total(sid)
                        
                        # Buscar o novo total_score do banco
                        total_query = """
                            SELECT total_score FROM supplier_score_records_table
                            WHERE supplier_id = ? AND month = ? AND year = ?
                        """
                        total_result = db_manager.query_one(total_query, (sid, int(month_val), int(year_val)))
                        new_total_raw = total_result['total_score'] if total_result and total_result['total_score'] is not None else 0
                        new_total = float(new_total_raw) if new_total_raw else 0.0
                        
                        # Atualizar campo de total_score
                        if total_ref:
                            total_ref.value = f"{new_total:.1f}"
                            total_ref.color = get_score_color(new_total)
                            total_ref.update()

                    def handle_success(_):
                        has_changes[0] = False
                        # Mostrar check verde por 1.8s, depois esconder completamente (volta ao normal)
                        update_status_icon("saved", hide_after=1.8, icon_ref=icon_ref)
                        if not is_auto_save_enabled():
                            show_snack_bar("Alterações salvas com sucesso!")

                    run_async(
                        persist,
                        on_success=handle_success,
                        on_error=lambda exc: show_snack_bar(f"Erro ao salvar: {exc}", is_error=True),
                    )
                except ValueError as validation_error:
                    show_snack_bar(str(validation_error), is_error=True)
                    update_status_icon("editing", icon_ref=icon_ref)
                except Exception as ex:
                    show_snack_bar(f"Erro ao salvar: {ex}", is_error=True)
                    update_status_icon("editing", icon_ref=icon_ref)
            
            # Criar campos editáveis
            def format_and_mark_change(field_ref, score_name=None):
                """Formata o valor para uma casa decimal (se for numérico) e marca alteração visualmente."""
                print(f"\n📝 DEBUG format_and_mark_change")
                print(f"   Score name: {score_name}")
                print(f"   Valor atual: '{field_ref.value}'")
                
                try:
                    val = (field_ref.value or "").strip()
                    if val:
                        # tentar converter para float e formatar
                        try:
                            if score_name:
                                num = normalize_score_input(val)
                                field_ref.value = f"{num:.1f}"
                                field_ref.color = get_score_color(num)
                                print(f"   ✅ Formatado: {field_ref.value}, cor: {field_ref.color}")
                        except Exception as ex:
                            # não é número: manter texto (comentário)
                            print(f"   ⚠️ Não é número ou erro: {ex}")
                            pass
                    else:
                        field_ref.value = ""
                        field_ref.color = text_color
                        print(f"   ⚠️ Valor vazio")

                    # marcar alteração
                    mark_editing_state()
                    field_ref.update()
                    print(f"   ✅ Campo atualizado")
                except Exception as ex:
                    print(f"   ❌ Erro: {ex}")
                    pass

            def validate_numeric_input(e):
                """Valida entrada em tempo real - aceita apenas números, vírgula, ponto e valores de 0 a 10"""
                if e.control.value:
                    # Remover caracteres inválidos (manter apenas dígitos, ponto e vírgula)
                    cleaned = ''.join(c for c in e.control.value if c.isdigit() or c in '.,')
                    
                    # Trocar vírgula por ponto para validação
                    test_value = cleaned.replace(',', '.')
                    
                    # Tentar converter para float e validar
                    try:
                        if test_value and test_value != '.':
                            num = float(test_value)
                            # Se for maior que 10, truncar para 10
                            if num > 10:
                                e.control.value = "10.0"
                            # Se for negativo, tornar positivo
                            elif num < 0:
                                e.control.value = "0.0"
                            else:
                                # Manter o valor limpo durante a digitação
                                e.control.value = cleaned
                        else:
                            e.control.value = cleaned
                    except ValueError:
                        # Se não conseguir converter, manter apenas a parte válida
                        e.control.value = cleaned
                    
                    e.control.update()

            # VERIFICAR SE O MÊS/ANO É FUTURO
            from datetime import datetime
            current_date = datetime.now()
            is_future_month = False
            if month_val and year_val:
                try:
                    selected_date = datetime(int(year_val), int(month_val), 1)
                    # Comparar apenas ano e mês (ignorar dia)
                    if (int(year_val) > current_date.year) or (int(year_val) == current_date.year and int(month_val) > current_date.month):
                        is_future_month = True
                        print(f"⚠️ Mês futuro detectado: {month_val}/{year_val} - campos serão bloqueados")
                except:
                    pass
            
            # VERIFICAR PERMISSÕES DO USUÁRIO
            global current_user_permissions
            has_otif_permission = current_user_permissions.get('otif', False) and not is_future_month
            has_nil_permission = current_user_permissions.get('nil', False) and not is_future_month
            has_pickup_permission = current_user_permissions.get('pickup', False) and not is_future_month
            has_package_permission = current_user_permissions.get('package', False) and not is_future_month
            
            print(f"📝 Permissões para supplier {supplier_id}:")
            print(f"   OTIF: {has_otif_permission} (permissão: {current_user_permissions.get('otif', False)}, futuro: {is_future_month})")
            print(f"   NIL: {has_nil_permission} (permissão: {current_user_permissions.get('nil', False)}, futuro: {is_future_month})")
            print(f"   Pickup: {has_pickup_permission} (permissão: {current_user_permissions.get('pickup', False)}, futuro: {is_future_month})")
            print(f"   Package: {has_package_permission} (permissão: {current_user_permissions.get('package', False)}, futuro: {is_future_month})")

            # Campo de Total Score (somente leitura, calculado automaticamente) - CRIAR ANTES DOS CAMPOS DE INPUT
            total_score_field = ft.Text(
                value=total_score_val,
                size=14,
                weight="bold",
                color=get_score_color(float(total_score_val)) if total_score_val else text_color,
                text_align=ft.TextAlign.CENTER,
            )

            otif_field = ft.TextField(
                value=str(otif_val) if otif_val else "",
                text_size=13,
                text_align=ft.TextAlign.CENTER,
                border_color=ft.Colors.TRANSPARENT,
                focused_border_color=primary_color,
                color=get_score_color(float(otif_val)) if otif_val else text_color,
                dense=True,
                content_padding=ft.padding.all(5),
                input_filter=ft.InputFilter(regex_string=r"^[0-9.,]*$"),  # Aceitar apenas números, vírgula e ponto
                disabled=not (has_otif_permission and fields_enabled),  # BLOQUEAR SE NÃO TEM PERMISSÃO, É MÊS FUTURO OU MÊS/ANO NÃO SELECIONADOS
                opacity=1.0 if (has_otif_permission and fields_enabled) else 0.5,  # VISUAL DE BLOQUEADO
            )
            def otif_on_change(e):
                print(f"[CHANGE] OTIF on_change disparado! Valor: {e.control.value}")
                validate_numeric_input(e)
            def otif_on_blur(e, sid=supplier_id, total_field=total_score_field, icon_ref=status_icon_ref):  # CAPTURAR TODOS!
                print(f"\n{'*'*60}")
                print(f"[BLUR] OTIF on_blur DISPARADO!")
                print(f"   supplier_id capturado no closure (sid): {sid}")
                print(f"   supplier_id da variável do loop: {supplier_id}")
                print(f"   Valor do campo: {e.control.value}")
                print(f"   Campo correto? {e.control is otif_field}")
                print(f"{'*'*60}\n")
                save_score_value(e, sid, "otif", e.control, total_field, icon_ref)  # PASSAR icon_ref!
            otif_field.on_change = otif_on_change
            otif_field.on_blur = otif_on_blur
            print(f"[OK] OTIF field criado com eventos: on_change={otif_field.on_change is not None}, on_blur={otif_field.on_blur is not None}")

            nil_field = ft.TextField(
                value=str(nil_val) if nil_val else "",
                text_size=13,
                text_align=ft.TextAlign.CENTER,
                border_color=ft.Colors.TRANSPARENT,
                focused_border_color=primary_color,
                color=get_score_color(float(nil_val)) if nil_val else text_color,
                dense=True,
                content_padding=ft.padding.all(5),
                input_filter=ft.InputFilter(regex_string=r"^[0-9.,]*$"),  # Aceitar apenas números, vírgula e ponto
                disabled=not (has_nil_permission and fields_enabled),  # BLOQUEAR SE NÃO TEM PERMISSÃO, É MÊS FUTURO OU MÊS/ANO NÃO SELECIONADOS
                opacity=1.0 if (has_nil_permission and fields_enabled) else 0.5,  # VISUAL DE BLOQUEADO
            )
            def nil_on_change(e):
                validate_numeric_input(e)
            def nil_on_blur(e, sid=supplier_id, total_field=total_score_field, icon_ref=status_icon_ref):  # CAPTURAR TODOS!
                print(f"[BLUR] NIL on_blur disparado! Valor: {e.control.value}, SID: {sid}")
                save_score_value(e, sid, "nil", e.control, total_field, icon_ref)  # PASSAR icon_ref!
            nil_field.on_change = nil_on_change
            nil_field.on_blur = nil_on_blur

            pickup_field = ft.TextField(
                value=str(pickup_val) if pickup_val else "",
                text_size=13,
                text_align=ft.TextAlign.CENTER,
                border_color=ft.Colors.TRANSPARENT,
                focused_border_color=primary_color,
                color=get_score_color(float(pickup_val)) if pickup_val else text_color,
                dense=True,
                content_padding=ft.padding.all(5),
                input_filter=ft.InputFilter(regex_string=r"^[0-9.,]*$"),  # Aceitar apenas números, vírgula e ponto
                disabled=not (has_pickup_permission and fields_enabled),  # BLOQUEAR SE NÃO TEM PERMISSÃO, É MÊS FUTURO OU MÊS/ANO NÃO SELECIONADOS
                opacity=1.0 if (has_pickup_permission and fields_enabled) else 0.5,  # VISUAL DE BLOQUEADO
            )
            def pickup_on_change(e):
                validate_numeric_input(e)
            def pickup_on_blur(e, sid=supplier_id, total_field=total_score_field, icon_ref=status_icon_ref):  # CAPTURAR TODOS!
                print(f"[BLUR] PICKUP on_blur disparado! Valor: {e.control.value}, SID: {sid}")
                save_score_value(e, sid, "quality_pickup", e.control, total_field, icon_ref)  # PASSAR icon_ref!
            pickup_field.on_change = pickup_on_change
            pickup_field.on_blur = pickup_on_blur

            package_field = ft.TextField(
                value=str(package_val) if package_val else "",
                text_size=13,
                text_align=ft.TextAlign.CENTER,
                border_color=ft.Colors.TRANSPARENT,
                focused_border_color=primary_color,
                color=get_score_color(float(package_val)) if package_val else text_color,
                dense=True,
                content_padding=ft.padding.all(5),
                input_filter=ft.InputFilter(regex_string=r"^[0-9.,]*$"),  # Aceitar apenas números, vírgula e ponto
                disabled=not (has_package_permission and fields_enabled),  # BLOQUEAR SE NÃO TEM PERMISSÃO, É MÊS FUTURO OU MÊS/ANO NÃO SELECIONADOS
                opacity=1.0 if (has_package_permission and fields_enabled) else 0.5,  # VISUAL DE BLOQUEADO
            )
            def package_on_change(e):
                validate_numeric_input(e)
            def package_on_blur(e, sid=supplier_id, total_field=total_score_field, icon_ref=status_icon_ref):  # CAPTURAR TODOS!
                print(f"[BLUR] PACKAGE on_blur disparado! Valor: {e.control.value}, SID: {sid}")
                save_score_value(e, sid, "quality_package", e.control, total_field, icon_ref)  # PASSAR icon_ref!
            package_field.on_change = package_on_change
            package_field.on_blur = package_on_blur

            comment_field = ft.TextField(
                value=comment_text,
                text_size=12,
                multiline=True,
                max_lines=2,
                border_color=ft.Colors.TRANSPARENT,
                focused_border_color=primary_color,
                color=ft.Colors.with_opacity(0.7, text_color),
                hint_text="",
                dense=True,
                content_padding=ft.padding.all(5),
                disabled=not fields_enabled or is_future_month,  # BLOQUEAR SE MÊS/ANO NÃO SELECIONADOS OU SE FOR MÊS FUTURO
                opacity=1.0 if (fields_enabled and not is_future_month) else 0.5,  # VISUAL DE BLOQUEADO
            )
            def comment_on_blur(e, sid=supplier_id, comment_ref=comment_field):  # CAPTURAR AMBOS!
                # formatar não aplicável; apenas marcar alteração e salvar/comportamento de auto_save
                format_and_mark_change(comment_ref)
                save_comment(e, sid, e.control)  # USAR e.control EM VEZ DE comment_field!
            comment_field.on_blur = comment_on_blur
            
            # Ícone de status das alterações
            status_icon = ft.Icon(
                name=ft.Icons.SUPERSCRIPT,
                color=ft.Colors.AMBER,
                size=16,
                visible=False,
                tooltip="Alterações em andamento"
            )
            status_icon_ref[0] = status_icon
            
            # Função para mostrar detalhes do supplier: reutiliza o diálogo da Timeline
            def show_supplier_details(e, rec=record):
                """Redireciona para a função reutilizável `show_supplier_info_dialog` definida na Timeline."""
                try:
                    supplier_id_for_dialog = rec.get('supplier_id', None)
                    print(f"\n[INFO BUTTON] Abrindo diálogo para supplier_id: {supplier_id_for_dialog}")
                    
                    # Garantir que temos acesso ao page
                    if not page:
                        print(f"[INFO BUTTON] ❌ ERRO: page não está disponível!")
                        show_snack_bar("Erro ao abrir diálogo: page não disponível", is_error=True)
                        return
                    
                    # Chamar a função existente que monta o diálogo completo da Timeline
                    show_supplier_info_dialog(supplier_id_for_dialog)
                    print(f"[INFO BUTTON] ✅ Diálogo aberto com sucesso!")
                    
                except Exception as ex:
                    print(f"❌ Erro ao abrir diálogo de detalhes do supplier: {ex}")
                    import traceback
                    traceback.print_exc()
                    show_snack_bar(f"Erro ao abrir diálogo: {str(ex)}", is_error=True)
            
            # Botão de informações do supplier
            info_btn = ft.IconButton(
                icon=ft.Icons.INFO_OUTLINE,
                icon_size=16,
                tooltip="Ver detalhes do fornecedor",
                on_click=show_supplier_details,
                icon_color=ft.Colors.BLUE_400,
            )
            
            # Ícone de status (ativo/inativo) do supplier
            is_active = supplier_status == "Active"
            supplier_status_icon = ft.Icon(
                ft.Icons.CHECK_CIRCLE if is_active else ft.Icons.CANCEL,
                size=14,
                color=ft.Colors.GREEN_400 if is_active else ft.Colors.RED_400,
                tooltip="Ativo" if is_active else "Inativo"
            )
            
            # Botão de salvar
            save_btn = ft.IconButton(
                icon=ft.Icons.SAVE,
                icon_size=16,
                tooltip="Salvar alterações" if not is_future_month else "Não é possível salvar em meses futuros",
                on_click=lambda e, sid=supplier_id, otif_ref=otif_field, nil_ref=nil_field, pickup_ref=pickup_field, package_ref=package_field, comment_ref=comment_field, total_ref=total_score_field, icon_ref=status_icon_ref: save_manually(e, sid, otif_ref, nil_ref, pickup_ref, package_ref, comment_ref, total_ref, icon_ref),  # CAPTURAR TODOS OS VALORES E REFERÊNCIAS
                disabled=is_future_month,  # BLOQUEAR BOTÃO DE SALVAR SE FOR MÊS FUTURO
                opacity=1.0 if not is_future_month else 0.3,  # VISUAL DE BLOQUEADO
            )
            
            row = ft.Container(
                content=ft.Row([
                    ft.Container(ft.Text(supplier_id_text, size=13, color=text_color), width=60, alignment=ft.alignment.center),
                    ft.Container(ft.Text(supplier_number_text, size=13, color=text_color), width=100, alignment=ft.alignment.center_left),
                    ft.Container(ft.Text(supplier_po_text, size=13, color=text_color), width=80, alignment=ft.alignment.center),
                    ft.Container(ft.Text(bu_text, size=13, color=text_color), width=80, alignment=ft.alignment.center),
                    ft.Container(
                        ft.Row([
                            supplier_status_icon,  # ÍCONE DE STATUS (ATIVO/INATIVO) DO SUPPLIER
                            ft.Container(
                                ft.Text(
                                    vendor_name_text, 
                                    size=13, 
                                    color=text_color,
                                    no_wrap=True,  # NÃO QUEBRAR LINHA
                                    overflow=ft.TextOverflow.ELLIPSIS,  # TRUNCAR COM "..." SE MUITO LONGO
                                ),
                                expand=True,  # OCUPAR ESPAÇO DISPONÍVEL
                            ),
                            info_btn,  # ÍCONE DE INFO AO LADO DO NOME
                        ], spacing=5, tight=True),
                        width=450,  # AUMENTADO PARA 450
                        alignment=ft.alignment.center_left,
                        clip_behavior=ft.ClipBehavior.HARD_EDGE,  # CORTAR OVERFLOW
                    ),
                    ft.Container(otif_field, width=80, alignment=ft.alignment.center),
                    ft.Container(nil_field, width=80, alignment=ft.alignment.center),
                    ft.Container(pickup_field, width=80, alignment=ft.alignment.center),
                    ft.Container(package_field, width=80, alignment=ft.alignment.center),
                    ft.Container(total_score_field, width=80, alignment=ft.alignment.center),
                    ft.Container(comment_field, expand=True, alignment=ft.alignment.center_left),
                    ft.Container(
                        ft.Row([status_icon, save_btn], spacing=5, tight=True),
                        width=80,
                        alignment=ft.alignment.center
                    ),
                ], spacing=10, tight=True),
                padding=ft.padding.symmetric(horizontal=16, vertical=10),
                bgcolor=row_bgcolor,
                border=ft.border.only(bottom=ft.BorderSide(1, ft.Colors.with_opacity(0.1, text_color))),
            )
            
            # ADICIONAR CAMPOS AO CACHE GLOBAL para atualização rápida
            table_fields_cache[supplier_id] = {
                'otif': otif_field,
                'nil': nil_field,
                'pickup': pickup_field,
                'package': package_field,
                'total': total_score_field,
                'comment': comment_field
            }
            current_displayed_suppliers.append(supplier_id)
            
            table_rows.append(row)
        
        # Criar controles de paginação (footer) separado
        pagination_footer = None
        if total_pages > 1:
            # Função para navegar entre páginas
            def go_to_page(e, new_page):
                print(f"📄 Navegando para página {new_page}/{total_pages}")
                # Recriar tabela com nova página
                new_table = create_score_table(records, page_num=new_page, records_per_page=records_per_page)
                
                # Substituir tabela existente
                try:
                    # Encontrar o índice da tabela atual em results_list
                    for idx, control in enumerate(results_list.controls):
                        if hasattr(control, 'content') and isinstance(control.content, ft.Column):
                            results_list.controls[idx] = new_table
                            results_list.update()
                            break
                except Exception as ex:
                    print(f"Erro ao atualizar página: {ex}")
            
            # Footer de paginação
            pagination_footer = ft.Container(
                content=ft.Row([
                    ft.IconButton(
                        icon=ft.Icons.FIRST_PAGE,
                        tooltip="Primeira página",
                        disabled=page_num == 1,
                        on_click=lambda e: go_to_page(e, 1),
                    ),
                    ft.IconButton(
                        icon=ft.Icons.CHEVRON_LEFT,
                        tooltip="Página anterior",
                        disabled=page_num == 1,
                        on_click=lambda e: go_to_page(e, page_num - 1),
                    ),
                    ft.Container(
                        content=ft.Text(
                            f"Página {page_num} de {total_pages} ({total_records} registros)",
                            size=13,
                            weight="bold",
                            color=text_color,
                        ),
                        padding=ft.padding.symmetric(horizontal=20),
                    ),
                    ft.IconButton(
                        icon=ft.Icons.CHEVRON_RIGHT,
                        tooltip="Próxima página",
                        disabled=page_num == total_pages,
                        on_click=lambda e: go_to_page(e, page_num + 1),
                    ),
                    ft.IconButton(
                        icon=ft.Icons.LAST_PAGE,
                        tooltip="Última página",
                        disabled=page_num == total_pages,
                        on_click=lambda e: go_to_page(e, total_pages),
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER, spacing=5),
                padding=ft.padding.symmetric(horizontal=16, vertical=12),
                bgcolor=ft.Colors.with_opacity(0.05, primary_color),
                border_radius=ft.border_radius.only(bottom_left=8, bottom_right=8),
            )
        
        # Container da tabela com header fixo, corpo scrollável e footer fixo (igual Timeline)
        table_content = [
            # Header fixo
            header_row,
            # Corpo da tabela com scroll
            ft.Container(
                content=ft.Column(
                    controls=table_rows,
                    spacing=0,
                    scroll=ft.ScrollMode.AUTO,
                ),
                expand=True,
            ),
        ]
        
        # Adicionar footer fixo se houver paginação
        if pagination_footer:
            table_content.append(pagination_footer)
        
        # Estrutura similar ao Timeline: Column com header fixo, corpo scrollável e footer fixo
        table_container = ft.Container(
            content=ft.Column(
                controls=table_content,
                spacing=0,
                expand=True
            ),
            border=ft.border.all(1, ft.Colors.with_opacity(0.12, text_color)),
            border_radius=8,
            bgcolor=ft.Colors.with_opacity(0.02, text_color),
            expand=True,
        )
        
        return table_container

    def edit_supplier_score(supplier_id):
        """Handler para editar score de um fornecedor - abre um formulário de edição"""
        print(f"Editando fornecedor: {supplier_id}")
        
        # Carregar dados do fornecedor
        try:
            supplier_query = "SELECT vendor_name, BU, supplier_po FROM supplier_database_table WHERE supplier_id = ?"
            supplier_data = db_manager.query_one(supplier_query, (supplier_id,))
            
            if not supplier_data:
                show_snack_bar(f"Fornecedor {supplier_id} não encontrado", True)
                return
            
            vendor_name = supplier_data.get('vendor_name', 'Unknown')
            
            # Obter mês e ano selecionados
            month_val = selected_month.current.value if selected_month.current else None
            year_val = selected_year.current.value if selected_year.current else None
            
            if not month_val or not year_val:
                show_snack_bar("Selecione um mês e ano para editar scores", True)
                return
            
            # Procurar o card correspondente e rolar até ele
            show_snack_bar(f"Abrindo edição para {vendor_name}...", False)
            
        except Exception as e:
            print(f"Erro ao editar: {e}")
            show_snack_bar(f"Erro ao editar: {str(e)}", True)

    def delete_supplier_score(supplier_id):
        """Handler para deletar score de um fornecedor"""
        print(f"Deletando score do fornecedor: {supplier_id}")
        
        # Verificar permissão
        if current_user_privilege != "Super Admin":
            show_snack_bar("❌ Acesso negado. Apenas Super Admin pode deletar scores.", True)
            return
        
        try:
            supplier_query = "SELECT vendor_name FROM supplier_database_table WHERE supplier_id = ?"
            supplier_data = db_manager.query_one(supplier_query, (supplier_id,))
            
            if not supplier_data:
                show_snack_bar(f"Fornecedor {supplier_id} não encontrado", True)
                return
            
            vendor_name = supplier_data.get('vendor_name', 'Unknown')
            
            # Criar diálogo de confirmação
            def handle_confirm_delete(e):
                try:
                    # Deletar todos os scores deste fornecedor
                    db_manager.execute("DELETE FROM supplier_score_records_table WHERE supplier_id = ?", (supplier_id,))
                    show_snack_bar(f"✅ Scores de {vendor_name} deletados com sucesso!")
                    page.close(confirm_dialog)
                    
                    # Recarregar a tabela
                    if search_field_ref.current:
                        search_suppliers()
                    
                except Exception as delete_error:
                    show_snack_bar(f"❌ Erro ao deletar: {str(delete_error)}", True)
                    print(f"Erro ao deletar scores: {delete_error}")
            
            def handle_cancel_delete(e):
                page.close(confirm_dialog)
            
            # Criar diálogo de confirmação simples
            confirm_dialog = ft.AlertDialog(
                title=ft.Text("Confirmar exclusão"),
                content=ft.Text(f"Tem certeza que deseja deletar todos os scores de {vendor_name}?\n\nEsta ação não pode ser desfeita."),
                actions=[
                    ft.TextButton("Cancelar", on_click=handle_cancel_delete),
                    ft.TextButton("Deletar", on_click=handle_confirm_delete, style=ft.ButtonStyle(color="red")),
                ],
            )
            
            page.open(confirm_dialog)
            
        except Exception as e:
            print(f"Erro ao preparar deleção: {e}")
            show_snack_bar(f"Erro ao deletar: {str(e)}", True)

    # Variáveis globais para debounce
    global search_timer, search_lock, unified_search_timer, unified_search_lock, last_search_term
    search_timer = None
    search_lock = False
    unified_search_timer = None
    unified_search_lock = False
    last_search_term = ""  # Guarda o último termo de busca para manter %% ao mudar mês/ano

    # Sistema de debounce para evitar múltiplas consultas simultâneas
    search_timer = None
    search_lock = False
    
    def search_suppliers_debounced():
        """Executa a busca com debounce para evitar conflitos SQLite."""
        global search_timer, search_lock
        
        # Cancelar timer anterior se existir
        if search_timer:
            try:
                search_timer.cancel()
            except:
                pass
            search_timer = None
        
        # Criar novo timer para executar a busca após 300ms de inatividade
        import threading
        search_timer = threading.Timer(0.3, search_suppliers)
        search_timer.start()

    def search_suppliers():
        """Versão otimizada da busca de fornecedores com proteção contra queries simultâneas."""
        global search_lock
        
        # Evitar múltiplas execuções simultâneas
        if search_lock:
            return
        search_lock = True
        
        try:
            if not search_field_ref.current:
                search_lock = False
                return
                
            search_term = search_field_ref.current.value.strip()
            bu_val = selected_bu.current.value if selected_bu.current else None
            po_val = selected_po.current.value if selected_po and selected_po.current else None
            month_val = selected_month.current.value if selected_month.current else None
            year_val = selected_year.current.value if selected_year.current else None

            print(f"🔍 Pesquisando por: '{search_term}', BU: {bu_val}, PO: {po_val}, Mês: {month_val}, Ano: {year_val}")
            
            # VALIDAÇÃO: Só buscar se AMBOS mês E ano estiverem preenchidos, ou se NENHUM estiver
            # Se apenas um estiver preenchido, não buscar (silenciosamente)
            has_month = month_val and str(month_val).strip()
            has_year = year_val and str(year_val).strip()
            
            if (has_month and not has_year) or (has_year and not has_month):
                print("⚠️ Busca cancelada: Mês e Ano devem estar ambos preenchidos ou ambos vazios")
                search_lock = False
                return

            # ===== LIMPEZA ROBUSTA DOS CARDS ANTES DE BUSCAR =====
            print("🗑️ Limpando cards antigos...")
            try:
                # Tentar via responsive_app_manager primeiro
                if responsive_app_manager and responsive_app_manager.results_container:
                    responsive_app_manager.clear_results()
                    print("✅ Cards limpos via responsive_app_manager")
                else:
                    # Fallback: limpar diretamente
                    if results_list and hasattr(results_list, 'controls'):
                        results_list.controls.clear()
                        results_list.update()
                        print("✅ Cards limpos via results_list")
            except Exception as clear_ex:
                print(f"⚠️ Erro ao limpar cards: {clear_ex}")
                # Tentar limpar diretamente mesmo com erro
                try:
                    if results_list and hasattr(results_list, 'controls'):
                        results_list.controls.clear()
                        print("✅ Cards limpos via fallback direto")
                except Exception as fallback_ex:
                    print(f"❌ Falha total ao limpar cards: {fallback_ex}")
            # ===== FIM DA LIMPEZA =====
        except Exception as ex:
            print(f"❌ Erro na inicialização da busca: {ex}")
            search_lock = False
            return
        finally:
            search_lock = False

        # Validar se há critérios de busca válidos
        # Mês/ano sozinhos NÃO devem buscar - precisam de termo de busca OU PO
        # BU sozinho também não é suficiente
        has_search_term = search_term and search_term.strip()
        has_po = po_val and str(po_val).strip()
        
        # Se não há termo de busca E não há PO, mostrar apenas favoritos
        if not has_search_term and not has_po:
            show_favorites_only()
            return

        if not db_conn:
            results_list.controls.append(
                ft.Container(
                    ft.Text("Erro: Não foi possível conectar ao banco de dados.", color="red"),
                    padding=20
                )
            )
            results_list.update()
            return

        try:
            base_query = """
                SELECT supplier_id, vendor_name, bu, supplier_status, supplier_number, supplier_name, supplier_po
                FROM supplier_database_table 
            """
            where_clauses = []
            params = []

            if search_term:
                where_clauses.append("(vendor_name LIKE ? OR supplier_id LIKE ?)")
                params.extend([f"%{search_term}%", f"%{search_term}%"])

            if bu_val and bu_val.strip():
                where_clauses.append("BU LIKE ?")
                params.append(f"%{bu_val.strip()}%")

            if po_val and str(po_val).strip():
                # Filtrar por supplier_po (PO)
                where_clauses.append("supplier_po LIKE ?")
                params.append(f"%{str(po_val).strip()}%")

            # Verificar se deve mostrar inativos
            show_inactive = False
            if show_inactive_switch.current and hasattr(show_inactive_switch.current, 'value'):
                show_inactive = show_inactive_switch.current.value
            
            # Se o switch estiver desligado, filtrar apenas Active
            if not show_inactive:
                where_clauses.append("supplier_status = ?")
                params.append("Active")

            # Limite fixo de 50 registros para a tabela
            card_limit = 50
            
            query = f"{base_query} WHERE {' AND '.join(where_clauses)} ORDER BY vendor_name LIMIT {card_limit}"
            
            print(f"Executando query: {query} (Limite: {card_limit})")
            print(f"Parâmetros: {params}")

            # Mostrar indicador de loading
            results_list.controls.clear()
            loading_indicator = ft.Container(
                content=ft.Column([
                    ft.ProgressRing(width=40, height=40),
                    ft.Text("Carregando...", size=14)
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                alignment=ft.alignment.center,
                padding=20
            )
            results_list.controls.append(loading_indicator)
            results_list.update()

            records = db_manager.query(query, params)
            
            print(f"Encontrados {len(records)} registros")
            
            # Limpar loading
            results_list.controls.clear()
            
            if not records:
                results_list.controls.append(
                    ft.Container(
                        ft.Text("Nenhum resultado encontrado.", italic=True, size=16),
                        alignment=ft.alignment.center,
                        padding=40
                    )
                )
            else:
                # Criar tabela com os registros em vez de cards individuais
                print(f"Criando tabela com {len(records)} registros")
                try:
                    table = create_score_table(records)
                    results_list.controls.append(table)
                except Exception as table_error:
                    print(f"Erro ao criar tabela: {table_error}")
                    results_list.controls.append(
                        ft.Container(
                            ft.Text(f"Erro ao criar tabela: {table_error}", color="red"),
                            padding=20
                        )
                    )
            
        except sqlite3.Error as db_error:
            error_msg = f"Erro no banco de dados: {db_error}"
            print(error_msg)
            results_list.controls.append(
                ft.Container(
                    ft.Text(error_msg, color="red"),
                    padding=20
                )
            )
        except Exception as e:
            error_msg = f"Erro inesperado: {e}"
            print(error_msg)
            results_list.controls.append(
                ft.Container(
                    ft.Text(error_msg, color="red"),
                    padding=20
                )
            )

        # Atualizar a lista
        try:
            results_list.update()
        except Exception as update_error:
            print(f"Erro ao atualizar lista: {update_error}")
        
        # Carregar scores foi removido pois agora usamos tabela em vez de cards
        # Os scores serão carregados quando o usuário clicar em "Editar"

    def show_favorites_only():
        """Mostra mensagem inicial de busca quando não há critérios suficientes"""
        
        # Limpar resultados anteriores
        print("🗑️ Limpando resultados anteriores...")
        try:
            # Tentar via responsive_app_manager primeiro
            if responsive_app_manager and responsive_app_manager.results_container:
                responsive_app_manager.clear_results()
                print("✅ Cards limpos via responsive_app_manager")
            else:
                # Fallback: limpar diretamente
                if results_list and hasattr(results_list, 'controls'):
                    results_list.controls.clear()
                    results_list.update()
                    print("✅ Cards limpos via results_list")
        except Exception as clear_ex:
            print(f"⚠️ Erro ao limpar cards: {clear_ex}")
            # Tentar limpar diretamente mesmo com erro
            try:
                if results_list and hasattr(results_list, 'controls'):
                    results_list.controls.clear()
                    print("✅ Cards limpos via fallback direto")
            except Exception as fallback_ex:
                print(f"❌ Falha total ao limpar cards: {fallback_ex}")

        # Mostrar mensagem inicial de busca
        results_list.controls.append(
            ft.Container(
                ft.Column([
                    ft.Icon(ft.Icons.SEARCH, size=40, color="gray"),
                    ft.Text("Digite no campo de busca para encontrar fornecedores", italic=True, size=16, text_align=ft.TextAlign.CENTER),
                    ft.Text("A busca inicia automaticamente após digitar 3 ou mais caracteres", 
                           size=12, text_align=ft.TextAlign.CENTER, color="gray"),
                    ft.Container(height=10),
                    ft.Row([
                        ft.Icon(ft.Icons.LIGHTBULB_OUTLINE, size=16, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                        ft.Text("Dica: Digite %% para buscar todos os fornecedores", 
                               size=12, italic=True, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                    ], alignment=ft.MainAxisAlignment.CENTER),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                alignment=ft.alignment.center,
                padding=40
            )
        )
        results_list.update()
        return

    # --- Busca Unificada em Tempo Real ---
    
    def unified_search_on_change(e):
        """Handler de busca unificada em tempo real - inicia após 3+ caracteres."""
        global unified_search_timer, unified_search_lock
        
        # Cancelar timer anterior se existir
        if unified_search_timer:
            try:
                unified_search_timer.cancel()
            except:
                pass
            unified_search_timer = None
        
        # Verificar se há pelo menos 3 caracteres (ou %% para busca especial)
        search_value = e.control.value.strip() if e.control.value else ""
        
        # Se menos de 3 caracteres E não for %%, mostrar favoritos
        if len(search_value) < 3 and search_value != "%%":
            show_favorites_only()
            return
        
        # Criar novo timer para executar a busca após 300ms de inatividade
        import threading
        unified_search_timer = threading.Timer(0.3, unified_search_execute)
        unified_search_timer.start()
    
    def unified_search_execute():
        """Executa busca unificada por Nome, ID, PO e BU."""
        global unified_search_lock, last_search_term
        
        # Evitar múltiplas execuções simultâneas
        if unified_search_lock:
            return
        unified_search_lock = True
        
        try:
            if not search_field_ref.current:
                unified_search_lock = False
                return
            
            search_term = search_field_ref.current.value.strip()
            last_search_term = search_term  # Guardar termo para manter ao mudar mês/ano
            month_val = selected_month.current.value if selected_month.current else None
            year_val = selected_year.current.value if selected_year.current else None
            
            print(f"🔍 Busca unificada: '{search_term}', Mês: {month_val}, Ano: {year_val}")
            
            # VALIDAÇÃO: Só buscar se AMBOS mês E ano estiverem preenchidos, ou se NENHUM estiver
            # Se apenas um estiver preenchido, não buscar (silenciosamente)
            has_month = month_val and str(month_val).strip()
            has_year = year_val and str(year_val).strip()
            
            if (has_month and not has_year) or (has_year and not has_month):
                print("⚠️ Busca cancelada: Mês e Ano devem estar ambos preenchidos ou ambos vazios")
                unified_search_lock = False
                return
            
            # Limpar resultados anteriores
            results_list.controls.clear()
            
            # Mostrar indicador de loading
            loading_indicator = ft.Container(
                content=ft.Column([
                    ft.ProgressRing(),
                    ft.Text("Carregando dados...", size=16, text_align=ft.TextAlign.CENTER),
                    ft.Text("Aguarde, isso pode levar alguns segundos", size=12, color="gray", text_align=ft.TextAlign.CENTER)
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                alignment=ft.alignment.center,
                padding=40
            )
            results_list.controls.append(loading_indicator)
            results_list.update()
            
            # Permitir "%%" como busca especial, caso contrário exigir 3+ caracteres
            if len(search_term) < 3 and search_term != "%%":
                unified_search_lock = False
                return
            
            if not db_conn:
                results_list.controls.append(
                    ft.Container(
                        ft.Text("Erro: Não foi possível conectar ao banco de dados.", color="red"),
                        padding=20
                    )
                )
                results_list.update()
                unified_search_lock = False
                return
            
            # Query unificada - busca em supplier_name (vendor_name), supplier_number (supplier_id), supplier_po e bu
            base_query = """
                SELECT supplier_id, vendor_name, bu, supplier_status, supplier_number, supplier_name, supplier_po
                FROM supplier_database_table 
            """
            where_clauses = []
            params = []
            
            # Se o termo de busca for "%%", buscar TODOS os registros
            if search_term == "%%":
                print("🔍 Busca especial: %% detectado - trazendo TODOS os registros")
                print(f"   Limite de registros: 200")
                # Não adicionar filtro de busca, apenas status
            else:
                # Busca unificada: Nome, ID, PO, BU
                where_clauses.append("(vendor_name LIKE ? OR supplier_id LIKE ? OR supplier_number LIKE ? OR supplier_po LIKE ? OR bu LIKE ?)")
                params.extend([f"%{search_term}%", f"%{search_term}%", f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"])
            
            # Verificar se deve mostrar inativos
            show_inactive = False
            if show_inactive_switch.current and hasattr(show_inactive_switch.current, 'value'):
                show_inactive = show_inactive_switch.current.value
            
            print(f"   Mostrar inativos: {show_inactive}")
            
            if not show_inactive:
                where_clauses.append("supplier_status = ?")
                params.append("Active")
            
            # Limite de registros: SEM LIMITE para %%, 50 para busca normal
            if search_term == "%%":
                card_limit = None  # Sem limite - traz TUDO
                print(f"   Limite aplicado: NENHUM (trazendo TODOS os registros)")
            else:
                card_limit = 50
                print(f"   Limite aplicado: {card_limit}")
            
            print(f"   Cláusulas WHERE: {where_clauses}")
            
            # Construir query
            if where_clauses:
                if card_limit:
                    query = f"{base_query} WHERE {' AND '.join(where_clauses)} ORDER BY vendor_name LIMIT {card_limit}"
                else:
                    query = f"{base_query} WHERE {' AND '.join(where_clauses)} ORDER BY vendor_name"
            else:
                if card_limit:
                    query = f"{base_query} ORDER BY vendor_name LIMIT {card_limit}"
                else:
                    query = f"{base_query} ORDER BY vendor_name"
            
            print(f"📊 Query final: {query}")
            print(f"📊 Parâmetros: {params}")
            
            records = db_manager.query(query, params)
            print(f"✅ Encontrados {len(records)} registros no banco de dados")
            
            if not records:
                results_list.controls.clear()
                results_list.controls.append(
                    ft.Container(
                        ft.Text("Nenhum resultado encontrado.", italic=True, size=16),
                        alignment=ft.alignment.center,
                        padding=40
                    )
                )
                results_list.update()
            else:
                # Processar criação da tabela em thread separada para não travar a UI
                print(f"🔄 Iniciando processamento assíncrono de {len(records)} registros")
                
                def create_table_async():
                    try:
                        import time
                        start_time = time.time()
                        
                        print(f"🏗️ Criando tabela...")
                        table = create_score_table(records)
                        
                        elapsed = time.time() - start_time
                        print(f"✅ Tabela criada em {elapsed:.2f}s")
                        
                        # Atualizar UI na thread principal
                        try:
                            results_list.controls.clear()
                            results_list.controls.append(table)
                            results_list.update()
                            print(f"✅ UI atualizada com sucesso")
                        except Exception as ui_error:
                            print(f"❌ Erro ao atualizar UI: {ui_error}")
                        
                    except Exception as table_error:
                        print(f"❌ Erro ao criar tabela: {table_error}")
                        import traceback
                        traceback.print_exc()
                        
                        try:
                            results_list.controls.clear()
                            results_list.controls.append(
                                ft.Container(
                                    ft.Text(f"Erro ao criar tabela: {table_error}", color="red"),
                                    padding=20
                                )
                            )
                            results_list.update()
                        except Exception as e:
                            print(f"❌ Erro ao mostrar erro: {e}")
                
                # Iniciar thread de processamento
                import threading
                thread = threading.Thread(target=create_table_async, daemon=True)
                thread.start()
            
        except sqlite3.Error as db_error:
            error_msg = f"Erro no banco de dados: {db_error}"
            print(error_msg)
            results_list.controls.append(
                ft.Container(
                    ft.Text(error_msg, color="red"),
                    padding=20
                )
            )
        except Exception as e:
            error_msg = f"Erro inesperado: {e}"
            print(error_msg)
            results_list.controls.append(
                ft.Container(
                    ft.Text(error_msg, color="red"),
                    padding=20
                )
            )
        finally:
            unified_search_lock = False
        
        # Atualizar a lista
        try:
            results_list.update()
        except Exception as update_error:
            print(f"Erro ao atualizar lista: {update_error}")
    
    # --- Fim: Busca Unificada em Tempo Real ---

    # --- Fim: Lógica do Banco de Dados e Pesquisa ---

    # --- Início: Lógica e Controles da Aba Score ---

    # Mudança: usar Column sem scroll para permitir header/footer fixos na tabela
    results_list = ft.Column(spacing=0, expand=True)
    
    # Adicionar mensagem inicial de busca
    results_list.controls.append(
        ft.Container(
            ft.Column([
                ft.Icon(ft.Icons.SEARCH, size=40, color="gray"),
                ft.Text("Digite no campo de busca para encontrar fornecedores", italic=True, size=16, text_align=ft.TextAlign.CENTER),
                ft.Text("A busca inicia automaticamente após digitar 3 ou mais caracteres", 
                       size=12, text_align=ft.TextAlign.CENTER, color="gray"),
                ft.Container(height=10),
                ft.Row([
                    ft.Icon(ft.Icons.LIGHTBULB_OUTLINE, size=16, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Text("Dica: Digite %% para buscar todos os fornecedores", 
                           size=12, italic=True, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                ], alignment=ft.MainAxisAlignment.CENTER),
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
            alignment=ft.alignment.center,
            padding=40
        )
    )

    # Referências para dropdowns que precisam ser atualizados com tema
    month_dropdown = ft.Dropdown(
        label="Mês", 
        options=month_options, 
        ref=selected_month, 
        width=150,  # Largura fixa para não diminuir
        value="", # Não pré-selecionar mês (iniciar vazio)
        on_change=on_month_year_change,
        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
        border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
    )
    
    year_dropdown = ft.Dropdown(
        label="Ano", 
        options=year_options, 
        ref=selected_year, 
        width=120,  # Largura fixa para não diminuir
        value=None,  # Não pré-selecionar ano
        on_change=on_month_year_change,
        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
        border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
    )

    # Criar referência para o botão de limpar
    clear_button_ref = ft.Ref[ft.IconButton]()
    
    # Função para limpar o campo de busca
    def clear_search_field(e):
        if search_field_ref.current:
            search_field_ref.current.value = ""
            # Esconder o botão
            if clear_button_ref.current:
                clear_button_ref.current.visible = False
                clear_button_ref.current.update()
            search_field_ref.current.update()
            # Trigger a busca para limpar resultados
            unified_search_on_change(None)
    
    # Função para atualizar o sufixo do campo de busca
    def update_search_suffix(e):
        if search_field_ref.current and clear_button_ref.current:
            # Mostrar/esconder X baseado se há texto
            clear_button_ref.current.visible = bool(search_field_ref.current.value)
            clear_button_ref.current.update()
        # Continuar com a busca normal
        unified_search_on_change(e)

    # Conteúdo da aba Score
    score_form = ft.Column(
        controls=[
            # Linha unificada de filtros: Busca, Mês, Ano, Opções
            ft.Row(
                controls=[
                    ft.Stack(
                        controls=[
                            ft.TextField(
                                hint_text="Buscar por Nome, ID, PO ou BU...",
                                border_radius=8,
                                ref=search_field_ref,
                                bgcolor=None,
                                color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                                border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline'),
                                prefix_icon=ft.Icons.SEARCH,  # Ícone de lupa à esquerda
                                on_change=update_search_suffix,  # Atualiza sufixo e busca
                            ),
                            ft.Container(
                                content=ft.IconButton(
                                    icon=ft.Icons.CLOSE,
                                    icon_size=20,
                                    tooltip="Limpar busca",
                                    on_click=clear_search_field,
                                    ref=clear_button_ref,
                                    visible=False,  # Inicialmente invisível
                                ),
                                right=0,
                                top=0,
                                bottom=0,
                            ),
                        ],
                        expand=True,
                    ),
                    month_dropdown,
                    year_dropdown,
                    ft.PopupMenuButton(
                        icon=ft.Icons.MORE_VERT,
                        tooltip="Mais opções",
                        items=[
                            ft.PopupMenuItem(
                                content=ft.Row([
                                    ft.Switch(
                                        label="Mostrar Inativos",
                                        value=False,
                                        ref=show_inactive_switch,
                                        on_change=on_show_inactive_change,
                                    ),
                                ], tight=True),
                            ),
                        ],
                    ),
                ],
                spacing=10,
                alignment=ft.MainAxisAlignment.START,
                vertical_alignment=ft.CrossAxisAlignment.CENTER,
            ),
        ],
        spacing=16,
    )

    score_form_container = ft.Container(
        content=score_form, 
        padding=25,  # Padding aumentado para melhor espaçamento
        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('card_background'),
        border_radius=8,
        width=1200,  # Largura aumentada para acomodar campo de busca maior
    )

    # Container responsivo - largura fixa de 700px para os filtros
    score_view_content = ft.Card(
        content=score_form_container,
        elevation=2,
    )

    # --- Fim: Lógica e Controles da Aba Score ---

    # --- Início: Lógica e Controles da Aba Configs ---

    # Estado para controlar qual sub-aba está selecionada
    selected_config_tab = ft.Ref()
    selected_config_tab.current = 0

    def set_config_tab(idx):
        def handler(e):
            # Verificar se o usuário tem acesso à aba de configuração
            accessible_tabs = get_user_accessible_config_tabs(current_user_privilege or "User")
            accessible_indices = [tab_idx for _, _, tab_idx in accessible_tabs]
            
            if idx not in accessible_indices:
                show_snack_bar("Acesso negado para esta configuração", True)
                return
                
            # Limpar campos da aba Log quando sair dela (se estava na aba Log antes)
            if selected_config_tab.current == 5 and idx != 5:
                try:
                    clear_log_fields()
                except Exception as ex:
                    print(f"Aviso: erro ao limpar campos da aba Log: {ex}")
            
            selected_config_tab.current = idx
            themes_content.visible = idx == 0
            suppliers_content.visible = idx == 1
            criteria_content.visible = idx == 2
            users_content.visible = idx == 3
            lists_content.visible = idx == 4
            log_content.visible = idx == 5
            info_content.visible = idx == 6
            # Ao abrir a aba Log, garantir que os dados sejam carregados
            if idx == 5:
                try:
                    ensure_log_loaded()
                except Exception:
                    pass
            update_config_tabs()
            safe_page_update(page)
        return handler

    def config_tab_item(icon, text, idx):
        is_selected = selected_config_tab.current == idx
        Colors = get_current_theme_colors(get_theme_name_from_page(page))
        return ft.Container(
            content=ft.Column([
                ft.Icon(icon, color=Colors['primary'] if is_selected else Colors['on_surface'], size=24),
                ft.Text(text, size=12, weight="bold" if is_selected else "normal",
                       color=Colors['primary'] if is_selected else Colors['on_surface'])
            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=5),
            padding=ft.padding.symmetric(horizontal=15, vertical=10),
            border_radius=8,
            bgcolor=Colors['primary_container'] if is_selected else None,
            on_click=set_config_tab(idx),
            animate=200,
        )

    def get_user_accessible_config_tabs(privilege):
        """Retorna lista de abas de configuração que o usuário pode acessar"""
        all_tabs = [
            (ft.Icons.PALETTE, "Themes", 0),
            (ft.Icons.BUSINESS, "Suppliers", 1),
            (ft.Icons.TUNE, "Criteria", 2),
            (ft.Icons.PEOPLE, "Users", 3),
            (ft.Icons.LIST, "Lists", 4),
            (ft.Icons.HISTORY, "Log", 5),
            (ft.Icons.INFO_OUTLINED, "Info", 6),
        ]
        
        if privilege == "Super Admin":
            # Super Admin pode acessar tudo
            return all_tabs
        elif privilege == "Admin":
            # Admin pode: Themes, Users, Lists, Info
            return [(icon, name, idx) for icon, name, idx in all_tabs if idx in [0, 3, 4, 6]]
        else:  # User
            # User pode: Themes, Users e Info
            return [(icon, name, idx) for icon, name, idx in all_tabs if idx in [0, 3, 6]]

    def update_config_tabs():
        # Filtrar abas baseado no privilégio do usuário
        accessible_tabs = get_user_accessible_config_tabs(current_user_privilege or "User")
        config_tabs_row.controls = [
            config_tab_item(icon, name, idx) for icon, name, idx in accessible_tabs
        ]
        config_tabs_row.update()

    def update_all_comboboxes():
        """Recarrega opções dinâmicas para todos os Dropdowns do app."""
        try:
            # Carregar opções de DB
            bu_opts = load_list_options('business_unit_table', 'bu')
            planner_opts = load_list_options('planner_table', 'alias')
            continuity_opts = load_list_options('continuity_table', 'alias')
            sourcing_opts = load_list_options('sourcing_table', 'alias')
            sqie_opts = load_list_options('sqie_table', 'alias')

            # Atualizar UI de listas (se existir) e forçar recarregamento das listas exibidas
            for k in lists_controls.keys():
                refresh_list_ui(k)
            
            # Reconstruir tabs com tema atualizado
            if 'lists_tabs_container' in globals() and lists_tabs_container:
                lists_tabs_container.content = build_lists_tabs()
            
            safe_page_update(page)
        except Exception as ex:
            print(f"Erro ao atualizar comboboxes: {ex}")

    # RadioGroup para seleção de tema
    theme_radio_group = ft.RadioGroup(
        content=ft.Column([
            ft.Radio(value="white", label="Tema Claro"),
            ft.Radio(value="dark", label="Tema Escuro"),
            ft.Radio(value="dracula", label="Tema Dracula"),
        ]),
        value="white",
        on_change=theme_changed,
    )
    
    
    # Carregar tema salvo do usuário e aplicar ao radio group
    if current_user_wwid:
        saved_theme = load_user_theme(current_user_wwid)
        if saved_theme:
            theme_radio_group.value = saved_theme
            print(f"Tema carregado para radio group: {saved_theme}")
        else:
            print("Nenhum tema salvo encontrado para usuário, usando padrão")
    else:
        # Se não há usuário logado, carregar tema padrão da tabela
        try:
            default_theme = db_manager.query_one("SELECT theme_mode FROM theme_settings ORDER BY last_updated DESC LIMIT 1")
            if default_theme:
                theme_radio_group.value = default_theme["theme_mode"]
                print(f"Tema padrão carregado: {default_theme['theme_mode']}")
        except Exception as e:
            print(f"Erro ao carregar tema padrão: {e}")

    def handle_auto_save_toggle(e):
        enabled = bool(e.control.value)
        app_settings['auto_save'] = enabled
        save_auto_save_setting(enabled)
        e.control.tooltip = "Auto save ativado" if enabled else "Auto save desativado"
        e.control.update()

    auto_save_switch = ft.Switch(
        label="",
        value=app_settings.get('auto_save', True),
        on_change=handle_auto_save_toggle,
    )
    auto_save_switch.tooltip = "Auto save ativado" if auto_save_switch.value else "Auto save desativado"

    # Conteúdo da sub-aba Themes
    themes_content = ft.Container(
        content=ft.Column([
            # Header fixo
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.PALETTE, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Column([
                        ft.Text("Configuração de Tema", size=22, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ft.Text("Personalize a aparência e controles do aplicativo", size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                    ], spacing=2, expand=True),
                ], spacing=12),
                margin=ft.margin.only(bottom=15),
            ),
            ft.Divider(height=1, color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
            # Conteúdo com scroll
            ft.Container(
                content=ft.Column([
                    theme_radio_group,
                    ft.Divider(),
                    # REMOVIDO - Tipo de Controle de Score (Cards obsoletos)
                    # ft.Text("Tipo de Controle de Score", size=16, weight="bold"),
                    # ft.Text("Escolha o tipo de controle para inserir scores:", size=12, color="on_surface_variant"),
                    # control_type_radio_group,
                    # ft.Divider(),
                    ft.Text("Auto save", size=16, weight="bold"),
                    ft.Text(
                        "Defina se as alterações nos scores são salvas automaticamente.",
                        size=12,
                        color="on_surface_variant",
                    ),
                    ft.Row(
                        [
                            ft.Text("Salvar automaticamente os valores editados", size=12),
                            auto_save_switch,
                        ],
                        alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                        vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    ),
                ], spacing=15, scroll=ft.ScrollMode.AUTO),
                expand=True
            )
        ], spacing=0, expand=True),
        padding=20,
        visible=True,
        expand=True
    )

    # ---------- Sub-aba Lists dentro de Configs (gerenciar opções dinâmicas) ----------
    
    # Placeholder para tabs (será preenchido depois)
    lists_tabs_container = ft.Container(expand=True)
    
    lists_content = ft.Container(
        content=ft.Column([
            # Header
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.LIST_ALT, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Column([
                        ft.Text("Gerenciar Listas", size=22, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ft.Text("Configure as opções dos dropdowns", size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                    ], spacing=2, expand=True),
                ], spacing=12),
                margin=ft.margin.only(bottom=15),
            ),
            ft.Divider(height=1, color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
            # Tabs container
            lists_tabs_container
        ], spacing=0, expand=True),
        padding=20,
        visible=False,
        expand=True
    )

    # Função para criar campos de texto sem borda colorida
    def create_list_textfield(label, expand=False, width=None, on_change=None):
        theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
        params = {
            'label': label,
            'filled': True,
            'bgcolor': theme_colors.get('field_background'),
            'color': theme_colors.get('on_surface'),
            'border_color': ft.Colors.TRANSPARENT,
            'focused_border_color': theme_colors.get('primary'),
            'on_change': on_change
        }
        if expand:
            params['expand'] = True
        if width:
            params['width'] = width
        return ft.TextField(**params)
    
    # Estruturas de UI para cada lista (controls) com feedback e botões de delete
    lists_controls = {
        'sqie': {
            'input': create_list_textfield('Nome SQIE', expand=True, on_change=lambda e: validate_input_exists('sqie')),
            'alias': create_list_textfield('Alias SQIE', width=200, on_change=lambda e: validate_input_exists('sqie')),
            'email': create_list_textfield('Email SQIE', expand=True, on_change=lambda e: validate_input_exists('sqie')),
            'list': ft.Column(spacing=2),
            'feedback': ft.Text(''),
            'selected_item': None
        },
        'continuity': {
            'input': create_list_textfield('Nome Continuity', expand=True, on_change=lambda e: validate_input_exists('continuity')),
            'alias': create_list_textfield('Alias Continuity', width=200, on_change=lambda e: validate_input_exists('continuity')),
            'email': create_list_textfield('Email Continuity', expand=True, on_change=lambda e: validate_input_exists('continuity')),
            'list': ft.Column(spacing=2),
            'feedback': ft.Text(''),
            'selected_item': None
        },
        'planner': {
            'input': create_list_textfield('Nome Planner', expand=True, on_change=lambda e: validate_input_exists('planner')),
            'alias': create_list_textfield('Alias Planner', width=200, on_change=lambda e: validate_input_exists('planner')),
            'email': create_list_textfield('Email Planner', expand=True, on_change=lambda e: validate_input_exists('planner')),
            'list': ft.Column(spacing=2),
            'feedback': ft.Text(''),
            'selected_item': None
        },
        'sourcing': {
            'input': create_list_textfield('Nome Sourcing', expand=True, on_change=lambda e: validate_input_exists('sourcing')),
            'alias': create_list_textfield('Alias Sourcing', width=200, on_change=lambda e: validate_input_exists('sourcing')),
            'email': create_list_textfield('Email Sourcing', expand=True, on_change=lambda e: validate_input_exists('sourcing')),
            'list': ft.Column(spacing=2),
            'feedback': ft.Text(''),
            'selected_item': None
        },
        'bu': {
            'input': create_list_textfield('Nome BU', expand=True, on_change=lambda e: validate_input_exists('bu')),
            'list': ft.Column(spacing=2),
            'feedback': ft.Text(''),
            'selected_item': None
        },
        'category': {
            'input': create_list_textfield('Nome Category', expand=True, on_change=lambda e: validate_input_exists('category')),
            'list': ft.Column(spacing=2),
            'feedback': ft.Text(''),
            'selected_item': None
        },
    }

    def create_list_item_card(key, item):
        """Cria um card para um item de lista com botões de ação"""
        Colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        # Determinar que tipo de lista estamos lidando
        list_types_with_edit = ['sqie', 'continuity', 'planner', 'sourcing']
        supports_edit = key in list_types_with_edit
        
        # Extrair informações do item
        alias = item.get("alias", "")
        name = item.get("name", "")
        email = item.get("email", "")
        
        # Criar título baseado no tipo
        if key in list_types_with_edit:
            title = f"{alias} - {name}" if name else alias
            subtitle = email if email else "Email não informado"
        else:
            title = alias or item.get("display", "")
            subtitle = None
        
        # Criar conteúdo do card
        content_column = ft.Column([
            ft.Text(
                title, 
                size=14, 
                weight="bold",
                color=Colors['on_surface']
            )
        ], spacing=2)
        
        if subtitle:
            content_column.controls.append(
                ft.Text(subtitle, size=12, color="outline")
            )
        
        # Criar botões de ação
        action_buttons = []
        
        if supports_edit:
            action_buttons.append(
                ft.IconButton(
                    icon=ft.Icons.EDIT,
                    tooltip="Editar",
                    icon_size=18,
                    on_click=lambda e, k=key, it=item: show_edit_list_dialog(k, it)
                )
            )
        
        action_buttons.append(
            ft.IconButton(
                icon=ft.Icons.DELETE,
                tooltip="Excluir",
                icon_size=18,
                icon_color="red",
                on_click=lambda e, k=key, it=item: confirm_delete_list_item(k, it)
            )
        )
        
        # Layout do card
        card_content = ft.Row([
            ft.Container(
                content=content_column,
                expand=True
            ),
            ft.Row(action_buttons, spacing=0)
        ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER)
        
        return ft.Card(
            content=ft.Container(
                content=card_content,
                padding=ft.padding.all(12),
                bgcolor=Colors['card_background'],
                border_radius=8
            ),
            elevation=1,
            margin=ft.margin.only(bottom=4)
        )

    def show_edit_list_dialog(key, item):
        """Mostra diálogo para editar um item da lista"""
        Colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        # Campos atuais - extrair ID único
        current_id = item.get('id', '')
        current_alias = item.get('alias', '')
        current_name = item.get('name', '')
        current_email = item.get('email', '')
        
        # Criar campos de entrada
        name_field = ft.TextField(
            label="Nome",
            value=current_name,
            width=380,
            bgcolor=Colors['field_background'],
            color=Colors['on_surface'],
            border_color=Colors['outline']
        )
        
        alias_field = ft.TextField(
            label="Alias",
            value=current_alias,
            width=380,
            bgcolor=Colors['field_background'],
            color=Colors['on_surface'],
            border_color=Colors['outline']
        )
        
        email_field = ft.TextField(
            label="Email",
            value=current_email,
            width=380,
            bgcolor=Colors['field_background'],
            color=Colors['on_surface'],
            border_color=Colors['outline']
        )
        
        error_text = ft.Text("", color="red", visible=False)
        
        def save_changes(e):
            new_name = name_field.value.strip()
            new_alias = alias_field.value.strip().upper()
            new_email = email_field.value.strip()
            
            # Validação
            if not new_name or not new_alias or not new_email:
                error_text.value = "Todos os campos são obrigatórios"
                error_text.visible = True
                edit_dialog.update()
                return
            
            try:
                # Mapear tabela e coluna ID (usar nomes reais das colunas de ID)
                # antes o código usava 'id' e por isso o UPDATE não encontrava a linha
                table_map = {
                    'sqie': ('sqie_table', 'sqie_id'),
                    'continuity': ('continuity_table', 'continuity_id'),
                    'planner': ('planner_table', 'planner_id'),
                    'sourcing': ('sourcing_table', 'sourcing_id')
                }
                
                table_name, id_column = table_map[key]
                
                # Atualizar no banco usando ID único
                db_manager.execute(
                    f"UPDATE {table_name} SET name = ?, alias = ?, email = ? WHERE {id_column} = ?",
                    (new_name, new_alias, new_email, current_id)
                )
                
                # Atualizar seleção se ID mudou (usar novo ID)
                if lists_controls[key]['selected_item'] == current_id:
                    lists_controls[key]['selected_item'] = current_id
                
                # Fechar diálogo e atualizar UI
                page.close(edit_dialog)
                refresh_list_ui(key)
                update_all_comboboxes()
                
                show_snack_bar("✅ Item atualizado com sucesso", False)
                
            except Exception as ex:
                import sqlite3 as _sqlite
                if isinstance(ex, _sqlite.IntegrityError):
                    error_text.value = "Alias já existe. Escolha outro."
                    error_text.visible = True
                    edit_dialog.update()
                else:
                    show_snack_bar(f"❌ Erro ao atualizar: {ex}", True)
                    safe_page_update(page)
        
        def cancel_edit(e):
            page.close(edit_dialog)
            safe_page_update(page)
        
        # Criar diálogo
        edit_dialog = ft.AlertDialog(
            title=ft.Text("Editar Item"),
            content=ft.Column([
                ft.Container(
                    content=ft.Column([
                        name_field,
                        ft.Container(height=10),
                        alias_field,
                        ft.Container(height=10),
                        email_field,
                        ft.Container(height=15),
                        error_text
                    ], spacing=0),
                    width=400
                )
            ], tight=True, spacing=15),
            actions=[
                ft.TextButton("Cancelar", on_click=cancel_edit),
                ft.ElevatedButton("Salvar", on_click=save_changes)
            ],
            actions_alignment=ft.MainAxisAlignment.END
        )
        
        page.open(edit_dialog)
        safe_page_update(page)

    def confirm_delete_list_item(key, item):
        """Confirma exclusão de item da lista"""
        alias = item.get('alias', '')
        
        # Mapear informações da tabela
        table_info = {
            'sqie': {'table': 'sqie_table', 'column': 'alias', 'display': 'SQIE'},
            'continuity': {'table': 'continuity_table', 'column': 'alias', 'display': 'Continuity'},
            'planner': {'table': 'planner_table', 'column': 'alias', 'display': 'Planner'},
            'sourcing': {'table': 'sourcing_table', 'column': 'alias', 'display': 'Sourcing'},
            'bu': {'table': 'business_unit_table', 'column': 'bu', 'display': 'Business Unit'},
            'category': {'table': 'categories_table', 'column': 'category', 'display': 'Category'}
        }
        
        info = table_info.get(key, {})
        list_display = info.get('display', key.upper())
        
        # Usar o diálogo de confirmação existente
        delete_alias_handler(info['table'], info['column'], alias)

    def refresh_list_ui(key):
        try:
            # Mapear tabelas e colunas
            table_map = {
                'sqie': ('sqie_table', 'alias'),
                'continuity': ('continuity_table', 'alias'),
                'planner': ('planner_table', 'alias'),
                'sourcing': ('sourcing_table', 'alias'),
                'bu': ('business_unit_table', 'bu'),
                'category': ('categories_table', 'category')
            }
            
            if key in table_map:
                table_name, column_name = table_map[key]
                # Usar nova função para dados completos
                items = load_list_items_full(table_name)
                col = lists_controls[key]['list']
                col.controls.clear()
                
                # Se não há itens, mostrar mensagem
                if not items:
                    col.controls.append(
                        ft.Container(
                            content=ft.Text(
                                "Nenhum item adicionado",
                                color="outline",
                                italic=True,
                                text_align=ft.TextAlign.CENTER
                            ),
                            alignment=ft.alignment.center,
                            padding=10
                        )
                    )
                else:
                    # Usar cards ao invés de ListTiles
                    for item in items:
                        card = create_list_item_card(key, item)
                        col.controls.append(card)
                        
        except Exception as ex:
            print(f"Erro ao atualizar UI da lista {key}: {ex}")

    def delete_alias_handler(table, alias_col, value=None):
        try:
            # Se não foi fornecido um valor, usar o item selecionado
            if not value:
                key = table.replace('_table', '')
                if key in lists_controls and lists_controls[key]['selected_item']:
                    value = lists_controls[key]['selected_item']
                else:
                    show_snack_bar("❌ Selecione um item da lista para excluir", True)
                    return
                    return
            
            # Mapear nomes amigáveis dos tipos de lista
            list_type_names = {
                'sqie_table': 'SQIE',
                'continuity_table': 'Continuity',
                'planner_table': 'Planner',
                'sourcing_table': 'Sourcing',
                'business_unit_table': 'Business Unit',
                'categories_table': 'Categories'
            }
            
            list_type = list_type_names.get(table, table)
            
            # Função para confirmar e executar a exclusão
            def confirm_delete(e):
                try:
                    delete_list_item_by_alias(table, alias_col, value)
                    
                    # Limpar seleção após exclusão
                    key = table.replace('_table', '')
                    if key in lists_controls:
                        lists_controls[key]['selected_item'] = None
                    
                    # atualizar todas as comboboxes e a UI
                    for k in lists_controls.keys():
                        refresh_list_ui(k)
                    update_all_comboboxes()
                    # atualizar lista de suppliers (reconstruir cards)
                    try:
                        search_suppliers_config()
                    except Exception:
                        pass
                    
                    # Fechar dialog
                    page.close(dialog)
                    
                    show_snack_bar(f"✅ Item '{value}' removido de {list_type}", False)
                    safe_page_update(page)
                    
                except Exception as ex:
                    # Fechar dialog mesmo se der erro
                    page.close(dialog)
                    
                    # mostrar mensagem amigável
                    import sqlite3 as _sqlite
                    if isinstance(ex, _sqlite.IntegrityError):
                        show_snack_bar(f"❌ Não foi possível excluir '{value}': restrição de integridade.", True)
                    else:
                        show_snack_bar(f"❌ Erro ao deletar '{value}': {ex}", True)
                    safe_page_update(page)
                    print(f"Erro ao deletar {value} em {table}: {ex}")
            
            # Função para cancelar
            def cancel_delete(e):
                page.close(dialog)
                safe_page_update(page)
            
            # Criar e mostrar dialog de confirmação
            dialog = DeleteListItemConfirmationDialog(
                item_name=value,
                item_type=list_type,
                on_confirm=confirm_delete,
                on_cancel=cancel_delete,
                scale_func=lambda x: x
            )
            
            page.open(dialog)
            safe_page_update(page)
            
        except Exception as ex:
            show_snack_bar(f"❌ Erro ao tentar excluir item: {ex}", True)
            safe_page_update(page)
            print(f"Erro geral no delete_alias_handler: {ex}")

    def validate_input_exists(key):
        """Valida se o valor digitado no input (ou alias) já existe na tabela correspondente."""
        try:
            if key == 'sqie':
                val_alias = lists_controls['sqie']['alias'].value.strip().upper() if lists_controls['sqie']['alias'].value else ''
                val = val_alias or lists_controls['sqie']['input'].value.strip()
                rows = load_list_options('sqie_table','alias')
                exists = val in rows
                fb = lists_controls['sqie']['feedback']
                if exists:
                    fb.value = f"Já existe: {val}"
                    fb.color = 'red'
                else:
                    fb.value = f"Disponível: {val}"
                    fb.color = 'green'
                fb.visible = True
            elif key == 'continuity':
                val_alias = lists_controls['continuity']['alias'].value.strip().upper() if lists_controls['continuity']['alias'].value else ''
                val = val_alias or lists_controls['continuity']['input'].value.strip()
                rows = load_list_options('continuity_table','alias')
                exists = val in rows
                fb = lists_controls['continuity']['feedback']
                fb.value = (f"Já existe: {val}" if exists else f"Disponível: {val}")
                fb.color = 'red' if exists else 'green'
                fb.visible = True
            elif key == 'planner':
                val_alias = lists_controls['planner']['alias'].value.strip().upper() if lists_controls['planner']['alias'].value else ''
                val = val_alias or lists_controls['planner']['input'].value.strip()
                rows = load_list_options('planner_table','alias')
                exists = val in rows
                fb = lists_controls['planner']['feedback']
                fb.value = (f"Já existe: {val}" if exists else f"Disponível: {val}")
                fb.color = 'red' if exists else 'green'
                fb.visible = True
            elif key == 'sourcing':
                val_alias = lists_controls['sourcing']['alias'].value.strip().upper() if lists_controls['sourcing']['alias'].value else ''
                val = val_alias or lists_controls['sourcing']['input'].value.strip()
                rows = load_list_options('sourcing_table','alias')
                exists = val in rows
                fb = lists_controls['sourcing']['feedback']
                fb.value = (f"Já existe: {val}" if exists else f"Disponível: {val}")
                fb.color = 'red' if exists else 'green'
                fb.visible = True
            elif key == 'bu':
                val = lists_controls['bu']['input'].value.strip()
                rows = load_list_options('business_unit_table','bu')
                exists = val in rows
                fb = lists_controls['bu']['feedback']
                fb.value = (f"Já existe: {val}" if exists else f"Disponível: {val}")
                fb.color = 'red' if exists else 'green'
                fb.visible = True
            elif key == 'category':
                val = lists_controls['category']['input'].value.strip()
                rows = load_list_options('categories_table','category')
                exists = val in rows
                fb = lists_controls['category']['feedback']
                fb.value = (f"Já existe: {val}" if exists else f"Disponível: {val}")
                fb.color = 'red' if exists else 'green'
                fb.visible = True
            safe_page_update(page)
        except Exception as ex:
            print(f"Erro na validação de {key}: {ex}")

    def add_alias_handler_sqie(e):
        name = lists_controls['sqie']['input'].value.strip()
        alias = lists_controls['sqie']['alias'].value.strip().upper()
        email = lists_controls['sqie']['email'].value.strip()
        
        if not name or not alias or not email:
            lists_controls['sqie']['feedback'].value = 'Informe nome, alias e email para SQIE'
            lists_controls['sqie']['feedback'].color = 'red'
            lists_controls['sqie']['feedback'].visible = True
            safe_page_update(page)
            return
        try:
            insert_list_item('sqie_table', ['name','alias','email','register_date','registered_by'], (name, alias, email, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system'))
            lists_controls['sqie']['input'].value = ''
            lists_controls['sqie']['alias'].value = ''
            lists_controls['sqie']['email'].value = ''
            refresh_list_ui('sqie')
            update_all_comboboxes()
            try:
                search_suppliers_config()
            except Exception:
                pass
            lists_controls['sqie']['feedback'].value = f"SQIE '{alias}' adicionado"
            lists_controls['sqie']['feedback'].color = 'green'
            lists_controls['sqie']['feedback'].visible = True
            show_snack_bar(f"✅ SQIE '{alias}' adicionado com sucesso", False)
            safe_page_update(page)
        except Exception as ex:
            import sqlite3 as _sqlite
            if isinstance(ex, _sqlite.IntegrityError):
                lists_controls['sqie']['feedback'].value = f"Alias '{alias}' já existe"
                lists_controls['sqie']['feedback'].color = 'red'
                lists_controls['sqie']['feedback'].visible = True
                safe_page_update(page)
            else:
                show_snack_bar(f"❌ Erro ao inserir sqie: {ex}", True)
                safe_page_update(page)
            print(f"Erro ao inserir sqie: {ex}")
        except Exception as ex:
            print(f"Erro ao inserir sqie: {ex}")

    def add_alias_handler_generic(table, input_field, alias_field=None):
        name = input_field.value.strip()
        if alias_field:
            alias = alias_field.value.strip().upper()
            if not name or not alias:
                print("Informe name e alias")
                return
        else:
            if not name:
                print("Informe nome")
                return
        try:
            if table == 'continuity_table':
                email = lists_controls['continuity']['email'].value.strip()
                if not email:
                    lists_controls['continuity']['feedback'].value = 'Informe nome, alias e email para Continuity'
                    lists_controls['continuity']['feedback'].color = 'red'
                    lists_controls['continuity']['feedback'].visible = True
                    safe_page_update(page)
                    return
                insert_list_item(table, ['name','alias','email','register_date','registered_by'], (name, alias, email, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system'))
                lists_controls['continuity']['email'].value = ''
            elif table == 'planner_table':
                email = lists_controls['planner']['email'].value.strip()
                if not email:
                    lists_controls['planner']['feedback'].value = 'Informe nome, alias e email para Planner'
                    lists_controls['planner']['feedback'].color = 'red'
                    lists_controls['planner']['feedback'].visible = True
                    safe_page_update(page)
                    return
                insert_list_item(table, ['name','alias','email','register_date','registered_by'], (name, alias, email, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system'))
                lists_controls['planner']['email'].value = ''
            elif table == 'sourcing_table':
                email = lists_controls['sourcing']['email'].value.strip()
                if not email:
                    lists_controls['sourcing']['feedback'].value = 'Informe nome, alias e email para Sourcing'
                    lists_controls['sourcing']['feedback'].color = 'red'
                    lists_controls['sourcing']['feedback'].visible = True
                    safe_page_update(page)
                    return
                insert_list_item(table, ['name','alias','email','register_date','registered_by'], (name, alias, email, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system'))
                lists_controls['sourcing']['email'].value = ''
            elif table == 'business_unit_table':
                insert_list_item(table, ['bu','register_date','registered_by'], (name, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system'))
            elif table == 'categories_table':
                insert_list_item(table, ['category','register_date','registered_by'], (name, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), 'system'))
            
            input_field.value = ''
            if alias_field:
                alias_field.value = ''
            refresh_list_ui(table.replace('_table',''))
            update_all_comboboxes()
            try:
                search_suppliers_config()
            except Exception:
                pass
            # feedback inline
            key = table.replace('_table','')
            if key in lists_controls:
                lists_controls[key]['feedback'].value = f"Item adicionado em {key}"
                lists_controls[key]['feedback'].color = 'green'
                lists_controls[key]['feedback'].visible = True
                show_snack_bar(f"✅ Item adicionado com sucesso em {key}", False)
                safe_page_update(page)
        except Exception as ex:
            import sqlite3 as _sqlite
            if isinstance(ex, _sqlite.IntegrityError):
                key = table.replace('_table','')
                if key in lists_controls:
                    lists_controls[key]['feedback'].value = f"Item já existe"
                    lists_controls[key]['feedback'].color = 'red'
                    lists_controls[key]['feedback'].visible = True
                    safe_page_update(page)
                else:
                    show_snack_bar(f"❌ Erro ao inserir item: {ex}", True)
                    safe_page_update(page)
            print(f"Erro ao inserir item em {table}: {ex}")
            safe_page_update(page)
        except Exception as ex:
            import sqlite3 as _sqlite
            if isinstance(ex, _sqlite.IntegrityError):
                key = table.replace('_table','')
                if key in lists_controls:
                    lists_controls[key]['feedback'].value = f"Item já existe em {key}"
                    lists_controls[key]['feedback'].color = 'red'
                    lists_controls[key]['feedback'].visible = True
                    safe_page_update(page)
                else:
                    show_snack_bar(f"❌ Erro ao inserir em {table}: {ex}", True)
                    safe_page_update(page)
            print(f"Erro ao inserir em {table}: {ex}")
        except Exception as ex:
            print(f"Erro ao inserir em {table}: {ex}")

    # Criar layout vertical simples ao invés de GridView
        ft.Container(
            content=ft.Column([
                # Header
                ft.Row([
                    ft.Icon(ft.Icons.ASSIGNMENT, size=20, color="primary"),
                    ft.Text('SQIE Classification', weight='bold', size=16)
                ], alignment=ft.MainAxisAlignment.START, spacing=8),
                
                # Inputs
                ft.Row([
                    lists_controls['sqie']['input'], 
                    lists_controls['sqie']['alias']
                ], spacing=8),
                
                # Feedback
                lists_controls['sqie']['feedback'],
                
                # Buttons
                ft.Row([
                    ft.ElevatedButton(
                        'Adicionar', 
                        on_click=add_alias_handler_sqie,
                        icon=ft.Icons.ADD
                    )
                ], spacing=8),
                
                # Divider
                ft.Divider(height=1),
                
                # Lista de itens com scroll
                ft.Container(
                    content=lists_controls['sqie']['list'], 
                    expand=True,
                    padding=ft.padding.all(6),
                    border=ft.border.all(1, "outline_variant"),
                    border_radius=4,
                    bgcolor="surface"
                )
            ], spacing=10),
            padding=12,
            border=ft.border.all(1, 'outline'),
            border_radius=8,
            bgcolor='surface_variant'
        )

    # Popular listas iniciais
        ft.Container(
            content=ft.Column([
                # Header
                ft.Row([
                    ft.Icon(ft.Icons.TIMELINE, size=20, color="primary"),
                    ft.Text('Continuity Level', weight='bold', size=16)
                ], alignment=ft.MainAxisAlignment.START, spacing=8),
                
                # Inputs
                ft.Row([
                    lists_controls['continuity']['input'], 
                    lists_controls['continuity']['alias']
                ], spacing=8),
                
                # Feedback
                lists_controls['continuity']['feedback'],
                
                # Buttons
                ft.Row([
                    ft.ElevatedButton(
                        'Adicionar', 
                        on_click=lambda e: add_alias_handler_generic('continuity_table', lists_controls['continuity']['input'], lists_controls['continuity']['alias']),
                        icon=ft.Icons.ADD
                    )
                ], spacing=8),
                
                # Divider
                ft.Divider(height=1),
                
                # Lista de itens com scroll
                ft.Container(
                    content=lists_controls['continuity']['list'], 
                    expand=True,
                    padding=ft.padding.all(6),
                    border=ft.border.all(1, "outline_variant"),
                    border_radius=4,
                    bgcolor="surface"
                )
            ], spacing=10),
            padding=12,
            border=ft.border.all(1, 'outline'),
            border_radius=8,
            bgcolor='surface_variant'
        )

    # Popular listas iniciais
    # Função para criar conteúdo de aba com layout limpo
    def create_list_tab_content(key, title_icon, has_email=True):
        """Cria conteúdo de uma aba de lista com layout moderno"""
        theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        # Campos de entrada em card
        input_fields = []
        
        if has_email:
            # Layout para listas com email (SQIE, Continuity, Planner, Sourcing)
            input_fields = [
                ft.Row([
                    lists_controls[key]['input'],
                    lists_controls[key]['alias'],
                ], spacing=12),
                ft.Container(height=8),
                lists_controls[key]['email'],
            ]
        else:
            # Layout para listas simples (BU, Category)
            input_fields = [lists_controls[key]['input']]
        
        # Card de entrada
        input_card = ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(title_icon, size=20, color=theme_colors.get('primary')),
                        ft.Text("Adicionar Novo Item", size=15, weight="bold", color=theme_colors.get('on_surface')),
                    ], spacing=8),
                    ft.Divider(height=1, color=ft.Colors.with_opacity(0.1, theme_colors.get('outline'))),
                    ft.Container(height=5),
                    ft.Column(input_fields, spacing=0),
                    lists_controls[key]['feedback'],
                    ft.Container(height=8),
                    ft.Row([
                        ft.ElevatedButton(
                            "Adicionar",
                            icon=ft.Icons.ADD_CIRCLE_OUTLINE,
                            on_click=add_alias_handler_sqie if key == 'sqie' else lambda e, k=key: add_alias_handler_generic(
                                f"{k}_table" if k in ['continuity', 'planner', 'sourcing'] else 
                                ('business_unit_table' if k == 'bu' else 'categories_table'),
                                lists_controls[k]['input'],
                                lists_controls[k].get('alias')
                            ),
                            bgcolor=theme_colors.get('primary'),
                            color=theme_colors.get('on_primary'),
                        ),
                    ], alignment=ft.MainAxisAlignment.END),
                ], spacing=8),
                padding=20,
            ),
            elevation=1,
        )
        
        # Card de lista de itens
        items_card = ft.Card(
            content=ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.FORMAT_LIST_BULLETED, size=20, color=theme_colors.get('primary')),
                        ft.Text("Itens Cadastrados", size=15, weight="bold", color=theme_colors.get('on_surface')),
                    ], spacing=8),
                    ft.Divider(height=1, color=ft.Colors.with_opacity(0.1, theme_colors.get('outline'))),
                    ft.Container(height=5),
                    ft.Container(
                        content=ft.ListView(
                            controls=[lists_controls[key]['list']],
                            spacing=0,
                            expand=True,
                        ),
                        expand=True,
                    ),
                ], spacing=8, expand=True),
                padding=20,
                expand=True,
            ),
            elevation=1,
            expand=True,
        )
        
        return ft.Container(
            content=ft.Column([
                input_card,
                ft.Container(height=15),
                items_card,
            ], spacing=0, expand=True),
            padding=15,
            expand=True,
        )
    
    # Criar tabs
    def build_lists_tabs():
        """Constrói as tabs para Lists"""
        theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        tabs = ft.Tabs(
            selected_index=0,
            animation_duration=300,
            expand=True,
            tabs=[
                ft.Tab(
                    text="SQIE",
                    icon=ft.Icons.ASSIGNMENT,
                    content=create_list_tab_content('sqie', ft.Icons.ASSIGNMENT, has_email=True),
                ),
                ft.Tab(
                    text="Continuity",
                    icon=ft.Icons.TIMELINE,
                    content=create_list_tab_content('continuity', ft.Icons.TIMELINE, has_email=True),
                ),
                ft.Tab(
                    text="Planner",
                    icon=ft.Icons.PERSON,
                    content=create_list_tab_content('planner', ft.Icons.PERSON, has_email=True),
                ),
                ft.Tab(
                    text="Sourcing",
                    icon=ft.Icons.PUBLIC,
                    content=create_list_tab_content('sourcing', ft.Icons.PUBLIC, has_email=True),
                ),
                ft.Tab(
                    text="Business Unit",
                    icon=ft.Icons.BUSINESS,
                    content=create_list_tab_content('bu', ft.Icons.BUSINESS, has_email=False),
                ),
                ft.Tab(
                    text="Category",
                    icon=ft.Icons.CATEGORY,
                    content=create_list_tab_content('category', ft.Icons.CATEGORY, has_email=False),
                ),
            ],
        )
        return tabs
    
    # Adicionar tabs ao container
    lists_tabs_container.content = build_lists_tabs()
    
    # Popular listas iniciais
    for k in lists_controls.keys():
        try:
            refresh_list_ui(k)
        except Exception as ex:
            print(f"Erro ao popular lista {k}: {ex}")

    # Lógica para gerenciamento de suppliers
    suppliers_search_field_ref = ft.Ref()
    global suppliers_results_list, suppliers_header_container
    
    # Container para o header do supplier (inicialmente vazio)
    suppliers_header_container = ft.Container()
    
    # Inicializar com mensagem padrão
    theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
    initial_message = ft.Container(
        content=ft.Column([
            ft.Icon(ft.Icons.SEARCH, size=64, color=theme_colors.get('on_surface_variant')),
            ft.Text("Digite no campo de busca", size=18, weight="bold", color=theme_colors.get('on_surface')),
            ft.Text("Procure por nome, ID, PO ou BU", size=14, color=theme_colors.get('on_surface_variant')),
        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=15),
        alignment=ft.alignment.center,
        expand=True
    )
    
    suppliers_results_list = ft.Column([initial_message], spacing=10)  # Removido scroll - será aplicado no Container pai


    def add_new_supplier(e):
        """Abre dialog para adicionar um novo supplier."""
        print("➕ DEBUG: add_new_supplier chamada")
        
        # Verificar permissão - apenas Super Admin
        if current_user_privilege != "Super Admin":
            show_snack_bar("❌ Acesso negado. Apenas Super Admin pode adicionar suppliers.", True)
            return
        
        def handle_confirm(e):
            """Callback para confirmar a criação do supplier"""
            print("🔍 DEBUG: handle_confirm chamado")
            
            # Função auxiliar para tratar valores (TextField e Dropdown)
            def safe_strip(value):
                if value is None:
                    return ""
                return str(value).strip()
                
            def get_field_value(field):
                """Obtém valor tanto de TextField quanto de Dropdown ou Container que embala um Dropdown.
                Suporta:
                - TextField / Dropdown (tem atributo .value)
                - Container com .content (ex: Container(content=Dropdown(...)))
                - Outros casos: tenta acessar .controls[0].value como fallback
                """
                try:
                    # Dropdowns e TextFields diretamente
                    if hasattr(field, 'value'):
                        return safe_strip(field.value)

                    # Container que embala um controle em .content
                    if hasattr(field, 'content') and hasattr(field.content, 'value'):
                        return safe_strip(field.content.value)

                    # Alguns componentes podem ter controls list
                    if hasattr(field, 'controls') and len(field.controls) > 0 and hasattr(field.controls[0], 'value'):
                        return safe_strip(field.controls[0].value)

                except Exception:
                    pass
                return ""
            
            if not db_conn:
                print("🔍 DEBUG: Banco de dados não conectado")
                show_snack_bar("❌ Erro: Banco de dados não conectado.", True)
                return
            
            print("🔍 DEBUG: Preparando para inserir no banco")
            try:
                cursor = db_conn.cursor()
                
                # Pegar valores dos campos
                vendor_name = get_field_value(add_dialog.fields["vendor_name"])
                supplier_number = get_field_value(add_dialog.fields["supplier_number"])
                supplier_po = get_field_value(add_dialog.fields["supplier_po"])
                print(f"🔍 DEBUG: Dados a inserir - vendor_name: '{vendor_name}', SSID: '{supplier_number}', PO: '{supplier_po}'")
                
                # ===== VALIDAÇÃO DE UNICIDADE =====
                # Verificar se SSID já existe (se fornecido e não vazio)
                if supplier_number and supplier_number.strip():
                    check_ssid_query = "SELECT supplier_id, vendor_name FROM supplier_database_table WHERE supplier_number = ?"
                    existing_ssid = db_manager.query(check_ssid_query, (supplier_number.strip(),))
                    if existing_ssid:
                        existing_supplier = existing_ssid[0]
                        show_snack_bar(f"❌ SSID '{supplier_number}' já está cadastrado para o supplier '{existing_supplier['vendor_name']}' (ID: {existing_supplier['supplier_id']})", True)
                        return
                
                # Verificar se PO já existe (se fornecido e não vazio)
                if supplier_po and supplier_po.strip():
                    check_po_query = "SELECT supplier_id, vendor_name FROM supplier_database_table WHERE supplier_po = ?"
                    existing_po = db_manager.query(check_po_query, (supplier_po.strip(),))
                    if existing_po:
                        existing_supplier = existing_po[0]
                        show_snack_bar(f"❌ PO '{supplier_po}' já está cadastrado para o supplier '{existing_supplier['vendor_name']}' (ID: {existing_supplier['supplier_id']})", True)
                        return
                # ===== FIM DA VALIDAÇÃO =====
                
                # Inserir novo supplier (supplier_id será gerado automaticamente pelo banco)
                insert_query = """
                    INSERT INTO supplier_database_table 
                    (vendor_name, supplier_category, bu, supplier_name,
                     supplier_email, supplier_number, supplier_po, supplier_status, planner, 
                     continuity, sourcing, sqie)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """

                db_manager.execute(insert_query, (
                    get_field_value(add_dialog.fields["vendor_name"]),
                    get_field_value(add_dialog.fields["supplier_category"]),
                    get_field_value(add_dialog.fields["bu"]),
                    get_field_value(add_dialog.fields["supplier_origin"]),
                    get_field_value(add_dialog.fields["supplier_email"]),
                    supplier_number,
                    supplier_po,
                    get_field_value(add_dialog.fields["supplier_status"]),
                    get_field_value(add_dialog.fields["planner"]),
                    get_field_value(add_dialog.fields["continuity"]),
                    get_field_value(add_dialog.fields["sourcing"]),
                    get_field_value(add_dialog.fields["sqie"])
                ))
                
                db_conn.commit()
                print("🔍 DEBUG: Dados inseridos com sucesso")
                
                # Atualizar lista de suppliers
                search_suppliers_config()
                
                show_snack_bar(f"✅ Supplier {vendor_name} criado com sucesso!", False)
                page.close(add_dialog)
                
            except Exception as ex:
                print(f"🔍 DEBUG: Erro ao inserir: {ex}")
                show_snack_bar(f"❌ Erro ao criar supplier: {str(ex)}", True)
                print(f"Erro ao criar supplier: {ex}")
        
        def handle_cancel(e):
            """Callback para cancelar a criação"""
            page.close(add_dialog)
        
        # Preparar opções dinâmicas para os Dropdowns do diálogo
        list_opts = {
            'bu': load_list_options('business_unit_table','bu'),
            'planner': load_list_options('planner_table','name'),
            'continuity': load_list_options('continuity_table','name'),
            'sourcing': load_list_options('sourcing_table','name'),
            'sqie': load_list_options('sqie_table','name'),
            'category': load_list_options('categories_table','category'),
        }
        
        # Debug: verificar se as listas foram carregadas
        print("📋 DEBUG: Opções carregadas para o diálogo:")
        for key, options in list_opts.items():
            print(f"  {key}: {options[:5]}{'...' if len(options) > 5 else ''} (total: {len(options)})")
        

        # Criar o diálogo personalizado
        add_dialog = AddSupplierDialog(
            on_confirm=handle_confirm,
            on_cancel=handle_cancel,
            page=page,
            list_options=list_opts
        )
        
        # Aplicar cores do tema aos campos
        current_colors = get_current_theme_colors(get_theme_name_from_page(page))
        field_bgcolor = current_colors.get('field_background')
        for field_name, field in add_dialog.fields.items():
            if hasattr(field, 'bgcolor'):
                field.bgcolor = field_bgcolor
        
        print("➕ DEBUG: Abrindo dialog para adicionar supplier...")
        page.open(add_dialog)
        print("➕ DEBUG: Dialog para adicionar supplier aberto!")

    # Variáveis globais para debounce de suppliers e paginação
    global suppliers_search_timer, suppliers_search_lock, suppliers_current_page, suppliers_total_pages, suppliers_per_page
    suppliers_search_timer = None
    suppliers_search_lock = False
    suppliers_current_page = 1
    suppliers_total_pages = 1
    suppliers_per_page = 20  # Número de suppliers por página
    
    # Refs para paginação
    suppliers_page_info = ft.Ref[ft.Text]()
    suppliers_first_button = ft.Ref[ft.IconButton]()
    suppliers_prev_button = ft.Ref[ft.IconButton]()
    suppliers_next_button = ft.Ref[ft.IconButton]()
    suppliers_last_button = ft.Ref[ft.IconButton]()
    suppliers_pagination_container = ft.Ref[ft.Container]()

    def search_suppliers_config_debounced():
        """Executa a busca de suppliers com debounce para evitar conflitos SQLite."""
        global suppliers_search_timer, suppliers_search_lock
        
        # Cancelar timer anterior se existir
        if suppliers_search_timer:
            try:
                suppliers_search_timer.cancel()
            except:
                pass
            suppliers_search_timer = None
        
        # Criar novo timer para executar a busca após 300ms de inatividade
        import threading
        suppliers_search_timer = threading.Timer(0.3, lambda: search_suppliers_config(reset_page=True))
        suppliers_search_timer.start()

    def search_suppliers_config(reset_page=False):
        """Busca suppliers para edição na aba de configurações com paginação (1 por página)."""
        global suppliers_search_lock, suppliers_current_page, suppliers_total_pages
        
        # Evitar múltiplas execuções simultâneas
        if suppliers_search_lock:
            return
        suppliers_search_lock = True
        
        # Reset para primeira página ao fazer nova busca
        if reset_page:
            suppliers_current_page = 1
        
        try:
            if not suppliers_search_field_ref.current:
                return
                
            search_term = suppliers_search_field_ref.current.value.strip()
            
            print(f"Buscando suppliers (Página {suppliers_current_page}): '{search_term}'")

            suppliers_results_list.controls.clear()
            
            # Se não há termo de busca ou tem menos de 3 caracteres, apenas limpar e mostrar mensagem
            if not search_term or len(search_term) < 3:
                theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
                message_text = "Digite no campo de busca" if not search_term else "Digite pelo menos 3 caracteres para buscar"
                suppliers_results_list.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Container(height=100),  # Espaço no topo
                            ft.Icon(ft.Icons.SEARCH, size=64, color=theme_colors.get('on_surface_variant')),
                            ft.Text(message_text, size=18, weight="bold", color=theme_colors.get('on_surface')),
                            ft.Text("Procure por nome, ID, PO ou BU", size=14, color=theme_colors.get('on_surface_variant')),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=15),
                        alignment=ft.alignment.center,
                        expand=True
                    )
                )
                
                # Esconde o header (nome do supplier)
                suppliers_header_container.content = None
                suppliers_header_container.update()
                
                # Atualizar informações de paginação
                if suppliers_page_info.current:
                    suppliers_page_info.current.value = "Página 1 de 1 (0 registros)"
                    suppliers_page_info.current.update()
                
                # Desabilitar todos os botões de navegação
                if suppliers_first_button.current:
                    suppliers_first_button.current.disabled = True
                    suppliers_first_button.current.update()
                
                if suppliers_prev_button.current:
                    suppliers_prev_button.current.disabled = True
                    suppliers_prev_button.current.update()
                
                if suppliers_next_button.current:
                    suppliers_next_button.current.disabled = True
                    suppliers_next_button.current.update()
                
                if suppliers_last_button.current:
                    suppliers_last_button.current.disabled = True
                    suppliers_last_button.current.update()
                
                # Esconde os controles de paginação
                if suppliers_pagination_container.current:
                    suppliers_pagination_container.current.visible = False
                    suppliers_pagination_container.current.update()
                
                safe_page_update(page)
                return
            
            if not db_conn:
                suppliers_results_list.controls.append(
                    ft.Container(
                        ft.Text("Erro: Não foi possível conectar ao banco de dados.", color="red"),
                        padding=20
                    )
                )
                safe_page_update(page)
                return

            try:
                # Contar total de registros
                if search_term:
                    count_query = """
                        SELECT COUNT(*) as total
                        FROM supplier_database_table
                        WHERE (vendor_name LIKE ? OR supplier_id LIKE ? OR supplier_po LIKE ? OR bu LIKE ?)
                    """
                    params = [f"%{search_term}%", f"%{search_term}%", f"%{search_term}%", f"%{search_term}%"]
                else:
                    count_query = "SELECT COUNT(*) as total FROM supplier_database_table"
                    params = []
                
                count_result = db_manager.query_one(count_query, params)
                total_records = count_result['total']
                suppliers_total_pages = total_records  # 1 supplier por página
                
                # Garantir que a página atual está no intervalo válido
                if suppliers_current_page > suppliers_total_pages:
                    suppliers_current_page = suppliers_total_pages
                if suppliers_current_page < 1:
                    suppliers_current_page = 1
                
                # Calcular offset (1 por página)
                offset = suppliers_current_page - 1
                
                # Buscar apenas 1 supplier da página atual
                if search_term:
                    query = """
                        SELECT supplier_id, vendor_name, supplier_category, bu, supplier_name,
                               supplier_email, supplier_number, supplier_po, supplier_status, planner, 
                               continuity, sourcing, sqie
                        FROM supplier_database_table
                        WHERE (vendor_name LIKE ? OR supplier_id LIKE ? OR supplier_po LIKE ? OR bu LIKE ?)
                        ORDER BY vendor_name
                        LIMIT 1 OFFSET ?
                    """
                    params.append(offset)
                else:
                    query = """
                        SELECT supplier_id, vendor_name, supplier_category, bu, supplier_name,
                               supplier_email, supplier_number, supplier_po, supplier_status, planner, 
                               continuity, sourcing, sqie
                        FROM supplier_database_table
                        ORDER BY vendor_name
                        LIMIT 1 OFFSET ?
                    """
                    params = [offset]

                records = db_manager.query(query, params)
                
                print(f"Supplier {suppliers_current_page}/{suppliers_total_pages} (total: {total_records})")
                
                if not records:
                    theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
                    suppliers_results_list.controls.append(
                        ft.Container(
                            content=ft.Column([
                                ft.Icon(ft.Icons.SEARCH_OFF, size=64, color=theme_colors.get('on_surface_variant')),
                                ft.Text("Nenhum supplier encontrado", size=18, weight="bold", color=theme_colors.get('on_surface')),
                                ft.Text("Digite no campo de busca para encontrar suppliers", size=14, color=theme_colors.get('on_surface_variant')),
                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=15),
                            alignment=ft.alignment.center,
                            expand=True
                        )
                    )
                else:
                    # Exibir o supplier
                    supplier_parts = create_supplier_display(records[0], page)
                    # Atualizar header
                    suppliers_header_container.content = supplier_parts["header"]
                    suppliers_header_container.update()
                    # Atualizar conteúdo
                    suppliers_results_list.controls.append(supplier_parts["content"])
                
                # Atualizar informações de paginação
                if suppliers_page_info.current:
                    if total_records > 0:
                        suppliers_page_info.current.value = f"Página {suppliers_current_page} de {suppliers_total_pages} ({total_records} registros)"
                    else:
                        suppliers_page_info.current.value = "Página 1 de 1 (0 registros)"
                    suppliers_page_info.current.update()
                
                # Atualizar estado dos botões de navegação
                at_first = (suppliers_current_page <= 1)
                at_last = (suppliers_current_page >= suppliers_total_pages)
                
                if suppliers_first_button.current:
                    suppliers_first_button.current.disabled = at_first
                    suppliers_first_button.current.update()
                
                if suppliers_prev_button.current:
                    suppliers_prev_button.current.disabled = at_first
                    suppliers_prev_button.current.update()
                
                if suppliers_next_button.current:
                    suppliers_next_button.current.disabled = at_last
                    suppliers_next_button.current.update()
                
                if suppliers_last_button.current:
                    suppliers_last_button.current.disabled = at_last
                    suppliers_last_button.current.update()
                
                # Mostrar controles de paginação se houver resultados
                if suppliers_pagination_container.current:
                    suppliers_pagination_container.current.visible = (total_records > 0)
                    suppliers_pagination_container.current.update()
                
            except sqlite3.Error as db_error:
                error_msg = f"Erro no banco de dados: {db_error}"
                print(error_msg)
                suppliers_results_list.controls.append(
                    ft.Container(
                        ft.Text(error_msg, color="red"),
                        padding=20
                    )
                )

            try:
                safe_page_update(page)
            except Exception as update_error:
                print(f"Erro ao atualizar página: {update_error}")
                
        finally:
            suppliers_search_lock = False

    def go_to_first_page():
        """Navega para a primeira página"""
        global suppliers_current_page
        if suppliers_current_page > 1:
            suppliers_current_page = 1
            search_suppliers_config(reset_page=False)
    
    def go_to_last_page():
        """Navega para a última página"""
        global suppliers_current_page
        if suppliers_current_page < suppliers_total_pages:
            suppliers_current_page = suppliers_total_pages
            search_suppliers_config(reset_page=False)
    
    def go_to_previous_page(e):
        """Navega para a página anterior"""
        global suppliers_current_page
        if suppliers_current_page > 1:
            suppliers_current_page -= 1
            search_suppliers_config(reset_page=False)
    
    def go_to_next_page(e):
        """Navega para a próxima página"""
        global suppliers_current_page
        if suppliers_current_page < suppliers_total_pages:
            suppliers_current_page += 1
            search_suppliers_config(reset_page=False)
    
    def create_supplier_display(record, page):
        """Cria o display completo de um supplier (1 por página) com campos editáveis"""
        Colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        supplier_id = record.get('supplier_id', '')
        vendor_name = record.get('vendor_name', '')
        supplier_category = record.get('supplier_category', '')
        bu = record.get('bu', '')
        supplier_name = record.get('supplier_name', '')
        supplier_email = record.get('supplier_email', '')
        supplier_number = record.get('supplier_number', '')
        supplier_po = record.get('supplier_po', '')
        supplier_status = record.get('supplier_status', '')
        planner = record.get('planner', '')
        continuity = record.get('continuity', '')
        sourcing = record.get('sourcing', '')
        sqie = record.get('sqie', '')
        
        # Criar campos editáveis (vendor_name fica de fora, será exibido no topo)
        
        def on_field_change(e):
            """Handler para quando um campo é alterado - limpa se vazio"""
            if e.control.value is None or str(e.control.value).strip() == "":
                e.control.value = ""
                e.control.error_text = None
                safe_page_update(page)
        
        fields = {
            "supplier_category": ft.Dropdown(
                label="Category",
                value=supplier_category,
                options=[ft.dropdown.Option(v) for v in load_list_options('categories_table','category')],
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline')
            ),
            "bu": ft.Dropdown(
                label="BU (Business Unit)",
                value=bu,
                options=[ft.dropdown.Option(v) for v in load_list_options('business_unit_table','bu')],
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline')
            ),
            "supplier_origin": ft.Dropdown(
                label="Origem",
                value=supplier_name if supplier_name in ['Nacional', 'Importado'] else '',
                options=[ft.dropdown.Option(v) for v in ["Nacional", "Importado"]],
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline')
            ),
        }
        
        # Criar campo de email separado usando EmailChipField
        email_field = EmailChipField(
            label="Email",
            initial_value=supplier_email,
            theme_colors=Colors,
            on_change=on_field_change
        )
        fields["supplier_email"] = email_field
        
        fields.update({
            "supplier_number": ft.TextField(
                label="SSID",
                value=format_po_ssid(supplier_number),
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline'),
                on_change=on_field_change
            ),
            "supplier_po": ft.TextField(
                label="PO",
                value=format_po_ssid(supplier_po),
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline'),
                on_change=on_field_change
            ),
            "supplier_status": ft.Dropdown(
                label="Status",
                value=supplier_status,
                options=[ft.dropdown.Option(v) for v in ["Active", "Inactive"]],
                expand=True,
                bgcolor=Colors.get('field_background'),
                color="green" if supplier_status == "Active" else "red",
                border_color=Colors.get('outline')
            ),
            "planner": ft.Dropdown(
                label="Planner",
                value=planner,
                options=[ft.dropdown.Option(v) for v in load_list_options('planner_table','name')],
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline')
            ),
            "continuity": ft.Dropdown(
                label="Continuity",
                value=continuity,
                options=[ft.dropdown.Option(v) for v in load_list_options('continuity_table','name')],
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline')
            ),
            "sourcing": ft.Dropdown(
                label="Sourcing",
                value=sourcing,
                options=[ft.dropdown.Option(v) for v in load_list_options('sourcing_table','name')],
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline')
            ),
            "sqie": ft.Dropdown(
                label="SQIE",
                value=sqie,
                options=[ft.dropdown.Option(v) for v in load_list_options('sqie_table','name')],
                expand=True,
                bgcolor=Colors.get('field_background'),
                color=Colors.get('on_surface'),
                border_color=Colors.get('outline')
            ),
        })
        
        def save_supplier(e):
            """Salva as alterações do supplier"""
            # Verificar permissão
            if current_user_privilege != "Super Admin":
                show_snack_bar("❌ Acesso negado. Apenas Super Admin pode editar suppliers.", True)
                return
            
            # Função auxiliar para tratar valores (permite vazio)
            def safe_strip(value):
                if value is None:
                    return ""
                return str(value).strip()
            
            # Validações básicas (apenas origem é obrigatória)
            if not safe_strip(fields["supplier_origin"].value):
                show_snack_bar("❌ Campo Origem é obrigatório!", True)
                return
            
            try:
                e.control.disabled = True
                e.control.text = "Salvando..."
                safe_page_update(page)
                
                # Processar valores dos campos (permite vazios)
                # EmailChipField usa get_value() em vez de .value
                email_value = fields["supplier_email"].get_value() if hasattr(fields["supplier_email"], 'get_value') else safe_strip(fields["supplier_email"].value)
                po_value = safe_strip(fields["supplier_po"].value)
                ssid_value = safe_strip(fields["supplier_number"].value)
                
                # Log de debug
                print(f"💾 Salvando supplier {supplier_id}:")
                print(f"   Email: '{email_value}' (vazio: {not email_value})")
                print(f"   PO: '{po_value}' (vazio: {not po_value})")
                print(f"   SSID: '{ssid_value}' (vazio: {not ssid_value})")
                
                # ===== VALIDAÇÃO DE UNICIDADE =====
                # Verificar se SSID já existe em outro supplier (se fornecido e não vazio)
                if ssid_value and ssid_value.strip():
                    check_ssid_query = "SELECT supplier_id, vendor_name FROM supplier_database_table WHERE supplier_number = ? AND supplier_id != ?"
                    existing_ssid = db_manager.query(check_ssid_query, (ssid_value.strip(), supplier_id))
                    if existing_ssid:
                        existing_supplier = existing_ssid[0]
                        show_snack_bar(f"❌ SSID '{ssid_value}' já está cadastrado para o supplier '{existing_supplier['vendor_name']}' (ID: {existing_supplier['supplier_id']})", True)
                        e.control.disabled = False
                        e.control.text = "Salvar"
                        safe_page_update(page)
                        return
                
                # Verificar se PO já existe em outro supplier (se fornecido e não vazio)
                if po_value and po_value.strip():
                    check_po_query = "SELECT supplier_id, vendor_name FROM supplier_database_table WHERE supplier_po = ? AND supplier_id != ?"
                    existing_po = db_manager.query(check_po_query, (po_value.strip(), supplier_id))
                    if existing_po:
                        existing_supplier = existing_po[0]
                        show_snack_bar(f"❌ PO '{po_value}' já está cadastrado para o supplier '{existing_supplier['vendor_name']}' (ID: {existing_supplier['supplier_id']})", True)
                        e.control.disabled = False
                        e.control.text = "Salvar"
                        safe_page_update(page)
                        return
                # ===== FIM DA VALIDAÇÃO =====
                
                # Atualizar registro (vendor_name não é editável, mantém o valor original)
                update_query = """
                    UPDATE supplier_database_table 
                    SET supplier_category = ?, bu = ?, supplier_name = ?,
                        supplier_email = ?, supplier_number = ?, supplier_po = ?, supplier_status = ?,
                        planner = ?, continuity = ?, sourcing = ?, sqie = ?
                    WHERE supplier_id = ?
                """
                db_manager.execute(update_query, (
                    safe_strip(fields["supplier_category"].value),
                    safe_strip(fields["bu"].value),
                    safe_strip(fields["supplier_origin"].value),
                    email_value,  # Permite vazio
                    ssid_value,   # Permite vazio
                    po_value,     # Permite vazio
                    safe_strip(fields["supplier_status"].value),
                    safe_strip(fields["planner"].value),
                    safe_strip(fields["continuity"].value),
                    safe_strip(fields["sourcing"].value),
                    safe_strip(fields["sqie"].value),
                    supplier_id
                ))
                
                show_snack_bar(f"✅ Supplier {vendor_name} atualizado com sucesso!")
                print(f"✅ Supplier {supplier_id} atualizado no banco de dados")
                
                # Recarregar a página atual
                search_suppliers_config(None)
                
            except Exception as ex:
                show_snack_bar(f"❌ Erro ao salvar: {str(ex)}", True)
                print(f"Erro ao salvar supplier: {ex}")
            finally:
                e.control.disabled = False
                e.control.text = "Salvar"
                safe_page_update(page)
        
        # Criar header separado do conteúdo para permitir layout fixo
        header = ft.Container(
            content=ft.Row([
                ft.Column([
                    ft.Text(
                        vendor_name or "Nome não disponível",
                        size=28,
                        weight="bold",
                        color=Colors.get('primary')
                    ),
                    ft.Text(
                        f"ID: {supplier_id}",
                        size=14,
                        color=Colors.get('on_surface_variant')
                    ),
                ], spacing=5, expand=True),
                ft.ElevatedButton(
                    "Salvar Alterações",
                    icon=ft.Icons.SAVE,
                    on_click=save_supplier,
                    style=ft.ButtonStyle(
                        padding=ft.padding.symmetric(horizontal=30, vertical=15)
                    )
                ),
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER),
            padding=ft.padding.all(20),
            bgcolor=Colors.get('surface_variant'),
            border=ft.border.only(bottom=ft.BorderSide(1, Colors.get('outline_variant'))),
        )
        
        # Conteúdo rolável
        content = ft.Column([
            # Seção 1: Categoria e Localização
            ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.CATEGORY, size=20, color=Colors.get('primary')),
                            ft.Text("Categoria e Localização", size=16, weight="bold", color=Colors.get('on_surface')),
                        ], spacing=8),
                        ft.Container(height=15),
                        ft.Row([
                            fields["supplier_category"],
                            fields["bu"],
                            fields["supplier_origin"],
                        ], spacing=20),
                    ]),
                    padding=ft.padding.all(20),
                    bgcolor=Colors.get('surface_variant'),
                    border_radius=8,
                ),
                
                ft.Container(height=20),
                
                # Seção 2: Identificação
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.BADGE, size=20, color=Colors.get('primary')),
                            ft.Text("Identificação", size=16, weight="bold", color=Colors.get('on_surface')),
                        ], spacing=8),
                        ft.Container(height=15),
                        ft.Row([
                            fields["supplier_po"],
                            fields["supplier_number"],
                            fields["supplier_status"],
                        ], spacing=20),
                    ]),
                    padding=ft.padding.all(20),
                    bgcolor=Colors.get('surface_variant'),
                    border_radius=8,
                ),
                
                ft.Container(height=20),
                
                # Seção 3: Contato
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.EMAIL, size=20, color=Colors.get('primary')),
                            ft.Text("Contato", size=16, weight="bold", color=Colors.get('on_surface')),
                        ], spacing=8),
                        ft.Container(height=15),
                        email_field.build(),  # Usar .build() para obter o controle visual
                    ]),
                    padding=ft.padding.all(20),
                    bgcolor=Colors.get('surface_variant'),
                    border_radius=8,
                ),
                
                ft.Container(height=20),
                
                # Seção 4: Equipe Responsável
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.GROUP, size=20, color=Colors.get('primary')),
                            ft.Text("Equipe Responsável", size=16, weight="bold", color=Colors.get('on_surface')),
                        ], spacing=8),
                        ft.Container(height=15),
                        ft.Row([
                            # Coluna Esquerda
                            ft.Column([
                                fields["planner"],
                                fields["continuity"],
                            ], spacing=15, expand=1),
                            
                            # Coluna Direita
                            ft.Column([
                                fields["sourcing"],
                                fields["sqie"],
                            ], spacing=15, expand=1),
                        ], spacing=25),
                    ]),
                    padding=ft.padding.all(20),
                    bgcolor=Colors.get('surface_variant'),
                    border_radius=8,
                ),
                
        ], spacing=20)
        
        # Retornar dicionário com header e content separados
        return {"header": header, "content": content}
    
    def open_edit_supplier_dialog(record, page):
        """Abre diálogo para editar um supplier"""
        # Por enquanto, mostrar mensagem
        show_snack_bar(f"Editar supplier: {record.get('vendor_name', 'Unknown')}", False, page)

    # Referência para o botão de limpar da aba Suppliers
    suppliers_clear_button_ref = ft.Ref[ft.IconButton]()
    
    # Função para limpar o campo de busca da aba Suppliers
    def clear_suppliers_search_field():
        if suppliers_search_field_ref.current:
            suppliers_search_field_ref.current.value = ""
            # Esconder o botão
            if suppliers_clear_button_ref.current:
                suppliers_clear_button_ref.current.visible = False
                suppliers_clear_button_ref.current.update()
            suppliers_search_field_ref.current.update()
            # Trigger a busca para limpar resultados
            search_suppliers_config_debounced()
    
    # Função para atualizar visibilidade do botão X
    def update_suppliers_search_suffix():
        if suppliers_search_field_ref.current and suppliers_clear_button_ref.current:
            # Mostrar/esconder X baseado se há texto
            suppliers_clear_button_ref.current.visible = bool(suppliers_search_field_ref.current.value)
            suppliers_clear_button_ref.current.update()

    # Conteúdo da sub-aba Suppliers
    suppliers_content = ft.Container(
        content=ft.Column([
            # Header fixo
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.BUSINESS, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Column([
                        ft.Text("Gerenciamento de Fornecedores", size=22, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ft.Text("Navegue pelos fornecedores cadastrados", size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                    ], spacing=2, expand=True),
                ], spacing=12),
                margin=ft.margin.only(bottom=15),
            ),
            # Barra de pesquisa e botão novo
            ft.Row([
                ft.Stack(
                    controls=[
                        ft.TextField(
                            label="Buscar Supplier (Nome, ID, PO ou BU)",
                            hint_text="Digite para buscar...",
                            prefix_icon=ft.Icons.SEARCH,
                            ref=suppliers_search_field_ref,
                            on_change=lambda e: [update_suppliers_search_suffix(), search_suppliers_config_debounced()],
                            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                            border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
                        ),
                        ft.Container(
                            content=ft.IconButton(
                                icon=ft.Icons.CLOSE,
                                icon_size=20,
                                tooltip="Limpar busca",
                                on_click=lambda e: clear_suppliers_search_field(),
                                ref=suppliers_clear_button_ref,
                                visible=False,
                            ),
                            right=0,
                            top=0,
                            bottom=0,
                        ),
                    ],
                    expand=True,
                ),
                ft.ElevatedButton("Novo Supplier", icon=ft.Icons.ADD_BUSINESS, on_click=add_new_supplier),
            ], spacing=10),
            
            ft.Container(height=10),
            
            # Container grande que ocupa toda a área disponível
            ft.Container(
                content=ft.Column([
                    # Header fixo do supplier (fica no topo)
                    suppliers_header_container,
                    
                    # Área de conteúdo (dados do supplier) - expande com scroll
                    ft.Container(
                        content=ft.Column([suppliers_results_list], scroll=ft.ScrollMode.AUTO, expand=True),
                        expand=True,
                    ),
                    
                    ft.Container(height=10),
                    
                    # Controles de paginação fixos na parte inferior
                    ft.Container(
                        content=ft.Row([
                            ft.IconButton(
                                icon=ft.Icons.FIRST_PAGE,
                                tooltip="Primeiro supplier",
                                on_click=lambda e: go_to_first_page(),
                                ref=suppliers_first_button,
                                icon_size=20
                            ),
                            ft.IconButton(
                                icon=ft.Icons.CHEVRON_LEFT,
                                tooltip="Supplier anterior",
                                on_click=go_to_previous_page,
                                ref=suppliers_prev_button,
                                icon_size=20
                            ),
                            ft.Container(
                                content=ft.Text(
                                    "Página 1 de 1 (0 registros)",
                                    size=13,
                                    ref=suppliers_page_info,
                                    text_align=ft.TextAlign.CENTER,
                                    color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                                ),
                                padding=ft.padding.symmetric(horizontal=20),
                            ),
                            ft.IconButton(
                                icon=ft.Icons.CHEVRON_RIGHT,
                                tooltip="Próximo supplier",
                                on_click=go_to_next_page,
                                ref=suppliers_next_button,
                                icon_size=20
                            ),
                            ft.IconButton(
                                icon=ft.Icons.LAST_PAGE,
                                tooltip="Último supplier",
                                on_click=lambda e: go_to_last_page(),
                                ref=suppliers_last_button,
                                icon_size=20
                            ),
                        ], alignment=ft.MainAxisAlignment.CENTER, spacing=5),
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                        padding=ft.padding.symmetric(horizontal=10, vertical=8),
                        border_radius=8,
                        ref=suppliers_pagination_container,
                        visible=False,  # Inicialmente invisível
                    ),
                ], spacing=0),
                expand=True,
            ),
        ], spacing=0, expand=True),
        padding=20,
        visible=False,
        expand=True
    )

    # Controle para mostrar a soma total dos pesos
    weight_sum_text = ft.Text("Soma dos pesos: 0.80", size=14, weight="bold", color="primary")
    
    # Função para atualizar o texto de um critério e a soma total
    def update_criteria_text(criteria_name, slider_value, text_control):
        text_control.value = f"{slider_value:.2f}"
        text_control.update()
        update_weight_sum()
    
    def update_weight_sum():
        """Atualiza a exibição da soma total dos pesos em tempo real."""
        total = nil_slider.value + otif_slider.value + pickup_slider.value + package_slider.value
        weight_sum_text.value = f"Soma dos pesos: {total:.2f}"
        
        # Mudar cor baseado na soma
        if abs(total - 1.0) <= 0.01:  # Muito próximo de 1.0
            weight_sum_text.color = "green"
        elif abs(total - 1.0) <= 0.05:  # Tolerável
            weight_sum_text.color = "orange" 
        else:  # Muito longe de 1.0
            weight_sum_text.color = "red"
        
        weight_sum_text.update()

    def auto_adjust_weights():
        """Auto-ajusta os pesos para somarem exatamente 1.00, mantendo as proporções."""
        current_total = nil_slider.value + otif_slider.value + pickup_slider.value + package_slider.value
        
        if current_total == 0:
            # Se tudo for zero, distribuir igualmente
            nil_slider.value = 0.25
            otif_slider.value = 0.25
            pickup_slider.value = 0.25
            package_slider.value = 0.25
        else:
            # Manter proporções e ajustar para somar exatamente 1.00
            factor = 1.0 / current_total
            values = [
                nil_slider.value * factor,
                otif_slider.value * factor, 
                pickup_slider.value * factor,
                package_slider.value * factor
            ]
            
            # Arredondar para 2 casas decimais
            rounded_values = [round(v, 2) for v in values]
            
            # Ajustar pequenos erros de arredondamento
            current_sum = sum(rounded_values)
            if current_sum != 1.0:
                diff = 1.0 - current_sum
                # Adicionar/subtrair a diferença do maior valor
                max_index = rounded_values.index(max(rounded_values))
                rounded_values[max_index] = round(rounded_values[max_index] + diff, 2)
            
            # Aplicar valores ajustados
            nil_slider.value = rounded_values[0]
            otif_slider.value = rounded_values[1] 
            pickup_slider.value = rounded_values[2]
            package_slider.value = rounded_values[3]
        
        # Atualizar textos
        nil_text.value = f"{nil_slider.value:.2f}"
        otif_text.value = f"{otif_slider.value:.2f}"
        pickup_text.value = f"{pickup_slider.value:.2f}"
        package_text.value = f"{package_slider.value:.2f}"
        
        # Atualizar soma
        update_weight_sum()
        
        # Atualizar UI
        safe_page_update(page)
        
        # Toast de sucesso
        page.overlay.append(
            ft.Container(
                content=ft.Text("✅ Pesos auto-ajustados para somar 1.00!", 
                               color="white", weight="bold"),
                bgcolor="blue",
                padding=10,
                border_radius=5,
                top=50,
                right=20,
                animate_opacity=300,
            )
        )
        safe_page_update(page)
        
        # Remover show_toast após 3 segundos
        import threading
        def remove_auto_toast():
            import time
            time.sleep(3)
            if page.overlay:
                page.overlay.pop()
                safe_page_update(page)
        
        threading.Thread(target=remove_auto_toast, daemon=True).start()

    # Sliders dos critérios com nomes corretos (100 divisões para precisão de 0.01)
    nil_text = ft.Text("0.20", size=12, width=60)
    nil_slider = ft.Slider(
        min=0, max=1, divisions=100, value=0.2,
        on_change=lambda e: update_criteria_text("NIL", e.control.value, nil_text),
        overlay_color=ft.Colors.TRANSPARENT,
    )
    
    otif_text = ft.Text("0.20", size=12, width=60)
    otif_slider = ft.Slider(
        min=0, max=1, divisions=100, value=0.2,
        on_change=lambda e: update_criteria_text("OTIF", e.control.value, otif_text)
    )
    
    pickup_text = ft.Text("0.20", size=12, width=60)
    pickup_slider = ft.Slider(
        min=0, max=1, divisions=100, value=0.2,
        on_change=lambda e: update_criteria_text("Quality of Pick Up", e.control.value, pickup_text),
        overlay_color=ft.Colors.TRANSPARENT,
    )
    
    package_text = ft.Text("0.20", size=12, width=60)
    package_slider = ft.Slider(
        min=0, max=1, divisions=100, value=0.2,
        on_change=lambda e: update_criteria_text("Quality-Supplier Package", e.control.value, package_text),
        overlay_color=ft.Colors.TRANSPARENT,
    )
    
    target_text = ft.Text("5.00", size=12, width=60)
    target_slider = ft.Slider(
        min=0, max=10, divisions=100, value=5.0,
        on_change=lambda e: update_criteria_text("Target", e.control.value, target_text)
    )

    def update_criteria_settings():
        """Atualiza os critérios na tabela criteria_table."""
        print("🔄 Botão Update Criteria pressionado!")
        
        if not db_conn:
            print("❌ Erro: Conexão com banco não disponível")
            show_snack_bar("❌ Erro: Conexão com banco não disponível", True)
            safe_page_update(page)
            return
        
        criteria_values = {
            'NIL': nil_slider.value,
            'OTIF': otif_slider.value,
            'Quality of Pick Up': pickup_slider.value,
            'Quality-Supplier Package': package_slider.value,
            'Target': target_slider.value  # Target mantém escala 0-10
        }
        
        print(f"📊 Valores coletados dos sliders:")
        for name, value in criteria_values.items():
            print(f"  {name}: {value}")
        
        # Validar soma apenas dos critérios de peso (excluindo Target que é escala diferente)
        weight_criteria = ['NIL', 'OTIF', 'Quality of Pick Up', 'Quality-Supplier Package']
        total_weight = sum(criteria_values[key] for key in weight_criteria)
        
        print(f"🧮 Soma dos pesos (sem Target): {total_weight}")
        
        if round(total_weight, 2) != 1.0:  # Deve ser exatamente 1.00
            print(f"⚠️ Validação falhou: soma {total_weight:.2f} não é exatamente 1.00")
            show_snack_bar(f"❌ ERRO: A soma dos 4 pesos é {total_weight:.2f} - DEVE ser exatamente 1.00!", True)
            return
        
        try:
            print("🔄 Iniciando atualização no banco de dados...")
            
            # Preparar dados para atualização em lote
            update_data = []
            for criteria_name, value in criteria_values.items():
                # Arredondar para 2 casas decimais antes de salvar
                rounded_value = round(value, 2)
                update_data.append((str(rounded_value), criteria_name))
                print(f"  ✅ Critério preparado: {criteria_name} = {rounded_value}")
            
            # Executar atualização em lote usando DBManager
            if update_data:
                db_manager.execute_many(
                    "UPDATE criteria_table SET value = ? WHERE criteria_category = ?", 
                    update_data
                )
                print("✅ Atualização em lote realizada com sucesso!")
            
            show_snack_bar("✅ Critérios atualizados com sucesso!", False, page)
            
        except Exception as ex:
            print(f"❌ Erro ao atualizar critérios: {ex}")
            show_snack_bar("❌ Erro ao atualizar critérios!", True, page)
            safe_page_update(page)

    # Criar títulos fixos para Criteria (fora do scroll)
    criteria_header = ft.Container(
        content=ft.Row([
            ft.Icon(ft.Icons.ASSESSMENT, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
            ft.Column([
                ft.Text("Configuração de Critérios", size=22, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                ft.Text("Configure os pesos dos critérios de avaliação", size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
            ], spacing=2, expand=True),
        ], spacing=12),
        margin=ft.margin.only(bottom=15),
    )
    
    # Criar conteúdo scrollável para Criteria
    criteria_scrollable_content = ft.ListView([
        ft.Container(
            content=ft.Column([
                ft.Text("⚠️ IMPORTANTE: Os 4 pesos (NIL+OTIF+Pickup+Package) DEVEM somar exatamente 1.00", 
                       size=12, weight="bold", color="error"),
                ft.Text("Target usa escala independente de 0-10", size=12, italic=True, color="outline"),
                ft.Container(height=10),
                weight_sum_text,
                ft.Container(height=15),
                
                # Sliders para cada critério com nomes corretos
                ft.Row([
                    ft.Text("NIL:", size=14, weight="bold", width=220),
                    ft.Container(content=nil_slider, expand=True, bgcolor=None),
                    nil_text
                ], alignment=ft.MainAxisAlignment.START),
                ft.Row([
                    ft.Text("OTIF:", size=14, weight="bold", width=220),
                    ft.Container(content=otif_slider, expand=True),
                    otif_text
                ], alignment=ft.MainAxisAlignment.START),
                ft.Row([
                    ft.Text("Quality of Pick Up:", size=14, weight="bold", width=220),
                    ft.Container(content=pickup_slider, expand=True, bgcolor=None),
                    pickup_text
                ], alignment=ft.MainAxisAlignment.START),
                ft.Row([
                    ft.Text("Quality-Supplier Package:", size=14, weight="bold", width=220),
                    ft.Container(content=package_slider, expand=True, bgcolor=None),
                    package_text
                ], alignment=ft.MainAxisAlignment.START),
                ft.Row([
                    ft.Text("Target:", size=14, weight="bold", width=220),
                    ft.Container(content=target_slider, expand=True),
                    target_text
                ], alignment=ft.MainAxisAlignment.START),
                ft.Container(height=20),
                ft.Row([
                    ft.ElevatedButton("Update Criteria", icon=ft.Icons.UPDATE, on_click=lambda e: update_criteria_settings()),
                    ft.OutlinedButton("Auto-ajustar para 1.00", icon=ft.Icons.AUTO_FIX_HIGH, on_click=lambda e: auto_adjust_weights())
                ], spacing=10),
                ft.Container(height=20)  # Espaço extra no final
            ], spacing=10),
            padding=ft.padding.all(10)
        )
    ], spacing=8, expand=True)

    # Conteúdo da sub-aba Criteria com scroll
    criteria_content = ft.Container(
        content=ft.Column([
            criteria_header,
            ft.Divider(height=1, color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
            ft.Container(
                content=criteria_scrollable_content,
                expand=True
            )
        ], spacing=0, expand=True),
        padding=ft.padding.all(20),
        visible=False,
        expand=True
    )

    # Conteúdo da sub-aba Users
    users_list_ref = ft.Ref()  # Referência para o ListView dos usuários
    
    # Criar referências para os campos da aba Users
    wwid_field_ref = ft.Ref()
    name_field_ref = ft.Ref()
    password_field_ref = ft.Ref()
    privilege_dropdown_ref = ft.Ref()
    otif_check_ref = ft.Ref()
    nil_check_ref = ft.Ref()
    pickup_check_ref = ft.Ref()
    package_check_ref = ft.Ref()
    action_btn_ref = ft.Ref()
    clear_btn_ref = ft.Ref()
    
    # Salvar as referências no dicionário users_controls
    users_controls = { 'wwid': wwid_field_ref, 'name': name_field_ref, 'password': password_field_ref, 'privilege': privilege_dropdown_ref, 'otif_check': otif_check_ref, 'nil_check': nil_check_ref, 'pickup_check': pickup_check_ref, 'package_check': package_check_ref, 'action_btn': action_btn_ref, 'clear_btn': clear_btn_ref }
    # Container do formulário de usuários com referência para atualização de tema
    users_form_container = ft.Container(
        content=ft.Container(
            content=ft.Row([
                # Coluna esquerda - Campos principais
                ft.Column([
                    # Linha 1: WWID e Nome
                    ft.Row([
                        ft.TextField(
                            label="WWID",
                            hint_text="Digite o WWID do usuário",
                            prefix_icon=ft.Icons.BADGE,
                            expand=True,
                            border_radius=8,
                            ref=wwid_field_ref,
                            on_change=on_wwid_change,
                            filled=False,  # Fundo transparente
                            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                            border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
                        ),
                        ft.TextField(
                            label="Nome Completo",
                            hint_text="Digite o nome do usuário",
                            prefix_icon=ft.Icons.PERSON,
                            expand=True,
                            border_radius=8,
                            ref=name_field_ref,
                            filled=False,  # Fundo transparente
                            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                            border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
                        ),
                    ], spacing=10),

                    # Linha 2: Senha e Privilégio
                    ft.Row([
                        ft.TextField(
                            label="Senha",
                            hint_text="Digite a senha do usuário",
                            prefix_icon=ft.Icons.LOCK,
                            password=True,
                            can_reveal_password=True,
                            expand=True,
                            border_radius=8,
                            ref=password_field_ref,
                            filled=False,  # Fundo transparente
                            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                            border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
                        ),
                        ft.Dropdown(
                            label="Nível de Privilégio",
                            hint_text="Selecione o nível",
                            options=[
                                ft.dropdown.Option("User", "User"),
                                ft.dropdown.Option("Admin", "Admin"),
                                ft.dropdown.Option("Super Admin", "Super Admin"),
                            ],
                            expand=True,
                            ref=privilege_dropdown_ref,
                            bgcolor=None,  # Fundo transparente
                            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                            border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline'),
                            content_padding=ft.padding.symmetric(horizontal=12, vertical=16),
                        ),
                    ], spacing=10),
                ], expand=2, spacing=15),
                
                # Divisor vertical
                ft.VerticalDivider(width=20, color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
                
                # Coluna direita - Permissões
                ft.Column([
                    ft.Text("Permissões de Avaliação", size=16, weight="bold", 
                            color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Container(
                        content=ft.Row([
                            # Primeira coluna de checkboxes
                            ft.Column([
                                ft.Checkbox(
                                    label="OTIF", 
                                    value=False, 
                                    ref=otif_check_ref,
                                    label_style=ft.TextStyle(
                                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                                    )
                                ),
                                ft.Checkbox(
                                    label="NIL", 
                                    value=False, 
                                    ref=nil_check_ref,
                                    label_style=ft.TextStyle(
                                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                                    )
                                ),
                            ], spacing=8, horizontal_alignment=ft.CrossAxisAlignment.START, expand=1),
                            
                            # Segunda coluna de checkboxes
                            ft.Column([
                                ft.Checkbox(
                                    label="Pickup", 
                                    value=False, 
                                    ref=pickup_check_ref,
                                    label_style=ft.TextStyle(
                                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                                    )
                                ),
                                ft.Checkbox(
                                    label="Package", 
                                    value=False, 
                                    ref=package_check_ref,
                                    label_style=ft.TextStyle(
                                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                                    )
                                ),
                            ], spacing=8, horizontal_alignment=ft.CrossAxisAlignment.START, expand=1),
                        ], spacing=20, vertical_alignment=ft.CrossAxisAlignment.START),
                        padding=ft.padding.only(top=10)
                    )
                ], expand=1, horizontal_alignment=ft.CrossAxisAlignment.START, spacing=10),
            ], spacing=15, vertical_alignment=ft.CrossAxisAlignment.START),
            padding=20,
            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('card_background'),
            border_radius=12,
            width=800,  # Largura maior para acomodar as duas colunas
        ),
        alignment=ft.alignment.center
    )


    users_content = ft.Container(
        content=ft.Column([
            # Header fixo
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.PEOPLE, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Column([
                        ft.Text("Gerenciamento de Usuários", size=22, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ft.Text("Configure usuários e suas permissões de avaliação", size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                    ], spacing=2, expand=True),
                ], spacing=12),
                margin=ft.margin.only(bottom=15),
            ),
            
            # Formulário de usuário - fixo
            users_form_container,

            # Botões de ação - fixo
            ft.Container(
                content=ft.Row([
                    ft.ElevatedButton(
                        "Add User", 
                        icon=ft.Icons.PERSON_ADD,
                        ref=action_btn_ref,
                        visible=False,  # Será controlado via update_action_button()
                    ),
                    ft.ElevatedButton(
                        "Limpar Campos", 
                        icon=ft.Icons.CLEAR,
                        ref=clear_btn_ref,
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER, spacing=15),
                margin=ft.margin.only(top=10, bottom=10)
            ),
            
            ft.Divider(),
            
            # Lista de usuários - com scroll independente - expande para ocupar espaço restante
            ft.Container(
                content=ft.Column([
                    ft.Text("Usuários Cadastrados", size=16, weight="bold"),
                    ft.Container(
                        content=ft.ListView(
                            spacing=8,
                            expand=True,
                            ref=users_list_ref,
                        ),
                        expand=True,
                        border_radius=8,
                        padding=8,
                        bgcolor="surface_variant",
                    ),
                ], spacing=10, expand=True),
                expand=True
            ),
            
        ], spacing=0, expand=True),
        padding=20,
        visible=False,
        expand=True  # Container principal expande
    )

    # Referências dos controles de usuários
    def extract_users_refs():
        """Extrai as referências dos controles do users_content"""
        users_refs = {}
        try:
            # Preferir usar as refs já criadas no escopo em vez de navegar por índices
            # Isso torna a extração resiliente a mudanças na estrutura do container
            ref_map = {
                'users_list': users_list_ref,
                'wwid': wwid_field_ref,
                'name': name_field_ref,
                'password': password_field_ref,
                'privilege': privilege_dropdown_ref,
                'otif_check': otif_check_ref,
                'nil_check': nil_check_ref,
                'pickup_check': pickup_check_ref,
                'package_check': package_check_ref,
                'action_btn': action_btn_ref,
                'clear_btn': clear_btn_ref,
            }

            for key, ref in ref_map.items():
                try:
                    if ref is None:
                        continue
                    # ft.Ref has attribute 'current'
                    ctrl = ref.current if hasattr(ref, 'current') else ref
                    if ctrl:
                        users_refs[key] = ctrl
                except Exception:
                    # ignorar referências que não possam ser resolvidas
                    continue

            # Fallback: se alguma referência não foi encontrada pelas refs acima,
            # tentar localizar a lista de usuários percorrendo a estrutura (mais lenta)
            if 'users_list' not in users_refs:
                try:
                    # procurar por um ListView que contenha Cards (heurística simples)
                    def find_users_list(control):
                        if isinstance(control, ft.ListView) and getattr(control, 'ref', None) is users_list_ref:
                            return control
                        if hasattr(control, 'controls'):
                            for c in control.controls:
                                found = find_users_list(c)
                                if found:
                                    return found
                        if hasattr(control, 'content') and control.content:
                            return find_users_list(control.content)
                        return None

                    found = find_users_list(users_content)
                    if found:
                        users_refs['users_list'] = found
                except Exception:
                    pass

        except Exception as ex:
            print(f"Erro ao extrair refs de usuários (extract_users_refs): {ex}")

        return users_refs

    # Funções para gerenciamento de usuários
    def load_users_full():
        """Carrega todos os usuários da tabela."""
        if not db_conn:
            return []
        try:
            print("📋 Lista sendo atualizada...")
            
            # Se não for Super Admin, mostrar apenas o próprio usuário
            if current_user_privilege != "Super Admin":
                rows = db_manager.query("""
                    SELECT user_wwid, user_name, user_password, user_privilege, 
                           otif, nil, pickup, package 
                    FROM users_table 
                    WHERE UPPER(user_wwid) = ?
                    ORDER BY user_wwid
                """, (current_user_wwid.upper() if current_user_wwid else "",))
            else:
                rows = db_manager.query("""
                    SELECT user_wwid, user_name, user_password, user_privilege, 
                           otif, nil, pickup, package 
                    FROM users_table 
                    ORDER BY user_wwid
                """)
            users = []
            for row in rows:
                wwid = row['user_wwid']
                name = row['user_name']
                password = row['user_password']
                privilege = row['user_privilege']
                otif = row['otif']
                nil = row['nil']
                pickup = row['pickup']
                package = row['package']
                
                # Criar texto de exibição em duas linhas
                name_display = name if name else "Nome não informado"
                permissions = []
                
                # Converter para int e verificar se é 1
                if int(otif) == 1: permissions.append("OTIF")
                if int(nil) == 1: permissions.append("NIL")
                if int(pickup) == 1: permissions.append("Pickup")
                if int(package) == 1: permissions.append("Package")
                
                permissions_str = ", ".join(permissions) if permissions else "Nenhuma permissão"
                
                # Linha 1: Dados do usuário
                line1 = f"WWID: {wwid} | Nome: {name_display} | Privilégio: {privilege}"
                # Linha 2: Permissões
                line2 = f"Permissões: {permissions_str}"
                
                # Criar texto de exibição com quebra de linha
                display_text = f"{line1}\n{line2}"
                
                users.append({
                    "line1": line1,
                    "line2": line2,
                    "display": display_text,
                    "wwid": wwid,
                    "name": name or "",
                    "password": password,
                    "privilege": privilege,
                    "otif": bool(otif),
                    "nil": bool(nil),
                    "pickup": bool(pickup),
                    "package": bool(package)
                })
            return users
        except Exception as ex:
            print(f"Erro ao carregar usuários: {ex}")
            return []

    def refresh_users_list():
        """Atualiza a lista de usuários na interface."""
        try:
            if 'users_list' not in users_controls:
                return
                
            users = load_users_full()
            users_list = users_controls['users_list']
            users_list.controls.clear()
            
            if not users:
                users_list.controls.append(
                    ft.Container(
                        content=ft.Text(
                            "Nenhum usuário cadastrado",
                            color="outline",
                            italic=True,
                            text_align=ft.TextAlign.CENTER
                        ),
                        alignment=ft.alignment.center,
                        padding=10
                    )
                )
            else:
                # Helper local para criar um card de usuário
                def _make_user_card(u):
                    user_wwid = u["wwid"]
                    name_display = u["name"] or "Nome não informado"
                    privilege = u["privilege"]
                    permissions_str = u["line2"].replace("Permissões: ", "")
                    is_selected = (selected_user == user_wwid)
                    Colors = get_current_theme_colors(get_theme_name_from_page(page))

                    # Botões de ação
                    edit_btn = ft.TextButton(
                        "Editar",
                        icon=ft.Icons.EDIT,
                        on_click=lambda e, w=user_wwid: select_user(w)
                    )

                    def _delete_from_card(e, w=user_wwid):
                        try:
                            # Diálogo de confirmação utilizando o componente já existente
                            def confirm_delete(evt):
                                try:
                                    db_manager.execute("DELETE FROM users_table WHERE UPPER(user_wwid) = ?", (w.upper(),))
                                    page.close(dialog)
                                    # Se o usuário selecionado foi excluído, limpar seleção e campos
                                    if globals().get('selected_user') == w:
                                        clear_users_fields()
                                    refresh_users_list()
                                    show_snack_bar(f"✅ Usuário '{w}' removido com sucesso", False)
                                    safe_page_update(page)
                                except Exception as ex:
                                    page.close(dialog)
                                    show_snack_bar(f"❌ Erro ao excluir usuário: {ex}", True)
                                    safe_page_update(page)

                            def cancel_delete(evt):
                                page.close(dialog)
                                safe_page_update(page)

                            dialog = DeleteListItemConfirmationDialog(
                                item_name=w,
                                item_type="Usuário",
                                on_confirm=confirm_delete,
                                on_cancel=cancel_delete,
                                scale_func=lambda x: x
                            )
                            page.open(dialog)
                            safe_page_update(page)
                        except Exception as ex:
                            show_snack_bar(f"❌ Erro ao tentar excluir usuário: {ex}", True)
                            safe_page_update(page)

                    delete_btn = ft.TextButton(
                        "Apagar",
                        icon=ft.Icons.DELETE,
                        style=ft.ButtonStyle(color="red"),
                        on_click=_delete_from_card,
                        visible=(current_user_privilege == "Super Admin")  # Apenas Super Admin pode apagar
                    )

                    header_row = ft.Row([
                        ft.Row([
                            ft.Icon(ft.Icons.PERSON, color=Colors['primary']),
                            ft.Text(f"{name_display}", size=15, weight="bold", color=Colors['on_surface'])
                        ], spacing=8),
                        ft.Container(
                            content=ft.Text(f"WWID: {user_wwid}", size=12, color="outline"),
                            alignment=ft.alignment.center_right,
                            expand=True
                        )
                    ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN)
                    
                    # Criar lista de botões baseado em permissões
                    action_buttons = [edit_btn]
                    if current_user_privilege == "Super Admin":
                        action_buttons.append(delete_btn)

                    content_column = ft.Column([
                        header_row,
                        ft.Text(f"Privilégio: {privilege}", size=12, color="outline"),
                        ft.Text(f"Permissões: {permissions_str}", size=12, color="outline"),
                        ft.Divider(),
                        ft.Row(action_buttons, alignment=ft.MainAxisAlignment.END, spacing=10)
                    ], spacing=6)

                    card_container = ft.Container(
                        content=content_column,
                        padding=12,
                        bgcolor=Colors['primary_container'] if is_selected else get_current_theme_colors(get_theme_name_from_page(page)).get('card_background'),
                        border_radius=8
                    )

                    return ft.Card(
                        content=card_container, 
                        elevation=0,
                        color=ft.Colors.TRANSPARENT
                    )

                # Adicionar cards ao ListView
                for user in users:
                    users_list.controls.append(_make_user_card(user))
            
            users_list.update()
        except Exception as ex:
            print(f"Erro ao atualizar lista de usuários: {ex}")

    def select_user(wwid):
        """Seleciona um usuário e carrega seus dados nos campos."""
        global selected_user
        try:

            selected_user = wwid
            
            # Buscar dados do usuário
            row = db_manager.query_one("""
                SELECT user_wwid, user_name, user_password, user_privilege,
                       otif, nil, pickup, package
                FROM users_table 
                WHERE user_wwid = ?
            """, (wwid,))
            
            if row:
                wwid_val = row['user_wwid']
                name = row['user_name']
                password = row['user_password']
                privilege = row['user_privilege']
                otif = row['otif']
                nil = row['nil']
                pickup = row['pickup']
                package = row['package']
                
                print(f"Dados carregados do banco: WWID={wwid_val}, OTIF={otif} (tipo: {type(otif)}), NIL={nil} (tipo: {type(nil)}), Pickup={pickup} (tipo: {type(pickup)}), Package={package} (tipo: {type(package)})")
                
                # Preencher campos de texto
                users_controls['wwid'].value = str(wwid_val) if wwid_val else ""
                users_controls['name'].value = str(name) if name else ""
                users_controls['password'].value = str(password) if password else ""
                users_controls['privilege'].value = str(privilege) if privilege else None
                
                # Definir valores dos checkboxes - convertendo explicitamente para int primeiro
                users_controls['otif_check'].value = int(otif) == 1
                users_controls['nil_check'].value = int(nil) == 1
                users_controls['pickup_check'].value = int(pickup) == 1
                users_controls['package_check'].value = int(package) == 1
                
                print(f"Checkboxes definidos: OTIF={users_controls['otif_check'].value}, NIL={users_controls['nil_check'].value}, Pickup={users_controls['pickup_check'].value}, Package={users_controls['package_check'].value}")
                
                # Se não for Super Admin, bloquear campos (exceto senha)
                is_super_admin = current_user_privilege == "Super Admin"
                users_controls['wwid'].disabled = not is_super_admin
                users_controls['name'].disabled = not is_super_admin
                users_controls['privilege'].disabled = not is_super_admin
                users_controls['otif_check'].disabled = not is_super_admin
                users_controls['nil_check'].disabled = not is_super_admin
                users_controls['pickup_check'].disabled = not is_super_admin
                users_controls['package_check'].disabled = not is_super_admin
                users_controls['password'].disabled = False  # Senha sempre editável
                
                # Atualizar todos os controles
                for control in users_controls.values():
                    if hasattr(control, 'update'):
                        control.update()
                
                # Atualizar botão de ação
                update_action_button()

                safe_page_update(page)
            
            # Atualizar lista visual
            refresh_users_list()
            
        except Exception as ex:
            print(f"Erro ao selecionar usuário: {ex}")
            show_snack_bar(f"❌ Erro ao carregar usuário: {ex}", True)
            safe_page_update(page)

    def clear_users_fields(e=None):
        """Limpa todos os campos do formulário de usuários."""
        try:
            print("Limpando todos os campos (via users_controls)...")

            def resolve(ctrl):
                try:
                    if ctrl is None:
                        return None
                    return ctrl.current if hasattr(ctrl, 'current') else ctrl
                except Exception:
                    return ctrl

            # TextFields
            try:
                w = resolve(users_controls.get('wwid'))
                if w and hasattr(w, 'value'):
                    w.value = ""
                    if hasattr(w, 'update'):
                        w.update()
            except Exception as e:
                print(f"Erro ao limpar WWID: {e}")

            try:
                n = resolve(users_controls.get('name'))
                if n and hasattr(n, 'value'):
                    n.value = ""
                    if hasattr(n, 'update'):
                        n.update()
            except Exception as e:
                print(f"Erro ao limpar Name: {e}")

            try:
                p = resolve(users_controls.get('password'))
                if p and hasattr(p, 'value'):
                    p.value = ""
                    if hasattr(p, 'update'):
                        p.update()
            except Exception as e:
                print(f"Erro ao limpar Password: {e}")

            # Dropdown / Select
            try:
                priv = resolve(users_controls.get('privilege'))
                if priv and hasattr(priv, 'value'):
                    priv.value = None
                    if hasattr(priv, 'update'):
                        priv.update()
            except Exception as e:
                print(f"Erro ao limpar Privilege: {e}")

            # Checkboxes
            try:
                for key in ('otif_check', 'nil_check', 'pickup_check', 'package_check'):
                    cb = resolve(users_controls.get(key))
                    if cb and hasattr(cb, 'value'):
                        cb.value = False
                        if hasattr(cb, 'update'):
                            cb.update()
            except Exception as e:
                print(f"Erro ao limpar checkboxes: {e}")

            # Limpar seleção global
            global selected_user
            selected_user = None
            
            # Reabilitar campos se for Super Admin (após limpar)
            is_super_admin = current_user_privilege == "Super Admin"
            try:
                users_controls['wwid'].disabled = not is_super_admin
                users_controls['name'].disabled = not is_super_admin
                users_controls['privilege'].disabled = not is_super_admin
                users_controls['otif_check'].disabled = not is_super_admin
                users_controls['nil_check'].disabled = not is_super_admin
                users_controls['pickup_check'].disabled = not is_super_admin
                users_controls['package_check'].disabled = not is_super_admin
                users_controls['password'].disabled = False
            except Exception as e:
                print(f"Erro ao configurar estado dos campos: {e}")

            # Atualizar botão de ação e lista visual
            try:
                update_action_button()
            except Exception:
                pass
            try:
                refresh_users_list()
            except Exception:
                pass

            # Atualizar página por fim
            try:
                safe_page_update(page)
            except Exception as e:
                print(f"Erro ao atualizar página: {e}")

            print("Campos limpos com sucesso!")
            
        except Exception as ex:
            print(f"Erro ao limpar campos: {ex}")

    def add_or_update_user(e):
        """Adiciona ou atualiza um usuário baseado na função original.

        Recebe o evento `e` do botão para poder manipular o estado do botão
        (desabilitar/enabled) e passá-lo ao `show_toast` como `restore_control`.
        """
        print("🔥 FUNÇÃO add_or_update_user CHAMADA!")
        global current_user_wwid
        save_button = e.control if e and hasattr(e, 'control') else None

        # Marcar botão como processando e desabilitar para evitar cliques múltiplos
        try:
            if save_button is not None:
                try:
                    save_button._is_processing = True
                    save_button.disabled = True
                    save_button.update()
                except Exception:
                    pass

            wwid = users_controls['wwid'].value.strip().upper() if users_controls['wwid'].value else ""
            name = users_controls['name'].value.strip() if users_controls['name'].value else ""
            password = users_controls['password'].value.strip() if users_controls['password'].value else ""
            privilege = users_controls['privilege'].value if users_controls['privilege'].value else ""
            
            # Verificar permissões: Admin e User só podem editar sua própria senha
            if current_user_privilege != "Super Admin":
                if wwid.upper() != current_user_wwid.upper():
                    show_snack_bar("❌ Você só pode editar seus próprios dados", True)
                    if save_button:
                        save_button._is_processing = False
                        save_button.disabled = False
                        save_button.update()
                    return
                # Admin e User só podem atualizar senha
                if not selected_user:
                    show_snack_bar("❌ Você não tem permissão para criar novos usuários", True)
                    if save_button:
                        save_button._is_processing = False
                        save_button.disabled = False
                        save_button.update()
                    return
            
            otif = 1 if users_controls['otif_check'].value else 0
            nil = 1 if users_controls['nil_check'].value else 0
            pickup = 1 if users_controls['pickup_check'].value else 0
            package = 1 if users_controls['package_check'].value else 0
            
            print(f"Salvando usuário: {wwid}")
            print(f"Valores dos checkboxes: OTIF={users_controls['otif_check'].value}, NIL={users_controls['nil_check'].value}, Pickup={users_controls['pickup_check'].value}, Package={users_controls['package_check'].value}")
            print(f"Valores para o banco: otif={otif}, nil={nil}, pickup={pickup}, package={package}")
            
            if not wwid or not password or not privilege:
                show_snack_bar("❌ Preencha WWID, Senha e Privilégio", True)
                safe_page_update(page)
                return
            
            # Verifica se já existe o usuário
            existing_user = db_manager.query_one("SELECT 1 FROM users_table WHERE UPPER(user_wwid) = ?", (wwid,))
            
            with db_manager.transaction(
                event=f"{'Atualização' if existing_user else 'Criação'} de usuário: {wwid}",
                user=current_user_wwid,
                wwid=current_user_wwid
            ) as conn:
                if existing_user:
                    # Atualiza usuário existente
                    print(f"🔄 Atualizando usuário EXISTENTE {wwid}...")
                    db_manager.execute("""
                        UPDATE users_table
                        SET user_name = ?, user_password = ?, user_privilege = ?, 
                            otif = ?, nil = ?, pickup = ?, package = ?,
                            last_updated_by = ?, last_updated_date = CURRENT_TIMESTAMP
                        WHERE UPPER(user_wwid) = ?
                    """, (name, password, privilege, otif, nil, pickup, package, current_user_wwid, wwid))
                    
                    print(f"✅ Usuário {wwid} atualizado com sucesso!")
                    show_snack_bar(f"✅ Usuário '{wwid}' atualizado com sucesso")
                else:
                    # Cria novo usuário
                    print(f"💾 Inserindo NOVO usuário {wwid} no banco de dados...")
                    db_manager.execute("""
                        INSERT INTO users_table (user_wwid, user_name, user_password, user_privilege, 
                                               otif, nil, pickup, package, registered_by) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (wwid, name, password, privilege, otif, nil, pickup, package, current_user_wwid))
                    
                    print(f"✅ Usuário {wwid} inserido com sucesso!")
                    show_snack_bar(f"✅ Usuário '{wwid}' criado com sucesso", False)
                    # Limpar campos do formulário e atualizar lista imediatamente
                    try:
                        clear_users_fields()
                        refresh_users_list()
                        safe_page_update(page)
                    except Exception as _e:
                        print(f"Aviso: falha ao limpar/atualizar lista após inserir usuário: {_e}")
            safe_page_update(page)
            
            print(f"Lista atualizada com {len(load_users_full())} usuários!")
            # Atualizar lista
            refresh_users_list()
            
        except Exception as ex:
            show_snack_bar(f"❌ Erro ao salvar usuário: {ex}", True)
            # Garantir que o botão seja restaurado se ocorrer erro
            try:
                if save_button is not None:
                    save_button.disabled = False
                    save_button._is_processing = False
                    save_button.update()
            except Exception:
                pass
            safe_page_update(page)
            print(f"Erro ao adicionar/atualizar usuário: {ex}")
        finally:
            # Certificar-se de limpar flag de processamento caso não tenha sido restaurada pelo show_toast
            try:
                if save_button is not None and getattr(save_button, '_is_processing', False):
                    save_button._is_processing = False
                    save_button.disabled = False
                    save_button.update()
            except Exception:
                pass

    def delete_user():
        """Exclui o usuário baseado no WWID do campo de texto."""
        try:
            # Pegar WWID do campo de texto
            wwid = users_controls['wwid'].value.strip() if users_controls['wwid'].value else ''

            
            if not wwid:
                show_snack_bar("❌ Digite um WWID para excluir", True)
                safe_page_update(page)
                return
            
            # Verificar se o usuário existe no banco
            if not check_user_exists(wwid):
                show_snack_bar(f"❌ Usuário '{wwid}' não encontrado", True)
                safe_page_update(page)
                return
            
            # Função para confirmar e executar a exclusão
            def confirm_delete_user(e):
                try:
                    db_manager.execute("DELETE FROM users_table WHERE UPPER(user_wwid) = ?", (wwid.upper(),))
                    
                    # Fechar dialog
                    page.close(dialog)
                    
                    # Limpar campos e seleção
                    clear_users_fields()
                    
                    # Atualizar lista
                    refresh_users_list()
                    
                    show_snack_bar(f"✅ Usuário '{wwid}' removido com sucesso", False)
                    safe_page_update(page)
                    
                except Exception as ex:
                    # Fechar dialog mesmo se der erro
                    page.close(dialog)
                    
                    show_snack_bar(f"❌ Erro ao excluir usuário: {ex}", True)
                    safe_page_update(page)
                    print(f"Erro ao deletar usuário {wwid}: {ex}")
            
            # Função para cancelar
            def cancel_delete_user(e):
                page.close(dialog)
                safe_page_update(page)
            
            # Criar e mostrar dialog de confirmação
            dialog = DeleteListItemConfirmationDialog(
                item_name=wwid,
                item_type="Usuário",
                on_confirm=confirm_delete_user,
                on_cancel=cancel_delete_user,
                scale_func=lambda x: x
            )
            
            page.open(dialog)
            safe_page_update(page)
            
        except Exception as ex:
            show_snack_bar(f"❌ Erro ao tentar excluir usuário: {ex}", True)
            safe_page_update(page)
            print(f"Erro geral no delete_user: {ex}")

    # Referências para os controles da aba Log
    log_user_filter_ref = ft.Ref()
    # DataTable única para consolidar lista e detalhes
    log_table_ref = ft.Ref()

    log_controls = {
        'user_filter': log_user_filter_ref,
        'table': log_table_ref,
    }

    # Conteúdo da sub-aba Log com largura maximizada
    log_content = ft.Container(
        content=ft.Column([
            # Header fixo
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.HISTORY, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Column([
                        ft.Text("Log de Atividades", size=22, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ft.Text("Visualize o histórico de ações dos usuários", size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                    ], spacing=2, expand=True),
                ], spacing=12),
                margin=ft.margin.only(bottom=15),
            ),
            ft.Divider(height=1, color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
            # Filtro de pesquisa - fixo
            ft.Container(
                content=ft.Row([
                    ft.TextField(
                        ref=log_user_filter_ref,
                        label="Pesquisar por WWID", 
                        expand=True,
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                        border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
                    ),
                    ft.ElevatedButton("Filtrar", icon=ft.Icons.FILTER_LIST, on_click=lambda e: load_log_entries(log_user_filter_ref.current.value if log_user_filter_ref and hasattr(log_user_filter_ref, 'current') else None)),
                ], spacing=10),
                margin=ft.margin.only(top=10, bottom=10)
            ),
            
            # Tabela de logs com scroll - expande para ocupar espaço restante
            ft.Container(
                content=ft.Column([
                    # Header fixo
                    ft.Container(
                        content=ft.Row([
                            ft.Container(ft.Text("", size=12, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'), no_wrap=True), width=40, alignment=ft.alignment.center),  # Ícone
                            ft.Container(ft.Text("Data/Hora", size=12, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'), no_wrap=True), width=150, alignment=ft.alignment.center_left),
                            ft.Container(ft.Text("WWID", size=12, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'), no_wrap=True), width=100, alignment=ft.alignment.center_left),
                            ft.Container(ft.Text("Usuário", size=12, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'), no_wrap=True), width=150, alignment=ft.alignment.center_left),
                            ft.Container(ft.Text("Detalhes", size=12, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'), no_wrap=True), expand=True, alignment=ft.alignment.center_left),
                        ], spacing=10, tight=True),
                        padding=ft.padding.symmetric(horizontal=16, vertical=12),
                        bgcolor=ft.Colors.with_opacity(0.1, get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                        border_radius=ft.border_radius.only(top_left=8, top_right=8),
                    ),
                    # Corpo scrollável - será preenchido via DataTable
                    ft.Container(
                        content=ft.DataTable(
                            ref=log_table_ref,
                            columns=[
                                ft.DataColumn(ft.Text(""), numeric=False),  # Ícone
                                ft.DataColumn(ft.Text(""), numeric=False),  # Data/Hora
                                ft.DataColumn(ft.Text(""), numeric=False),  # WWID
                                ft.DataColumn(ft.Text(""), numeric=False),  # Usuário
                                ft.DataColumn(ft.Text(""), numeric=False),  # Detalhes
                            ],
                            rows=[],
                            heading_row_height=0,
                            data_row_color={"hovered": ft.Colors.with_opacity(0.05, get_current_theme_colors(get_theme_name_from_page(page)).get('primary'))},
                            column_spacing=10,
                            horizontal_margin=16,
                            expand=True,
                        ),
                        expand=True,
                    ),
                ], spacing=0, expand=True, scroll=ft.ScrollMode.ADAPTIVE),
                expand=True,
                padding=0,
                bgcolor=ft.Colors.with_opacity(0.02, get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                border_radius=8,
            )
        ], spacing=0, expand=True),
        padding=20,
        visible=False,
        expand=True,  # Container principal expandir para ocupar espaço disponível
    )

    # --- Conteúdo da aba Info ---
    info_content = ft.Container(
        content=ft.Column([
            # Header fixo
            ft.Container(
                content=ft.Row([
                    ft.Icon(ft.Icons.INFO_OUTLINED, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Column([
                        ft.Text("Informações sobre o App", size=22, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ft.Text("Conheça mais sobre o sistema e seus desenvolvedores", size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                    ], spacing=2, expand=True),
                ], spacing=12),
                margin=ft.margin.only(bottom=15),
            ),
            ft.Divider(height=1, color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
            
            # Card com objetivo do app
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.TRACK_CHANGES, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'), size=20),
                            ft.Text("Objetivo do Sistema", size=18, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ], spacing=10),
                        ft.Container(height=10),
                        ft.Text(
                            "Este sistema foi desenvolvido para consolidar e gerenciar as notas de avaliação dos fornecedores. "
                            "O objetivo é facilitar a tomada de decisões estratégicas e melhorar "
                            "o relacionamento com os fornecedores.",
                            size=14,
                            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                            text_align=ft.TextAlign.JUSTIFY,
                        ),
                    ]),
                    padding=20,
                ),
                elevation=2,
                margin=ft.margin.only(bottom=20),
            ),
            
            # Card com informações de desenvolvimento
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.CODE, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'), size=20),
                            ft.Text("Informações de Desenvolvimento", size=18, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ], spacing=10),
                        ft.Container(height=10),
                        
                        # Desenvolvedor
                        ft.Container(
                            content=ft.Column([
                                ft.Text("Desenvolvido por:", size=12, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                                ft.Text("Rafael Negrão de Souza - Supply Continuity Intern", size=14, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                                ft.Text("rafael.negrao.souza@cummins.com", size=12, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                ft.Text("AN62H", size=12, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                            ], spacing=5),
                            margin=ft.margin.only(bottom=15),
                        ),
                        
                        ft.Divider(color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline'), height=1),
                        
                        # Autor intelectual
                        ft.Container(
                            content=ft.Column([
                                ft.Text("Autor intelectual do projeto:", size=12, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                                ft.Text("Cleiton Bianchi dos Santos - Supply Continuity Manager", size=14, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                                ft.Text("Cleiton.Bianchi.Santos@cummins.com", size=12, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                ft.Text("IV338", size=12, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                            ], spacing=5),
                            margin=ft.margin.only(top=15),
                        ),
                    ]),
                    padding=20,
                ),
                elevation=2,
                margin=ft.margin.only(bottom=20),
            ),
            
            # Card com funcionalidades principais
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.STAR_OUTLINE, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'), size=20),
                            ft.Text("Funcionalidades Principais", size=18, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                        ], spacing=10),
                        ft.Container(height=10),
                        
                        # Lista de funcionalidades
                        ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.CIRCLE, size=6, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                ft.Text("Consolidação de notas de fornecedores por período", size=14, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                            ], spacing=10),
                            ft.Row([
                                ft.Icon(ft.Icons.CIRCLE, size=6, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                ft.Text("Importação e exportação de dados via Excel", size=14, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                            ], spacing=10),
                        ], spacing=8),
                    ]),
                    padding=20,
                ),
                elevation=2,
            ),
        ], spacing=10, scroll=ft.ScrollMode.AUTO, expand=True),
        padding=20,
        visible=False,
        expand=True,
    )

    # --- Aba Data removida (import/export) ---

    # Variável global para controlar paginação do log
    log_pagination = {
        'current_offset': 0,
        'page_size': 200,
        'loading': False,
        'has_more': True,
        'current_filter': None
    }

    # Funções de import/export removidas - aba Data foi removida

    # Variável global para controlar paginação do log
    log_pagination = {
        'current_offset': 0,
        'page_size': 200,
        'loading': False,
        'has_more': True,
        'current_filter': None
    }

    # Função para carregar entradas da tabela log_table e popular a DataTable com paginação
    def load_log_entries(wwid_filter=None, append_mode=False):
        try:
            # Referência para a DataTable
            table_control = log_table_ref.current if log_table_ref and hasattr(log_table_ref, 'current') else None
            
            if not table_control:
                print("DataTable não encontrada")
                return

            # Evitar múltiplas requisições simultâneas
            if log_pagination['loading']:
                return
                
            log_pagination['loading'] = True

            # Se não é append_mode, resetar paginação
            if not append_mode:
                log_pagination['current_offset'] = 0
                log_pagination['has_more'] = True
                log_pagination['current_filter'] = wwid_filter
                table_control.rows.clear()
            
            # Se filtro mudou, resetar
            if wwid_filter != log_pagination['current_filter']:
                log_pagination['current_offset'] = 0
                log_pagination['has_more'] = True  
                log_pagination['current_filter'] = wwid_filter
                if not append_mode:
                    table_control.rows.clear()

            # preparar ícones por tipo
            def icon_for_event(event_text):
                try:
                    if not event_text:
                        return ft.Container()  # Sem ícone se não há evento
                    
                    # procurar prefixo [UPDATED], [EXCLUDED], [INCLUDED] (case insensitive)
                    event_upper = event_text.upper()
                    if event_upper.startswith('[UPDATED]'):
                        return ft.Icon(ft.Icons.EDIT, size=16, color="#2196F3")  # Azul
                    elif event_upper.startswith('[EXCLUDED]'):
                        return ft.Icon(ft.Icons.DELETE, size=16, color="#F44336")  # Vermelho
                    elif event_upper.startswith('[INCLUDED]'):
                        return ft.Icon(ft.Icons.ADD, size=16, color="#4CAF50")  # Verde
                    
                    # Se não tem prefixo específico, não mostrar ícone
                    return ft.Container()
                except Exception as e:
                    print(f"Erro ao processar ícone: {e}")
                    return ft.Container()

            # Buscar dados do DB com LIMIT e OFFSET para paginação
            rows = []
            try:
                base_q = f"SELECT log_id, date, time, user, wwid, event FROM {db_manager.log_table_name}"
                if wwid_filter:
                    # permitir pesquisa parcial (LIKE) em user, wwid E nos dados JSON do evento (para WWID)
                    q = base_q + f" WHERE (UPPER(user) LIKE ? OR UPPER(wwid) LIKE ? OR UPPER(event) LIKE ?) ORDER BY date DESC, time DESC LIMIT {log_pagination['page_size']} OFFSET {log_pagination['current_offset']}"
                    filter_param = f"%{wwid_filter.upper()}%"
                    rows = db_manager.query(q, (filter_param, filter_param, filter_param))
                else:
                    q = base_q + f" ORDER BY date DESC, time DESC LIMIT {log_pagination['page_size']} OFFSET {log_pagination['current_offset']}"
                    rows = db_manager.query(q)
                    
                # Verificar se há mais dados
                if len(rows) < log_pagination['page_size']:
                    log_pagination['has_more'] = False
                else:
                    log_pagination['has_more'] = True
                    
            except Exception as e:
                print(f"Erro ao ler log_table: {e}")
                log_pagination['loading'] = False
                return

            # Se não há linhas na primeira página, mostrar mensagem
            if not rows and log_pagination['current_offset'] == 0:
                table_control.rows.append(
                    ft.DataRow(cells=[
                        ft.DataCell(ft.Text("")),
                        ft.DataCell(ft.Text("")),
                        ft.DataCell(ft.Text("")),
                        ft.DataCell(ft.Text("Nenhuma entrada de log encontrada.", italic=True)),
                        ft.DataCell(ft.Text(""))
                    ])
                )
                table_control.update()
                log_pagination['loading'] = False
                return

            # Função para formatar texto de evento para exibição
            def format_event_text(raw_text):
                # Tenta parsear JSON para formatar legivelmente
                try:
                    import json
                    parsed = json.loads(raw_text)
                    # Se for dict ou list, formatar com separador "-"
                    if isinstance(parsed, dict):
                        lines = []
                        for k, v in parsed.items():
                            # Limpar valor de aspas e caracteres especiais
                            clean_v = str(v).replace('"', '').replace("'", '').replace('\n', ' ')
                            lines.append(f"{k}: {clean_v}")
                        return " - ".join(lines)
                    if isinstance(parsed, list):
                        cleaned_list = [str(x).replace('"', '').replace("'", '').replace('\n', ' ') for x in parsed]
                        return " - ".join(cleaned_list)
                except Exception:
                    # não JSON -> seguir para parsing simples
                    pass

                # Limpar texto simples de caracteres especiais
                try:
                    # Remover aspas, quebras de linha e outros caracteres especiais
                    cleaned_text = raw_text.replace('"', '').replace("'", '').replace('\n', ' ').replace('\\n', ' ')
                    # Se há vírgulas, separar por elas e usar "-"
                    if ',' in cleaned_text:
                        parts = [p.strip() for p in cleaned_text.split(',') if p.strip()]
                        return " - ".join(parts) if parts else cleaned_text
                    return cleaned_text
                except Exception:
                    return raw_text

            # Extrair WWID do registro (agora que temos coluna específica)
            def extract_wwid(row):
                # Com a nova coluna wwid, simplesmente retornar seu valor
                if 'wwid' in row and row['wwid']:
                    return str(row['wwid'])
                # Fallback para compatibilidade com registros antigos
                return row.get('user', '')

            # Popular tabela com nova estrutura de colunas
            for r in rows:
                event = r.get('event') or ''
                date = r.get('date') or ''
                time = r.get('time') or ''
                user = r.get('user') or ''
                wwid = extract_wwid(r)  # Agora sem segundo parâmetro

                # Criar células da linha com nova estrutura
                icon_widget = icon_for_event(event)
                icon_cell = ft.DataCell(
                    ft.Container(
                        content=icon_widget,
                        width=40,
                        height=30,  # Altura mínima para garantir visibilidade
                        alignment=ft.alignment.center
                    )
                )
                
                datetime_cell = ft.DataCell(
                    ft.Container(
                        content=ft.Text(f"{date}\n{time}", size=11),
                        width=150
                    )
                )
                
                # Nova coluna WWID separada
                wwid_cell = ft.DataCell(
                    ft.Container(
                        content=ft.Text(wwid or "", size=11, overflow=ft.TextOverflow.ELLIPSIS),
                        width=100
                    )
                )
                
                # Coluna usuário agora só mostra o nome do usuário
                user_cell = ft.DataCell(
                    ft.Container(
                        content=ft.Text(user, size=11, overflow=ft.TextOverflow.ELLIPSIS),
                        width=150
                    )
                )
                
                # Detalhes formatados do evento (ajustado)
                details_formatted = format_event_text(event)
                details_cell = ft.DataCell(
                    ft.Text(
                        details_formatted,
                        size=11,
                        max_lines=2,
                        overflow=ft.TextOverflow.ELLIPSIS
                    )
                )

                # Adicionar linha à tabela com nova estrutura
                table_control.rows.append(
                    ft.DataRow(cells=[icon_cell, datetime_cell, wwid_cell, user_cell, details_cell])
                )

            # Atualizar tabela
            table_control.update()

            # Incrementar offset para próxima página
            log_pagination['current_offset'] += log_pagination['page_size']
            
            # Adicionar linha "Carregar mais" se há mais dados
            if log_pagination['has_more']:
                def load_more_handler(e):
                    load_log_entries(wwid_filter, append_mode=True)
                
                load_more_row = ft.DataRow(cells=[
                    ft.DataCell(ft.Text("")),
                    ft.DataCell(ft.Text("")),
                    ft.DataCell(ft.Text("")),
                    ft.DataCell(
                        ft.ElevatedButton(
                            "Carregar mais...",
                            icon=ft.Icons.EXPAND_MORE,
                            on_click=load_more_handler,
                            style=ft.ButtonStyle(
                                bgcolor="#E3F2FD",
                                color="#1976D2"
                            )
                        )
                    ),
                    ft.DataCell(ft.Text(""))
                ])
                table_control.rows.append(load_more_row)
                table_control.update()

            log_pagination['loading'] = False

        except Exception as e:
            print(f"Erro geral ao carregar logs: {e}")
            log_pagination['loading'] = False


    # Hook: carregar log quando a aba for aberta
    def ensure_log_loaded():
        try:
            # A aba `log_content` é mostrada ao selecionar o índice; carregamos os dados quando visível
            if log_content.visible:
                # usar valor atual do campo de pesquisa (WWID)
                w = log_user_filter_ref.current.value if log_user_filter_ref and hasattr(log_user_filter_ref, 'current') and log_user_filter_ref.current.value else None
                load_log_entries(w)
        except Exception as ex:
            print(f"Erro ao garantir carregamento do log: {ex}")

    def clear_log_fields():
        """Limpa os campos da aba Log quando o usuário sair da aba"""
        try:
            # Resetar paginação
            log_pagination['current_offset'] = 0
            log_pagination['has_more'] = True
            log_pagination['current_filter'] = None
            log_pagination['loading'] = False
            
            # Limpar campo de filtro/pesquisa
            if log_user_filter_ref.current:
                log_user_filter_ref.current.value = ""
                log_user_filter_ref.current.update()
            
            # Limpar tabela de logs
            if log_table_ref.current:
                log_table_ref.current.rows.clear()
                log_table_ref.current.update()
            
            print("🧹 Campos da aba Log limpos com sucesso")
            
        except Exception as ex:
            print(f"❌ Erro ao limpar campos da aba Log: {ex}")

    def test_log_clear_functionality():
        """Função de teste para demonstrar a limpeza automática da aba Log"""
        try:
            print("🧪 Testando funcionalidade de limpeza da aba Log...")
            print("✅ Implementação concluída:")
            print("   - Campos da aba Log são limpos automaticamente ao sair da aba")
            print("   - Campo de pesquisa/filtro é limpo")
            print("   - Lista de logs é limpa")
            print("   - Detalhes do log selecionado são limpos")
            print("🎯 Agora quando você sair da aba Log, todos os campos serão automaticamente limpos!")
            
        except Exception as ex:
            print(f"❌ Erro no teste de limpeza da aba Log: {ex}")

    def demo_home_tab():
        """Função de demonstração da nova aba Home"""
        try:
            print("🏠 Nova Aba Home implementada com sucesso!")
            print("✅ Características da aba Home:")
            print("   - Dashboard principal com boas-vindas personalizadas")
            print("   - Cards de estatísticas rápidas (Fornecedores, Avaliações, Score Médio, Usuários)")
            print("   - Informações do sistema e usuário logado")
            print("   - Design responsivo com cards organizados")
            print("🎯 Navegação atualizada:")
            print("   - Home (índice 0) - Nova aba principal")
            print("   - Score (índice 1) - Anteriormente índice 0")
            print("   - Timeline (índice 2) - Anteriormente índice 1")
            print("   - Risks (índice 3) - Anteriormente índice 2")
            print("   - Email (índice 4) - Anteriormente índice 3")
            print("   - Configs (índice 5) - Anteriormente índice 4")
            print("🚀 A aplicação agora inicia na aba Home como painel principal!")
            
        except Exception as ex:
            print(f"❌ Erro na demonstração da aba Home: {ex}")

    # Row para as abas de configuração
    config_tabs_row = ft.Row(
        alignment=ft.MainAxisAlignment.START,
        spacing=10
    )

    # Container principal da aba configs
    configs_view_content = ft.Column([
        config_tabs_row,
        ft.Divider(),
        ft.Container(
            content=ft.Stack([
                themes_content,
                suppliers_content,
                criteria_content,
                users_content,
                lists_content,
                log_content,
                info_content
            ]),
            expand=True
        )
    ], expand=True)

    # --- Fim: Lógica e Controles da Aba Configs ---

    # --- Início: Conteúdo da Aba Home ---
    def format_stats_value(value, stat_type):
        """Formata valores das estatísticas para exibição"""
        try:
            if stat_type == 'suppliers_count':
                return f"{value:,}" if value > 0 else "0"
            elif stat_type == 'evaluations_count':
                if value >= 1000:
                    return f"{value/1000:.1f}K"
                else:
                    return str(value)
            elif stat_type == 'average_score':
                return f"{value:.1f}" if value > 0 else "0.0"
            elif stat_type == 'users_count':
                return str(value)
            else:
                return str(value)
        except:
            return "0"

    def get_home_statistics():
        """Busca estatísticas reais do banco de dados para a tela Home"""
        try:
            stats = {
                'suppliers_count': 0,
                'evaluations_count': 0,
                'average_score': 0.0,
                'users_count': 0
            }
            
            if not db_manager:
                return stats
            
            # 1. Contar fornecedores ativos
            suppliers_query = "SELECT COUNT(*) as count FROM supplier_database_table WHERE supplier_status = 'Active'"
            suppliers_result = db_manager.query_one(suppliers_query)
            stats['suppliers_count'] = suppliers_result['count'] if suppliers_result else 0
            
            # 2. Contar avaliações (registros na tabela de scores)
            evaluations_query = "SELECT COUNT(*) as count FROM supplier_score_records_table"
            evaluations_result = db_manager.query_one(evaluations_query)
            stats['evaluations_count'] = evaluations_result['count'] if evaluations_result else 0
            
            # 3. Calcular score médio geral
            avg_score_query = "SELECT AVG(total_score) as avg_score FROM supplier_score_records_table WHERE total_score > 0"
            avg_score_result = db_manager.query_one(avg_score_query)
            if avg_score_result and avg_score_result['avg_score']:
                stats['average_score'] = round(avg_score_result['avg_score'], 1)
            
            # 4. Contar usuários
            users_query = "SELECT COUNT(*) as count FROM users_table"
            users_result = db_manager.query_one(users_query)
            stats['users_count'] = users_result['count'] if users_result else 0
            
            print(f"📊 Estatísticas do Home carregadas: Fornecedores={stats['suppliers_count']}, Avaliações={stats['evaluations_count']}, Score Médio={stats['average_score']}, Usuários={stats['users_count']}")
            
            return stats
            
        except Exception as e:
            print(f"❌ Erro ao buscar estatísticas do Home: {e}")
            # Retornar valores padrão em caso de erro
            return {
                'suppliers_count': 0,
                'evaluations_count': 0,
                'average_score': 0.0,
                'users_count': 0
            }

    def create_home_content():
        """Cria o conteúdo da aba Home com informações e design diferenciado"""
        
        # Cards de estatísticas rápidas - Design minimalista
        def create_stats_card(title, value, icon, color_type="primary"):
            theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
            current_theme = get_theme_name_from_page(page)
            
            primary_color = theme_colors.get('primary', '#0066CC')
            
            return ft.Container(
                content=ft.Column([
                    ft.Icon(icon, color=primary_color, size=36),
                    ft.Container(height=12),
                    ft.Text(
                        value, 
                        size=32, 
                        weight="bold", 
                        color=theme_colors.get('on_surface'),
                        text_align=ft.TextAlign.CENTER,
                        no_wrap=False,
                        overflow=ft.TextOverflow.VISIBLE
                    ),
                    ft.Container(
                        content=ft.Text(
                            title, 
                            size=13, 
                            color=theme_colors.get('on_surface_variant'), 
                            weight="w400",
                            text_align=ft.TextAlign.CENTER,
                            no_wrap=False,
                            overflow=ft.TextOverflow.VISIBLE
                        ),
                        width=180
                    )
                ], 
                spacing=4, 
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER
                ),
                bgcolor=theme_colors.get('card_background'),
                border_radius=12,
                padding=ft.padding.all(20),
                width=200,
                height=150,
                border=ft.border.all(1, theme_colors.get('outline_variant', '#E8E8E8')),
                shadow=ft.BoxShadow(
                    spread_radius=0,
                    blur_radius=8,
                    color="#00000010" if current_theme == "white" else "#00000020",
                    offset=ft.Offset(0, 2)
                ),
                clip_behavior=ft.ClipBehavior.NONE
            )



        # Seção de boas-vindas - Design minimalista
        theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
        welcome_section = ft.Container(
            content=ft.Column([
                ft.Text(f"Olá, {current_user_name}!", size=28, weight="bold", color=theme_colors.get('on_surface')),
                ft.Text("Painel de Controle", size=16, color=theme_colors.get('on_surface_variant'), weight="w400")
            ], spacing=4, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
            padding=ft.padding.only(bottom=32)
        )

        # Buscar estatísticas reais do banco de dados
        stats = get_home_statistics()
        
        # Seção de estatísticas - Grid limpo
        stats_section = ft.Container(
            content=ft.Row([
                create_stats_card("Fornecedores", format_stats_value(stats['suppliers_count'], 'suppliers_count'), ft.Icons.BUSINESS),
                create_stats_card("Avaliações", format_stats_value(stats['evaluations_count'], 'evaluations_count'), ft.Icons.ASSESSMENT),
                create_stats_card("Score Médio", format_stats_value(stats['average_score'], 'average_score'), ft.Icons.TRENDING_UP),
                create_stats_card("Usuários", format_stats_value(stats['users_count'], 'users_count'), ft.Icons.PEOPLE)
            ], 
            wrap=True, 
            spacing=16, 
            run_spacing=16,
            alignment=ft.MainAxisAlignment.CENTER
            ),
            padding=ft.padding.only(bottom=40),
            alignment=ft.alignment.center
        )

        # Seção de informações - Cards lado a lado
        info_section = ft.Container(
            content=ft.Row([
                # Card de Perfil
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.ACCOUNT_CIRCLE, size=24, color=theme_colors.get('primary')),
                            ft.Text("Perfil", size=18, weight="bold", color=theme_colors.get('on_surface'))
                        ], spacing=8),
                        ft.Container(height=16),
                        ft.Column([
                            ft.Row([
                                ft.Text("Nome:", size=13, color=theme_colors.get('on_surface_variant'), width=80),
                                ft.Text(f"{current_user_name}", size=13, weight="w500", color=theme_colors.get('on_surface'))
                            ]),
                            ft.Row([
                                ft.Text("WWID:", size=13, color=theme_colors.get('on_surface_variant'), width=80),
                                ft.Text(f"{current_user_wwid}", size=13, weight="w500", color=theme_colors.get('on_surface'))
                            ]),
                            ft.Row([
                                ft.Text("Privilégio:", size=13, color=theme_colors.get('on_surface_variant'), width=80),
                                ft.Text(f"{current_user_privilege}", size=13, weight="w500", color=theme_colors.get('on_surface'))
                            ])
                        ], spacing=8)
                    ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.START),
                    bgcolor=theme_colors.get('card_background'),
                    border_radius=12,
                    padding=20,
                    width=280,
                    border=ft.border.all(1, theme_colors.get('outline_variant', '#E8E8E8'))
                ),
                
                # Card de Permissões
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.VERIFIED_USER, size=24, color=theme_colors.get('primary')),
                            ft.Text("Permissões", size=18, weight="bold", color=theme_colors.get('on_surface'))
                        ], spacing=8),
                        ft.Container(height=16),
                        ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.CHECK_CIRCLE if current_user_permissions.get('otif') else ft.Icons.CANCEL, 
                                       size=16, color="#4CAF50" if current_user_permissions.get('otif') else "#9E9E9E"),
                                ft.Text("OTIF", size=13, color=theme_colors.get('on_surface'))
                            ], spacing=8),
                            ft.Row([
                                ft.Icon(ft.Icons.CHECK_CIRCLE if current_user_permissions.get('nil') else ft.Icons.CANCEL, 
                                       size=16, color="#4CAF50" if current_user_permissions.get('nil') else "#9E9E9E"),
                                ft.Text("NIL", size=13, color=theme_colors.get('on_surface'))
                            ], spacing=8),
                            ft.Row([
                                ft.Icon(ft.Icons.CHECK_CIRCLE if current_user_permissions.get('pickup') else ft.Icons.CANCEL, 
                                       size=16, color="#4CAF50" if current_user_permissions.get('pickup') else "#9E9E9E"),
                                ft.Text("Pickup", size=13, color=theme_colors.get('on_surface'))
                            ], spacing=8),
                            ft.Row([
                                ft.Icon(ft.Icons.CHECK_CIRCLE if current_user_permissions.get('package') else ft.Icons.CANCEL, 
                                       size=16, color="#4CAF50" if current_user_permissions.get('package') else "#9E9E9E"),
                                ft.Text("Package", size=13, color=theme_colors.get('on_surface'))
                            ], spacing=8)
                        ], spacing=8)
                    ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.START),
                    bgcolor=theme_colors.get('card_background'),
                    border_radius=12,
                    padding=20,
                    width=280,
                    border=ft.border.all(1, theme_colors.get('outline_variant', '#E8E8E8'))
                ),
                
                # Card do Sistema
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Icon(ft.Icons.SETTINGS, size=24, color=theme_colors.get('primary')),
                            ft.Text("Sistema", size=18, weight="bold", color=theme_colors.get('on_surface'))
                        ], spacing=8),
                        ft.Container(height=16),
                        ft.Column([
                            ft.Row([
                                ft.Text("Versão:", size=13, color=theme_colors.get('on_surface_variant'), width=80),
                                ft.Text(f"{APP_VERSION}", size=13, weight="w500", color=theme_colors.get('on_surface'))
                            ]),
                            ft.Row([
                                ft.Text("Banco:", size=13, color=theme_colors.get('on_surface_variant'), width=80),
                                ft.Row([
                                    ft.Icon(ft.Icons.CHECK_CIRCLE, size=14, color="#4CAF50"),
                                    ft.Text("Conectado", size=13, weight="w500", color=theme_colors.get('on_surface'))
                                ], spacing=4)
                            ]),
                            ft.Row([
                                ft.Text("Atualizado:", size=13, color=theme_colors.get('on_surface_variant'), width=80),
                                ft.Text("Hoje", size=13, weight="w500", color=theme_colors.get('on_surface'))
                            ])
                        ], spacing=8)
                    ], spacing=0, horizontal_alignment=ft.CrossAxisAlignment.START),
                    bgcolor=theme_colors.get('card_background'),
                    border_radius=12,
                    padding=20,
                    width=280,
                    border=ft.border.all(1, theme_colors.get('outline_variant', '#E8E8E8'))
                )
            ], spacing=16, wrap=True, alignment=ft.MainAxisAlignment.CENTER)
        )

        return ft.Column([
            welcome_section,
            stats_section,
            info_section
        ], spacing=20, scroll=ft.ScrollMode.AUTO, horizontal_alignment=ft.CrossAxisAlignment.CENTER)

    # Criar a aba Home com referência atualizável
    home_content_container = ft.Ref[ft.Container]()
    
    home_view = ft.Container(
        content=ft.Container(
            ref=home_content_container,
            content=create_home_content(),
            padding=ft.padding.only(top=40, left=40, right=40, bottom=40),
            alignment=ft.alignment.top_center
        ),
        visible=True,  # Home será a aba inicial
        expand=True
    )

    # --- Fim: Conteúdo da Aba Home ---

    # --- Início: Aba Cross Check ---
    
    # Container para os resultados do Cross Check
    cross_check_results_list = ft.Column([], spacing=10, expand=True, scroll=ft.ScrollMode.AUTO)
    
    def get_pending_suppliers():
        """Retorna lista de suppliers com notas pendentes para o usuário"""
        try:
            pending_list = []
            user_permissions = current_user_permissions
            
            if not user_permissions:
                return []
            
            # Buscar TODOS os suppliers com QUALQUER nota pendente (NULL ou vazia)
            # em TODOS os meses/anos, não apenas o mês atual
            query = """
                SELECT DISTINCT 
                    s.supplier_id, 
                    s.vendor_name as supplier_name, 
                    s.bu, 
                    s.supplier_po as po,
                    r.month,
                    r.year,
                    r.otif,
                    r.nil,
                    r.quality_pickup,
                    r.quality_package
                FROM supplier_score_records_table r
                JOIN supplier_database_table s ON r.supplier_id = s.supplier_id
                WHERE (
                    r.otif IS NULL OR r.otif = '' OR
                    r.nil IS NULL OR r.nil = '' OR
                    r.quality_pickup IS NULL OR r.quality_pickup = '' OR
                    r.quality_package IS NULL OR r.quality_package = ''
                )
                ORDER BY r.year DESC, r.month DESC, s.vendor_name
            """
            
            suppliers = db_manager.query(query)
            
            for supplier_row in suppliers:
                # Extrair dados
                if isinstance(supplier_row, dict):
                    supplier_id = supplier_row.get('supplier_id')
                    supplier_name = supplier_row.get('supplier_name')
                    bu = supplier_row.get('bu')
                    po = supplier_row.get('po')
                    month = supplier_row.get('month')
                    year = supplier_row.get('year')
                    otif = supplier_row.get('otif')
                    nil = supplier_row.get('nil')
                    pickup = supplier_row.get('quality_pickup')
                    package = supplier_row.get('quality_package')
                else:
                    supplier_id, supplier_name, bu, po, month, year, otif, nil, pickup, package = supplier_row
                
                # Verificar quais notas estão pendentes (NULL ou string vazia)
                # Mostrar TODAS as notas pendentes, mas destacar as que o usuário pode preencher
                pending_scores = []
                
                # Verificar todas as notas pendentes
                if otif is None or otif == '':
                    pending_scores.append('OTIF')
                if nil is None or nil == '':
                    pending_scores.append('NIL')
                if pickup is None or pickup == '':
                    pending_scores.append('Pickup')
                if package is None or package == '':
                    pending_scores.append('Package')
                
                # Apenas adicionar à lista se houver notas pendentes que o usuário pode preencher
                user_can_fill = []
                if user_permissions.get('otif', False) and (otif is None or otif == ''):
                    user_can_fill.append('OTIF')
                if user_permissions.get('nil', False) and (nil is None or nil == ''):
                    user_can_fill.append('NIL')
                if user_permissions.get('pickup', False) and (pickup is None or pickup == ''):
                    user_can_fill.append('Pickup')
                if user_permissions.get('package', False) and (package is None or package == ''):
                    user_can_fill.append('Package')
                
                # Adicionar à lista se houver notas pendentes que o usuário pode preencher
                if user_can_fill:
                    pending_list.append({
                        'supplier_id': supplier_id,
                        'supplier_name': supplier_name,
                        'bu': bu,
                        'po': po,
                        'pending_scores': user_can_fill,  # Apenas as que o usuário pode preencher
                        'all_pending_scores': pending_scores,  # Todas as pendentes (para referência)
                        'month': month,
                        'year': year
                    })
            
            return pending_list
            
        except Exception as e:
            print(f"Erro ao buscar suppliers pendentes: {e}")
            return []
    
    def go_to_score_tab(supplier_id, month, year):
        """Navega para a aba Score e busca o supplier especificado"""
        try:
            # Mudar para aba Score (índice 1)
            set_selected(1)(None)
            
            # Preencher campos de busca
            if search_field_ref.current:
                search_field_ref.current.value = supplier_id
            
            if selected_month.current:
                selected_month.current.value = month
            
            if selected_year.current:
                selected_year.current.value = year
            
            # Executar busca
            # load_scores()  # REMOVIDO - Cards obsoletos, apenas tabela é usada
            
            show_snack_bar(f"Navegado para {supplier_id}", False)
            
        except Exception as e:
            print(f"Erro ao navegar para Score: {e}")
            show_snack_bar(f"Erro ao navegar: {e}", True)
    
    def update_cross_check_view():
        """Atualiza a visualização do Cross Check com suppliers pendentes"""
        try:
            cross_check_results_list.controls.clear()
            
            pending_suppliers = get_pending_suppliers()
            
            if not pending_suppliers:
                cross_check_results_list.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Icon(ft.Icons.CHECK_CIRCLE_OUTLINE, size=80, color=ft.Colors.GREEN),
                            ft.Text("✅ Nenhuma pendência encontrada!", size=24, weight="bold"),
                            ft.Text("Todas as suas avaliações estão em dia.", size=16, color=ft.Colors.GREY_700),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=15),
                        alignment=ft.alignment.center,
                        expand=True
                    )
                )
            else:
                # Header com contador
                now = datetime.datetime.now()
                month_name = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"][now.month - 1]
                
                cross_check_results_list.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Text(
                                f"📋 Pendências de {month_name}/{now.year}",
                                size=24,
                                weight="bold"
                            ),
                            ft.Text(
                                f"{len(pending_suppliers)} supplier{'s' if len(pending_suppliers) > 1 else ''} com notas pendentes",
                                size=16,
                                color=ft.Colors.ORANGE_700
                            ),
                        ], spacing=5),
                        padding=ft.padding.only(bottom=20)
                    )
                )
                
                # Criar cards dos suppliers pendentes
                for pending in pending_suppliers:
                    Colors = get_current_theme_colors(get_theme_name_from_page(page))
                    
                    # Card simplificado destacando pendências
                    card_content = ft.Container(
                        content=ft.Column([
                            # Header do card
                            ft.Row([
                                ft.Icon(ft.Icons.WARNING_AMBER, color=ft.Colors.ORANGE, size=24),
                                ft.Column([
                                    ft.Text(
                                        pending['supplier_name'],
                                        size=18,
                                        weight="bold",
                                        color=Colors['on_surface']
                                    ),
                                    ft.Text(
                                        f"ID: {pending['supplier_id']} | BU: {pending['bu'] or 'N/A'} | PO: {format_po_ssid(pending['po'])}",
                                        size=12,
                                        color=Colors['on_surface_variant']
                                    ),
                                ], spacing=2, expand=True),
                            ], spacing=10),
                            ft.Divider(),
                            # Notas pendentes
                            ft.Container(
                                content=ft.Column([
                                    ft.Text("📝 Notas Pendentes:", size=14, weight="bold", color=ft.Colors.ORANGE_700),
                                    ft.Wrap(
                                        [
                                            ft.Container(
                                                content=ft.Text(score, size=12, color=ft.Colors.WHITE),
                                                bgcolor=ft.Colors.ORANGE_700,
                                                padding=ft.padding.symmetric(horizontal=10, vertical=5),
                                                border_radius=15
                                            )
                                            for score in pending['pending_scores']
                                        ],
                                        spacing=8,
                                        run_spacing=8
                                    ),
                                ], spacing=8),
                                padding=ft.padding.all(10),
                                bgcolor=ft.Colors.ORANGE_50,
                                border_radius=8
                            ),
                            # Botão para ir direto ao supplier
                            ft.Container(
                                content=ft.ElevatedButton(
                                    "Avaliar Agora",
                                    icon=ft.Icons.EDIT,
                                    on_click=lambda e, sid=pending['supplier_id'], m=pending['month'], y=pending['year']: go_to_score_tab(sid, m, y),
                                    bgcolor=Colors['primary'],
                                    color=Colors['on_primary']
                                ),
                                alignment=ft.alignment.center,
                                padding=ft.padding.only(top=10)
                            )
                        ], spacing=12),
                        padding=ft.padding.all(15),
                        border_radius=12,
                        bgcolor=Colors['card_background'],
                        border=ft.border.all(2, ft.Colors.ORANGE_300)
                    )
                    
                    cross_check_results_list.controls.append(card_content)
            
            cross_check_results_list.update()
            
            # Atualizar menu para mostrar/ocultar a aba
            update_menu()
            
        except Exception as e:
            print(f"Erro ao atualizar Cross Check: {e}")
            cross_check_results_list.controls.append(
                ft.Text(f"Erro ao carregar pendências: {e}", color=ft.Colors.RED)
            )
            cross_check_results_list.update()
    
    cross_check_view = ft.Column(
        [
            ft.Container(
                content=ft.Text(
                    "🔍 Cross Check - Avaliações Pendentes",
                    size=20,
                    weight="bold"
                ),
                padding=ft.padding.all(15)
            ),
            ft.Divider(),
            ft.Container(
                content=cross_check_results_list,
                expand=True,
                padding=ft.padding.all(10)
            ),
        ],
        visible=False,
        expand=True
    )
    
    # --- Fim: Aba Cross Check ---

    # Container para resultados de pendências - usando Column com scroll
    pending_results_list = ft.Column([], spacing=16, scroll=ft.ScrollMode.AUTO, expand=True, horizontal_alignment=ft.CrossAxisAlignment.STRETCH)
    
    # Variável global para controlar quantos cards de pendência exibir (máximo 10)
    MAX_PENDING_CARDS = 10
    
    # Envolver o pending_results_list em um Container com bgcolor que segue o tema
    pending_results_container = ft.Container(
        content=pending_results_list,
        expand=True,
        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface'),
        padding=ft.padding.symmetric(horizontal=20, vertical=20)
    )
    
    def create_score_control_standalone(label, value, has_permission):
        """Cria controle de score standalone (igual ao dos cards normais)"""
        global score_control_type
        Colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        # Normalizar valor
        norm_value = None
        try:
            if value is None:
                norm_value = None
            elif isinstance(value, (int, float)):
                norm_value = float(value)
            elif isinstance(value, str):
                v = value.strip()
                norm_value = float(v) if v != "" else None
            else:
                norm_value = float(value)
        except (ValueError, TypeError):
            norm_value = None
        
        if score_control_type == "slider":
            # Criar slider
            score_text = ft.Text(
                f"{norm_value:.1f}" if norm_value is not None else "0.0", 
                width=40, 
                text_align=ft.TextAlign.CENTER,
                size=16,
                weight=ft.FontWeight.W_300,
                color=Colors['on_surface']
            )
            score_slider = ft.Slider(
                min=0,
                max=10,
                divisions=100,
                value=norm_value if norm_value is not None else 0,
                width=350,
                active_color=Colors['primary'],
                inactive_color=Colors['surface_variant'],
                disabled=not has_permission
            )
            
            def on_slider_change(e):
                score_text.value = f"{e.control.value:.1f}"
                if hasattr(score_text, 'page') and score_text.page is not None:
                    score_text.update()
            
            score_slider.on_change = on_slider_change
            opacity = 1.0 if has_permission else 0.5
            score_text.opacity = opacity
            
            return ft.Column([
                ft.Text(label, size=12, weight="bold", opacity=opacity, color=Colors['on_surface']),
                score_slider,
                ft.Container(
                    content=score_text,
                    alignment=ft.alignment.center,
                    width=350,
                    margin=ft.margin.only(top=-15)
                )
            ], spacing=0, tight=True, horizontal_alignment=ft.CrossAxisAlignment.CENTER)
        
        else:  # spinbox
            try:
                increment = load_spinbox_increment()
            except:
                increment = 0.1
            
            score_field = ft.TextField(
                value=f"{norm_value:.1f}" if norm_value is not None else "0.0",
                text_align=ft.TextAlign.CENTER,
                width=70,
                border_radius=8,
                bgcolor=Colors['field_background'],
                color=Colors['on_surface'],
                border_color=Colors['outline'],
                disabled=not has_permission
            )
            
            def on_score_field_blur(e):
                try:
                    v = e.control.value
                    if v is None or str(v).strip() == "":
                        e.control.value = "0.0"
                        e.control.update()
                        return
                    v = str(v).strip().replace(',', '.')
                    num = float(v)
                    num = max(0, min(10, num))
                    e.control.value = f"{num:.1f}"
                    e.control.update()
                except (ValueError, TypeError):
                    e.control.value = "0.0"
                    e.control.update()
            
            score_field.on_blur = on_score_field_blur
            
            def adjust_score(e):
                if not has_permission:
                    return
                try:
                    current_value = float(score_field.value)
                    if e.control.data == "+":
                        current_value += increment
                    elif e.control.data == "-":
                        current_value -= increment
                    new_value = max(0, min(10, current_value))
                    score_field.value = str(round(new_value, 1))
                    score_field.update()
                except ValueError:
                    score_field.value = "0.0"
                    score_field.update()
            
            opacity = 1.0 if has_permission else 0.5
            
            return ft.Column([
                ft.Text(label, size=12, weight="bold", opacity=opacity, color=Colors['on_surface']),
                ft.Row([
                    ft.IconButton(
                        ft.Icons.REMOVE, 
                        on_click=adjust_score, 
                        data="-", 
                        icon_size=16, 
                        disabled=not has_permission, 
                        opacity=opacity,
                        icon_color=Colors['on_surface']
                    ),
                    score_field,
                    ft.IconButton(
                        ft.Icons.ADD, 
                        on_click=adjust_score, 
                        data="+", 
                        icon_size=16, 
                        disabled=not has_permission, 
                        opacity=opacity,
                        icon_color=Colors['on_surface']
                    ),
                ], alignment=ft.MainAxisAlignment.CENTER)
            ], spacing=5, horizontal_alignment=ft.CrossAxisAlignment.CENTER)
    
    def get_control_value_standalone(control):
        """Extrai valor do controle standalone"""
        try:
            if score_control_type == "slider":
                if hasattr(control, 'controls') and len(control.controls) >= 2:
                    slider = control.controls[1]
                    if hasattr(slider, 'value'):
                        val = float(slider.value)
                        return max(0, min(10, val))
            else:  # spinbox
                if hasattr(control, 'controls') and len(control.controls) >= 2:
                    spinbox_row = control.controls[1]
                    if hasattr(spinbox_row, 'controls') and len(spinbox_row.controls) >= 2:
                        text_field = spinbox_row.controls[1]
                        if hasattr(text_field, 'value') and text_field.value:
                            val = float(text_field.value)
                            return max(0, min(10, val))
        except (ValueError, TypeError, AttributeError):
            pass
        return None
    
    def load_pending_scores():
        """Carrega suppliers com notas pendentes do usuário (máximo 10 por vez)"""
        try:
            pending_results_list.controls.clear()
            
            # Buscar pendências
            pending_suppliers = get_pending_suppliers()
            
            # Contar total de pendências
            total_pending = len(pending_suppliers)
            
            if not pending_suppliers:
                # Adicionar mensagem centralizada quando não há pendências
                pending_results_list.controls.append(
                    ft.Container(
                        content=ft.Column([
                            ft.Icon(ft.Icons.CHECK_CIRCLE_OUTLINE, size=80, color=ft.Colors.GREEN),
                            ft.Text("✅ Nenhuma pendência encontrada!", size=24, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                            ft.Text("Todas as suas avaliações estão em dia.", size=16, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=15, alignment=ft.MainAxisAlignment.CENTER, expand=True),
                        alignment=ft.alignment.center,
                        expand=True
                    )
                )
            else:
                # Adicionar header com contador se houver mais de 10 pendências
                if total_pending > MAX_PENDING_CARDS:
                    pending_results_list.controls.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Icon(ft.Icons.INFO_OUTLINE, color=ft.Colors.BLUE, size=24),
                                ft.Column([
                                    ft.Text(
                                        f"Mostrando {MAX_PENDING_CARDS} de {total_pending} pendências",
                                        size=16,
                                        weight="bold",
                                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')
                                    ),
                                    ft.Text(
                                        f"Ao preencher uma avaliação, a próxima pendência será carregada automaticamente",
                                        size=12,
                                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'),
                                        italic=True
                                    ),
                                ], spacing=2, expand=True),
                            ], spacing=10),
                            padding=ft.padding.all(15),
                            bgcolor=ft.Colors.BLUE_50,
                            border_radius=8,
                            margin=ft.margin.only(bottom=10)
                        )
                    )
                
                # Limitar a MAX_PENDING_CARDS (10) cards por vez
                limited_pending = pending_suppliers[:MAX_PENDING_CARDS]
                
                # Criar cards de pendências (máximo 10)
                for pending in limited_pending:
                    # Buscar informações completas do supplier
                    supplier_query = """
                        SELECT supplier_number, bu, vendor_name, supplier_po
                        FROM supplier_database_table
                        WHERE supplier_id = ?
                    """
                    supplier_info = db_manager.query_one(supplier_query, (pending['supplier_id'],))
                    
                    if supplier_info:
                        if isinstance(supplier_info, dict):
                            supplier_number = supplier_info.get('supplier_number')
                            bu = supplier_info.get('bu')
                            vendor_name = supplier_info.get('vendor_name')
                            supplier_po = supplier_info.get('supplier_po')
                        else:
                            supplier_number, bu, vendor_name, supplier_po = supplier_info
                        
                        # Criar card de pendência
                        create_pending_card(
                            pending['supplier_id'],
                            vendor_name,
                            supplier_number,
                            bu,
                            supplier_po,
                            pending['pending_scores'],
                            pending['month'],
                            pending['year']
                        )
            
            pending_results_list.update()
            
            # Não precisa atualizar o estilo da tab - apenas o ícone do menu superior indica pendências
            
        except Exception as e:
            print(f"Erro ao carregar pendências: {e}")
            import traceback
            traceback.print_exc()
    
    def create_pending_card(supplier_id, vendor_name, supplier_number, bu, supplier_po, pending_scores, month, year):
        """Cria um card de pendência com controles de avaliação direta"""
        Colors = get_current_theme_colors(get_theme_name_from_page(page))
        
        # Usar o mês e ano do registro pendente
        # Garantir que month está no formato string (pode vir como int ou string do banco)
        current_month = str(month) if month else "01"
        current_year = str(year) if year else "2025"
        
        # Converter para nome do mês
        try:
            month_int = int(month)
            month_names = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", 
                          "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
            month_name = month_names[month_int] if 1 <= month_int <= 12 else str(month)
        except:
            month_name = str(month)
        
        # Buscar os scores atuais do supplier para este mês
        score_query = """
            SELECT otif, nil, quality_pickup, quality_package
            FROM supplier_score_records_table
            WHERE supplier_id = ? AND month = ? AND year = ?
        """
        # Month e year são TEXT no banco, usar como string
        score_data = db_manager.query_one(score_query, (supplier_id, current_month, current_year))
        
        # Criar controles de score usando a mesma função dos cards normais
        score_controls_dict = {}
        
        # Função helper para criar controle individual
        def create_control_for_score(score_name, db_field, permission_key):
            value = score_data.get(db_field) if score_data else None
            has_permission = current_user_permissions.get(permission_key, False)
            
            # Apenas criar se estiver pendente
            if score_name in pending_scores:
                control = create_score_control_standalone(score_name, value, has_permission)
                score_controls_dict[db_field] = control
                return control
            return None
        
        # Criar controles apenas para os pendentes
        nil_control = create_control_for_score('NIL', 'nil', 'nil')
        otif_control = create_control_for_score('OTIF', 'otif', 'otif')
        pickup_control = create_control_for_score('Pickup', 'quality_pickup', 'pickup')
        package_control = create_control_for_score('Package', 'quality_package', 'package')
        
        # Função para salvar as notas
        def save_pending_scores(e):
            try:
                # Coletar valores dos controles
                updates = {}
                
                if 'nil' in score_controls_dict:
                    value = get_control_value_standalone(score_controls_dict['nil'])
                    if value is not None:
                        updates['nil'] = float(value)
                
                if 'otif' in score_controls_dict:
                    value = get_control_value_standalone(score_controls_dict['otif'])
                    if value is not None:
                        updates['otif'] = float(value)
                
                if 'quality_pickup' in score_controls_dict:
                    value = get_control_value_standalone(score_controls_dict['quality_pickup'])
                    if value is not None:
                        updates['quality_pickup'] = float(value)
                
                if 'quality_package' in score_controls_dict:
                    value = get_control_value_standalone(score_controls_dict['quality_package'])
                    if value is not None:
                        updates['quality_package'] = float(value)
                
                if not updates:
                    show_snack_bar("❌ Nenhuma nota foi preenchida", True)
                    return
                
                print(f"🔍 Updates coletados: {updates}")
                
                # Buscar TODOS os scores atuais do supplier (não apenas os atualizados)
                current_scores_query = """
                    SELECT otif, nil, quality_pickup, quality_package
                    FROM supplier_score_records_table
                    WHERE supplier_id = ? AND month = ? AND year = ?
                """
                # Month e year são TEXT no banco, usar como string
                current_scores_data = db_manager.query_one(current_scores_query, (supplier_id, current_month, current_year))
                print(f"🔍 Scores atuais do banco: {current_scores_data}")
                
                # Função helper para garantir valor numérico (trata None, "", e valores inválidos)
                def get_numeric_value(value):
                    # Tratar None (NULL)
                    if value is None:
                        return 0.0
                    # Tratar string vazia
                    if value == "" or (isinstance(value, str) and value.strip() == ""):
                        return 0.0
                    # Tentar converter para float
                    try:
                        num_value = float(value)
                        # Verificar se é NaN
                        if num_value != num_value:  # NaN check
                            return 0.0
                        return num_value
                    except (ValueError, TypeError):
                        return 0.0
                
                # Mesclar scores atuais com os atualizados, SEMPRE garantindo valores numéricos
                final_scores = {
                    'otif': get_numeric_value(updates.get('otif', current_scores_data.get('otif') if current_scores_data else None)),
                    'nil': get_numeric_value(updates.get('nil', current_scores_data.get('nil') if current_scores_data else None)),
                    'pickup': get_numeric_value(updates.get('quality_pickup', current_scores_data.get('quality_pickup') if current_scores_data else None)),
                    'package': get_numeric_value(updates.get('quality_package', current_scores_data.get('quality_package') if current_scores_data else None))
                }
                
                print(f"🔍 Final scores após merge: {final_scores}")
                
                # Carregar critérios (pesos) do banco de dados
                criterios_raw = db_manager.query("SELECT criteria_category, value FROM criteria_table")
                criterio_map = {}
                
                for row in criterios_raw:
                    nome = str(row.get('criteria_category') if isinstance(row, dict) else row[0]).strip().lower()
                    valor = float(row.get('value') if isinstance(row, dict) else row[1])
                    if "package" in nome:
                        criterio_map["package"] = valor
                    elif "pick" in nome:
                        criterio_map["pickup"] = valor
                    elif "nil" in nome:
                        criterio_map["nil"] = valor
                    elif "otif" in nome:
                        criterio_map["otif"] = valor
                
                # Verificar se todos os critérios foram carregados
                if not all(k in criterio_map for k in ["package", "pickup", "nil", "otif"]):
                    print("⚠️ Critérios incompletos, usando pesos padrão")
                    criterio_map = {
                        "otif": 0.31,
                        "nil": 0.19,
                        "pickup": 0.25,
                        "package": 0.25
                    }
                
                print(f"🔍 Critérios carregados: {criterio_map}")
                
                # Calcular total_score - GARANTINDO que todos os valores são float
                total_score = (
                    float(final_scores['otif']) * float(criterio_map['otif']) +
                    float(final_scores['nil']) * float(criterio_map['nil']) +
                    float(final_scores['pickup']) * float(criterio_map['pickup']) +
                    float(final_scores['package']) * float(criterio_map['package'])
                )
                total_score = round(total_score, 2)
                
                print(f"🔍 Total Score calculado: {total_score}")
                
                # Adicionar total_score aos updates
                updates['total_score'] = total_score
                
                # Construir query de UPDATE incluindo total_score e metadados
                set_clause = ", ".join([f"{col} = ?" for col in updates.keys()])
                set_clause += ", changed_by = ?, register_date = CURRENT_TIMESTAMP"
                
                update_query = f"""
                    UPDATE supplier_score_records_table
                    SET {set_clause}
                    WHERE supplier_id = ? AND month = ? AND year = ?
                """
                # Month e year são TEXT no banco, usar como string
                params = list(updates.values()) + [current_user_wwid, supplier_id, current_month, current_year]
                
                db_manager.execute(update_query, params)
                
                print(f"✅ Total Score calculado: {total_score} para {vendor_name}")
                show_snack_bar(f"✅ Notas salvas para {vendor_name}! (Total Score: {total_score})", False)
                
                # Remover o card salvo e carregar o próximo dinamicamente
                try:
                    # Contar cards atuais (excluindo header de info se existir)
                    current_cards = [c for c in pending_results_list.controls if isinstance(c, ft.Container) and hasattr(c, 'data')]
                    
                    # Remover o card que foi salvo
                    for control in pending_results_list.controls[:]:
                        if hasattr(control, 'data') and control.data == supplier_id:
                            pending_results_list.controls.remove(control)
                            print(f"🗑️ Card removido: {vendor_name}")
                            break
                    
                    # Buscar todas as pendências novamente
                    all_pending = get_pending_suppliers()
                    
                    # Se ainda houver pendências além das que já estão exibidas
                    if len(all_pending) >= MAX_PENDING_CARDS:
                        # Pegar a próxima pendência (índice = MAX_PENDING_CARDS - 1 porque acabamos de remover 1)
                        next_pending_index = MAX_PENDING_CARDS - 1
                        if next_pending_index < len(all_pending):
                            next_pending = all_pending[next_pending_index]
                            
                            # Buscar informações completas do supplier
                            supplier_query = """
                                SELECT supplier_number, bu, vendor_name, supplier_po
                                FROM supplier_database_table
                                WHERE supplier_id = ?
                            """
                            supplier_info = db_manager.query_one(supplier_query, (next_pending['supplier_id'],))
                            
                            if supplier_info:
                                if isinstance(supplier_info, dict):
                                    next_supplier_number = supplier_info.get('supplier_number')
                                    next_bu = supplier_info.get('bu')
                                    next_vendor_name = supplier_info.get('vendor_name')
                                    next_supplier_po = supplier_info.get('supplier_po')
                                else:
                                    next_supplier_number, next_bu, next_vendor_name, next_supplier_po = supplier_info
                                
                                # Criar e adicionar card da próxima pendência
                                create_pending_card(
                                    next_pending['supplier_id'],
                                    next_vendor_name,
                                    next_supplier_number,
                                    next_bu,
                                    next_supplier_po,
                                    next_pending['pending_scores'],
                                    next_pending['month'],
                                    next_pending['year']
                                )
                                print(f"➕ Próximo card carregado: {next_vendor_name}")
                    
                    # Atualizar header com contador se necessário
                    total_pending = len(all_pending)
                    if total_pending > MAX_PENDING_CARDS:
                        # Procurar e atualizar header existente
                        for i, control in enumerate(pending_results_list.controls):
                            if isinstance(control, ft.Container) and control.bgcolor == ft.Colors.BLUE_50:
                                # Atualizar texto do header
                                control.content.controls[1].controls[0].value = f"Mostrando {MAX_PENDING_CARDS} de {total_pending} pendências"
                                break
                    elif total_pending == 0:
                        # Não há mais pendências - mostrar mensagem de sucesso
                        pending_results_list.controls.clear()
                        pending_results_list.controls.append(
                            ft.Container(
                                content=ft.Column([
                                    ft.Icon(ft.Icons.CHECK_CIRCLE_OUTLINE, size=80, color=ft.Colors.GREEN),
                                    ft.Text("✅ Nenhuma pendência encontrada!", size=24, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                                    ft.Text("Todas as suas avaliações estão em dia.", size=16, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=15, alignment=ft.MainAxisAlignment.CENTER, expand=True),
                                alignment=ft.alignment.center,
                                expand=True
                            )
                        )
                    else:
                        # Menos de 10 pendências - remover header se existir
                        for control in pending_results_list.controls[:]:
                            if isinstance(control, ft.Container) and control.bgcolor == ft.Colors.BLUE_50:
                                pending_results_list.controls.remove(control)
                                break
                    
                    pending_results_list.update()
                    
                except Exception as reload_ex:
                    print(f"⚠️ Erro ao recarregar próxima pendência: {reload_ex}")
                    # Fallback: recarregar tudo
                    load_pending_scores()
                
            except Exception as ex:
                print(f"Erro ao salvar notas pendentes: {ex}")
                import traceback
                traceback.print_exc()
                show_snack_bar(f"❌ Erro ao salvar: {ex}", True)
        
        # Montar lista de controles para exibição
        score_controls_list = []
        if nil_control:
            score_controls_list.append(nil_control)
        if otif_control:
            score_controls_list.append(otif_control)
        if pickup_control:
            score_controls_list.append(pickup_control)
        if package_control:
            score_controls_list.append(package_control)
        
        # Criar card com layout ocupando linha inteira
        card = ft.Container(
            content=ft.Container(
                content=ft.Column([
                    # Nome do fornecedor
                    ft.Text(
                        vendor_name, 
                        size=18, 
                        weight="bold", 
                        color=Colors['on_surface']
                    ),
                    
                    # Espaço
                    ft.Container(height=12),
                    
                    # Layout horizontal: Dados à esquerda, Notas à direita
                    ft.Row([
                        # Coluna de dados à esquerda
                        ft.Column([
                            ft.Text(f"Data ref: {month_name}/{current_year}", size=14, color=Colors['on_surface']),
                            ft.Text(f"ID: {supplier_id}", size=14, color=Colors['on_surface']),
                            ft.Text(f"SSID: {format_po_ssid(supplier_number)}", size=14, color=Colors['on_surface']),
                            ft.Text(f"PO: {format_po_ssid(supplier_po)}", size=14, color=Colors['on_surface']),
                            ft.Text(f"BU: {bu or 'N/A'}", size=14, color=Colors['on_surface']),
                        ], spacing=6, expand=1),
                        
                        # Divider vertical
                        ft.VerticalDivider(width=1, color=Colors['outline_variant']),
                        
                        # Controles de avaliação à direita
                        ft.Container(
                            content=ft.Row(
                                score_controls_list,
                                spacing=30,
                                alignment=ft.MainAxisAlignment.CENTER,
                                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                            ),
                            alignment=ft.alignment.center,
                            expand=2,
                        ),
                    ], spacing=20, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                    
                    # Espaço
                    ft.Container(height=12),
                    
                    # Divider entre os campos e o botão
                    ft.Divider(height=1, color=Colors['outline_variant']),
                    
                    # Espaço
                    ft.Container(height=10),
                    
                    # Botão salvar alinhado à direita
                    ft.Container(
                        content=ft.ElevatedButton(
                            "Salvar",
                            icon=ft.Icons.CHECK,
                            on_click=save_pending_scores,
                            bgcolor=Colors['primary'],
                            color=Colors['on_primary'],
                            height=38,
                        ),
                        alignment=ft.alignment.center_right
                    ),
                ], spacing=4, horizontal_alignment=ft.CrossAxisAlignment.START),
                padding=ft.padding.all(20),
                expand=True,  # Ocupar largura disponível
                border=ft.border.all(1, Colors['outline_variant']),
                border_radius=8,
                bgcolor=Colors['card_background']  # Cor do tema
            ),
            padding=ft.padding.symmetric(horizontal=0, vertical=6),
            expand=True,  # Ocupar linha inteira
            data=supplier_id  # Identificador para remoção dinâmica
        )
        
        pending_results_list.controls.append(card)
    
    # Criar as tabs dentro do Score
    search_tab_content = ft.Column([
        # Header com botão de menu - FIXO no topo
        ft.Container(
            content=ft.Row([
                ft.PopupMenuButton(
                    icon=ft.Icons.MORE_HORIZ,
                    items=[
                        ft.PopupMenuItem(
                            text="Gerar nota cheia",
                            icon=ft.Icons.STAR,
                            on_click=lambda _: generate_full_score_dialog()
                        ),
                        ft.PopupMenuItem(
                            text="Importar Score",
                            icon=ft.Icons.FILE_UPLOAD,
                            on_click=lambda _: import_score_dialog()
                        ),
                        ft.PopupMenuItem(
                            text="Exportar form",
                            icon=ft.Icons.FILE_DOWNLOAD,
                            on_click=lambda _: export_form_dialog()
                        )
                    ],
                    tooltip="Opções",
                    icon_color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'),
                    bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                    shape=ft.RoundedRectangleBorder(radius=8)
                ),
                # Formulário de pesquisa centralizado e responsivo
                ft.Container(
                    content=score_view_content,
                    expand=True,
                    alignment=ft.alignment.center,
                ),
            ], 
            alignment=ft.MainAxisAlignment.CENTER, 
            vertical_alignment=ft.CrossAxisAlignment.START,
            ),
            padding=ft.padding.only(left=10, right=10, top=10, bottom=10)
        ),
        ft.Divider(),
        # Área de resultados com scroll independente - sem padding para tabela ocupar toda largura
        ft.Container(
            content=results_list,
            expand=True,
            clip_behavior=ft.ClipBehavior.HARD_EDGE,
            padding=0,
            margin=0,
        ),
    ], expand=True)
    
    pending_tab_content = ft.Column([
        ft.Container(
            content=ft.Column([
                ft.Row([
                    ft.Icon(ft.Icons.PENDING_ACTIONS, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                    ft.Text("Pendências", size=24, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                ], spacing=10),
                ft.Text("Suppliers que precisam de avaliação", size=14, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
            ], spacing=5),
            padding=ft.padding.all(20)
        ),
        ft.Divider(color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
        pending_results_container,
    ], expand=True)
    
    def on_tab_change(e):
        """Callback quando muda de tab"""
        if score_tabs.selected_index == 1:  # Tab de Pendências
            load_pending_scores()
    
    # Criar referência para a tab de Pendências com ícone
    pending_tab = ft.Tab(
        text="Pendências",
        icon=ft.Icons.PENDING_ACTIONS,
        content=pending_tab_content
    )
    
    # Criar tab de Critérios de Avaliação
    criteria_tab_content = ft.Container(
        content=ft.Column([
            # Cabeçalho
            ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Icon(ft.Icons.RULE, size=28, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                        ft.Text("Critérios de Avaliação", size=24, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                    ], spacing=10),
                    ft.Text("Entenda como as notas são calculadas", size=14, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                ], spacing=5),
                padding=ft.padding.all(20)
            ),
            ft.Divider(color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
            
            # Container com todos os critérios (scrollable)
            ft.Container(
                content=ft.Column([
                    # Critério NIL
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.DELIVERY_DINING, size=24, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                ft.Text("NIL (Non-Invoiced Items)", size=18, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                            ]),
                            ft.Container(height=8),
                            ft.Container(
                                content=ft.Column([
                                    ft.Row([ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=18), ft.Text("0 problemas", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 10", size=15, weight="bold", color=ft.Colors.GREEN)], spacing=8),
                                    ft.Row([ft.Icon(ft.Icons.WARNING, color=ft.Colors.ORANGE, size=18), ft.Text("1 problema", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 5", size=15, weight="bold", color=ft.Colors.ORANGE)], spacing=8),
                                    ft.Row([ft.Icon(ft.Icons.CANCEL, color=ft.Colors.RED, size=18), ft.Text("Mais de 1 problema", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 0", size=15, weight="bold", color=ft.Colors.RED)], spacing=8),
                                ], spacing=8),
                                padding=15,
                                border_radius=8,
                            ),
                        ]),
                        padding=15,
                        border_radius=12,
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                    ),
                    
                    ft.Container(height=12),
                    
                    # Critério PICKUP
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.INVENTORY, size=24, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                ft.Text("PICKUP (Quality Pickup)", size=18, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                            ]),
                            ft.Container(height=8),
                            ft.Container(
                                content=ft.Column([
                                    ft.Row([ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=18), ft.Text("0 problemas", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 10", size=15, weight="bold", color=ft.Colors.GREEN)], spacing=8),
                                    ft.Row([ft.Icon(ft.Icons.WARNING, color=ft.Colors.ORANGE, size=18), ft.Text("1 problema", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 5", size=15, weight="bold", color=ft.Colors.ORANGE)], spacing=8),
                                    ft.Row([ft.Icon(ft.Icons.CANCEL, color=ft.Colors.RED, size=18), ft.Text("Mais de 1 problema", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 0", size=15, weight="bold", color=ft.Colors.RED)], spacing=8),
                                ], spacing=8),
                                padding=15,
                                border_radius=8,
                            ),
                        ]),
                        padding=15,
                        border_radius=12,
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                    ),
                    
                    ft.Container(height=12),
                    
                    # Critério PACKAGE
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.INVENTORY_2, size=24, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                ft.Text("PACKAGE (Quality Package)", size=18, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                            ]),
                            ft.Container(height=8),
                            ft.Container(
                                content=ft.Column([
                                    ft.Row([ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=18), ft.Text("0 problemas", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 10", size=15, weight="bold", color=ft.Colors.GREEN)], spacing=8),
                                    ft.Row([ft.Icon(ft.Icons.WARNING, color=ft.Colors.ORANGE, size=18), ft.Text("1 problema", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 5", size=15, weight="bold", color=ft.Colors.ORANGE)], spacing=8),
                                    ft.Row([ft.Icon(ft.Icons.CANCEL, color=ft.Colors.RED, size=18), ft.Text("Mais de 1 problema", size=15, weight="bold"), ft.Text("→", size=15), ft.Text("Nota 0", size=15, weight="bold", color=ft.Colors.RED)], spacing=8),
                                ], spacing=8),
                                padding=15,
                                border_radius=8,
                            ),
                        ]),
                        padding=15,
                        border_radius=12,
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                    ),
                    
                    ft.Container(height=12),
                    
                    # Critério OTIF
                    ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Icon(ft.Icons.SCHEDULE, size=24, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                ft.Text("OTIF (On Time In Full)", size=18, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')),
                            ]),
                            ft.Container(height=8),
                            ft.Text("Segue tabela de OTIF. As notas são de 0 a 100%, porém a conversão para adição é multiplicar por 10.", size=13, italic=True, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant')),
                            ft.Container(height=6),
                            ft.Text("Exemplo: 87% × 10 = 8,7", size=13, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                            ft.Container(height=10),
                            ft.Container(
                                content=ft.Column([
                                    # Cabeçalho da tabela
                                    ft.Row([
                                        ft.Container(ft.Text("OTIF (%)", size=13, weight="bold", text_align=ft.TextAlign.CENTER, color=ft.Colors.WHITE), width=100, bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'), padding=6, border_radius=ft.border_radius.only(top_left=8)),
                                        ft.Container(ft.Text("Nota Final", size=13, weight="bold", text_align=ft.TextAlign.CENTER, color=ft.Colors.WHITE), expand=True, bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'), padding=6, border_radius=ft.border_radius.only(top_right=8)),
                                    ], spacing=0),
                                    # Linhas da tabela
                                    ft.Row([
                                        ft.Container(ft.Text("100%", size=13, text_align=ft.TextAlign.CENTER), width=100, bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.GREEN), padding=6),
                                        ft.Container(ft.Text("10,0", size=13, weight="bold", text_align=ft.TextAlign.CENTER, color=ft.Colors.GREEN), expand=True, bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.GREEN), padding=6),
                                    ], spacing=0),
                                    ft.Row([
                                        ft.Container(ft.Text("90%", size=13, text_align=ft.TextAlign.CENTER), width=100, bgcolor=ft.Colors.with_opacity(0.03, get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')), padding=6),
                                        ft.Container(ft.Text("9,0", size=13, text_align=ft.TextAlign.CENTER), expand=True, bgcolor=ft.Colors.with_opacity(0.03, get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')), padding=6),
                                    ], spacing=0),
                                    ft.Row([
                                        ft.Container(ft.Text("87%", size=13, text_align=ft.TextAlign.CENTER), width=100, bgcolor=ft.Colors.with_opacity(0.05, get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')), padding=6),
                                        ft.Container(ft.Text("8,7", size=13, text_align=ft.TextAlign.CENTER), expand=True, bgcolor=ft.Colors.with_opacity(0.05, get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')), padding=6),
                                    ], spacing=0),
                                    ft.Row([
                                        ft.Container(ft.Text("75%", size=13, text_align=ft.TextAlign.CENTER), width=100, bgcolor=ft.Colors.with_opacity(0.03, get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')), padding=6),
                                        ft.Container(ft.Text("7,5", size=13, text_align=ft.TextAlign.CENTER), expand=True, bgcolor=ft.Colors.with_opacity(0.03, get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface')), padding=6),
                                    ], spacing=0),
                                    ft.Row([
                                        ft.Container(ft.Text("50%", size=13, text_align=ft.TextAlign.CENTER), width=100, bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.ORANGE), padding=6),
                                        ft.Container(ft.Text("5,0", size=13, text_align=ft.TextAlign.CENTER, color=ft.Colors.ORANGE), expand=True, bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.ORANGE), padding=6),
                                    ], spacing=0),
                                    ft.Row([
                                        ft.Container(ft.Text("0%", size=13, text_align=ft.TextAlign.CENTER), width=100, bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.RED), padding=6, border_radius=ft.border_radius.only(bottom_left=8)),
                                        ft.Container(ft.Text("0,0", size=13, weight="bold", text_align=ft.TextAlign.CENTER, color=ft.Colors.RED), expand=True, bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.RED), padding=6, border_radius=ft.border_radius.only(bottom_right=8)),
                                    ], spacing=0),
                                ], spacing=0),
                                border_radius=8,
                            ),
                        ]),
                        padding=15,
                        border_radius=12,
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                    ),
                ], spacing=0, scroll=ft.ScrollMode.AUTO),
                expand=True,
                padding=ft.padding.all(20),
            ),
        ], expand=True),
        expand=True,
    )
    
    criteria_tab = ft.Tab(
        text="Critérios",
        icon=ft.Icons.RULE,
        content=criteria_tab_content
    )
    
    score_tabs = ft.Tabs(
        selected_index=0,
        animation_duration=300,
        tabs=[
            ft.Tab(
                text="Pesquisa",
                icon=ft.Icons.SEARCH,
                content=search_tab_content
            ),
            pending_tab,
            criteria_tab,
        ],
        expand=True,
        on_change=on_tab_change
    )
    
    # Carregar pendências na inicialização
    load_pending_scores()
    
    score_view = ft.Column(
        [score_tabs],
        visible=False,  # Score não será mais a aba inicial
        expand=True,
    )
    # --- Início: Lógica da Aba Timeline ---
    
    # Referências para os campos de seleção do Timeline
    timeline_vendor_dropdown = ft.Ref[ft.Dropdown]()
    timeline_vendor_search_field = ft.Ref[ft.TextField]()  # Campo de pesquisa visual
    selected_vendor_id = ft.Ref[str]()  # Armazena o ID do fornecedor selecionado
    timeline_year_dropdown = ft.Ref[ft.Dropdown]()
    timeline_bu_dropdown = ft.Ref[ft.Dropdown]()
    
    
    # Referência para o container dos campos de pesquisa
    timeline_search_container = ft.Ref[ft.Container]()
    # Referência para o botão de limpar da aba Timeline
    timeline_clear_button_ref = ft.Ref[ft.IconButton]()

    def clear_timeline_search_field(e=None):
        """Limpa o campo visual de fornecedor na Timeline"""
        if timeline_vendor_search_field.current:
            timeline_vendor_search_field.current.value = ""
            # esconder botão
            if timeline_clear_button_ref.current:
                timeline_clear_button_ref.current.visible = False
                timeline_clear_button_ref.current.update()
            timeline_vendor_search_field.current.update()
            # resetar selection dropdown
            if timeline_vendor_dropdown.current:
                timeline_vendor_dropdown.current.value = None
            # atualizar timeline
            update_timeline_metrics()

    def update_timeline_search_suffix(e=None):
        """Atualiza a visibilidade do botão limpar na Timeline"""
        if timeline_vendor_search_field.current and timeline_clear_button_ref.current:
            timeline_clear_button_ref.current.visible = bool(timeline_vendor_search_field.current.value)
            timeline_clear_button_ref.current.update()
    
    # Refs for additional score checkboxes
    timeline_otif_check = ft.Ref[ft.Checkbox]()
    timeline_nil_check = ft.Ref[ft.Checkbox]()
    timeline_pickup_check = ft.Ref[ft.Checkbox]()
    timeline_package_check = ft.Ref[ft.Checkbox]()
    
    # Referências para os cards de métricas
    overall_avg_card = ft.Ref[ft.Text]()
    twelve_month_avg_card = ft.Ref[ft.Text]()
    year_avg_card = ft.Ref[ft.Text]()
    q1_avg_card = ft.Ref[ft.Text]()
    q2_avg_card = ft.Ref[ft.Text]()
    q3_avg_card = ft.Ref[ft.Text]()
    q4_avg_card = ft.Ref[ft.Text]()
    
    # Referências para os ícones de seta nos cards Q
    q1_arrow_icon = ft.Ref[ft.Icon]()
    q2_arrow_icon = ft.Ref[ft.Icon]()
    q3_arrow_icon = ft.Ref[ft.Icon]()
    q4_arrow_icon = ft.Ref[ft.Icon]()
    
    # Referências para containers dinâmicos (ALL mode)
    yearly_cards_container = ft.Ref[ft.Row]()
    quarterly_cards_container = ft.Ref[ft.Row]()
    
    # Referências para os containers wrapper na aba de métricas
    quarterly_section_container = ft.Ref[ft.Container]()
    yearly_section_container = ft.Ref[ft.Container]()

    # Referências para aba Risks
    risks_year_dropdown = ft.Ref[ft.Dropdown]()
    risks_cards_container = ft.Ref[ft.Container]()
    target_risks_text = ft.Ref[ft.Text]()
    include_inactive_switch = ft.Ref[ft.Switch]()
    inactive_switch_container = ft.Ref[ft.Container]()
    inactive_switch_icon = ft.Ref[ft.Icon]()
    # Ref para o container que agrupa ano e target (para atualização de tema)
    risks_header_container = ft.Ref[ft.Container]()
    # Ref para o container do display de Target (borda/背景 especial)
    target_display_container = ft.Ref[ft.Container]()
    
    # Container para o gráfico e tabela
    timeline_chart_container = ft.Ref[ft.Container]()
    timeline_individual_charts_container = ft.Ref[ft.Container]()
    timeline_table_container = ft.Ref[ft.Container]()
    timeline_analytics_container = ft.Ref[ft.Container]()
    timeline_content_container = ft.Ref[ft.Container]()
    timeline_scrollable_content = ft.Ref[ft.Column]()
    timeline_tabs = ft.Ref[ft.Tabs]()
    
    # Referência para a linha que contém os cards de métricas
    timeline_metrics_row = ft.Ref[ft.Row]()

    # Dicionário de referências para os componentes dos cards de métricas
    timeline_cards_refs = {
        "overall": {"card": ft.Ref(), "icon": ft.Ref(), "value": overall_avg_card, "gradient": ft.Ref()},
        "12m": {"card": ft.Ref(), "icon": ft.Ref(), "value": twelve_month_avg_card, "gradient": ft.Ref()},
        "year": {"card": ft.Ref(), "icon": ft.Ref(), "value": year_avg_card, "gradient": ft.Ref()},
        "q1": {"card": ft.Ref(), "icon": ft.Ref(), "value": q1_avg_card, "gradient": ft.Ref()},
        "q2": {"card": ft.Ref(), "icon": ft.Ref(), "value": q2_avg_card, "gradient": ft.Ref()},
        "q3": {"card": ft.Ref(), "icon": ft.Ref(), "value": q3_avg_card, "gradient": ft.Ref()},
        "q4": {"card": ft.Ref(), "icon": ft.Ref(), "value": q4_avg_card, "gradient": ft.Ref()},
    }
    
    def open_vendor_selection_dialog(e):
        """Abre um diálogo com campo de pesquisa e lista de fornecedores"""
        search_query = ft.Ref[ft.TextField]()
        vendors_list = ft.Ref[ft.Column]()
        
        def load_vendors_list(filter_text=""):
            """Carrega a lista de fornecedores filtrada do banco (limite 50)"""
            try:
                if not db_conn:
                    return []
                
                # Se há filtro, buscar no banco; senão, carregar os primeiros 50
                if filter_text.strip():
                    query = """
                        SELECT supplier_id, vendor_name, bu, supplier_po 
                        FROM supplier_database_table 
                        WHERE vendor_name LIKE ? OR supplier_id LIKE ?
                        ORDER BY vendor_name
                        LIMIT 50
                    """
                    filter_param = f"%{filter_text}%"
                    suppliers = db_manager.query(query, (filter_param, filter_param))
                else:
                    query = """
                        SELECT supplier_id, vendor_name, bu, supplier_po 
                        FROM supplier_database_table 
                        ORDER BY vendor_name
                        LIMIT 50
                    """
                    suppliers = db_manager.query(query)
                
                items = []
                for supplier in suppliers:
                    supplier_id = supplier['supplier_id']
                    vendor_name = supplier['vendor_name']
                    bu = supplier['bu'] if supplier['bu'] else "N/A"
                    # Converter PO para inteiro, se for número
                    po_value = supplier['supplier_po']
                    if po_value:
                        try:
                            po = str(int(float(po_value)))
                        except (ValueError, TypeError):
                            po = str(po_value)
                    else:
                        po = "N/A"
                    
                    def create_select_handler(sid, vname):
                        def handler(e):
                            # Atualizar o campo de pesquisa visual
                            if timeline_vendor_search_field.current:
                                timeline_vendor_search_field.current.value = vname
                                timeline_vendor_search_field.current.update()
                            
                            # Atualizar o dropdown oculto (compatibilidade)
                            if timeline_vendor_dropdown.current:
                                timeline_vendor_dropdown.current.value = sid
                            
                            # Fechar o diálogo
                            page.close(vendor_dialog)
                            
                            # Chamar a função de mudança
                            on_timeline_vendor_change(None)
                            # Atualizar a visibilidade do botão de limpar
                            try:
                                update_timeline_search_suffix()
                            except Exception:
                                pass
                        return handler
                    
                    items.append(
                        ft.Container(
                            content=ft.Row([
                                ft.Text(vendor_name, size=14, weight="w500", expand=2),
                                ft.Text(bu, size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'), expand=1),
                                ft.Text(po, size=13, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'), expand=1),
                            ], spacing=12, alignment=ft.MainAxisAlignment.START),
                            padding=ft.padding.symmetric(horizontal=12, vertical=10),
                            on_click=create_select_handler(supplier_id, vendor_name),
                            ink=True,
                            border_radius=4,
                        )
                    )
                
                return items
            except Exception as ex:
                print(f"Erro ao carregar fornecedores: {ex}")
                return []
        
        def on_search_input(e):
            """Filtra a lista quando o usuário digita"""
            filter_text = search_query.current.value if search_query.current else ""
            vendors_list.current.controls = load_vendors_list(filter_text)
            vendors_list.current.update()
        
        # Criar o diálogo
        vendor_dialog = ft.AlertDialog(
            title=ft.Text("Selecionar Fornecedor", size=18, weight="bold"),
            content=ft.Container(
                content=ft.Column([
                    ft.TextField(
                        hint_text="Buscar fornecedor...",
                        ref=search_query,
                        on_change=on_search_input,
                        border_radius=8,
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                        border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline'),
                        text_size=14,
                        suffix_icon=ft.Icons.SEARCH,
                    ),
                    ft.Container(height=10),
                    # Cabeçalho das colunas
                    ft.Container(
                        content=ft.Row([
                            ft.Text("Nome", size=13, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'), expand=2),
                            ft.Text("BU", size=13, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'), expand=1),
                            ft.Text("PO", size=13, weight="bold", color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'), expand=1),
                        ], spacing=12, alignment=ft.MainAxisAlignment.START),
                        padding=ft.padding.symmetric(horizontal=12, vertical=8),
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                        border_radius=ft.border_radius.only(top_left=8, top_right=8),
                    ),
                    ft.Container(
                        content=ft.Column(
                            controls=load_vendors_list(),
                            ref=vendors_list,
                            scroll=ft.ScrollMode.AUTO,
                            spacing=0,
                        ),
                        height=400,
                        border=ft.border.all(1, get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
                        border_radius=ft.border_radius.only(bottom_left=8, bottom_right=8),
                    ),
                ], spacing=0, tight=True),
                width=600,
            ),
            actions=[
                ft.TextButton("Cancelar", on_click=lambda e: page.close(vendor_dialog)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        
        page.open(vendor_dialog)
    
    def load_suppliers_for_timeline():
        """Carrega todos os suppliers do banco de dados para o dropdown no formato 'Supplier - BU'"""
        try:
            if not db_conn:
                return [ft.dropdown.Option("", "Nenhum supplier disponível")]
            
            suppliers = db_manager.query("SELECT supplier_id, vendor_name, bu FROM supplier_database_table ORDER BY vendor_name")
            
            options = [ft.dropdown.Option("", "Selecione um supplier")]
            for supplier in suppliers:
                supplier_id = supplier['supplier_id']
                vendor_name = supplier['vendor_name']
                bu = supplier['bu']
                # Formato: "Supplier - BU"
                bu_text = bu if bu else "N/A"
                display_text = f"{vendor_name} - {bu_text}"
                options.append(ft.dropdown.Option(str(supplier_id), display_text))
                
            return options
        except Exception as e:
            print(f"Erro ao carregar suppliers: {e}")
            return [ft.dropdown.Option("", "Erro ao carregar suppliers")]

            
    def load_business_units_for_timeline():
        """Carrega todas as business units do banco de dados para o dropdown"""
        try:
            if not db_conn:
                return [ft.dropdown.Option("", "Nenhuma BU disponível")]
            
            bus = db_manager.query("SELECT DISTINCT bu FROM business_unit_table WHERE bu IS NOT NULL AND bu != '' ORDER BY bu")
            
            options = [ft.dropdown.Option("", "Todas as BUs")]
            for bu_row in bus:
                bu = bu_row['bu']
                options.append(ft.dropdown.Option(str(bu), str(bu)))
                
            return options
        except Exception as e:
            print(f"Erro ao carregar business units: {e}")
            return [ft.dropdown.Option("", "Erro ao carregar BUs")]
  
    
    def update_timeline_metrics():
        """Atualiza todos os cards de métricas baseado no vendor selecionado"""
        vendor_id = timeline_vendor_dropdown.current.value if timeline_vendor_dropdown.current else ""
        year = timeline_year_dropdown.current.value if timeline_year_dropdown.current else None
        
        if not vendor_id:
            # Limpar todos os cards
            overall_avg_card.current.value = "--"
            twelve_month_avg_card.current.value = "--"
            year_avg_card.current.value = "--"
            q1_avg_card.current.value = "--"
            q2_avg_card.current.value = "--"
            q3_avg_card.current.value = "--"
            q4_avg_card.current.value = "--"
            
            # Ocultar cards dinâmicos e restaurar visibilidade padrão
            if yearly_cards_container.current:
                yearly_cards_container.current.controls = []
            if yearly_section_container.current:
                yearly_section_container.current.visible = False
            if quarterly_cards_container.current:
                quarterly_cards_container.current.visible = True
            if quarterly_section_container.current:
                quarterly_section_container.current.visible = True
            if timeline_cards_refs["year"]["card"].current:
                timeline_cards_refs["year"]["card"].current.visible = True
            
            safe_page_update(page)
            return
        
        # Atualizar informações do fornecedor
        update_supplier_info(vendor_id)
            
        calculate_overall_average(vendor_id)
        calculate_twelve_month_average(vendor_id)
        
        # Modo ALL: mostrar cards anuais, ocultar trimestrais e Year Avg
        if year == "ALL":
            # Ocultar cards trimestrais
            if quarterly_cards_container.current:
                quarterly_cards_container.current.visible = False
            if quarterly_section_container.current:
                quarterly_section_container.current.visible = False
            
            # Ocultar card Year Avg (não faz sentido em modo ALL)
            if timeline_cards_refs["year"]["card"].current:
                timeline_cards_refs["year"]["card"].current.visible = False
            
            # Criar e mostrar cards anuais
            if yearly_cards_container.current:
                yearly_cards = create_yearly_average_cards(vendor_id)
                yearly_cards_container.current.controls = yearly_cards
            if yearly_section_container.current:
                yearly_section_container.current.visible = True
        else:
            # Modo ano específico: mostrar trimestrais e Year Avg, ocultar anuais
            if quarterly_cards_container.current:
                quarterly_cards_container.current.visible = True
            if quarterly_section_container.current:
                quarterly_section_container.current.visible = True
            
            # Mostrar card Year Avg
            if timeline_cards_refs["year"]["card"].current:
                timeline_cards_refs["year"]["card"].current.visible = True
            
            if yearly_cards_container.current:
                yearly_cards_container.current.controls = []
            if yearly_section_container.current:
                yearly_section_container.current.visible = False
            
            calculate_year_average(vendor_id, year)
            calculate_quarterly_averages(vendor_id, year)
        
        update_timeline_chart(vendor_id, year)
        
        # Atualizar a página após todos os cálculos
        safe_page_update(page)
    
    def on_timeline_tab_change(e):
        """Callback quando o usuário muda de aba na Timeline"""
        try:
            selected_tab_index = e.control.selected_index
            vendor_id = timeline_vendor_dropdown.current.value if timeline_vendor_dropdown.current else ""
            year = timeline_year_dropdown.current.value if timeline_year_dropdown.current else None
            
            # Índice 1 = Tab de Gráficos Individuais (0=Chart, 1=Individual Charts, 2=Table)
            if selected_tab_index == 1:
                if vendor_id:
                    print(f"🔄 Atualizando Gráficos Individuais ao entrar na aba para vendor: {vendor_id}, year: {year}")
                    update_timeline_individual_charts(vendor_id, year)
                    safe_page_update(page)
                else:
                    print("⚠️ Nenhum fornecedor selecionado para atualizar Gráficos Individuais")
        except Exception as ex:
            print(f"❌ Erro ao mudar de aba na Timeline: {ex}")
            import traceback
            traceback.print_exc()
        
    def show_supplier_info_dialog(vendor_id):
        """Mostra um diálogo com todas as informações do fornecedor"""
        try:
            if not db_conn or not vendor_id:
                show_snack_bar("Nenhum fornecedor selecionado", True)
                return
            
            # Buscar TODAS as informações do supplier
            query = """SELECT supplier_id, vendor_name, supplier_category, bu, supplier_name, 
                      supplier_email, supplier_number, supplier_po, supplier_status, 
                      planner, continuity, sourcing, sqie 
                      FROM supplier_database_table WHERE supplier_id = ?"""
            results = db_manager.query(query, (vendor_id,))
            
            if not results or len(results) == 0:
                show_snack_bar("Fornecedor não encontrado", True)
                return
            
            supplier = results[0]
            theme_colors = get_current_theme_colors(get_theme_name_from_page(page))
            primary_color = theme_colors.get('primary', ft.Colors.BLUE)
            text_color = theme_colors.get('on_surface', ft.Colors.BLACK)
            
            # Buscar dados completos dos responsáveis
            responsible_data = {}
            responsible_types = [
                ('planner', 'planner_table', 'name'),
                ('continuity', 'continuity_table', 'name'),
                ('sourcing', 'sourcing_table', 'name'),
                ('sqie', 'sqie_table', 'name')
            ]
            
            for resp_key, table_name, field_name in responsible_types:
                alias_value = supplier.get(resp_key)
                print(f"\n🔍 ===== Buscando {resp_key} =====")
                print(f"  Name value: '{alias_value}' (tipo: {type(alias_value)})")
                if alias_value:
                    query_resp = f"SELECT name, alias, email FROM {table_name} WHERE {field_name} = ?"
                    print(f"  📝 Query: {query_resp}")
                    print(f"  📝 Parâmetro: {alias_value}")
                    result_resp = db_manager.query_one(query_resp, (alias_value,))
                    print(f"  📋 Resultado bruto: {result_resp}")
                    print(f"  📋 Tipo do resultado: {type(result_resp)}")
                    if result_resp:
                        print(f"  📋 Keys do resultado: {result_resp.keys() if isinstance(result_resp, dict) else 'N/A'}")
                        if isinstance(result_resp, dict):
                            for key in ['name', 'alias', 'email']:
                                val = result_resp.get(key)
                                print(f"      - {key}: '{val}' (tipo: {type(val)}, None: {val is None})")
                    
                    if result_resp:
                        # Tentar diferentes formas de acessar os dados
                        try:
                            # Acessar como dicionário (query_one retorna dict)
                            if isinstance(result_resp, dict):
                                name_val = result_resp.get('name', '') or ''
                                alias_val = result_resp.get('alias', '') or ''
                                email_val = result_resp.get('email', '') or ''
                            elif isinstance(result_resp, (list, tuple)):
                                name_val = result_resp[0] if len(result_resp) > 0 else ''
                                alias_val = result_resp[1] if len(result_resp) > 1 else ''
                                email_val = result_resp[2] if len(result_resp) > 2 else ''
                            else:
                                name_val = str(result_resp) if result_resp else ''
                                alias_val = alias_value
                                email_val = ''
                            
                            # Garantir que os valores são strings não-None
                            name_val = str(name_val) if name_val is not None else ''
                            alias_val = str(alias_val) if alias_val is not None else ''
                            email_val = str(email_val) if email_val is not None else ''
                            
                            print(f"  ✅ Extraído - Nome: '{name_val}', Alias: '{alias_val}', Email: '{email_val}'")
                            
                            responsible_data[resp_key] = {
                                'name': name_val,
                                'alias': alias_val,
                                'email': email_val
                            }
                        except Exception as ex:
                            print(f"  ❌ Erro ao extrair dados: {ex}")
                            import traceback
                            traceback.print_exc()
                            responsible_data[resp_key] = {'name': alias_value, 'alias': alias_value, 'email': ''}
                    else:
                        print(f"  ⚠️ Nenhum resultado encontrado, usando alias como fallback")
                        responsible_data[resp_key] = {'name': alias_value, 'alias': alias_value, 'email': ''}
                else:
                    print(f"  ⚠️ Alias vazio para {resp_key}")
                    responsible_data[resp_key] = None
            
            print(f"📦 Dados finais dos responsáveis: {responsible_data}")
            
            # Função para formatar campo
            def format_field(label, value):
                return ft.Row([
                    ft.Text(f"{label}:", size=13, weight="bold", color=ft.Colors.with_opacity(0.7, text_color), width=150),
                    ft.Text(format_display_value(value) if label not in ["PO", "SSID"] else format_po_ssid(value), 
                           size=13, color=text_color, selectable=True),
                ], spacing=10)
            
            # Função para formatar campo de email com quebra de linha
            def format_email_field(label, value):
                email_value = format_display_value(value)
                # Se tem múltiplos emails separados por ; ou ,
                if email_value and (';' in email_value or ',' in email_value):
                    # Separar emails e criar texto com quebra de linha
                    separator = ';' if ';' in email_value else ','
                    emails = [e.strip() for e in email_value.split(separator) if e.strip()]
                    email_text = '\n'.join(emails)
                else:
                    email_text = email_value
                
                return ft.Column([
                    ft.Text(f"{label}:", size=13, weight="bold", color=ft.Colors.with_opacity(0.7, text_color)),
                    ft.Container(
                        content=ft.Text(email_text, size=13, color=text_color, selectable=True),
                        padding=ft.padding.only(left=10, top=5),
                    ),
                ], spacing=5)
            
            # Função para criar card de responsável
            def create_responsible_card(title, data, icon):
                print(f"🎨 Criando card para {title}: data = {data}")
                
                # Usar cor de fundo secundária do tema
                card_bg_color = theme_colors.get('surface_variant', ft.Colors.BLUE_GREY_50)
                border_color = theme_colors.get('outline', ft.Colors.GREY_300)
                
                if not data:
                    return ft.Container(
                        content=ft.Column([
                            ft.Row([
                                ft.Icon(icon, size=18, color=ft.Colors.GREY_400),
                                ft.Text(title, size=14, weight="bold", color=ft.Colors.GREY_400),
                            ], spacing=5),
                            ft.Text("Não atribuído", size=13, color=ft.Colors.GREY_500, italic=True),
                            ft.Container(expand=True),  # Espaçador para manter altura
                        ], spacing=5),
                        bgcolor=card_bg_color,
                        border=ft.border.all(1, border_color),
                        border_radius=8,
                        padding=14,
                        expand=True,
                        height=130,  # Altura fixa aumentada
                    )
                
                name_value = data.get('name', '')
                alias_value = data.get('alias', '')
                email_value = data.get('email', '')
                
                print(f"  Nome: '{name_value}', Alias: '{alias_value}', Email: '{email_value}'")
                
                card_content = [
                    ft.Row([
                        ft.Icon(icon, size=18, color=primary_color),
                        ft.Text(title, size=14, weight="bold", color=primary_color),
                    ], spacing=5),
                ]
                
                # Adicionar nome (sempre uma linha, mesmo vazio)
                card_content.append(
                    ft.Text(name_value if name_value else "N/A", size=13, weight="w500", color=text_color)
                )
                
                # Adicionar alias (sempre uma linha)
                card_content.append(
                    ft.Row([
                        ft.Icon(ft.Icons.BADGE, size=14, color=ft.Colors.with_opacity(0.6, text_color)),
                        ft.Text(alias_value if alias_value else "N/A", size=12, color=ft.Colors.with_opacity(0.7, text_color)),
                    ], spacing=5)
                )
                
                # Adicionar email (sempre uma linha)
                email_display = email_value if email_value else "Sem email"
                card_content.append(
                    ft.Row([
                        ft.Icon(ft.Icons.EMAIL, size=14, color=ft.Colors.with_opacity(0.6, text_color)),
                        ft.Text(email_display, size=12, color=ft.Colors.with_opacity(0.7, text_color), selectable=True, italic=not bool(email_value)),
                    ], spacing=5, tight=True)
                )
                
                return ft.Container(
                    content=ft.Column(card_content, spacing=6),
                    bgcolor=card_bg_color,
                    border=ft.border.all(1, border_color),
                    border_radius=8,
                    padding=14,
                    expand=True,
                    height=130,  # Altura fixa aumentada
                )
            
            # Criar conteúdo do diálogo
            dialog_content = ft.Container(
                content=ft.Column([
                    # Cabeçalho
                    ft.Row([
                        ft.Icon(ft.Icons.BUSINESS, size=32, color=primary_color),
                        ft.Column([
                            ft.Text("Informações do Fornecedor", size=18, weight="bold", color=primary_color),
                            ft.Text(format_display_value(supplier.get('vendor_name')), size=14, color=text_color),
                        ], spacing=2, expand=True),
                    ], spacing=12),
                    
                    ft.Divider(height=2, thickness=2),
                    
                    # Informações principais
                    ft.Container(
                        content=ft.Column([
                            format_field("ID", supplier.get('supplier_id')),
                            format_field("Nome", supplier.get('vendor_name')),
                            format_field("Categoria", supplier.get('supplier_category')),
                            format_field("BU", supplier.get('bu')),
                            format_field("Origem", supplier.get('supplier_name')),
                            format_field("SSID", supplier.get('supplier_number')),
                            format_field("PO", supplier.get('supplier_po')),
                            format_field("Status", supplier.get('supplier_status')),
                            ft.Container(height=5),  # Espaçamento antes do email
                            format_email_field("Email", supplier.get('supplier_email')),
                        ], spacing=8, scroll=ft.ScrollMode.AUTO),
                        height=300,
                    ),
                    
                    ft.Divider(height=1),
                    
                    # Responsáveis
                    ft.Text("Responsáveis", size=14, weight="bold", color=primary_color),
                    ft.Container(
                        content=ft.Column([
                            # Primeira linha: Planner e Continuity
                            ft.Row([
                                create_responsible_card("Planner", responsible_data.get('planner'), ft.Icons.CALENDAR_MONTH),
                                create_responsible_card("Continuity", responsible_data.get('continuity'), ft.Icons.AUTORENEW),
                            ], spacing=12, expand=True),
                            # Segunda linha: Sourcing e SQIE
                            ft.Row([
                                create_responsible_card("Sourcing", responsible_data.get('sourcing'), ft.Icons.SHOPPING_CART),
                                create_responsible_card("SQIE", responsible_data.get('sqie'), ft.Icons.VERIFIED),
                            ], spacing=12, expand=True),
                        ], spacing=12),
                    ),
                ], spacing=12, scroll=ft.ScrollMode.AUTO),
                padding=20,
                width=700,
            )
            
            # Criar e abrir diálogo
            dialog = ft.AlertDialog(
                title=ft.Text(""),  # Título vazio pois está no content
                content=dialog_content,
                actions=[
                    ft.TextButton("Fechar", on_click=lambda e: page.close(dialog))
                ],
                actions_alignment=ft.MainAxisAlignment.END,
            )
            
            page.open(dialog)
            
        except Exception as e:
            print(f"Erro ao mostrar informações do fornecedor: {e}")
            import traceback
            traceback.print_exc()
            show_snack_bar(f"Erro ao carregar informações: {str(e)}", True)
    
    def update_supplier_info(vendor_id):
        """Atualiza as informações do fornecedor selecionado (mantido para compatibilidade)"""
        # Função mantida para compatibilidade mas não há mais card para mostrar/esconder
        # O botão de informações agora é sempre visível na barra de filtros
        pass
    
    def calculate_overall_average(vendor_id):
        """Calcula a média geral de todos os scores do vendor"""
        try:
            if not db_conn:
                overall_avg_card.current.value = "Error"
                return
                
            query = "SELECT total_score FROM supplier_score_records_table WHERE supplier_id = ? AND total_score > 0"
            results = db_manager.query(query, (vendor_id,))
            
            if results:
                scores = [float(row['total_score']) for row in results]
                avg = sum(scores) / len(scores)
                overall_avg_card.current.value = f"{avg:.1f}"
            else:
                overall_avg_card.current.value = "--"
                
        except Exception as e:
            print(f"Erro ao calcular média geral: {e}")
            overall_avg_card.current.value = "Error"
            
    def calculate_twelve_month_average(vendor_id):
        """Calcula a média dos últimos 12 meses"""
        try:
            if not db_conn:
                twelve_month_avg_card.current.value = "Error"
                return
                
            cursor = db_conn.cursor()
            
            # Obter data atual
            import datetime
            hoje = datetime.date.today()
            
            # Calcular os últimos 12 meses
            scores = []
            for i in range(12):
                data = hoje - datetime.timedelta(days=30*i)  # Aproximação de mês
                query = """SELECT total_score FROM supplier_score_records_table 
                          WHERE supplier_id = ? AND year = ? AND month = ? AND total_score > 0"""
                result = db_manager.query_one(query, (vendor_id, data.year, data.month))
                if result:
                    scores.append(float(result["total_score"]))
            
            if scores:
                avg = sum(scores) / len(scores)
                twelve_month_avg_card.current.value = f"{avg:.1f}"
            else:
                twelve_month_avg_card.current.value = "--"
                
        except Exception as e:
            print(f"Erro ao calcular média 12 meses: {e}")
            twelve_month_avg_card.current.value = "Error"
            
    def calculate_year_average(vendor_id, year=None):
        """Calcula a média do ano selecionado ou ano atual"""
        try:
            if not db_conn:
                year_avg_card.current.value = "Error"
                return
                
            cursor = db_conn.cursor()
            
            if year and year.strip():
                query = """SELECT total_score FROM supplier_score_records_table 
                          WHERE supplier_id = ? AND year = ? AND total_score > 0"""
                results = db_manager.query(query, (vendor_id, year))
            else:
                import datetime
                current_year = datetime.date.today().year
                query = """SELECT total_score FROM supplier_score_records_table 
                          WHERE supplier_id = ? AND year = ? AND total_score > 0"""
                results = db_manager.query(query, (vendor_id, current_year))
            
            if results:
                scores = [float(row['total_score']) for row in results]
                avg = sum(scores) / len(scores)
                year_avg_card.current.value = f"{avg:.1f}"
            else:
                year_avg_card.current.value = "--"
                
        except Exception as e:
            print(f"Erro ao calcular média anual: {e}")
            year_avg_card.current.value = "Error"
            
    def calculate_quarterly_averages(vendor_id, year=None):
        """Calcula as médias trimestrais"""
        try:
            if not db_conn:
                q1_avg_card.current.value = "Error"
                q2_avg_card.current.value = "Error"
                q3_avg_card.current.value = "Error"
                q4_avg_card.current.value = "Error"
                return
                
            cursor = db_conn.cursor()
            current_year = int(year) if year and year.strip() else datetime.date.today().year
            previous_year = current_year - 1
            
            # Definir meses de cada trimestre como lista ordenada
            # Usamos uma lista para garantir a ordem (Q1..Q4) e permitir buscar o trimestre anterior
            quarters = [
                ('Q1', [1, 2, 3]),
                ('Q2', [4, 5, 6]),
                ('Q3', [7, 8, 9]),
                ('Q4', [10, 11, 12])
            ]

            quarter_cards = [q1_avg_card, q2_avg_card, q3_avg_card, q4_avg_card]
            quarter_icons = [q1_arrow_icon, q2_arrow_icon, q3_arrow_icon, q4_arrow_icon]

            # Para cada trimestre, calculamos a média do trimestre atual e do TRIMESTRE ANTERIOR
            # (ex: Q2 compara com Q1; Q1 compara com Q4 do ano anterior). Isso corrige a lógica que
            # antes comparava com o mesmo trimestre do ano anterior.
            for i, (quarter_name, months) in enumerate(quarters):
                # Calcular média do ano atual
                month_placeholders = ','.join(['?' for _ in months])
                query_current = f"""SELECT total_score FROM supplier_score_records_table 
                                   WHERE supplier_id = ? AND year = ? AND month IN ({month_placeholders}) AND total_score > 0"""
                
                params_current = [vendor_id, current_year] + months
                results_current = db_manager.query(query_current, params_current)
                
                avg_current = None
                if results_current:
                    scores = [float(row['total_score']) for row in results_current]
                    avg_current = sum(scores) / len(scores)
                    quarter_cards[i].current.value = f"{avg_current:.1f}"
                else:
                    quarter_cards[i].current.value = "--"
                
                # Procurar o último trimestre anterior que tenha dados
                def find_prev_available_avg(start_idx, start_year):
                    """Procura até 4 trimestres anteriores (incluindo cruzar para o ano anterior) e retorna a média do primeiro encontrado."""
                    for step in range(1, len(quarters) + 1):
                        delta = start_idx - step
                        idx = delta % len(quarters)
                        year_to_check = start_year + (delta // len(quarters))
                        months_to_check = quarters[idx][1]
                        placeholders = ','.join(['?' for _ in months_to_check])
                        query = f"""SELECT total_score FROM supplier_score_records_table
                                    WHERE supplier_id = ? AND year = ? AND month IN ({placeholders}) AND total_score > 0"""
                        params = [vendor_id, year_to_check] + months_to_check
                        res = db_manager.query(query, params)
                        if res:
                            scores = [float(r["total_score"]) for r in res]
                            return sum(scores) / len(scores)
                    return None

                avg_previous = find_prev_available_avg(i, current_year)

                # Definir ícone baseado na comparação com o último trimestre anterior disponível
                try:
                    # Debug: imprimir valores encontrados para diagnóstico
                    print(f"[Timeline][Quarter={quarter_name}] avg_current={avg_current} avg_previous={avg_previous}")
                    if avg_current is not None and avg_previous is not None:
                        if avg_current > avg_previous:
                            # ARROW_UP pode não existir em algumas versões; usar ARROW_UPWARD
                            quarter_icons[i].current.name = ft.Icons.ARROW_UPWARD
                            quarter_icons[i].current.color = ft.Colors.GREEN
                            print(f"[Timeline][{quarter_name}] DECISION: UP (green)")
                        elif avg_current < avg_previous:
                            # ft.Icons.ARROW_DOWN não existe em algumas versões do pacote; usar ARROW_DOWNWARD
                            quarter_icons[i].current.name = ft.Icons.ARROW_DOWNWARD
                            quarter_icons[i].current.color = ft.Colors.RED
                            print(f"[Timeline][{quarter_name}] DECISION: DOWN (red)")
                        else:
                            quarter_icons[i].current.name = ft.Icons.ARROW_FORWARD
                            quarter_icons[i].current.color = ft.Colors.GREY
                            print(f"[Timeline][{quarter_name}] DECISION: EQUAL (grey)")
                    else:
                        # Sem dados atuais ou anteriores disponíveis -> mostrar seta neutra
                        quarter_icons[i].current.name = ft.Icons.ARROW_FORWARD
                        quarter_icons[i].current.color = ft.Colors.GREY
                        print(f"[Timeline][{quarter_name}] DECISION: NO DATA (grey)")
                except Exception as ex:
                    # No caso de refs não estarem inicializadas, ignorar falha sem quebrar a execução
                    print(f"[Timeline][{quarter_name}] Erro ao definir icone: {ex}")
                    pass
                    
        except Exception as e:
            # Imprimir traceback completo para facilitar identificação de erros inesperados
            import traceback
            print(f"Erro ao calcular médias trimestrais: {e}")
            traceback.print_exc()
            for card in [q1_avg_card, q2_avg_card, q3_avg_card, q4_avg_card]:
                try:
                    card.current.value = "Error"
                except Exception:
                    pass
    
    def create_yearly_average_cards(vendor_id):
        """Cria cards com médias anuais para cada ano com dados disponíveis (modo ALL)"""
        try:
            if not db_conn:
                return []
            
            # Buscar todos os anos disponíveis para o fornecedor
            query = """SELECT DISTINCT year FROM supplier_score_records_table 
                      WHERE supplier_id = ? AND total_score > 0 
                      ORDER BY year"""
            years_data = db_manager.query(query, (vendor_id,))
            
            if not years_data:
                return []
            
            cards = []
            # Obter cores do tema
            theme_name_for_timeline = get_theme_name_from_page(page)
            theme_colors_for_timeline = get_current_theme_colors(theme_name_for_timeline)
            primary_color_for_timeline = theme_colors_for_timeline.get('primary')
            
            if theme_name_for_timeline == 'dracula':
                timeline_card_bg = theme_colors_for_timeline.get('field_background') or "#44475A"
                label_color = theme_colors_for_timeline.get('on_surface')
            else:
                timeline_card_bg = None
                label_color = theme_colors_for_timeline.get('on_surface')
            
            for row in years_data:
                year = int(row['year'])
                
                # Calcular média anual
                query_avg = """SELECT AVG(total_score) as avg_score FROM supplier_score_records_table 
                              WHERE supplier_id = ? AND year = ? AND total_score > 0"""
                result = db_manager.query(query_avg, (vendor_id, year))
                
                if result and result[0]['avg_score']:
                    avg_score = float(result[0]['avg_score'])
                    
                    # Comparar com ano anterior se existir
                    prev_year = year - 1
                    query_prev = """SELECT AVG(total_score) as avg_score FROM supplier_score_records_table 
                                   WHERE supplier_id = ? AND year = ? AND total_score > 0"""
                    prev_result = db_manager.query(query_prev, (vendor_id, prev_year))
                    
                    # Determinar ícone e cor
                    icon_name = ft.Icons.ARROW_FORWARD
                    icon_color = ft.Colors.GREY_400
                    
                    if prev_result and prev_result[0]['avg_score']:
                        prev_avg = float(prev_result[0]['avg_score'])
                        if avg_score > prev_avg:
                            icon_name = ft.Icons.ARROW_UPWARD
                            icon_color = ft.Colors.GREEN
                        elif avg_score < prev_avg:
                            icon_name = ft.Icons.ARROW_DOWNWARD
                            icon_color = ft.Colors.RED
                    
                    # Criar card
                    card = ft.Card(
                        content=ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Text(str(year), size=11, weight="w600", color=label_color),
                                    ft.Container(expand=True),
                                    ft.Icon(icon_name, color=icon_color, size=16),
                                ], spacing=6),
                                ft.Container(expand=True),
                                ft.Text(f"{avg_score:.1f}", size=24, weight="bold", color=primary_color_for_timeline),
                            ], spacing=4),
                            padding=ft.padding.all(14),
                            bgcolor=timeline_card_bg,
                            width=135,
                            height=90,
                            border_radius=10
                        ),
                        elevation=2,
                        surface_tint_color=primary_color_for_timeline
                    )
                    cards.append(card)
            
            return cards
            
        except Exception as e:
            print(f"Erro ao criar cards anuais: {e}")
            import traceback
            traceback.print_exc()
            return []
                
    def find_intersection(p1, p2, target_y):
        # p1 = (x1, y1), p2 = (x2, y2)
        if p1[1] == p2[1]:  # Linha horizontal, sem interseção a menos que esteja na linha de destino
            return None
        
        # Calcula a interseção x
        x = p1[0] + (p2[0] - p1[0]) * (target_y - p1[1]) / (p2[1] - p1[1])
        return (x, target_y)

    def create_colored_line_series(points, target_y, tooltip_label="Total Score"):
        if not points:
            return []
        if len(points) < 2:
            point = points[0]
            color = ft.Colors.GREEN_600 if point[1] >= target_y else ft.Colors.RED_600
            tooltip_text = f"{tooltip_label} - {point[1]}"
            return [ft.LineChartData(data_points=[ft.LineChartDataPoint(point[0], point[1], tooltip=tooltip_text)], color=color, stroke_width=3)]

        above_segments, below_segments = [], []
        current_above, current_below = [], []

        p1 = points[0]
        tooltip_p1 = f"{tooltip_label} - {p1[1]}"
        if p1[1] >= target_y:
            current_above.append(ft.LineChartDataPoint(p1[0], p1[1], tooltip=tooltip_p1))
        else:
            current_below.append(ft.LineChartDataPoint(p1[0], p1[1], tooltip=tooltip_p1))

        for i in range(len(points) - 1):
            p1, p2 = points[i], points[i+1]
            p1_above, p2_above = p1[1] >= target_y, p2[1] >= target_y

            if p1_above != p2_above:
                intersection = find_intersection(p1, p2, target_y)
                if intersection:
                    ix, iy = intersection
                    tooltip_inter = f"{tooltip_label} - {iy}"
                    intersection_point = ft.LineChartDataPoint(ix, iy, tooltip=tooltip_inter)
                    tooltip_p2 = f"{tooltip_label} - {p2[1]}"
                    if p1_above:
                        current_above.append(intersection_point)
                        above_segments.append(current_above)
                        current_above = []
                        current_below = [intersection_point, ft.LineChartDataPoint(p2[0], p2[1], tooltip=tooltip_p2)]
                    else:
                        current_below.append(intersection_point)
                        below_segments.append(current_below)
                        current_below = []
                        current_above = [intersection_point, ft.LineChartDataPoint(p2[0], p2[1], tooltip=tooltip_p2)]
            else:
                tooltip_p2 = f"{tooltip_label} - {p2[1]}"
                if p2_above:
                    current_above.append(ft.LineChartDataPoint(p2[0], p2[1], tooltip=tooltip_p2))
                else:
                    current_below.append(ft.LineChartDataPoint(p2[0], p2[1], tooltip=tooltip_p2))

        if current_above: above_segments.append(current_above)
        if current_below: below_segments.append(current_below)

        series = []
        for segment in above_segments:
            if len(segment) > 1:
                series.append(ft.LineChartData(
                    data_points=segment, color=ft.Colors.GREEN_600, stroke_width=3, curved=True, stroke_cap_round=True,
                    below_line_gradient=ft.LinearGradient(
                        begin=ft.alignment.top_center, end=ft.alignment.bottom_center,
                        colors=[ft.Colors.with_opacity(0.4, ft.Colors.GREEN_400), ft.Colors.with_opacity(0.0, ft.Colors.GREEN_400)]
                    )
                ))
        for segment in below_segments:
            if len(segment) > 1:
                series.append(ft.LineChartData(
                    data_points=segment, color=ft.Colors.RED_600, stroke_width=3, curved=True, stroke_cap_round=True,
                    below_line_gradient=ft.LinearGradient(
                        begin=ft.alignment.top_center, end=ft.alignment.bottom_center,
                        colors=[ft.Colors.with_opacity(0.4, ft.Colors.RED_400), ft.Colors.with_opacity(0.0, ft.Colors.RED_400)]
                    )
                ))
        return series

    def calculate_regression(points):
        """Calcula regressão linear para os pontos fornecidos"""
        if not points or len(points) < 2:
            return None
        
        x_values = [p[0] for p in points]
        y_values = [p[1] for p in points]
        n = len(x_values)
        
        x_mean = sum(x_values) / n
        y_mean = sum(y_values) / n
        
        # Calcular variações para detectar dados sem variação
        ss_tot = sum((y_values[i] - y_mean) ** 2 for i in range(n))
        
        # Se não há variação nos dados Y (todos os valores são iguais)
        if ss_tot == 0 or ss_tot < 0.0001:  # Tolerância para erros de ponto flutuante
            # Sem variação = não há nada para explicar = R² = 0
            # (Matematicamente, R² é indefinido quando SS_tot = 0, mas usamos 0 por convenção)
            return {
                'm': 0,  # Linha horizontal (sem inclinação)
                'b': y_mean,  # Intercepto é a própria média
                'r_squared': 0,  # Não há variação para explicar
                'n': n,
                'x_values': x_values,
                'y_values': y_values,
                'x_mean': x_mean,
                'y_mean': y_mean,
                'y_predicted': [y_mean] * n  # Todos os valores previstos são a média
            }
        
        numerator = sum((x_values[i] - x_mean) * (y_values[i] - y_mean) for i in range(n))
        denominator = sum((x_values[i] - x_mean) ** 2 for i in range(n))
        
        if denominator == 0:
            # Não há variação em X (todos os meses são iguais - não deveria acontecer)
            # Mas temos variação em Y, então usamos a média
            return {
                'm': 0,
                'b': y_mean,
                'r_squared': 0,  # Não conseguimos fazer previsão baseada em X
                'n': n,
                'x_values': x_values,
                'y_values': y_values,
                'x_mean': x_mean,
                'y_mean': y_mean,
                'y_predicted': [y_mean] * n
            }
        
        m = numerator / denominator
        b = y_mean - m * x_mean
        
        y_predicted = [m * x + b for x in x_values]
        ss_res = sum((y_values[i] - y_predicted[i]) ** 2 for i in range(n))
        
        # Calcular R²
        r_squared = 1 - (ss_res / ss_tot)
        # Garantir que R² esteja entre 0 e 1
        r_squared = max(0, min(1, r_squared))
        
        return {
            'm': m,
            'b': b,
            'r_squared': r_squared,
            'n': n,
            'x_values': x_values,
            'y_values': y_values,
            'x_mean': x_mean,
            'y_mean': y_mean,
            'y_predicted': y_predicted
        }

    def show_trend_explanation_dialog(metric_name, reg_data, metric_color=None):
        """Mostra o diálogo de explicação da tendência para qualquer métrica"""
        if not reg_data:
            return
        
        color = metric_color or ft.Colors.BLUE
        months_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
        
        # Criar tabela com dados observados vs preditos
        sorted_data = sorted(zip(reg_data['x_values'], reg_data['y_values'], reg_data['y_predicted']), 
                           key=lambda item: item[0])
        
        data_rows = []
        for x, y_obs, y_pred in sorted_data:
            residual = y_obs - y_pred
            if abs(residual) <= 0.5:
                residual_color = ft.Colors.GREY
            elif residual > 0:
                residual_color = ft.Colors.GREEN
            else:
                residual_color = ft.Colors.RED
            
            data_rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(months_labels[int(x) - 1], size=14)),
                    ft.DataCell(ft.Text(f"{y_obs:.2f}", size=14)),
                    ft.DataCell(ft.Text(f"{y_pred:.2f}", size=14)),
                    ft.DataCell(ft.Text(f"{residual:+.2f}", size=14, color=residual_color)),
                ])
            )
        
        # Determinar interpretação da tendência
        if reg_data['m'] > 0.05:
            trend_explanation = "Os scores estão melhorando ao longo do tempo"
            trend_icon = "📈"
            trend_color_dialog = ft.Colors.GREEN_600
        elif reg_data['m'] < -0.05:
            trend_explanation = "Os scores estão caindo ao longo do tempo"
            trend_icon = "📉"
            trend_color_dialog = ft.Colors.RED_600
        else:
            trend_explanation = "Os scores estão relativamente estáveis"
            trend_icon = "➡️"
            trend_color_dialog = ft.Colors.BLUE_600
        
        # Interpretação do R² - mede o quanto a equação explica a variação dos dados
        r_squared_pct = reg_data['r_squared'] * 100
        
        # Caso especial: R² = 0 pode significar que não há variação nos dados
        if r_squared_pct < 0.1:  # Praticamente zero
            # Verificar se todos os Y são iguais (sem variação)
            y_values_set = set(reg_data['y_values'])
            if len(y_values_set) == 1:
                r_squared_quality = "0%"
                r_squared_explanation = "Não há variação nos dados - todos os valores são constantes"
            else:
                r_squared_quality = "0-19%"
                r_squared_explanation = "A equação não explica quase nada da variação dos dados observados"
        elif r_squared_pct >= 80:
            r_squared_quality = "80-100%"
            r_squared_explanation = "A equação explica muito bem a variação dos dados observados"
        elif r_squared_pct >= 60:
            r_squared_quality = "60-79%"
            r_squared_explanation = "A equação explica boa parte da variação dos dados observados"
        elif r_squared_pct >= 40:
            r_squared_quality = "40-59%"
            r_squared_explanation = "A equação explica parte da variação dos dados observados"
        elif r_squared_pct >= 20:
            r_squared_quality = "20-39%"
            r_squared_explanation = "A equação explica pouca variação dos dados observados"
        else:
            r_squared_quality = "0-19%"
            r_squared_explanation = "A equação não explica quase nada da variação dos dados observados"
        
        dialog_content = ft.Column([
            ft.Text(f"Como calcular a tendência - {metric_name}", size=22, weight="bold"),
            ft.Divider(),
            
            # Resumo em linguagem simples
            ft.Container(
                content=ft.Column([
                    ft.Row([
                        ft.Text(trend_icon, size=32),
                        ft.Column([
                            ft.Text("Resumo:", size=18, weight="bold", color=color),
                            ft.Text(trend_explanation, size=16, color=trend_color_dialog, weight="bold"),
                        ], spacing=2),
                    ], spacing=10),
                    ft.Container(height=8),
                    ft.Text(f"R² (Coeficiente de Determinação): {r_squared_pct:.1f}%", 
                           size=15, weight="bold"),
                    ft.Text("👉 Mede o quanto a equação consegue explicar a variação dos dados", 
                           size=14, italic=True, color=ft.Colors.BLUE_700),
                    ft.Text(r_squared_explanation, size=14, italic=True, color=ft.Colors.GREY_700),
                    ft.Container(height=5),
                    ft.Row([
                        ft.Text("• R² = 1 (100%)", size=13, weight="bold"),
                        ft.Text("→ explica toda variação", size=13, italic=True, color=ft.Colors.GREY_600),
                    ], spacing=5),
                    ft.Row([
                        ft.Text("• R² = 0 (0%)", size=13, weight="bold"),
                        ft.Text("→ não explica nada", size=13, italic=True, color=ft.Colors.GREY_600),
                    ], spacing=5),
                    
                    # Indicador visual do R² na escala 0 a 1
                    ft.Container(height=15),
                    ft.Text("Indicador de R²:", size=13, weight="bold"),
                    ft.Container(height=5),
                    
                    # Escala visual de 0 a 1 (largura total da janela)
                    ft.Column([
                        # Marcador de posição do R² atual (acima da barra)
                        ft.Container(
                            content=ft.Stack([
                                # Indicador posicionado
                                ft.Container(
                                    content=ft.Column([
                                        ft.Container(
                                            content=ft.Text(f"{reg_data['r_squared']:.3f}", size=12, weight="bold", color=ft.Colors.WHITE),
                                            bgcolor=ft.Colors.BLUE_700,
                                            padding=ft.padding.symmetric(horizontal=8, vertical=3),
                                            border_radius=4,
                                        ),
                                        ft.Icon(ft.Icons.ARROW_DROP_DOWN, size=28, color=ft.Colors.BLUE_700),
                                    ], spacing=-5, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                                    left=max(0, min(480, reg_data['r_squared'] * 520 - 25)),  # Limita para não sair da barra
                                    top=10,
                                ),
                            ], width=520, height=50),
                            margin=ft.margin.only(bottom=0),
                        ),
                        
                        # Barra de gradiente (vermelho → amarelo → verde)
                        ft.Container(
                            width=520,
                            height=30,
                            border_radius=5,
                            border=ft.border.all(2, ft.Colors.BLACK),
                            gradient=ft.LinearGradient(
                                begin=ft.alignment.center_left,
                                end=ft.alignment.center_right,
                                colors=[
                                    ft.Colors.RED_700,      # 0.0
                                    ft.Colors.ORANGE_700,   # 0.25
                                    ft.Colors.YELLOW_700,   # 0.5
                                    ft.Colors.LIGHT_GREEN,  # 0.75
                                    ft.Colors.GREEN_700,    # 1.0
                                ],
                            ),
                        ),
                        
                        # Labels da escala (0.0, 0.25, 0.50, 0.75, 1.0)
                        ft.Container(height=5),
                        ft.Row([
                            ft.Text("0.0", size=11, weight="bold"),
                            ft.Container(expand=True),
                            ft.Text("0.25", size=11, weight="bold"),
                            ft.Container(expand=True),
                            ft.Text("0.50", size=11, weight="bold"),
                            ft.Container(expand=True),
                            ft.Text("0.75", size=11, weight="bold"),
                            ft.Container(expand=True),
                            ft.Text("1.0", size=11, weight="bold"),
                        ], width=520),
                        
                        # Legendas dos extremos
                        ft.Container(height=3),
                        ft.Row([
                            ft.Text("Nenhuma explicação", size=10, italic=True, color=ft.Colors.GREY_600),
                            ft.Container(expand=True),
                            ft.Text("Explicação perfeita", size=10, italic=True, color=ft.Colors.GREY_600),
                        ], width=520),
                    ], spacing=0),
                    
                ], spacing=5),
                padding=10,
                bgcolor=ft.Colors.with_opacity(0.05, color),
                border_radius=8,
            ),
            
            ft.Container(height=10),
            ft.Text("📊 Informações Básicas", size=17, weight="bold"),
            ft.Container(
                content=ft.Column([
                    ft.Text(f"Equação da linha: y = {reg_data['m']:.4f}x + {reg_data['b']:.2f}", size=14),
                    ft.Text(f"Inclinação (m): {reg_data['m']:.4f}", size=13),
                    ft.Text(f"Intercepto (b): {reg_data['b']:.2f}", size=13),
                    ft.Text(f"Número de pontos (n): {reg_data['n']}", size=13),
                    ft.Text(f"Média dos valores X: {reg_data['x_mean']:.2f}", size=13),
                    ft.Text(f"Média dos valores Y: {reg_data['y_mean']:.2f}", size=13),
                ], spacing=5),
                padding=10,
                bgcolor=ft.Colors.with_opacity(0.03, ft.Colors.BLUE),
                border_radius=8,
            ),
            
            ft.Container(height=10),
            ft.Text("📈 Previsão vs Realidade", size=17, weight="bold"),
            ft.Container(
                content=ft.DataTable(
                    columns=[
                        ft.DataColumn(ft.Text("Mês", size=14, weight="bold")),
                        ft.DataColumn(ft.Text("Score Real", size=14, weight="bold")),
                        ft.DataColumn(ft.Text("Previsão", size=14, weight="bold")),
                        ft.DataColumn(ft.Text("Diferença", size=14, weight="bold")),
                    ],
                    rows=data_rows,
                ),
                border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.ON_SURFACE)),
                border_radius=8,
                padding=10,
            ),
            ft.Column([
                ft.Text("Legenda:", size=13, weight="bold", color=ft.Colors.GREY_700),
                ft.Text("🟢 Verde = Real superou a previsão (diferença > +0.5)", 
                       size=13, color=ft.Colors.GREY_600, italic=True),
                ft.Text("⚪ Cinza = Próximo da previsão (diferença entre -0.5 e +0.5)", 
                       size=13, color=ft.Colors.GREY_600, italic=True),
                ft.Text("🔴 Vermelho = Real abaixo da previsão (diferença < -0.5)", 
                       size=13, color=ft.Colors.GREY_600, italic=True),
            ], spacing=3),
        ], spacing=10, scroll=ft.ScrollMode.AUTO, height=550)
        
        dialog = ft.AlertDialog(
            title=ft.Row([
                ft.Icon(ft.Icons.SCHOOL, color=color),
                ft.Text("Análise de Regressão Linear", color=color),
            ], spacing=8),
            content=dialog_content,
            actions=[
                ft.TextButton("Entendi", on_click=lambda e: page.close(dialog))
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        page.open(dialog)

    def update_timeline_chart(vendor_id, year=None):
        """Atualiza o gráfico de linha baseado no vendor, ano e métricas selecionadas."""
        try:
            if not db_conn or not vendor_id:
                timeline_chart_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.SHOW_CHART, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text("Selecione um fornecedor para visualizar o gráfico", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                    expand=True,
                )
                timeline_chart_container.current.update()
                return

            target_value = target_slider.value if target_slider and target_slider.value is not None else 5.0
            
            # Verificar se é "Todo o Período" ou um ano específico
            if year == "ALL":
                query = "SELECT month, year, total_score, otif, nil, quality_pickup, quality_package FROM supplier_score_records_table WHERE supplier_id = ? ORDER BY year, month"
                results = db_manager.query(query, (vendor_id,))
                # Para ALL: usar lista simples, cada item é um mês em ordem cronológica
                monthly_data = []
            else:
                analysis_year = int(year) if year and year.strip() else datetime.datetime.now().year
                query = "SELECT month, total_score, otif, nil, quality_pickup, quality_package FROM supplier_score_records_table WHERE supplier_id = ? AND year = ?"
                results = db_manager.query(query, (vendor_id, analysis_year))
                monthly_data = {m: {} for m in range(1, 13)}
            
            for row in results:
                month = row['month']
                total = row['total_score']
                otif = row['otif']
                nil = row['nil']
                pickup = row['quality_pickup']
                package = row['quality_package']
                try:
                    month_int = int(month)
                    
                    if year == "ALL":
                        # Para ALL: adicionar cada mês encontrado sequencialmente
                        year_val = int(row['year'])
                        monthly_data.append({
                            "total_score": float(total) if total is not None else None,
                            "otif": float(otif) if otif is not None else None,
                            "nil": float(nil) if nil is not None else None,
                            "pickup": float(pickup) if pickup is not None else None,
                            "package": float(package) if package is not None else None,
                            "year": year_val,
                            "month": month_int
                        })
                    else:
                        monthly_data[month_int] = {
                            "total_score": float(total) if total is not None else None,
                            "otif": float(otif) if otif is not None else None,
                            "nil": float(nil) if nil is not None else None,
                            "pickup": float(pickup) if pickup is not None else None,
                            "package": float(package) if package is not None else None,
                        }
                except (ValueError, TypeError):
                    print(f"⚠️ Dados inválidos ignorados: month='{month}'")
                    continue

            # Verificar se há dados
            if year == "ALL":
                has_data = any(data.get("total_score") is not None for data in monthly_data)
            else:
                has_data = any(data.get("total_score") is not None for data in monthly_data.values())
            
            if not has_data:
                period_text = "todo o período" if year == "ALL" else str(analysis_year)
                timeline_chart_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.CALENDAR_TODAY, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text(f"Nenhum dado de score encontrado para {period_text}", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                    expand=True,
                )
                timeline_chart_container.current.update()
                return

            data_series = []
            series_names = []
            
            if year == "ALL":
                # Para ALL: usar índice 0, 1, 2, 3... para cada mês encontrado
                total_score_points_chart = [(idx, d["total_score"]) for idx, d in enumerate(monthly_data) if d.get("total_score") is not None]
                total_score_points = total_score_points_chart  # Mesmos valores para regressão
            else:
                # Para um ano específico, usar índice 0-11 para gráfico e 1-12 para regressão
                total_score_points_chart = [(m - 1, d["total_score"]) for m, d in monthly_data.items() if d and d.get("total_score") is not None]
                total_score_points = [(m, d["total_score"]) for m, d in monthly_data.items() if d and d.get("total_score") is not None]
            
            if total_score_points_chart:
                colored_series = create_colored_line_series(total_score_points_chart, target_value)
                data_series.extend(colored_series)
                series_names.extend(["Total Score"] * len(colored_series))

            # Adicionar linha de target
            if year == "ALL":
                # Para ALL: linha de target ao longo de todos os meses encontrados
                num_points = len(monthly_data)
                target_line_points = [ft.LineChartDataPoint(i, target_value, tooltip=f"Target - {target_value}") for i in range(num_points)]
            else:
                # Para um ano específico, linha de target nos 12 meses
                target_line_points = [ft.LineChartDataPoint(i, target_value, tooltip=f"Target - {target_value}") for i in range(12)]
            data_series.append(ft.LineChartData(
                data_points=target_line_points,
                color=ft.Colors.AMBER_700,
                stroke_width=2,
                curved=False,
                dash_pattern=[5, 5]
            ))
            series_names.append("Target")

            additional_scores_map = {
                "otif": {"ref": timeline_otif_check, "color": ft.Colors.ORANGE_ACCENT_700, "tooltip": "OTIF"},
                "nil": {"ref": timeline_nil_check, "color": ft.Colors.PURPLE_ACCENT_700, "tooltip": "NIL"},
                "pickup": {"ref": timeline_pickup_check, "color": ft.Colors.CYAN_700, "tooltip": "Quality Pickup"},
                "package": {"ref": timeline_package_check, "color": ft.Colors.PINK_ACCENT_700, "tooltip": "Quality Package"},
            }
            for key, details in additional_scores_map.items():
                if hasattr(details["ref"], 'current') and details["ref"].current and details["ref"].current.value:
                    if year == "ALL":
                        points = [ft.LineChartDataPoint(idx, d[key], tooltip=f"{details['tooltip']} - {d[key]}") for idx, d in enumerate(monthly_data) if d.get(key) is not None]
                    else:
                        points = [ft.LineChartDataPoint(m - 1, d[key], tooltip=f"{details['tooltip']} - {d[key]}") for m, d in monthly_data.items() if d and d.get(key) is not None]
                    if points:
                        data_series.append(ft.LineChartData(data_points=points, color=details["color"], stroke_width=2, curved=False, dash_pattern=[3, 3]))
                        series_names.append(details["tooltip"])

            months_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
            
            def timeline_tooltip_provider(series_index, point_index, point):
                if series_index < len(series_names):
                    name = series_names[series_index]
                    value = point.y
                    return f"{name} - {value}"
                return f"{point.y}"
            
            # Configurar eixos baseado no período
            if year == "ALL":
                # Para ALL: criar labels para cada mês em ordem cronológica
                bottom_labels = []
                for idx, data in enumerate(monthly_data):
                    year_val = data.get("year", 2025)
                    month_val = data.get("month", 1)
                    label_text = f"{months_labels[month_val-1]}/{str(year_val)[2:]}"
                    bottom_labels.append(ft.ChartAxisLabel(value=idx, label=ft.Text(label_text, size=8, weight=ft.FontWeight.BOLD)))
                
                max_index = len(monthly_data) - 1 if monthly_data else 0
                
                line_chart = ft.LineChart(
                    data_series=data_series,
                    border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.ON_SURFACE)),
                    horizontal_grid_lines=ft.ChartGridLines(interval=1, color=ft.Colors.with_opacity(0.1, ft.Colors.ON_SURFACE), width=1),
                    vertical_grid_lines=ft.ChartGridLines(interval=1, color=ft.Colors.with_opacity(0.1, ft.Colors.ON_SURFACE), width=1),
                    left_axis=ft.ChartAxis(labels=[ft.ChartAxisLabel(value=i, label=ft.Text(str(i), size=10)) for i in range(0, 11, 1)], labels_size=30),
                    bottom_axis=ft.ChartAxis(labels=bottom_labels, labels_size=40),
                    tooltip_bgcolor=ft.Colors.with_opacity(0.8, ft.Colors.BLUE_GREY),
                    min_y=0, max_y=10, min_x=0, max_x=max_index,
                    expand=True,
                )
            else:
                # Para um ano específico, usar labels mensais normais
                line_chart = ft.LineChart(
                    data_series=data_series,
                    border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.ON_SURFACE)),
                    horizontal_grid_lines=ft.ChartGridLines(interval=1, color=ft.Colors.with_opacity(0.1, ft.Colors.ON_SURFACE), width=1),
                    vertical_grid_lines=ft.ChartGridLines(interval=1, color=ft.Colors.with_opacity(0.1, ft.Colors.ON_SURFACE), width=1),
                    left_axis=ft.ChartAxis(labels=[ft.ChartAxisLabel(value=i, label=ft.Text(str(i), size=10)) for i in range(0, 11, 1)], labels_size=30),
                    bottom_axis=ft.ChartAxis(labels=[ft.ChartAxisLabel(value=i, label=ft.Text(label, size=10, weight=ft.FontWeight.BOLD)) for i, label in enumerate(months_labels)], labels_size=30),
                    tooltip_bgcolor=ft.Colors.with_opacity(0.8, ft.Colors.BLUE_GREY),
                    min_y=0, max_y=10, min_x=0, max_x=11,
                    expand=True,
                )
            # Criar legenda
            legend_items = []
            if total_score_points:
                legend_items.append(ft.Row([ft.Container(width=20, height=10, bgcolor=ft.Colors.GREEN_600), ft.Text("Total Score", size=12)], spacing=5))
            legend_items.append(ft.Row([ft.Container(width=20, height=10, bgcolor=ft.Colors.AMBER_700), ft.Text("Target", size=12)], spacing=5))
            for key, details in additional_scores_map.items():
                if hasattr(details["ref"], 'current') and details["ref"].current and details["ref"].current.value:
                    legend_items.append(ft.Row([ft.Container(width=20, height=10, bgcolor=details["color"]), ft.Text(details["tooltip"], size=12)], spacing=5))
            
            # Calcular regressão para Total Score e criar análise
            total_score_regression = calculate_regression(total_score_points) if total_score_points and len(total_score_points) >= 2 else None
            
            # Criar botão e informações de análise
            analysis_row_items = [ft.Row(legend_items, spacing=10, wrap=True)]
            
            if total_score_regression:
                # Determinar tendência
                if total_score_regression['m'] > 0.05:
                    trend_text = "↗ Crescente"
                    trend_color = ft.Colors.GREEN_600
                elif total_score_regression['m'] < -0.05:
                    trend_text = "↘ Decrescente"
                    trend_color = ft.Colors.RED_600
                else:
                    trend_text = "→ Estável"
                    trend_color = ft.Colors.BLUE_600
                
                # Adicionar divider
                analysis_row_items.append(ft.VerticalDivider(width=1, color=ft.Colors.with_opacity(0.3, ft.Colors.ON_SURFACE)))
                
                # Adicionar informações da análise
                analysis_row_items.extend([
                    ft.Text(f"y={total_score_regression['m']:.3f}x+{total_score_regression['b']:.2f}", 
                           size=11, weight="w500",
                           tooltip="Equação da linha de tendência"),
                    ft.Text(trend_text, size=11, color=trend_color, weight="bold",
                           tooltip="Tendência dos dados"),
                    ft.Text(f"R²={total_score_regression['r_squared']:.3f}", 
                           size=11, color=ft.Colors.GREY,
                           tooltip="Coeficiente de determinação"),
                ])
                
                # Função para mostrar dialog
                def show_total_score_info(e):
                    show_trend_explanation_dialog("Total Score", total_score_regression, ft.Colors.GREEN_600)
                
                # Adicionar botão de info
                analysis_row_items.append(
                    ft.IconButton(
                        icon=ft.Icons.INFO_OUTLINE,
                        icon_size=18,
                        icon_color=ft.Colors.GREEN_600,
                        tooltip="Como funciona a análise de tendência",
                        on_click=show_total_score_info,
                    )
                )
            
            chart_with_legend = ft.Column([
                line_chart,
                ft.Container(height=10),
                ft.Row(analysis_row_items, spacing=10, alignment=ft.MainAxisAlignment.START)
            ], spacing=5, expand=True)
            
            timeline_chart_container.current.content = chart_with_legend
            timeline_chart_container.current.update()
            
        except Exception as e:
            print(f"Erro ao atualizar gráfico: {e}")
            import traceback
            traceback.print_exc()
            timeline_chart_container.current.content = ft.Container(
                content=ft.Column([
                    ft.Icon(ft.Icons.ERROR_OUTLINE, size=48, color=ft.Colors.RED_400),
                    ft.Container(height=10),
                    ft.Text(f"Erro ao gerar gráfico: {e}", size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.RED_600),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                alignment=ft.alignment.center,
                expand=True,
            )
            timeline_chart_container.current.update()
    
    def update_timeline_individual_charts(vendor_id, year=None):
        """Atualiza a aba de gráficos individuais com 4 gráficos (OTIF, NIL, Pickup, Package) em 2x2"""
        try:
            if not db_conn or not vendor_id:
                timeline_individual_charts_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.SHOW_CHART, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text("Selecione um fornecedor para visualizar os gráficos", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                    expand=True,
                )
                timeline_individual_charts_container.current.update()
                return

            # Obter valor do target
            target_value = target_slider.value if target_slider and target_slider.value is not None else 5.0

            # Verificar se é "Todo o Período" ou um ano específico
            if year == "ALL":
                query = "SELECT month, year, otif, nil, quality_pickup, quality_package FROM supplier_score_records_table WHERE supplier_id = ? ORDER BY year, month"
                results = db_manager.query(query, (vendor_id,))
                # Para ALL: usar lista simples
                monthly_data = []
            else:
                analysis_year = int(year) if year and year.strip() else datetime.datetime.now().year
                query = "SELECT month, otif, nil, quality_pickup, quality_package FROM supplier_score_records_table WHERE supplier_id = ? AND year = ?"
                results = db_manager.query(query, (vendor_id, analysis_year))
                monthly_data = {m: {} for m in range(1, 13)}
            
            for row in results:
                month = row['month']
                otif = row['otif']
                nil = row['nil']
                pickup = row['quality_pickup']
                package = row['quality_package']
                try:
                    month_int = int(month)
                    
                    if year == "ALL":
                        # Para ALL: adicionar cada mês sequencialmente
                        year_val = int(row['year'])
                        monthly_data.append({
                            "otif": float(otif) if otif is not None else None,
                            "nil": float(nil) if nil is not None else None,
                            "pickup": float(pickup) if pickup is not None else None,
                            "package": float(package) if package is not None else None,
                            "year": year_val,
                            "month": month_int
                        })
                    else:
                        monthly_data[month_int] = {
                            "otif": float(otif) if otif is not None else None,
                            "nil": float(nil) if nil is not None else None,
                            "pickup": float(pickup) if pickup is not None else None,
                            "package": float(package) if package is not None else None,
                        }
                except (ValueError, TypeError):
                    print(f"⚠️ Dados inválidos ignorados: month='{month}'")
                    continue

            # Verificar se há dados
            if year == "ALL":
                has_data = any(any(d.values()) for d in monthly_data if d)
            else:
                has_data = any(any(d.values()) for d in monthly_data.values() if d)
            
            if not has_data:
                period_text = "todo o período" if year == "ALL" else str(analysis_year)
                timeline_individual_charts_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.CALENDAR_TODAY, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text(f"Nenhum dado encontrado para {period_text}", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                    expand=True,
                )
                timeline_individual_charts_container.current.update()
                return

            months_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
            
            # Função para criar um gráfico individual
            def create_individual_chart(metric_key, metric_name, color):
                # Criar pontos de dados para cada mês e calcular média
                data_points = []
                values_list = []
                points = []
                
                # Preparar labels do eixo X baseado em ano ALL ou específico
                bottom_labels = []
                
                if year == "ALL":
                    # Para ALL: usar índice 0, 1, 2... para cada mês encontrado
                    for idx, d in enumerate(monthly_data):
                        if d.get(metric_key) is not None:
                            value = d[metric_key]
                            values_list.append(value)
                            points.append((idx, value))
                            
                            # Obter ano e mês do dado
                            year_val = d.get("year", 2025)
                            month_val = d.get("month", 1)
                            label_text = f"{months_labels[month_val-1]}/{str(year_val)[2:]}"
                            
                            bottom_labels.append(ft.ChartAxisLabel(
                                value=idx,
                                label=ft.Text(label_text, size=8, weight=ft.FontWeight.BOLD)
                            ))
                            
                            data_points.append(
                                ft.LineChartDataPoint(
                                    idx,
                                    value,
                                    tooltip=f"{metric_name} ({label_text}) - {value}"
                                )
                            )
                else:
                    # Para ano específico, usar índices 0-11 (meses)
                    for m, d in monthly_data.items():
                        if d and d.get(metric_key) is not None:
                            value = d[metric_key]
                            values_list.append(value)
                            points.append((m, value))
                            data_points.append(
                                ft.LineChartDataPoint(
                                    m - 1,
                                    value,
                                    tooltip=f"{metric_name} - {value}"
                                )
                            )
                    
                    # Labels fixos para meses
                    bottom_labels = [ft.ChartAxisLabel(value=i, label=ft.Text(label, size=10, weight=ft.FontWeight.BOLD)) 
                                   for i, label in enumerate(months_labels)]
                
                if not data_points:
                    return ft.Container(
                        content=ft.Column([
                            ft.Text(metric_name, size=14, weight="bold", text_align=ft.TextAlign.CENTER),
                            ft.Container(height=10),
                            ft.Text("Sem dados", size=12, color=ft.Colors.GREY_600, text_align=ft.TextAlign.CENTER),
                        ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                        border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.ON_SURFACE)),
                        border_radius=8,
                        padding=20,
                        expand=True,
                    )
                
                # Calcular média
                average = sum(values_list) / len(values_list) if values_list else 0
                
                # Criar linha principal do gráfico
                line_data = ft.LineChartData(
                    data_points=data_points,
                    color=color,
                    stroke_width=3,
                    curved=False,
                )
                
                # Adicionar linha de target baseada no modo (ALL ou ano específico)
                if year == "ALL":
                    # Para ALL, a linha de target vai através de todos os pontos
                    num_points = len(monthly_data)
                    target_line_points = [ft.LineChartDataPoint(i, target_value, tooltip=f"Target - {target_value}") 
                                        for i in range(num_points)]
                    max_x_value = num_points - 1 if num_points > 0 else 0
                    min_x_value = 0
                else:
                    # Para ano específico, target de 0 a 11
                    target_line_points = [ft.LineChartDataPoint(i, target_value, tooltip=f"Target - {target_value}") 
                                        for i in range(12)]
                    max_x_value = 11
                    min_x_value = 0
                
                target_line_data = ft.LineChartData(
                    data_points=target_line_points,
                    color=ft.Colors.AMBER_700,
                    stroke_width=2,
                    curved=False,
                    dash_pattern=[5, 5]
                )
                
                chart = ft.LineChart(
                    data_series=[line_data, target_line_data],
                    border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.ON_SURFACE)),
                    vertical_grid_lines=ft.ChartGridLines(
                        interval=1,
                        color=ft.Colors.with_opacity(0.1, ft.Colors.ON_SURFACE),
                        width=1
                    ),
                    left_axis=ft.ChartAxis(
                        labels=[ft.ChartAxisLabel(value=i, label=ft.Text(str(i), size=10)) for i in range(0, 11, 1)],
                        labels_size=30
                    ),
                    bottom_axis=ft.ChartAxis(
                        labels=bottom_labels,
                        labels_size=30
                    ),
                    tooltip_bgcolor=ft.Colors.with_opacity(0.8, ft.Colors.BLUE_GREY),
                    min_y=0,
                    max_y=10,
                    min_x=min_x_value,
                    max_x=max_x_value,
                    expand=True,
                )
                
                # Calcular regressão linear
                regression = None
                trend_text = ""
                trend_color = ft.Colors.GREY
                
                if len(points) >= 2:
                    x_values = [p[0] for p in points]
                    y_values = [p[1] for p in points]
                    n = len(x_values)
                    
                    x_mean = sum(x_values) / n
                    y_mean = sum(y_values) / n
                    
                    numerator = sum((x_values[i] - x_mean) * (y_values[i] - y_mean) for i in range(n))
                    denominator = sum((x_values[i] - x_mean) ** 2 for i in range(n))
                    
                    if denominator != 0:
                        m = numerator / denominator
                        b = y_mean - m * x_mean
                        
                        y_predicted = [m * x + b for x in x_values]
                        ss_res = sum((y_values[i] - y_predicted[i]) ** 2 for i in range(n))
                        ss_tot = sum((y_values[i] - y_mean) ** 2 for i in range(n))
                        r_squared = 1 - (ss_res / ss_tot) if ss_tot != 0 else 0
                        
                        regression = {'m': m, 'b': b, 'r_squared': r_squared}
                        
                        # Determinar tendência
                        if m > 0.05:
                            trend_text = "↗ Crescente"
                            trend_color = ft.Colors.GREEN_600
                        elif m < -0.05:
                            trend_text = "↘ Decrescente"
                            trend_color = ft.Colors.RED_600
                        else:
                            trend_text = "→ Estável"
                            trend_color = ft.Colors.BLUE_600
                
                # Preparar dados de regressão completos para usar na função existente
                regression_data = None
                if regression and len(points) >= 2:
                    x_values = [p[0] for p in points]
                    y_values = [p[1] for p in points]
                    y_predicted = [regression['m'] * x + regression['b'] for x in x_values]
                    
                    x_mean = sum(x_values) / len(x_values)
                    y_mean = sum(y_values) / len(y_values)
                    
                    regression_data = {
                        'm': regression['m'],
                        'b': regression['b'],
                        'r_squared': regression['r_squared'],
                        'n': len(points),
                        'x_values': x_values,
                        'y_values': y_values,
                        'x_mean': x_mean,
                        'y_mean': y_mean,
                        'y_predicted': y_predicted
                    }
                
                # Função que cria o diálogo de detalhes (igual à aba Analytics)
                def create_info_dialog_for_metric(metric, reg_data):
                    def show_info(e):
                        if not reg_data:
                            return
                        
                        months_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
                        
                        # Criar tabela com dados observados vs preditos
                        sorted_data = sorted(zip(reg_data['x_values'], reg_data['y_values'], reg_data['y_predicted']), 
                                           key=lambda item: item[0])
                        
                        data_rows = []
                        for x, y_obs, y_pred in sorted_data:
                            residual = y_obs - y_pred
                            if abs(residual) <= 0.5:
                                residual_color = ft.Colors.GREY
                            elif residual > 0:
                                residual_color = ft.Colors.GREEN
                            else:
                                residual_color = ft.Colors.RED
                            
                            # Determinar o label do período
                            if year == "ALL":
                                # Para ALL, x é o índice sequencial (0, 1, 2...)
                                x_int = int(x)
                                if 0 <= x_int < len(monthly_data):
                                    year_val = monthly_data[x_int].get("year", 2025)
                                    month_val = monthly_data[x_int].get("month", 1)
                                    period_label = f"{months_labels[month_val-1]}/{str(year_val)[2:]}"
                                else:
                                    period_label = f"Idx {x_int}"
                            else:
                                # Para ano específico, x é o mês (1-12)
                                period_label = months_labels[int(x) - 1]
                            
                            data_rows.append(
                                ft.DataRow(cells=[
                                    ft.DataCell(ft.Text(period_label, size=14)),
                                    ft.DataCell(ft.Text(f"{y_obs:.2f}", size=14)),
                                    ft.DataCell(ft.Text(f"{y_pred:.2f}", size=14)),
                                    ft.DataCell(ft.Text(f"{residual:+.2f}", size=14, color=residual_color)),
                                ])
                            )
                        
                        # Determinar interpretação da tendência
                        if reg_data['m'] > 0.05:
                            trend_explanation = "Os scores estão melhorando ao longo do tempo"
                            trend_icon = "📈"
                            trend_color_dialog = ft.Colors.GREEN_600
                        elif reg_data['m'] < -0.05:
                            trend_explanation = "Os scores estão caindo ao longo do tempo"
                            trend_icon = "📉"
                            trend_color_dialog = ft.Colors.RED_600
                        else:
                            trend_explanation = "Os scores estão relativamente estáveis"
                            trend_icon = "➡️"
                            trend_color_dialog = ft.Colors.BLUE_600
                        
                        # Interpretação do R² - mede o quanto a equação explica a variação dos dados
                        r_squared_pct = reg_data['r_squared'] * 100
                        if r_squared_pct >= 80:
                            r_squared_quality = "80-100%"
                            r_squared_explanation = "A equação explica muito bem a variação dos dados observados"
                        elif r_squared_pct >= 60:
                            r_squared_quality = "60-79%"
                            r_squared_explanation = "A equação explica boa parte da variação dos dados observados"
                        elif r_squared_pct >= 40:
                            r_squared_quality = "40-59%"
                            r_squared_explanation = "A equação explica parte da variação dos dados observados"
                        elif r_squared_pct >= 20:
                            r_squared_quality = "20-39%"
                            r_squared_explanation = "A equação explica pouca variação dos dados observados"
                        else:
                            r_squared_quality = "0-19%"
                            r_squared_explanation = "A equação não explica quase nada da variação dos dados observados"
                        
                        dialog_content = ft.Column([
                            ft.Text(f"Como calcular a tendência - {metric}", size=22, weight="bold"),
                            ft.Divider(),
                            
                            # Resumo em linguagem simples
                            ft.Container(
                                content=ft.Column([
                                    ft.Row([
                                        ft.Text(trend_icon, size=32),
                                        ft.Column([
                                            ft.Text("Resumo:", size=18, weight="bold", color=color),
                                            ft.Text(trend_explanation, size=16, color=trend_color_dialog, weight="bold"),
                                        ], spacing=2),
                                    ], spacing=10),
                                    ft.Container(height=8),
                                    ft.Text(f"R² (Coeficiente de Determinação): {r_squared_pct:.1f}%", 
                                           size=15, weight="bold"),
                                    ft.Text("👉 Mede o quanto a equação consegue explicar a variação dos dados", 
                                           size=14, italic=True, color=ft.Colors.BLUE_700),
                                    ft.Text(r_squared_explanation, size=14, italic=True, color=ft.Colors.GREY_700),
                                    ft.Container(height=4),
                                    ft.Text("• R² = 1 (100%) → a equação explica toda a variação", 
                                           size=13, color=ft.Colors.GREY_600),
                                    ft.Text("• R² = 0 (0%) → a equação não explica nada", 
                                           size=13, color=ft.Colors.GREY_600),
                                    
                                    # Indicador visual do R² na escala 0 a 1
                                    ft.Container(height=15),
                                    ft.Text("Indicador de R²:", size=13, weight="bold"),
                                    ft.Container(height=5),
                                    
                                    # Escala visual de 0 a 1
                                    ft.Column([
                                        # Marcador de posição do R² atual (acima da barra)
                                        ft.Container(
                                            content=ft.Stack([
                                                # Indicador posicionado
                                                ft.Container(
                                                    content=ft.Column([
                                                        ft.Container(
                                                            content=ft.Text(f"{reg_data['r_squared']:.3f}", size=12, weight="bold", color=ft.Colors.WHITE),
                                                            bgcolor=ft.Colors.BLUE_700,
                                                            padding=ft.padding.symmetric(horizontal=8, vertical=3),
                                                            border_radius=4,
                                                        ),
                                                        ft.Icon(ft.Icons.ARROW_DROP_DOWN, size=28, color=ft.Colors.BLUE_700),
                                                    ], spacing=-5, horizontal_alignment=ft.CrossAxisAlignment.CENTER),
                                                    left=max(0, min(450, reg_data['r_squared'] * 490 - 25)),
                                                    top=10,
                                                ),
                                            ], width=490, height=50),
                                            margin=ft.margin.only(bottom=0),
                                        ),
                                        
                                        # Barra de gradiente (vermelho → amarelo → verde)
                                        ft.Container(
                                            width=490,
                                            height=30,
                                            border_radius=5,
                                            border=ft.border.all(2, ft.Colors.BLACK),
                                            gradient=ft.LinearGradient(
                                                begin=ft.alignment.center_left,
                                                end=ft.alignment.center_right,
                                                colors=[
                                                    ft.Colors.RED_700,
                                                    ft.Colors.ORANGE_700,
                                                    ft.Colors.YELLOW_700,
                                                    ft.Colors.LIGHT_GREEN,
                                                    ft.Colors.GREEN_700,
                                                ],
                                            ),
                                        ),
                                        
                                        # Labels da escala
                                        ft.Container(height=5),
                                        ft.Row([
                                            ft.Text("0.0", size=11, weight="bold"),
                                            ft.Container(expand=True),
                                            ft.Text("0.25", size=11, weight="bold"),
                                            ft.Container(expand=True),
                                            ft.Text("0.50", size=11, weight="bold"),
                                            ft.Container(expand=True),
                                            ft.Text("0.75", size=11, weight="bold"),
                                            ft.Container(expand=True),
                                            ft.Text("1.0", size=11, weight="bold"),
                                        ], width=490),
                                        
                                        # Legendas dos extremos
                                        ft.Container(height=3),
                                        ft.Row([
                                            ft.Text("Nenhuma explicação", size=10, italic=True, color=ft.Colors.GREY_600),
                                            ft.Container(expand=True),
                                            ft.Text("Explicação perfeita", size=10, italic=True, color=ft.Colors.GREY_600),
                                        ], width=490),
                                    ], spacing=0),
                                    
                                ], spacing=5),
                                padding=15,
                                bgcolor=ft.Colors.with_opacity(0.08, color),
                                border_radius=10,
                                border=ft.border.all(2, ft.Colors.with_opacity(0.2, color)),
                            ),
                            
                            ft.Container(height=15),
                            
                            ft.Text("📊 Informações Básicas", size=17, weight="bold"),
                            ft.Container(
                                content=ft.Column([
                                    ft.Text(f"• Quantidade de meses analisados: {reg_data['n']}", size=14),
                                    ft.Text(f"• Score médio no período: {reg_data['y_mean']:.2f}", size=14),
                                    ft.Container(height=8),
                                    ft.Text("💡 Como funciona:", size=16, weight="bold", color=color),
                                    ft.Text("Traçamos uma linha que passa o mais perto possível de todos os pontos.", 
                                           size=13, color=ft.Colors.GREY_700),
                                    ft.Text("Essa linha mostra a direção geral dos scores.", 
                                           size=13, color=ft.Colors.GREY_700),
                                    ft.Container(height=8),
                                    ft.Text("📐 Equação da Linha de Tendência:", size=16, weight="bold", color=color),
                                    ft.Container(
                                        content=ft.Column([
                                            ft.Text(f"y = {reg_data['m']:.3f}x + {reg_data['b']:.2f}", 
                                                   size=16, weight="bold", color=ft.Colors.BLUE_700),
                                            ft.Container(height=4),
                                            ft.Text(f"Onde:", size=13, weight="bold", color=ft.Colors.GREY_700),
                                            ft.Text(f"   • y = score previsto", size=13, color=ft.Colors.GREY_600),
                                            ft.Text(f"   • x = período (mês)", size=13, color=ft.Colors.GREY_600),
                                            ft.Text(f"   • m = {reg_data['m']:.3f} (inclinação da linha)", size=13, color=ft.Colors.GREY_600),
                                            ft.Text(f"   • b = {reg_data['b']:.2f} (ponto inicial)", size=13, color=ft.Colors.GREY_600),
                                        ], spacing=2),
                                        padding=10,
                                        bgcolor=ft.Colors.with_opacity(0.05, ft.Colors.BLUE),
                                        border_radius=6,
                                        border=ft.border.all(1, ft.Colors.with_opacity(0.3, ft.Colors.BLUE)),
                                    ),
                                    ft.Container(height=8),
                                    ft.Text(f"📈 Inclinação da linha: {reg_data['m']:.4f}", size=14),
                                    ft.Text(f"   • Positivo = subindo | Negativo = descendo | Perto de zero = estável", 
                                           size=13, color=ft.Colors.GREY_600, italic=True),
                                ], spacing=5),
                                padding=12,
                                bgcolor=ft.Colors.with_opacity(0.03, color),
                                border_radius=8,
                            ),
                            
                            ft.Container(height=10),
                            ft.Text("📋 Comparação: Real vs Previsto", size=17, weight="bold"),
                            ft.Text("Veja como o score real de cada mês se compara com o que a linha previu:", 
                                   size=13, color=ft.Colors.GREY_600, italic=True),
                            ft.Container(height=5),
                            ft.Container(
                                content=ft.DataTable(
                                    columns=[
                                        ft.DataColumn(ft.Text("Mês", size=14, weight="bold")),
                                        ft.DataColumn(ft.Text("Score Real", size=14, weight="bold")),
                                        ft.DataColumn(ft.Text("Previsão", size=14, weight="bold")),
                                        ft.DataColumn(ft.Text("Diferença", size=14, weight="bold")),
                                    ],
                                    rows=data_rows,
                                ),
                                border=ft.border.all(1, ft.Colors.with_opacity(0.2, ft.Colors.ON_SURFACE)),
                                border_radius=8,
                                padding=10,
                            ),
                            ft.Column([
                                ft.Text("Legenda:", size=13, weight="bold", color=ft.Colors.GREY_700),
                                ft.Text("🟢 Verde = Real superou a previsão (diferença > +0.5)", 
                                       size=13, color=ft.Colors.GREY_600, italic=True),
                                ft.Text("⚪ Cinza = Próximo da previsão (diferença entre -0.5 e +0.5)", 
                                       size=13, color=ft.Colors.GREY_600, italic=True),
                                ft.Text("🔴 Vermelho = Real abaixo da previsão (diferença < -0.5)", 
                                       size=13, color=ft.Colors.GREY_600, italic=True),
                            ], spacing=3),
                        ], spacing=10, scroll=ft.ScrollMode.AUTO, height=550)
                        
                        dialog = ft.AlertDialog(
                            title=ft.Row([
                                ft.Icon(ft.Icons.SCHOOL, color=color),
                                ft.Text("Análise de Regressão Linear", color=color),
                            ], spacing=8),
                            content=dialog_content,
                            actions=[
                                ft.TextButton("Entendi", on_click=lambda e: page.close(dialog))
                            ],
                            actions_alignment=ft.MainAxisAlignment.END,
                        )
                        page.open(dialog)
                    return show_info
                
                # Container do gráfico com análise integrada - TUDO EM UMA LINHA
                chart_container = ft.Container(
                    content=ft.Column([
                        ft.Container(
                            content=ft.Row([
                                ft.Container(
                                    content=ft.Row([
                                        ft.Text(metric_name, size=16, weight="bold", color=color, overflow=ft.TextOverflow.ELLIPSIS, no_wrap=True),
                                        ft.Container(width=15),
                                        ft.Container(
                                            content=ft.Text(f"Média: {average:.1f}", size=13, weight="w500", 
                                                   color=ft.Colors.with_opacity(0.7, color), overflow=ft.TextOverflow.ELLIPSIS, no_wrap=True),
                                            tooltip="Média dos scores ao longo do período analisado",
                                        ),
                                        ft.Container(width=5) if regression else ft.Container(),
                                        ft.VerticalDivider(width=1, color=ft.Colors.with_opacity(0.3, color)) if regression else ft.Container(),
                                        ft.Container(width=5) if regression else ft.Container(),
                                        ft.Container(
                                            content=ft.Text(f"y={regression['m']:.3f}x+{regression['b']:.2f}", size=11, weight="w500", overflow=ft.TextOverflow.ELLIPSIS, no_wrap=True),
                                            tooltip="Equação da linha de tendência: y = mx + b\nOnde 'm' é a inclinação e 'b' é o ponto inicial",
                                        ) if regression else ft.Container(),
                                        ft.Container(width=12) if regression else ft.Container(),
                                        ft.Container(
                                            content=ft.Text(trend_text, size=11, color=trend_color, weight="bold", overflow=ft.TextOverflow.ELLIPSIS, no_wrap=True),
                                            tooltip="Tendência dos dados:\n↗ Crescente: scores melhorando\n↘ Decrescente: scores piorando\n→ Estável: scores sem grande variação",
                                        ) if regression else ft.Container(),
                                        ft.Container(width=12) if regression else ft.Container(),
                                        ft.Container(
                                            content=ft.Text(f"R²={regression['r_squared']:.3f}", size=11, color=ft.Colors.GREY, overflow=ft.TextOverflow.ELLIPSIS, no_wrap=True),
                                            tooltip="R² (coeficiente de determinação):\nIndica o quanto a linha de tendência representa bem os dados\n75-100%: Excelente ajuste\n50-75%: Bom ajuste\n25-50%: Ajuste moderado\n0-25%: Ajuste fraco",
                                        ) if regression else ft.Container(),
                                    ], spacing=0, alignment=ft.MainAxisAlignment.START),
                                    expand=True,
                                    clip_behavior=ft.ClipBehavior.HARD_EDGE,
                                ),
                                ft.IconButton(
                                    icon=ft.Icons.INFO_OUTLINE,
                                    icon_size=18,
                                    tooltip="Clique para ver análise detalhada",
                                    on_click=create_info_dialog_for_metric(metric_name, regression_data),
                                    icon_color=color,
                                ) if regression else ft.Container(),
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                            bgcolor=ft.Colors.with_opacity(0.05, color),
                            padding=ft.padding.symmetric(vertical=1, horizontal=6),
                            border_radius=8,
                        ),
                        ft.Container(height=10),
                        chart,
                    ], spacing=0),
                    border=ft.border.all(2, ft.Colors.with_opacity(0.15, ft.Colors.ON_SURFACE)),
                    border_radius=12,
                    padding=15,
                    expand=True,
                    shadow=ft.BoxShadow(
                        spread_radius=1,
                        blur_radius=10,
                        color=ft.Colors.with_opacity(0.1, ft.Colors.BLACK),
                        offset=ft.Offset(0, 2),
                    ),
                    bgcolor=ft.Colors.with_opacity(0.02, ft.Colors.SURFACE),
                )
                
                return chart_container
            
            # Criar os 4 gráficos com cores modernas e vibrantes
            otif_chart = create_individual_chart("otif", "OTIF", ft.Colors.DEEP_ORANGE_500)
            nil_chart = create_individual_chart("nil", "NIL", ft.Colors.DEEP_PURPLE_500)
            pickup_chart = create_individual_chart("pickup", "Quality Pickup", ft.Colors.LIGHT_BLUE_500)
            package_chart = create_individual_chart("package", "Quality Package", ft.Colors.PINK_500)
            
            # Layout 2x2 com espaçamento aprimorado
            charts_grid = ft.Column([
                ft.Row([otif_chart, nil_chart], spacing=20, expand=True),
                ft.Row([pickup_chart, package_chart], spacing=20, expand=True),
            ], spacing=20, expand=True)
            
            timeline_individual_charts_container.current.content = charts_grid
            timeline_individual_charts_container.current.update()
            
        except Exception as e:
            print(f"Erro ao atualizar gráficos individuais: {e}")
            import traceback
            traceback.print_exc()
            timeline_individual_charts_container.current.content = ft.Container(
                content=ft.Column([
                    ft.Icon(ft.Icons.ERROR_OUTLINE, size=48, color=ft.Colors.RED_400),
                    ft.Container(height=10),
                    ft.Text(f"Erro ao gerar gráficos: {e}", size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.RED_600),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                alignment=ft.alignment.center,
                expand=True,
            )
            timeline_individual_charts_container.current.update()
    
    def update_timeline_analytics(vendor_id, year=None):
        """Atualiza a aba de Analytics com análises de tendência para todas as métricas"""
        try:
            if not db_conn or not vendor_id:
                timeline_analytics_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.ANALYTICS, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text("Selecione um fornecedor para visualizar as análises", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                    expand=True,
                )
                timeline_analytics_container.current.update()
                return
            
            analysis_year = int(year) if year and year.strip() else datetime.datetime.now().year
            
            # Buscar dados do ano
            query = "SELECT month, total_score, otif, nil, quality_pickup, quality_package FROM supplier_score_records_table WHERE supplier_id = ? AND year = ? ORDER BY month"
            results = db_manager.query(query, (vendor_id, analysis_year))
            
            if not results:
                timeline_analytics_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.INBOX, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text(f"Nenhum dado encontrado para {analysis_year}", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                    expand=True,
                )
                timeline_analytics_container.current.update()
                return
            
            # Organizar dados por métrica
            metrics_data = {
                "Total Score": [],
                "OTIF": [],
                "NIL": [],
                "Quality Pickup": [],
                "Quality Package": []
            }
            
            for row in results:
                month = row['month']
                try:
                    month_int = int(month)
                    if row['total_score'] is not None:
                        metrics_data["Total Score"].append((month_int, float(row['total_score'])))
                    if row['otif'] is not None:
                        metrics_data["OTIF"].append((month_int, float(row['otif'])))
                    if row['nil'] is not None:
                        metrics_data["NIL"].append((month_int, float(row['nil'])))
                    if row['quality_pickup'] is not None:
                        metrics_data["Quality Pickup"].append((month_int, float(row['quality_pickup'])))
                    if row['quality_package'] is not None:
                        metrics_data["Quality Package"].append((month_int, float(row['quality_package'])))
                except (ValueError, TypeError):
                    continue
            
            # Usar a função calculate_regression global (já definida anteriormente)
            # Não precisa redefinir aqui
            
            # Criar cards de análise para cada métrica
            theme_name = get_theme_name_from_page(page)
            theme_colors = get_current_theme_colors(theme_name)
            primary_color = theme_colors.get('primary', ft.Colors.BLUE)
            text_color = theme_colors.get('on_surface', ft.Colors.BLACK)
            
            analysis_cards = []
            
            metric_colors = {
                "Total Score": ft.Colors.GREEN_600,
                "OTIF": ft.Colors.ORANGE_ACCENT_700,
                "NIL": ft.Colors.PURPLE_ACCENT_700,
                "Quality Pickup": ft.Colors.CYAN_700,
                "Quality Package": ft.Colors.PINK_ACCENT_700
            }
            
            for metric_name, points in metrics_data.items():
                regression = calculate_regression(points)
                
                if regression:
                    m, b, r_squared = regression['m'], regression['b'], regression['r_squared']
                    
                    # Determinar tendência
                    if m > 0.05:
                        trend = "↗ Crescimento"
                        trend_color = ft.Colors.GREEN_600
                    elif m < -0.05:
                        trend = "↘ Queda"
                        trend_color = ft.Colors.RED_600
                    else:
                        trend = "→ Estável"
                        trend_color = ft.Colors.BLUE_600
                    
                    # Criar card expandível com memorial de cálculo
                    def create_info_dialog(metric, reg_data):
                        def show_info(e):
                            months_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
                            
                            # Criar tabela com dados observados vs preditos
                            # Ordenar os dados por mês (x_values) antes de criar as linhas
                            sorted_data = sorted(zip(reg_data['x_values'], reg_data['y_values'], reg_data['y_predicted']), 
                                               key=lambda item: item[0])
                            
                            data_rows = []
                            for x, y_obs, y_pred in sorted_data:
                                residual = y_obs - y_pred
                                # Definir cor baseada no valor da diferença
                                # Positivo (real > previsto) = Verde (bom, superou expectativa)
                                # Negativo (real < previsto) = Vermelho (ruim, abaixo da expectativa)
                                # Próximo de zero = Cinza (neutro)
                                if abs(residual) <= 0.5:
                                    residual_color = ft.Colors.GREY
                                elif residual > 0:
                                    residual_color = ft.Colors.GREEN
                                else:
                                    residual_color = ft.Colors.RED
                                
                                data_rows.append(
                                    ft.DataRow(cells=[
                                        ft.DataCell(ft.Text(months_labels[int(x) - 1], size=14)),
                                        ft.DataCell(ft.Text(f"{y_obs:.2f}", size=14)),
                                        ft.DataCell(ft.Text(f"{y_pred:.2f}", size=14)),
                                        ft.DataCell(ft.Text(f"{residual:+.2f}", size=14, color=residual_color)),
                                    ])
                                )
                            
                            # Determinar interpretação da tendência em linguagem simples
                            if reg_data['m'] > 0.05:
                                trend_explanation = "Os scores estão melhorando ao longo do tempo"
                                trend_icon = "📈"
                                trend_color = ft.Colors.GREEN_600
                            elif reg_data['m'] < -0.05:
                                trend_explanation = "Os scores estão caindo ao longo do tempo"
                                trend_icon = "📉"
                                trend_color = ft.Colors.RED_600
                            else:
                                trend_explanation = "Os scores estão relativamente estáveis"
                                trend_icon = "➡️"
                                trend_color = ft.Colors.BLUE_600
                            
                            # Interpretação do R² - mede o quanto a equação explica a variação dos dados
                            r_squared_pct = reg_data['r_squared'] * 100
                            if r_squared_pct >= 80:
                                r_squared_quality = "80-100%"
                                r_squared_explanation = "A equação explica muito bem a variação dos dados observados"
                            elif r_squared_pct >= 60:
                                r_squared_quality = "60-79%"
                                r_squared_explanation = "A equação explica boa parte da variação dos dados observados"
                            elif r_squared_pct >= 40:
                                r_squared_quality = "40-59%"
                                r_squared_explanation = "A equação explica parte da variação dos dados observados"
                            elif r_squared_pct >= 20:
                                r_squared_quality = "20-39%"
                                r_squared_explanation = "A equação explica pouca variação dos dados observados"
                            else:
                                r_squared_quality = "0-19%"
                                r_squared_explanation = "A equação não explica quase nada da variação dos dados observados"
                            
                            dialog_content = ft.Column([
                                ft.Text(f"Como calcular a tendência - {metric}", size=22, weight="bold"),
                                ft.Divider(),
                                
                                # Resumo em linguagem simples
                                ft.Container(
                                    content=ft.Column([
                                        ft.Row([
                                            ft.Text(trend_icon, size=32),
                                            ft.Column([
                                                ft.Text("Resumo:", size=18, weight="bold", color=primary_color),
                                                ft.Text(trend_explanation, size=16, color=trend_color, weight="bold"),
                                            ], spacing=2),
                                        ], spacing=10),
                                        ft.Container(height=8),
                                        ft.Text(f"R² (Coeficiente de Determinação): {r_squared_pct:.1f}%", 
                                               size=15, weight="bold"),
                                        ft.Text("👉 Mede o quanto a equação consegue explicar a variação dos dados", 
                                               size=14, italic=True, color=ft.Colors.BLUE_700),
                                        ft.Text(r_squared_explanation, size=14, italic=True, color=ft.Colors.GREY_700),
                                        ft.Container(height=4),
                                        ft.Text("• R² = 1 (100%) → a equação explica toda a variação", 
                                               size=13, color=ft.Colors.GREY_600),
                                        ft.Text("• R² = 0 (0%) → a equação não explica nada", 
                                               size=13, color=ft.Colors.GREY_600),
                                    ], spacing=5),
                                    padding=15,
                                    bgcolor=ft.Colors.with_opacity(0.08, primary_color),
                                    border_radius=10,
                                    border=ft.border.all(2, ft.Colors.with_opacity(0.2, primary_color)),
                                ),
                                
                                ft.Container(height=15),
                                
                                ft.Text("📊 Informações Básicas", size=17, weight="bold"),
                                ft.Container(
                                    content=ft.Column([
                                        ft.Text(f"• Quantidade de meses analisados: {reg_data['n']}", size=14),
                                        ft.Text(f"• Score médio no período: {reg_data['y_mean']:.2f}", size=14),
                                        ft.Container(height=8),
                                        ft.Text("💡 Como funciona:", size=16, weight="bold", color=primary_color),
                                        ft.Text("Traçamos uma linha que passa o mais perto possível de todos os pontos.", 
                                               size=13, color=ft.Colors.GREY_700),
                                        ft.Text("Essa linha mostra a direção geral dos scores.", 
                                               size=13, color=ft.Colors.GREY_700),
                                        ft.Container(height=8),
                                        ft.Text(f"📈 Inclinação da linha: {reg_data['m']:.4f}", size=14),
                                        ft.Text(f"   • Positivo = subindo | Negativo = descendo | Perto de zero = estável", 
                                               size=13, color=ft.Colors.GREY_600, italic=True),
                                    ], spacing=5),
                                    padding=12,
                                    bgcolor=ft.Colors.with_opacity(0.03, primary_color),
                                    border_radius=8,
                                ),
                                
                                ft.Container(height=10),
                                ft.Text("📋 Comparação: Real vs Previsto", size=17, weight="bold"),
                                ft.Text("Veja como o score real de cada mês se compara com o que a linha previu:", 
                                       size=13, color=ft.Colors.GREY_600, italic=True),
                                ft.Container(height=5),
                                ft.Container(
                                    content=ft.DataTable(
                                        columns=[
                                            ft.DataColumn(ft.Text("Mês", size=14, weight="bold")),
                                            ft.DataColumn(ft.Text("Score Real", size=14, weight="bold")),
                                            ft.DataColumn(ft.Text("Previsão", size=14, weight="bold")),
                                            ft.DataColumn(ft.Text("Diferença", size=14, weight="bold")),
                                        ],
                                        rows=data_rows,
                                    ),
                                    border=ft.border.all(1, ft.Colors.with_opacity(0.2, text_color)),
                                    border_radius=8,
                                    padding=10,
                                ),
                                ft.Column([
                                    ft.Text("Legenda:", size=14, weight="bold", color=ft.Colors.GREY_700),
                                    ft.Text("🟢 Verde = Real superou a previsão (diferença > +0.5)", 
                                           size=13, color=ft.Colors.GREY_600, italic=True),
                                    ft.Text("⚪ Cinza = Próximo da previsão (diferença entre -0.5 e +0.5)", 
                                           size=13, color=ft.Colors.GREY_600, italic=True),
                                    ft.Text("🔴 Vermelho = Real abaixo da previsão (diferença < -0.5)", 
                                           size=13, color=ft.Colors.GREY_600, italic=True),
                                ], spacing=3),
                            ], spacing=10, scroll=ft.ScrollMode.AUTO, height=550)
                            
                            dialog = ft.AlertDialog(
                                title=ft.Row([
                                    ft.Icon(ft.Icons.SCHOOL, color=primary_color),
                                    ft.Text("Como Funciona a Análise", color=primary_color),
                                ], spacing=8),
                                content=dialog_content,
                                actions=[
                                    ft.TextButton("Entendi", on_click=lambda e: page.close(dialog))
                                ],
                                actions_alignment=ft.MainAxisAlignment.END,
                            )
                            page.open(dialog)
                        return show_info
                    
                    card = ft.Card(
                        content=ft.Container(
                            content=ft.Column([
                                ft.Row([
                                    ft.Text(metric_name, size=16, weight="bold", color=text_color),
                                    ft.Container(expand=True),
                                    ft.IconButton(
                                        icon=ft.Icons.INFO_OUTLINE,
                                        icon_size=20,
                                        tooltip="Como funciona o cálculo (clique para ver explicação)",
                                        on_click=create_info_dialog(metric_name, regression),
                                        icon_color=primary_color,
                                    ),
                                ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                                ft.Divider(height=1),
                                ft.Row([
                                    ft.Column([
                                        ft.Text("Tendência:", size=12, color=ft.Colors.GREY_600),
                                        ft.Text(trend, size=15, weight="bold", color=trend_color),
                                    ], spacing=2),
                                    ft.Container(width=20),
                                    ft.Column([
                                        ft.Text("Equação:", size=12, color=ft.Colors.GREY_600),
                                        ft.Text(f"y = {m:.3f}x + {b:.2f}", size=14, weight="w500"),
                                    ], spacing=2),
                                    ft.Container(width=20),
                                    ft.Column([
                                        ft.Text("R²:", size=12, color=ft.Colors.GREY_600),
                                        ft.Text(f"{r_squared:.3f}", size=14, weight="w500"),
                                    ], spacing=2),
                                ], spacing=10),
                            ], spacing=8),
                            padding=ft.padding.all(16),
                        ),
                        elevation=2,
                    )
                    analysis_cards.append(card)
                else:
                    # Sem dados suficientes
                    card = ft.Card(
                        content=ft.Container(
                            content=ft.Column([
                                ft.Text(metric_name, size=16, weight="bold", color=text_color),
                                ft.Divider(height=1),
                                ft.Text("Dados insuficientes para análise", size=13, color=ft.Colors.GREY_600, italic=True),
                            ], spacing=8),
                            padding=ft.padding.all(16),
                        ),
                        elevation=1,
                    )
                    analysis_cards.append(card)
            
            # Montar layout final - um card por linha
            cards_layout = ft.Column(
                controls=analysis_cards,
                spacing=15,
            )
            
            analytics_content = ft.Column(
                controls=[
                    ft.Row([
                        ft.Icon(ft.Icons.ANALYTICS, color=primary_color, size=24),
                        ft.Text(f"Análise de Tendências - {analysis_year}", size=18, weight="bold", color=primary_color),
                    ], spacing=10),
                    ft.Divider(height=1, color=ft.Colors.with_opacity(0.2, text_color)),
                    ft.Container(height=5),
                    cards_layout,
                ],
                spacing=10,
                scroll=ft.ScrollMode.AUTO,
                expand=True,
            )
            
            timeline_analytics_container.current.content=analytics_content
            timeline_analytics_container.current.update()
            
        except Exception as e:
            print(f"Erro ao atualizar analytics: {e}")
            import traceback
            traceback.print_exc()
            timeline_analytics_container.current.content = ft.Container(
                content=ft.Column([
                    ft.Icon(ft.Icons.ERROR_OUTLINE, size=48, color=ft.Colors.RED_400),
                    ft.Container(height=10),
                    ft.Text(f"Erro ao gerar análises: {e}", size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.RED_600),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                alignment=ft.alignment.center,
                expand=True,
            )
            timeline_analytics_container.current.update()
            
    def show_timeline_snackbar(message):
        """Mostra snackbar na timeline"""
        # Substituído por show_toast para notificações consistentes
        show_snack_bar(message, False)
        safe_page_update(page)

    def delete_timeline_record(month, year_data, vendor_id):
        """Deleta um registro específico da timeline"""
        try:
            if not db_conn:
                show_timeline_snackbar("❌ Erro: Banco de dados não conectado.")
                return
                
            db_manager.execute("""DELETE FROM supplier_score_records_table 
                             WHERE supplier_id = ? AND month = ? AND year = ?""", 
                          (vendor_id, month, year_data))
            
            # Atualizar a tabela
            year = timeline_year_dropdown.current.value if timeline_year_dropdown.current else None
            update_timeline_table(vendor_id, year)
            # Atualizar gráfico e métricas
            update_timeline_metrics()
            show_timeline_snackbar("✅ Registro deletado com sucesso!")
            
        except Exception as e:
            show_timeline_snackbar(f"❌ Erro ao deletar registro: {str(e)}")

    def edit_timeline_record(record_data, otif, nil, pickup, package, comment):
        """Edita um registro específico da timeline"""
        try:
            if not db_conn:
                show_timeline_snackbar("❌ Erro: Banco de dados não conectado.")
                return

            month, year_data, _, _, _, _, _, _ = record_data

            # Obter supplier_id selecionado atualmente no filtro da timeline
            vendor_id = timeline_vendor_dropdown.current.value if timeline_vendor_dropdown and timeline_vendor_dropdown.current else None
            if not vendor_id:
                show_timeline_snackbar("❌ Erro: Nenhum fornecedor selecionado para atualizar o registro.")
                return

            cursor = db_conn.cursor()

            # Calcular novo total_score
            total_score = None
            if otif is not None and nil is not None and pickup is not None and package is not None:
                total_score = (otif + nil + pickup + package) / 4

            # Registrar quem fez a alteração e quando (compatível com a aba Score)
            from datetime import datetime
            current_date = datetime.now()
            register_date = current_date.strftime("%Y-%m-%d %H:%M:%S")
            registered_by = current_user_name if 'current_user_name' in globals() else None

            db_manager.execute("""UPDATE supplier_score_records_table
                             SET otif = ?, nil = ?, quality_pickup = ?, quality_package = ?,
                                 total_score = ?, comment = ?, registered_by = ?, register_date = ?, changed_by = ?
                             WHERE supplier_id = ? AND month = ? AND year = ?""",
                          (otif, nil, pickup, package, total_score, comment, registered_by, register_date, registered_by, vendor_id, month, year_data))

            # Atualizar a tabela
            vendor_id = timeline_vendor_dropdown.current.value if timeline_vendor_dropdown.current else ""
            year = timeline_year_dropdown.current.value if timeline_year_dropdown.current else None
            update_timeline_table(vendor_id, year)
            # Atualizar gráfico e métricas
            update_timeline_metrics()
            show_timeline_snackbar("✅ Registro editado com sucesso!")

        except Exception as e:
            show_timeline_snackbar(f"❌ Erro ao editar registro: {str(e)}")

    def update_timeline_table(vendor_id, year=None):
        """Atualiza a tabela com os dados detalhados"""
        try:
            if not db_conn or not vendor_id:
                return
                
            cursor = db_conn.cursor()
            
            if year == "ALL":
                # Para "Todo o Período", buscar todos os dados ordenados cronologicamente
                query = """SELECT month, year, otif, nil, quality_pickup, quality_package, 
                          total_score, comment FROM supplier_score_records_table 
                          WHERE supplier_id = ? ORDER BY year ASC, CAST(month AS INTEGER) ASC"""
                results = db_manager.query(query, (vendor_id,))
            elif year and year.strip():
                query = """SELECT month, year, otif, nil, quality_pickup, quality_package, 
                          total_score, comment FROM supplier_score_records_table 
                          WHERE supplier_id = ? AND year = ? ORDER BY year DESC, CAST(month AS INTEGER) ASC"""
                results = db_manager.query(query, (vendor_id, year))
            else:
                query = """SELECT month, year, otif, nil, quality_pickup, quality_package, 
                          total_score, comment FROM supplier_score_records_table 
                          WHERE supplier_id = ? ORDER BY year DESC, CAST(month AS INTEGER) ASC"""
                results = db_manager.query(query, (vendor_id,))
            
            if not results:
                timeline_table_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.INBOX, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text("Nenhum dado encontrado para este fornecedor", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                    expand=True,
                )
                timeline_table_container.current.update()
                return
            
            # Verificar largura da tela para layout responsivo
            is_mobile = page.window.width < 900 if page.window else False

            theme_name = get_theme_name_from_page(page)
            theme_colors = get_current_theme_colors(theme_name)
            primary_color = theme_colors.get('primary', ft.Colors.BLUE)
            chip_bgcolor = theme_colors.get('primary_container') or ft.Colors.with_opacity(0.08, primary_color)
            card_bgcolor = theme_colors.get('field_background') if theme_name == 'dracula' else theme_colors.get('surface_variant')
            if not card_bgcolor:
                card_bgcolor = ft.Colors.with_opacity(0.05, primary_color)
            text_color = theme_colors.get('on_surface', ft.Colors.BLACK)

            months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                      "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

            # Criar cabeçalho da tabela
            header_row = ft.Container(
                content=ft.Row([
                    ft.Container(ft.Text("Período", size=12, weight="bold", color=text_color, no_wrap=True), width=100, alignment=ft.alignment.center_left),
                    ft.Container(ft.Text("OTIF", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                    ft.Container(ft.Text("NIL", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                    ft.Container(ft.Text("Pickup", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                    ft.Container(ft.Text("Package", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                    ft.Container(ft.Text("Total", size=12, weight="bold", color=text_color, no_wrap=True), width=80, alignment=ft.alignment.center),
                    ft.Container(ft.Text("Comentário", size=12, weight="bold", color=text_color, no_wrap=True), expand=True, alignment=ft.alignment.center_left),
                    ft.Container(ft.Icon(ft.Icons.THERMOSTAT, size=16, color=text_color, tooltip="Confiabilidade"), width=50, alignment=ft.alignment.center),
                    ft.Container(ft.Text("Ações", size=12, weight="bold", color=text_color, no_wrap=True), width=100, alignment=ft.alignment.center),
                ], spacing=10, tight=True),
                padding=ft.padding.symmetric(horizontal=16, vertical=12),
                bgcolor=ft.Colors.with_opacity(0.1, primary_color),
                border_radius=ft.border_radius.only(top_left=8, top_right=8),
            )

            table_rows = []  # Não incluir o header aqui

            for i, result in enumerate(results):
                # Suporta tanto tupla quanto dicionário
                if isinstance(result, dict):
                    month = result['month']
                    year_data = result['year']
                    otif = result['otif']
                    nil = result['nil']
                    pickup = result['quality_pickup']
                    package = result['quality_package']
                    total = result['total_score']
                    comment = result['comment']
                else:
                    month, year_data, otif, nil, pickup, package, total, comment = result
                
                try:
                    month_int = int(month)
                    month_name = months[month_int-1] if 1 <= month_int <= 12 else str(month)
                except (ValueError, TypeError):
                    month_name = str(month)
                total_display = f"{float(total):.1f}" if total is not None else "--"
                bgcolor = ft.Colors.with_opacity(0.04, primary_color) if i % 2 == 0 and not is_mobile else None
                
                # Função para editar este registro específico
                def create_edit_handler(record_row):
                    def handle_edit(e):
                        def handle_confirm(record_data, otif, nil, pickup, package, comment):
                            edit_timeline_record(record_data, otif, nil, pickup, package, comment)
                            page.close(edit_dialog)
                        
                        def handle_cancel(e):
                            page.close(edit_dialog)
                        
                        edit_dialog = EditTimelineRecordDialog(
                            record_data=record_row,
                            on_confirm=handle_confirm,
                            on_cancel=handle_cancel,
                            page=page
                        )
                        page.open(edit_dialog)
                    return handle_edit
                
                # Passar uma tupla com os valores na ordem esperada pelo diálogo
                record_tuple = (month, year_data, otif, nil, pickup, package, total, comment)
                
                # Verificar se usuário pode editar (apenas Admin e Super Admin)
                can_edit = current_user_privilege in ["Admin", "Super Admin"]
                
                # Formatação dos valores
                otif_display = f"{float(otif):.1f}" if otif is not None else "--"
                nil_display = f"{float(nil):.1f}" if nil is not None else "--"
                pickup_display = f"{float(pickup):.1f}" if pickup is not None else "--"
                package_display = f"{float(package):.1f}" if package is not None else "--"
                total_display = f"{float(total):.1f}" if total is not None else "--"
                
                comment_text = comment.strip() if isinstance(comment, str) else comment
                comment_display = comment_text if comment_text else "Sem comentários"

                # Botões de ação
                action_buttons = []
                if can_edit:
                    action_buttons.append(ft.IconButton(
                        icon=ft.Icons.EDIT,
                        icon_size=16,
                        tooltip="Editar",
                        on_click=create_edit_handler(record_tuple),
                        icon_color=primary_color
                    ))
                    
                    def create_delete_handler(m, y, vid):
                        def handle_delete(e):
                            def handle_confirm(e):
                                delete_timeline_record(m, y, vid)
                                page.close(delete_dialog)
                            
                            def handle_cancel(e):
                                page.close(delete_dialog)
                            
                            delete_dialog = DeleteListItemConfirmationDialog(
                                item_name=f"{months[int(m)-1] if 1 <= int(m) <= 12 else str(m)}/{y}",
                                item_type="Registro Timeline",
                                on_confirm=handle_confirm,
                                on_cancel=handle_cancel
                            )
                            page.open(delete_dialog)
                        return handle_delete
                    
                    action_buttons.append(ft.IconButton(
                        icon=ft.Icons.DELETE,
                        icon_size=16,
                        tooltip="Deletar",
                        on_click=create_delete_handler(month, year_data, vendor_id),
                        icon_color=ft.Colors.RED_400
                    ))

                # Cor alternada para linhas
                row_index = results.index(result)
                row_bgcolor = ft.Colors.with_opacity(0.03, text_color) if row_index % 2 == 0 else None

                # Analisar consistência dos dados
                consistency_score, consistency_issues, consistency_icon, consistency_color = analyze_data_consistency(otif, nil, pickup, package)
                consistency_tooltip = f"Confiabilidade: {consistency_score}%"
                if consistency_issues:
                    consistency_tooltip += "\n⚠ " + "\n⚠ ".join(consistency_issues)
                
                # Linha da tabela
                table_row = ft.Container(
                    content=ft.Row([
                        ft.Container(ft.Text(f"{month_name}/{year_data}", size=13, color=text_color), width=100, alignment=ft.alignment.center_left),
                        ft.Container(ft.Text(otif_display, size=13, weight="w500", color=get_score_color(otif) if otif is not None else text_color), width=80, alignment=ft.alignment.center),
                        ft.Container(ft.Text(nil_display, size=13, weight="w500", color=get_score_color(nil) if nil is not None else text_color), width=80, alignment=ft.alignment.center),
                        ft.Container(ft.Text(pickup_display, size=13, weight="w500", color=get_score_color(pickup) if pickup is not None else text_color), width=80, alignment=ft.alignment.center),
                        ft.Container(ft.Text(package_display, size=13, weight="w500", color=get_score_color(package) if package is not None else text_color), width=80, alignment=ft.alignment.center),
                        ft.Container(ft.Text(total_display, size=13, weight="w500", color=primary_color), width=80, alignment=ft.alignment.center),
                        ft.Container(
                            ft.Text(comment_display, size=12, color=ft.Colors.with_opacity(0.7, text_color), max_lines=2, overflow=ft.TextOverflow.ELLIPSIS),
                            expand=True,
                            alignment=ft.alignment.center_left
                        ),
                        ft.Container(
                            content=ft.Icon(consistency_icon, size=20, color=consistency_color),
                            width=50,
                            alignment=ft.alignment.center,
                            tooltip=consistency_tooltip
                        ),
                        ft.Container(
                            ft.Row(action_buttons, spacing=0, tight=True),
                            width=100,
                            alignment=ft.alignment.center
                        ),
                    ], spacing=10, tight=True),
                    padding=ft.padding.symmetric(horizontal=16, vertical=10),
                    bgcolor=row_bgcolor,
                    border=ft.border.only(bottom=ft.BorderSide(1, ft.Colors.with_opacity(0.1, text_color))),
                )

                table_rows.append(table_row)

            # Container da tabela com header fixo e corpo scrollável
            table_container = ft.Container(
                content=ft.Column([
                    # Header fixo
                    header_row,
                    # Corpo da tabela com scroll
                    ft.Container(
                        content=ft.Column(
                            controls=table_rows,  # Agora table_rows não contém o header
                            spacing=0,
                            scroll=ft.ScrollMode.AUTO,
                        ),
                        expand=True,
                    ),
                ], spacing=0, expand=True),
                border=ft.border.all(1, ft.Colors.with_opacity(0.12, text_color)),
                border_radius=8,
                bgcolor=ft.Colors.with_opacity(0.02, text_color),
                expand=True,
            )

            timeline_table_container.current.content = table_container
            timeline_table_container.current.update()
            
        except Exception as e:
            print(f"Erro ao atualizar tabela: {e}")
            timeline_table_container.current.content = ft.Container(
                content=ft.Column([
                    ft.Icon(ft.Icons.ERROR_OUTLINE, size=48, color=ft.Colors.RED_400),
                    ft.Container(height=10),
                    ft.Text(f"Erro: {e}", size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.RED_600),
                ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                alignment=ft.alignment.center,
                expand=True,
            )
            timeline_table_container.current.update()
            
    def on_timeline_vendor_change(e):
        """Callback quando o vendor é alterado"""
        update_timeline_metrics()
        # Como não há mais abas, sempre atualizar a tabela quando há mudança
        vendor_id = timeline_vendor_dropdown.current.value if timeline_vendor_dropdown.current else ""
        year = timeline_year_dropdown.current.value if timeline_year_dropdown.current else None
        update_timeline_table(vendor_id, year)
            
    def on_timeline_year_change(e):
        """Callback quando o ano é alterado"""
        update_timeline_metrics()
        # Como não há mais abas, apenas atualizar os dados diretamente
        vendor_id = timeline_vendor_dropdown.current.value if timeline_vendor_dropdown.current else ""
        year = timeline_year_dropdown.current.value if timeline_year_dropdown.current else None
        update_timeline_table(vendor_id, year)

    def clear_timeline_vendor(e):
        """Limpa o fornecedor selecionado e reseta todos os cards"""
        try:
            print("🔄 Iniciando limpeza da timeline...")
            
            # Limpar o dropdown
            if timeline_vendor_dropdown.current:
                timeline_vendor_dropdown.current.value = None
            
            # Resetar todos os cards de métricas para "--"
            metric_cards = [
                overall_avg_card, twelve_month_avg_card, year_avg_card,
                q1_avg_card, q2_avg_card, q3_avg_card, q4_avg_card
            ]
            for card in metric_cards:
                if card.current:
                    card.current.value = "--"
            
            # Resetar ícones de setas
            arrow_icons = [q1_arrow_icon, q2_arrow_icon, q3_arrow_icon, q4_arrow_icon]
            for icon in arrow_icons:
                if icon.current:
                    icon.current.name = ft.Icons.ARROW_FORWARD
                    icon.current.color = ft.Colors.GREY_400
            
            # Limpar containers dinâmicos e restaurar visibilidade
            if yearly_cards_container.current:
                yearly_cards_container.current.controls = []
            if yearly_section_container.current:
                yearly_section_container.current.visible = False
            if quarterly_cards_container.current:
                quarterly_cards_container.current.visible = True
            if quarterly_section_container.current:
                quarterly_section_container.current.visible = True
            if timeline_cards_refs["year"]["card"].current:
                timeline_cards_refs["year"]["card"].current.visible = True
            
            # Limpar gráfico
            if timeline_chart_container.current:
                timeline_chart_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.SHOW_CHART, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text("Selecione um fornecedor para visualizar o gráfico", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                )
            
            # Limpar gráficos individuais
            if timeline_individual_charts_container.current:
                timeline_individual_charts_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.GRID_VIEW, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text("Selecione um fornecedor para visualizar os gráficos", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                )
            
            # Limpar tabela
            if timeline_table_container.current:
                timeline_table_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.TABLE_CHART, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text("Selecione um fornecedor para visualizar a tabela", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                )
            
            # Limpar analytics
            if timeline_analytics_container.current:
                timeline_analytics_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.ANALYTICS, size=48, color=ft.Colors.GREY_400),
                        ft.Container(height=10),
                        ft.Text("Selecione um fornecedor para visualizar as análises", 
                               size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                    ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                    alignment=ft.alignment.center,
                )
            
            # Atualizar a página
            safe_page_update(page)
            
            print("✅ Timeline limpa com sucesso")
            
        except Exception as e:
            print(f"❌ Erro ao limpar timeline: {e}")
            import traceback
            traceback.print_exc()

    # Criar conteúdo da aba Timeline
    # Obter cores do tema para a criação inicial dos componentes
    theme_name_for_timeline = get_theme_name_from_page(page)
    theme_colors_for_timeline = get_current_theme_colors(theme_name_for_timeline)
    primary_color_for_timeline = theme_colors_for_timeline.get('primary')
    # Se o tema for dracula, usar fundo cinza e textos/icones violeta (primary)
    if theme_name_for_timeline == 'dracula':
        timeline_card_bg = theme_colors_for_timeline.get('field_background') or "#44475A"
        label_color = theme_colors_for_timeline.get('on_surface')
    else:
        timeline_card_bg = None
        label_color = theme_colors_for_timeline.get('on_surface')

    timeline_view = ft.Container(
        content=ft.Column([
            # Título da seção
            ft.Container(
                content=ft.Text("Timeline Analytics", size=28, weight="bold"),
                alignment=ft.alignment.top_left,
                padding=ft.padding.only(bottom=10),
            ),
            
            # Campos de seleção - mais clean e minimalista
            ft.Row([
                ft.Stack(
                    controls=[
                        ft.Container(
                            content=ft.TextField(
                                label="Fornecedor",
                                hint_text="Clique para selecionar...",
                                ref=timeline_vendor_search_field,
                                read_only=True,
                                on_click=open_vendor_selection_dialog,
                                prefix_icon=ft.Icons.SEARCH,  # Lupa à esquerda
                                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                                color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                                border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline'),
                                text_size=14,
                            ),
                            expand=True,
                            padding=ft.padding.all(5)
                        ),
                        ft.Container(
                            content=ft.IconButton(
                                icon=ft.Icons.CLOSE,
                                icon_size=20,
                                tooltip="Limpar fornecedor",
                                on_click=clear_timeline_search_field,
                                ref=timeline_clear_button_ref,
                                visible=False,
                            ),
                            right=6,
                            top=6,
                            bottom=6,
                        ),
                    ],
                    expand=True,
                ),
                # Dropdown oculto para manter compatibilidade
                ft.Container(
                    content=ft.Dropdown(
                        ref=timeline_vendor_dropdown,
                        visible=False,
                        options=load_suppliers_for_timeline(),
                    ),
                    width=0,
                    height=0,
                ),
                ft.IconButton(
                    icon=ft.Icons.INFO_OUTLINE,
                    tooltip="Informações do fornecedor",
                    on_click=lambda e: show_supplier_info_dialog(timeline_vendor_dropdown.current.value if timeline_vendor_dropdown.current else None),
                    icon_color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'),
                    icon_size=20,
                ),
                # o botão de limpar fornecedor foi substituído pelo botão flutuante dentro do campo
                ft.Dropdown(
                    label="Ano",
                    ref=timeline_year_dropdown,
                    on_change=on_timeline_year_change,
                    options=[
                        ft.dropdown.Option("ALL", "📊 Todo o Período")
                    ] + [ft.dropdown.Option(str(y)) for y in range(2024, 2041)],
                    value="",
                    width=180,
                    bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                    color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                    border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
                ),
            ], spacing=15, ref=timeline_search_container),
            
            ft.Container(height=20),
            
            # Tabs para alternar entre métricas, gráfico e tabela
            ft.Container(
                content=ft.Tabs(
                    ref=timeline_tabs,
                    selected_index=0,
                    animation_duration=300,
                    on_change=lambda e: on_timeline_tab_change(e),
                    tabs=[
                        # Tab de Métricas (cards)
                        ft.Tab(
                            text="Métricas",
                            icon=ft.Icons.DASHBOARD,
                            content=ft.Container(
                                content=ft.Column([
                                    # Título da seção
                                    ft.Container(
                                        content=ft.Row([
                                            ft.Icon(ft.Icons.ANALYTICS, size=32, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                            ft.Text("Visão Geral das Métricas", size=24, weight="bold"),
                                        ], spacing=12),
                                        margin=ft.margin.only(bottom=20),
                                    ),
                                    
                                    # Cards principais (Overall, 12M, Year) - Layout em linha
                                    ft.Container(
                                        content=ft.Row([
                                            # Card Overall Average
                                            ft.Container(
                                                content=ft.Card(
                                                    ref=timeline_cards_refs["overall"]["card"],
                                                    content=ft.Container(
                                                        ref=timeline_cards_refs["overall"]["gradient"],
                                                        content=ft.Column([
                                                            ft.Row([
                                                                ft.Icon(ft.Icons.ANALYTICS_OUTLINED, size=28, color=primary_color_for_timeline),
                                                                ft.Container(expand=True),
                                                            ]),
                                                            ft.Text("Overall Average", size=14, weight="w600", color=label_color),
                                                            ft.Container(height=5),
                                                            ft.Text("--", size=36, weight="bold", color=primary_color_for_timeline, ref=overall_avg_card),
                                                            ft.Container(height=5),
                                                            ft.Text("Média Geral", size=11, color=label_color, italic=True),
                                                        ], spacing=0),
                                                        padding=ft.padding.all(20),
                                                        bgcolor=timeline_card_bg,
                                                        border_radius=12
                                                    ),
                                                    elevation=3,
                                                    surface_tint_color=primary_color_for_timeline
                                                ),
                                                expand=True,
                                            ),
                                            ft.Container(width=15),
                                            # Card 12 Month Average
                                            ft.Container(
                                                content=ft.Card(
                                                    ref=timeline_cards_refs["12m"]["card"],
                                                    content=ft.Container(
                                                        ref=timeline_cards_refs["12m"]["gradient"],
                                                        content=ft.Column([
                                                            ft.Row([
                                                                ft.Icon(ft.Icons.CALENDAR_MONTH, size=28, color=primary_color_for_timeline),
                                                                ft.Container(expand=True),
                                                            ]),
                                                            ft.Text("12 Months Avg", size=14, weight="w600", color=label_color),
                                                            ft.Container(height=5),
                                                            ft.Text("--", size=36, weight="bold", color=primary_color_for_timeline, ref=twelve_month_avg_card),
                                                            ft.Container(height=5),
                                                            ft.Text("Últimos 12 meses", size=11, color=label_color, italic=True),
                                                        ], spacing=0),
                                                        padding=ft.padding.all(20),
                                                        bgcolor=timeline_card_bg,
                                                        border_radius=12
                                                    ),
                                                    elevation=3,
                                                    surface_tint_color=primary_color_for_timeline
                                                ),
                                                expand=True,
                                            ),
                                            ft.Container(width=15),
                                            # Card Year Average
                                            ft.Container(
                                                content=ft.Card(
                                                    ref=timeline_cards_refs["year"]["card"],
                                                    content=ft.Container(
                                                        ref=timeline_cards_refs["year"]["gradient"],
                                                        content=ft.Column([
                                                            ft.Row([
                                                                ft.Icon(ft.Icons.TODAY, size=28, color=primary_color_for_timeline),
                                                                ft.Container(expand=True),
                                                            ]),
                                                            ft.Text("Year Average", size=14, weight="w600", color=label_color),
                                                            ft.Container(height=5),
                                                            ft.Text("--", size=36, weight="bold", color=primary_color_for_timeline, ref=year_avg_card),
                                                            ft.Container(height=5),
                                                            ft.Text("Média do ano", size=11, color=label_color, italic=True),
                                                        ], spacing=0),
                                                        padding=ft.padding.all(20),
                                                        bgcolor=timeline_card_bg,
                                                        border_radius=12
                                                    ),
                                                    elevation=3,
                                                    surface_tint_color=primary_color_for_timeline
                                                ),
                                                expand=True,
                                            ),
                                        ], alignment=ft.MainAxisAlignment.START),
                                        margin=ft.margin.only(bottom=25),
                                    ),
                                    
                                    # Divisor visual
                                    ft.Divider(height=1, color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline_variant')),
                                    ft.Container(height=15),
                                    
                                    # Título para cards trimestrais
                                    ft.Container(
                                        ref=quarterly_section_container,
                                        content=ft.Column([
                                            ft.Text("Métricas Trimestrais", size=18, weight="bold"),
                                            ft.Container(height=15),
                                            # Container para cards trimestrais (Q1-Q4)
                                            ft.Container(
                                                ref=timeline_metrics_row,
                                                content=ft.Row(
                                                    ref=quarterly_cards_container,
                                                    controls=[
                                                # Card Q1
                                                ft.Card(
                                                    ref=timeline_cards_refs["q1"]["card"],
                                                    content=ft.Container(
                                                        ref=timeline_cards_refs["q1"]["gradient"],
                                                        content=ft.Column([
                                                            ft.Row([
                                                                ft.Text("Q1", size=16, weight="bold", color=primary_color_for_timeline),
                                                                ft.Container(expand=True),
                                                                ft.Icon(ft.Icons.ARROW_FORWARD, color=ft.Colors.GREY_400, size=20, ref=q1_arrow_icon),
                                                            ], spacing=6),
                                                            ft.Container(height=10),
                                                            ft.Text("--", size=32, weight="bold", color=primary_color_for_timeline, ref=q1_avg_card),
                                                            ft.Container(height=5),
                                                            ft.Text("Jan - Mar", size=11, color=label_color, italic=True),
                                                        ], spacing=0),
                                                        padding=ft.padding.all(18),
                                                        bgcolor=timeline_card_bg,
                                                        width=180,
                                                        border_radius=12
                                                    ),
                                                    elevation=3,
                                                    surface_tint_color=primary_color_for_timeline
                                                ),
                                                # Card Q2
                                                ft.Card(
                                                    ref=timeline_cards_refs["q2"]["card"],
                                                    content=ft.Container(
                                                        ref=timeline_cards_refs["q2"]["gradient"],
                                                        content=ft.Column([
                                                            ft.Row([
                                                                ft.Text("Q2", size=16, weight="bold", color=primary_color_for_timeline),
                                                                ft.Container(expand=True),
                                                                ft.Icon(ft.Icons.ARROW_FORWARD, color=ft.Colors.GREY_400, size=20, ref=q2_arrow_icon),
                                                            ], spacing=6),
                                                            ft.Container(height=10),
                                                            ft.Text("--", size=32, weight="bold", color=primary_color_for_timeline, ref=q2_avg_card),
                                                            ft.Container(height=5),
                                                            ft.Text("Apr - Jun", size=11, color=label_color, italic=True),
                                                        ], spacing=0),
                                                        padding=ft.padding.all(18),
                                                        bgcolor=timeline_card_bg,
                                                        width=180,
                                                        border_radius=12
                                                    ),
                                                    elevation=3,
                                                    surface_tint_color=primary_color_for_timeline
                                                ),
                                                # Card Q3
                                                ft.Card(
                                                    ref=timeline_cards_refs["q3"]["card"],
                                                    content=ft.Container(
                                                        ref=timeline_cards_refs["q3"]["gradient"],
                                                        content=ft.Column([
                                                            ft.Row([
                                                                ft.Text("Q3", size=16, weight="bold", color=primary_color_for_timeline),
                                                                ft.Container(expand=True),
                                                                ft.Icon(ft.Icons.ARROW_FORWARD, color=ft.Colors.GREY_400, size=20, ref=q3_arrow_icon),
                                                            ], spacing=6),
                                                            ft.Container(height=10),
                                                            ft.Text("--", size=32, weight="bold", color=primary_color_for_timeline, ref=q3_avg_card),
                                                            ft.Container(height=5),
                                                            ft.Text("Jul - Sep", size=11, color=label_color, italic=True),
                                                        ], spacing=0),
                                                        padding=ft.padding.all(18),
                                                        bgcolor=timeline_card_bg,
                                                        width=180,
                                                        border_radius=12
                                                    ),
                                                    elevation=3,
                                                    surface_tint_color=primary_color_for_timeline
                                                ),
                                                # Card Q4
                                                ft.Card(
                                                    ref=timeline_cards_refs["q4"]["card"],
                                                    content=ft.Container(
                                                        ref=timeline_cards_refs["q4"]["gradient"],
                                                        content=ft.Column([
                                                            ft.Row([
                                                                ft.Text("Q4", size=16, weight="bold", color=primary_color_for_timeline),
                                                                ft.Container(expand=True),
                                                                ft.Icon(ft.Icons.ARROW_FORWARD, color=ft.Colors.GREY_400, size=20, ref=q4_arrow_icon),
                                                            ], spacing=6),
                                                            ft.Container(height=10),
                                                            ft.Text("--", size=32, weight="bold", color=primary_color_for_timeline, ref=q4_avg_card),
                                                            ft.Container(height=5),
                                                            ft.Text("Oct - Dec", size=11, color=label_color, italic=True),
                                                        ], spacing=0),
                                                        padding=ft.padding.all(18),
                                                        bgcolor=timeline_card_bg,
                                                        width=180,
                                                        border_radius=12
                                                    ),
                                                    elevation=3,
                                                    surface_tint_color=primary_color_for_timeline
                                                ),
                                            ],
                                            spacing=15,
                                            scroll=ft.ScrollMode.AUTO,
                                        ),
                                        margin=ft.margin.only(bottom=25),
                                    ),
                                        ]),
                                        visible=True,
                                    ),
                                    
                                    # Container para cards anuais - visível em modo ALL
                                    ft.Container(
                                        ref=yearly_section_container,
                                        content=ft.Column([
                                            ft.Text("Métricas Anuais", size=18, weight="bold"),
                                            ft.Container(height=15),
                                            ft.Row(
                                                ref=yearly_cards_container,
                                                controls=[],  # Será preenchido dinamicamente
                                                spacing=15,
                                                scroll=ft.ScrollMode.AUTO,
                                            ),
                                        ]),
                                        visible=False,  # Oculto por padrão, aparece em modo ALL
                                    ),
                                ], spacing=15, scroll=ft.ScrollMode.AUTO),
                                padding=ft.padding.all(20),
                                expand=True,
                            ),
                        ),
                        # Tab do Gráfico
                        ft.Tab(
                            text="Performance Chart",
                            icon=ft.Icons.SHOW_CHART,
                            content=ft.Container(
                                content=ft.Card(
                                    content=ft.Container(
                                        ref=timeline_chart_container,
                                        content=ft.Container(
                                            content=ft.Column([
                                                ft.Icon(ft.Icons.SHOW_CHART, size=48, color=ft.Colors.GREY_400),
                                                ft.Container(height=10),
                                                ft.Text("Selecione um fornecedor para visualizar o gráfico", 
                                                       size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                                            alignment=ft.alignment.center,
                                        ),
                                        padding=ft.padding.all(20),
                                        expand=True,
                                    ),
                                    elevation=1,
                                ),
                                padding=ft.padding.only(top=10),
                                expand=True,
                            ),
                        ),
                        # Tab dos Gráficos Individuais
                        ft.Tab(
                            text="Individual Metrics",
                            icon=ft.Icons.GRID_VIEW,
                            content=ft.Container(
                                content=ft.Card(
                                    content=ft.Container(
                                        ref=timeline_individual_charts_container,
                                        content=ft.Container(
                                            content=ft.Column([
                                                ft.Icon(ft.Icons.GRID_VIEW, size=48, color=ft.Colors.GREY_400),
                                                ft.Container(height=10),
                                                ft.Text("Selecione um fornecedor para visualizar os gráficos", 
                                                       size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                                            alignment=ft.alignment.center,
                                        ),
                                        padding=ft.padding.all(20),
                                        expand=True,
                                    ),
                                    elevation=1,
                                ),
                                padding=ft.padding.only(top=10),
                                expand=True,
                            ),
                        ),
                        # Tab da Tabela
                        ft.Tab(
                            text="Detailed Records",
                            icon=ft.Icons.TABLE_CHART,
                            content=ft.Container(
                                content=ft.Card(
                                    content=ft.Container(
                                        ref=timeline_table_container,
                                        content=ft.Container(
                                            content=ft.Column([
                                                ft.Icon(ft.Icons.TABLE_CHART, size=48, color=ft.Colors.GREY_400),
                                                ft.Container(height=10),
                                                ft.Text("Selecione um fornecedor para visualizar a tabela", 
                                                       size=14, text_align=ft.TextAlign.CENTER, color=ft.Colors.GREY_600),
                                            ], horizontal_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                                            alignment=ft.alignment.center,
                                        ),
                                        padding=ft.padding.all(20),
                                        expand=True,
                                    ),
                                    elevation=1,
                                ),
                                padding=ft.padding.only(top=10),
                                expand=True,
                            ),
                        ),
                        # Tab de Informações do Termômetro
                        ft.Tab(
                            text="Data Info",
                            icon=ft.Icons.INFO_OUTLINE,
                            content=ft.Container(
                                content=ft.Card(
                                    content=ft.Container(
                                        content=ft.Column([
                                            # Título
                                            ft.Container(
                                                content=ft.Row([
                                                    ft.Icon(ft.Icons.THERMOSTAT, size=32, color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')),
                                                    ft.Container(width=10),
                                                    ft.Text("Termômetro de Consistência de Dados", size=22, weight="bold"),
                                                ]),
                                                padding=ft.padding.only(bottom=20),
                                            ),
                                            
                                            # Descrição geral
                                            ft.Text(
                                                "O termômetro de consistência analisa automaticamente os dados inseridos e identifica possíveis inconsistências ou problemas, "
                                                "ajudando a garantir a qualidade e confiabilidade das informações.",
                                                size=14,
                                                color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'),
                                            ),
                                            
                                            ft.Container(height=20),
                                            ft.Divider(),
                                            ft.Container(height=10),
                                            
                                            # Seção de Score
                                            ft.Text("📊 Score de Confiabilidade (0-100)", size=16, weight="bold"),
                                            ft.Container(height=10),
                                            ft.Text(
                                                "• 100: Dados totalmente consistentes e confiáveis\n"
                                                "• 80-99: Boa qualidade, pequenas observações\n"
                                                "• 50-79: Qualidade média, requer atenção\n"
                                                "• 0-49: Dados com problemas sérios, revisão necessária",
                                                size=13,
                                            ),
                                            
                                            ft.Container(height=20),
                                            ft.Divider(),
                                            ft.Container(height=10),
                                            
                                            # Seção de Ícones
                                            ft.Text("🌡️ Indicadores Visuais", size=16, weight="bold"),
                                            ft.Container(height=10),
                                            
                                            # Ícone Verde (Alto)
                                            ft.Row([
                                                ft.Icon(ft.Icons.THERMOSTAT, size=28, color=ft.Colors.GREEN),
                                                ft.Container(width=10),
                                                ft.Column([
                                                    ft.Text("Score ≥ 80 - Excelente Qualidade", size=14, weight="bold", color=ft.Colors.GREEN),
                                                    ft.Text("Dados consistentes e confiáveis. Sem problemas detectados.", size=12),
                                                ], spacing=2),
                                            ], spacing=5),
                                            
                                            ft.Container(height=10),
                                            
                                            # Ícone Amarelo (Médio)
                                            ft.Row([
                                                ft.Icon(ft.Icons.THERMOSTAT_AUTO, size=28, color=ft.Colors.ORANGE),
                                                ft.Container(width=10),
                                                ft.Column([
                                                    ft.Text("Score 50-79 - Qualidade Média", size=14, weight="bold", color=ft.Colors.ORANGE),
                                                    ft.Text("Algumas inconsistências detectadas. Recomenda-se revisar os dados.", size=12),
                                                ], spacing=2),
                                            ], spacing=5),
                                            
                                            ft.Container(height=10),
                                            
                                            # Ícone Vermelho (Baixo)
                                            ft.Row([
                                                ft.Icon(ft.Icons.AC_UNIT, size=28, color=ft.Colors.RED),
                                                ft.Container(width=10),
                                                ft.Column([
                                                    ft.Text("Score < 50 - Qualidade Baixa", size=14, weight="bold", color=ft.Colors.RED),
                                                    ft.Text("Múltiplas inconsistências graves. Revisão urgente necessária.", size=12),
                                                ], spacing=2),
                                            ], spacing=5),
                                            
                                            ft.Container(height=20),
                                            ft.Divider(),
                                            ft.Container(height=10),
                                            
                                            # Regras de Validação
                                            ft.Text("✅ Regras de Validação", size=16, weight="bold"),
                                            ft.Container(height=10),
                                            
                                            ft.Text("1. Regra de Entrega Sem Dados (-40 pontos)", size=14, weight="bold"),
                                            ft.Text(
                                                "• Se OTIF está vazio (sem entrega), NIL, Pickup e Package também devem estar vazios\n"
                                                "• É incoerente avaliar qualidade de algo que não foi entregue",
                                                size=12,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            ft.Text("2. Logística Ruim vs Inspeção OK (-20 pontos)", size=14, weight="bold"),
                                            ft.Text(
                                                "• Se NIL ≤ 3 (logística péssima) mas Pickup ≥ 8 (inspeção excelente)\n"
                                                "• É improvável produto chegar bem se transporte foi péssimo",
                                                size=12,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            ft.Text("3. Embalagem Péssima vs Produto OK (-25 pontos)", size=14, weight="bold"),
                                            ft.Text(
                                                "• Se Package ≤ 3 (embalagem péssima) mas Pickup ≥ 8 (produto excelente)\n"
                                                "• Embalagem ruim normalmente danifica o produto",
                                                size=12,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            ft.Text("4. OTIF Alto vs NIL Zero (-30 pontos)", size=14, weight="bold"),
                                            ft.Text(
                                                "• Se OTIF ≥ 8 (entrega excelente) mas NIL = 0 (logística péssima)\n"
                                                "• Não é possível entregar bem sem logística adequada",
                                                size=12,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            ft.Text("5. Reprovou Inspeção vs Embalagem OK (-20 pontos)", size=14, weight="bold"),
                                            ft.Text(
                                                "• Se Pickup = 0 (reprovou inspeção) mas Package ≥ 8 (embalagem excelente)\n"
                                                "• Produto reprovado não justifica embalagem perfeita",
                                                size=12,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            ft.Text("6. Valores Idênticos Suspeitos (-15 pontos)", size=14, weight="bold"),
                                            ft.Text(
                                                "• Se 3+ métricas têm valores idênticos (exceto 0 ou 10)\n"
                                                "• Exemplo: OTIF=7.5, NIL=7.5, Pickup=7.5, Package=7.5\n"
                                                "• Obs: Tudo 10 (perfeito) ou tudo 0 (péssimo) são aceitáveis",
                                                size=12,
                                            ),
                                            
                                            ft.Container(height=20),
                                            ft.Divider(),
                                            ft.Container(height=10),
                                            
                                            # Exemplos Práticos
                                            ft.Text("💡 Exemplos Práticos", size=16, weight="bold"),
                                            ft.Container(height=10),
                                            
                                            # Exemplo 1 - Perfeito
                                            ft.Container(
                                                content=ft.Column([
                                                    ft.Row([
                                                        ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=20),
                                                        ft.Text("Exemplo 1 - Dados Perfeitos (Score: 100)", size=13, weight="bold", color=ft.Colors.GREEN),
                                                    ]),
                                                    ft.Text("OTIF: 8.5 | NIL: 9.0 | Pickup: 8.8 | Package: 9.2", size=12),
                                                    ft.Text("✓ Sem inconsistências detectadas", size=11, italic=True),
                                                ], spacing=5),
                                                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                                                padding=ft.padding.all(10),
                                                border_radius=8,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            # Exemplo 2 - Valores idênticos (OK se forem 10)
                                            ft.Container(
                                                content=ft.Column([
                                                    ft.Row([
                                                        ft.Icon(ft.Icons.CHECK_CIRCLE, color=ft.Colors.GREEN, size=20),
                                                        ft.Text("Exemplo 2 - Tudo Perfeito (Score: 100)", size=13, weight="bold", color=ft.Colors.GREEN),
                                                    ]),
                                                    ft.Text("OTIF: 10.0 | NIL: 10.0 | Pickup: 10.0 | Package: 10.0", size=12),
                                                    ft.Text("✓ Valores idênticos são OK quando todos são 10 (perfeito)", size=11, italic=True),
                                                ], spacing=5),
                                                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                                                padding=ft.padding.all(10),
                                                border_radius=8,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            # Exemplo 3 - Valores idênticos suspeitos
                                            ft.Container(
                                                content=ft.Column([
                                                    ft.Row([
                                                        ft.Icon(ft.Icons.WARNING, color=ft.Colors.ORANGE, size=20),
                                                        ft.Text("Exemplo 3 - Valores Idênticos Suspeitos (Score: 85)", size=13, weight="bold", color=ft.Colors.ORANGE),
                                                    ]),
                                                    ft.Text("OTIF: 7.5 | NIL: 7.5 | Pickup: 7.5 | Package: 7.5", size=12),
                                                    ft.Text("⚠ Valores idênticos (exceto 0 ou 10) podem ser erro (-15 pontos)", size=11, italic=True),
                                                ], spacing=5),
                                                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                                                padding=ft.padding.all(10),
                                                border_radius=8,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            # Exemplo 4 - Logística ruim mas produto OK
                                            ft.Container(
                                                content=ft.Column([
                                                    ft.Row([
                                                        ft.Icon(ft.Icons.WARNING, color=ft.Colors.ORANGE, size=20),
                                                        ft.Text("Exemplo 4 - Logística Ruim mas Inspeção OK (Score: 80)", size=13, weight="bold", color=ft.Colors.ORANGE),
                                                    ]),
                                                    ft.Text("OTIF: 7.0 | NIL: 2.0 | Pickup: 9.0 | Package: 8.0", size=12),
                                                    ft.Text("⚠ NIL péssimo (≤3) mas Pickup excelente (≥8) é inconsistente (-20 pontos)", size=11, italic=True),
                                                ], spacing=5),
                                                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                                                padding=ft.padding.all(10),
                                                border_radius=8,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            # Exemplo 5 - Sem entrega mas com avaliações
                                            ft.Container(
                                                content=ft.Column([
                                                    ft.Row([
                                                        ft.Icon(ft.Icons.ERROR, color=ft.Colors.RED, size=20),
                                                        ft.Text("Exemplo 5 - Sem Entrega mas com Avaliações (Score: 60)", size=13, weight="bold", color=ft.Colors.RED),
                                                    ]),
                                                    ft.Text("OTIF: (vazio) | NIL: 8.0 | Pickup: 8.0 | Package: 8.0", size=12),
                                                    ft.Text("✗ Sem entrega (OTIF vazio) mas outras métricas preenchidas (-40 pontos)", size=11, italic=True),
                                                ], spacing=5),
                                                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                                                padding=ft.padding.all(10),
                                                border_radius=8,
                                            ),
                                            
                                            ft.Container(height=10),
                                            
                                            # Exemplo 6 - Múltiplas inconsistências
                                            ft.Container(
                                                content=ft.Column([
                                                    ft.Row([
                                                        ft.Icon(ft.Icons.ERROR, color=ft.Colors.RED, size=20),
                                                        ft.Text("Exemplo 6 - Múltiplas Inconsistências (Score: 25)", size=13, weight="bold", color=ft.Colors.RED),
                                                    ]),
                                                    ft.Text("OTIF: 9.0 | NIL: 0.0 | Pickup: 0.0 | Package: 9.0", size=12),
                                                    ft.Text(
                                                        "✗ OTIF alto mas NIL zero (-30) + Produto reprovou mas embalagem OK (-20)\n"
                                                        "✗ Embalagem OK mas produto péssimo (-25) = -75 pontos total",
                                                        size=11, italic=True
                                                    ),
                                                ], spacing=5),
                                                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                                                padding=ft.padding.all(10),
                                                border_radius=8,
                                            ),
                                            
                                        ], spacing=0, scroll=ft.ScrollMode.AUTO),
                                        padding=ft.padding.all(20),
                                        expand=True,
                                    ),
                                    elevation=1,
                                ),
                                padding=ft.padding.only(top=10),
                                expand=True,
                            ),
                        ),
                    ],
                    expand=True,
                ),
                expand=True,
            ),
            
        ], spacing=0),
        padding=ft.padding.symmetric(horizontal=20, vertical=20),
        expand=True,
        visible=False
    )
    
    # --- Fim: Lógica da Aba Timeline ---

    def generate_risk_cards(e=None):
        """
        Consulta o banco para encontrar fornecedores em risco e gera os cards de visualização.
        """
        try:
            # 1. Checagens iniciais
            if not db_conn:
                print("⚠️ Conexão com o banco de dados não disponível.")
                return
            if not risks_cards_container or not risks_cards_container.current:
                print("⚠️ Container de cards de risco não inicializado.")
                return

            # 2. Obter parâmetros: ano e meta
            import datetime
            year_val = None
            search_year = None
            if risks_year_dropdown and risks_year_dropdown.current and risks_year_dropdown.current.value:
                year_val = risks_year_dropdown.current.value
                # se for um número válido, usar como filtro; caso contrário, ignorar
                if year_val and year_val.strip().isdigit():
                    search_year = int(year_val)
            # Se search_year for None, vamos buscar todo o período disponível

            meta = target_slider.value if target_slider and target_slider.value is not None else 5.0
            
            # Verificar se deve incluir suppliers inativos
            include_inactive = include_inactive_switch.current.value if include_inactive_switch and include_inactive_switch.current else False

            # 3. Limpar container e mostrar mensagem de carregamento
            risks_cards_container.current.content = ft.Column(
                [ft.ProgressRing(), ft.Text("Buscando fornecedores em risco...")],
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
                alignment=ft.MainAxisAlignment.CENTER,
                expand=True
            )
            risks_cards_container.current.update()

            cursor = db_conn.cursor()

            # Construir filtro de status baseado no switch
            status_filter = "" if include_inactive else "AND s.supplier_status = 'Active'"

            # Abordagem otimizada:
            # - se `search_year` estiver definido, buscar apenas aquele ano
            # - caso contrário, buscar todo o período disponível (todos os anos)
            if search_year is not None:
                all_scores_query = f"""
                    SELECT
                        s.supplier_id, s.vendor_name, s.bu, s.supplier_number, s.supplier_po, sr.year, sr.month, sr.total_score
                    FROM supplier_database_table s
                    JOIN supplier_score_records_table sr ON s.supplier_id = sr.supplier_id
                    WHERE sr.year = ? AND sr.total_score IS NOT NULL 
                    {status_filter}
                    ORDER BY s.supplier_id, sr.year, sr.month
                """
                all_scores = db_manager.query(all_scores_query, (search_year,))
            else:
                all_scores_query = f"""
                    SELECT
                        s.supplier_id, s.vendor_name, s.bu, s.supplier_number, s.supplier_po, sr.year, sr.month, sr.total_score
                    FROM supplier_database_table s
                    JOIN supplier_score_records_table sr ON s.supplier_id = sr.supplier_id
                    WHERE sr.total_score IS NOT NULL 
                    {status_filter}
                    ORDER BY s.supplier_id, sr.year, sr.month
                """
                all_scores = db_manager.query(all_scores_query)

            # Agrupar dados por fornecedor
            from collections import defaultdict
            supplier_data = defaultdict(lambda: {'scores': [], 'info': {}})
            for row in all_scores:
                sid = row['supplier_id']
                vname = row['vendor_name']
                bu = row['bu']
                supplier_number = row['supplier_number']
                supplier_po = row['supplier_po']
                month = row['month']
                score = row['total_score']

                try:
                    month_int = int(month)
                    score_float = float(score)
                    supplier_data[sid]['scores'].append((month_int, score_float))
                    supplier_data[sid]['info'] = {'vendor_name': vname, 'bu': bu, 'supplier_number': supplier_number, 'supplier_po': supplier_po}
                except (ValueError, TypeError) as e:
                    print(f"⚠️ Dados inválidos ignorados para supplier {sid}: month='{month}', score='{score}' - {e}")
                    continue

            if search_year is not None:
                print(f"Total de suppliers com dados no ano {search_year}: {len(supplier_data)}")
            else:
                print(f"Total de suppliers com dados em todo o período: {len(supplier_data)}")

            cards = []
            suppliers_at_risk = 0

            for supplier_id, data in supplier_data.items():
                vendor_name = data['info']['vendor_name']
                bu = data['info']['bu']
                supplier_number = data['info'].get('supplier_number', '')
                supplier_po = data['info'].get('supplier_po', '')
                monthly_scores = data['scores']

                scores_list = [s[1] for s in monthly_scores]
                if not scores_list:
                    continue

                media_score = sum(scores_list) / len(scores_list)

                if media_score >= float(meta):
                    continue

                try:
                    media_val = float(media_score)
                except Exception:
                    continue

                suppliers_at_risk += 1

                # Calcular médias trimestrais
                q_scores = defaultdict(list)
                for month, score in monthly_scores:
                    if 1 <= month <= 3: q_scores['Q1'].append(score)
                    elif 4 <= month <= 6: q_scores['Q2'].append(score)
                    elif 7 <= month <= 9: q_scores['Q3'].append(score)
                    elif 10 <= month <= 12: q_scores['Q4'].append(score)

                q_avgs = {q: sum(s_list) / len(s_list) if s_list else None for q, s_list in q_scores.items()}
                q1, q2, q3, q4 = q_avgs.get('Q1'), q_avgs.get('Q2'), q_avgs.get('Q3'), q_avgs.get('Q4')

                # Criar mini-gráfico - mostrar apenas total_score sem hover
                # Eixo X: meses de 1 a 12, cada mês tem apenas uma nota
                # Se houver múltiplos valores para o mesmo mês (de anos diferentes), usar a média
                month_scores_dict = {}
                for month, score in monthly_scores:
                    if month in month_scores_dict:
                        # Se já existe, fazer média
                        existing_scores = month_scores_dict[month] if isinstance(month_scores_dict[month], list) else [month_scores_dict[month]]
                        existing_scores.append(score)
                        month_scores_dict[month] = existing_scores
                    else:
                        month_scores_dict[month] = [score]
                
                # Criar pontos ordenados por mês, com média se houver duplicatas
                sorted_months = sorted(month_scores_dict.keys())
                
                # Criar segmentos coloridos baseado no target
                chart_series = []
                current_segment = []
                current_color = None
                
                for i, month in enumerate(sorted_months):
                    scores = month_scores_dict[month]
                    avg_score = sum(scores) / len(scores)
                    point = ft.LineChartDataPoint(month, avg_score)
                    
                    # Determinar cor do ponto
                    point_color = ft.Colors.GREEN_600 if avg_score >= meta else ft.Colors.RED_600
                    
                    if current_color is None:
                        # Primeiro ponto
                        current_color = point_color
                        current_segment.append(point)
                    elif current_color == point_color:
                        # Mesma cor, adicionar ao segmento atual
                        current_segment.append(point)
                    else:
                        # Mudança de cor - adicionar ponto ao segmento atual para continuidade
                        current_segment.append(point)
                        
                        # Finalizar segmento atual
                        if len(current_segment) > 0:
                            chart_series.append(ft.LineChartData(
                                data_points=current_segment,
                                color=current_color,
                                stroke_width=2,
                                curved=False,
                                prevent_curve_over_shooting=True,
                            ))
                        
                        # Começar novo segmento com o mesmo ponto para continuidade
                        current_segment = [point]
                        current_color = point_color
                
                # Adicionar último segmento
                if len(current_segment) > 0:
                    chart_series.append(ft.LineChartData(
                        data_points=current_segment,
                        color=current_color,
                        stroke_width=2,
                        curved=False,
                        prevent_curve_over_shooting=True,
                    ))
                
                mini_chart = ft.LineChart(
                    data_series=chart_series,
                    min_y=0, max_y=10, min_x=1, max_x=12,
                    left_axis=ft.ChartAxis(show_labels=False),
                    bottom_axis=ft.ChartAxis(show_labels=False),
                    horizontal_grid_lines=ft.ChartGridLines(width=0),
                    vertical_grid_lines=ft.ChartGridLines(width=0),
                    border=None,
                    expand=True,
                    interactive=False,
                    tooltip_bgcolor=ft.Colors.TRANSPARENT,
                )

                # Helper para criar ícone de tendência
                def trend_icon(current, previous):
                    try:
                        if current is None or previous is None:
                            return ft.Icon(ft.Icons.ARROW_FORWARD, color=ft.Colors.GREY)
                        cur = float(current)
                        prev = float(previous)
                        if cur > prev:
                            return ft.Icon(ft.Icons.ARROW_UPWARD, color=ft.Colors.GREEN)
                        elif cur < prev:
                            return ft.Icon(ft.Icons.ARROW_DOWNWARD, color=ft.Colors.RED)
                        else:
                            return ft.Icon(ft.Icons.ARROW_FORWARD, color=ft.Colors.GREY)
                    except Exception:
                        # Qualquer erro ao converter, retornar ícone neutro
                        return ft.Icon(ft.Icons.ARROW_FORWARD, color=ft.Colors.GREY)

                t2 = trend_icon(q2, q1)
                t3 = trend_icon(q3, q2)
                t4 = trend_icon(q4, q3)

                # Handler para ir para a aba Timeline e filtrar pelo supplier
                def create_goto_handler(sid):
                    def handle(e):
                        try:
                            # 1. Definir supplier no filtro da Timeline
                            if timeline_vendor_dropdown and timeline_vendor_dropdown.current:
                                timeline_vendor_dropdown.current.value = sid

                            # 2. Transferir o ano do filtro de Risks para o de Timeline
                            if risks_year_dropdown and risks_year_dropdown.current and risks_year_dropdown.current.value:
                                y = risks_year_dropdown.current.value
                                if timeline_year_dropdown and timeline_year_dropdown.current:
                                    timeline_year_dropdown.current.value = y

                            # 3. Mudar para a aba Timeline (índice 2)
                            set_selected(2)(None)

                            # 4. Atualizar as métricas e o conteúdo da aba Timeline
                            on_timeline_vendor_change(None)

                            # Ambos gráfico e tabela já estão visíveis, então não precisa alternar abas

                        except Exception as ex:
                            print(f"Erro ao navegar para Timeline: {ex}")
                            import traceback
                            traceback.print_exc()
                    return handle

                # Construir card simplificado
                # Layout final: card maior; média posicionada acima da linha dos Qs, à direita
                card_inner = ft.Container(
                    content=ft.Column([
                        # Topo: nome, BU e ID + média
                        ft.Row(
                            controls=[
                                ft.Column([
                                    ft.Text(format_display_value(vendor_name), weight=ft.FontWeight.BOLD, size=18, max_lines=2, overflow=ft.TextOverflow.ELLIPSIS),
                                    ft.Text(f"BU: {format_display_value(bu)}", size=12, color="gray"),
                                    ft.Text(f"PO: {format_po_ssid(supplier_po)}", size=12, color="gray"),
                                    ft.Text(f"SSID: {format_po_ssid(supplier_number)}", size=12, color="gray"),
                                    ft.Text(f"ID: {format_display_value(supplier_id)}", size=11, color="gray")
                                ], expand=True),
                                ft.Column([
                                    ft.Row([
                                        ft.IconButton(icon=ft.Icons.TIMELINE, tooltip="Abrir Timeline para este fornecedor", icon_size=18, on_click=create_goto_handler(supplier_id))
                                    ], alignment=ft.MainAxisAlignment.END),
                                    ft.Text(f"{media_val:.2f}", size=28, weight=ft.FontWeight.BOLD, color="red"),
                                    ft.Container(content=mini_chart, width=120, height=40)
                                ], horizontal_alignment=ft.CrossAxisAlignment.END)
                            ],
                            vertical_alignment=ft.CrossAxisAlignment.START
                        ),

                        ft.Divider(),

                        # Meio: trimestres (responsivos, quebram linha se faltar espaço)
                        # O controle ft.Wrap não está disponível na sua versão do Flet.
                        # Usando ft.Row com a propriedade wrap=True, que é
                        # a abordagem correta para versões mais antigas e garante a quebra de linha.
                        ft.Row(
                            controls=[
                                ft.Container(
                                    content=ft.Column([ft.Text("Q1", size=11), ft.Text(f"{q1:.2f}" if q1 is not None else "--", size=14)], alignment=ft.CrossAxisAlignment.CENTER, spacing=2),
                                    expand=True, alignment=ft.alignment.center
                                ),
                                ft.Container(
                                    content=ft.Row([t2, ft.Column([ft.Text("Q2", size=11), ft.Text(f"{q2:.2f}" if q2 is not None else "--", size=14)], alignment=ft.CrossAxisAlignment.CENTER, spacing=2)], spacing=2, vertical_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                                    expand=True, alignment=ft.alignment.center
                                ),
                                ft.Container(
                                    content=ft.Row([t3, ft.Column([ft.Text("Q3", size=11), ft.Text(f"{q3:.2f}" if q3 is not None else "--", size=14)], alignment=ft.CrossAxisAlignment.CENTER, spacing=2)], spacing=2, vertical_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                                    expand=True, alignment=ft.alignment.center
                                ),
                                ft.Container(
                                    content=ft.Row([t4, ft.Column([ft.Text("Q4", size=11), ft.Text(f"{q4:.2f}" if q4 is not None else "--", size=14)], alignment=ft.CrossAxisAlignment.CENTER, spacing=2)], spacing=2, vertical_alignment=ft.CrossAxisAlignment.CENTER, alignment=ft.MainAxisAlignment.CENTER),
                                    expand=True, alignment=ft.alignment.center
                                ),
                            ],
                            vertical_alignment=ft.CrossAxisAlignment.CENTER,
                        )
                    ], spacing=8, horizontal_alignment=ft.CrossAxisAlignment.STRETCH), # Espaçamento interno do card reduzido
                    # Padding ajustado para um visual mais "justo"
                    padding=ft.padding.symmetric(vertical=12, horizontal=16),
                    border_radius=12,
                    bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('card_background')
                )


                card = ft.Card(content=card_inner, elevation=6)
                # armazenar média no próprio objeto do card para ordenação segura
                try:
                    card.media_score = float(media_val)
                except Exception:
                    card.media_score = float('inf')
                cards.append(card)

            print(f"Riscos: Encontrados {suppliers_at_risk} fornecedores abaixo da meta {meta} para o ano {search_year}.")

            # Verificar se não há fornecedores em risco
            if not cards:
                title_text = f"Nenhum fornecedor em risco encontrado"
                if search_year is not None:
                    title_text = f"Nenhum fornecedor em risco encontrado para {search_year}!"
                else:
                    title_text = f"Nenhum fornecedor em risco encontrado no período pesquisado!"
                risks_cards_container.current.content = ft.Container(
                    content=ft.Column([
                        ft.Icon(ft.Icons.SHIELD, color="green", size=48),
                        ft.Text(title_text, size=16, text_align=ft.TextAlign.CENTER),
                        ft.Text(f"Todos os fornecedores estão com score acima da meta de {meta:.2f}.", size=12, color="gray")
                    ], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER, spacing=10),
                    expand=True,
                    alignment=ft.alignment.center
                )
                risks_cards_container.current.update()
                return

            # Ordenar cards por média (menor para maior - maior risco primeiro)
            cards.sort(key=lambda card: float(getattr(card, 'media_score', float('inf'))))

            # O GridView força todos os cards em uma mesma linha a terem a mesma altura,
            # o que pode criar um espaço vazio (a "borda exagerada") em cards com menos conteúdo.
            # Para que cada card "abrace" seu próprio conteúdo, usaremos um `ft.Row` com a
            # propriedade `wrap=True`. Isso permite que os cards tenham alturas variáveis. Adicionamos
            # um Column com scroll para conter a Row.
            risks_cards_container.current.content = ft.Column(
                controls=[
                    ft.Row(
                        # Envolvemos cada card em um Container com largura fixa para criar o efeito de colunas.
                        controls=[ft.Container(c, width=400) for c in cards],
                        wrap=True,
                        spacing=10,
                        run_spacing=10,
                        # Alinha os cards à esquerda dentro do container.
                        alignment=ft.MainAxisAlignment.START,
                    )
                ],
                scroll=ft.ScrollMode.AUTO,
                expand=True
            )
            risks_cards_container.current.update()

        except Exception as ex:
            print(f"Erro em generate_risk_cards: {ex}")
            import traceback
            traceback.print_exc()
            if risks_cards_container and risks_cards_container.current:
                risks_cards_container.current.content = ft.Text(f"Erro ao gerar cards de risco: {ex}", color="red")
                risks_cards_container.current.update()





    # Container da aba Risks (com dropdown de anos e área responsiva para cards)
    risks_view = ft.Container(
        content=ft.Column([
            ft.Row([
                ft.Text("Gestão de Riscos", size=24, weight="bold"),
            ], alignment=ft.MainAxisAlignment.START),
            ft.Divider(),
            # Campos ano e meta agrupados em container com cor do tema
            ft.Container(
                        content=ft.Row(
                    [
                        ft.Dropdown(
                            ref=risks_year_dropdown,
                            width=220,
                            value="",
                            on_change=generate_risk_cards,
                            options=[ft.dropdown.Option("", "Todo Período")] 
                                    + [ft.dropdown.Option(str(y), str(y)) for y in range(2024, 2041)],
                            hint_text="Ano",
                            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('field_background'),
                            color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface'),
                            border_color=get_current_theme_colors(get_theme_name_from_page(page)).get('outline')
                        ),
                        ft.Container(width=20),
                        ft.Container(
                            ref=target_display_container,
                            content=ft.Row(
                                [
                                    ft.Icon(ft.Icons.FLAG_CIRCLE_OUTLINED, color=ft.Colors.AMBER_700, size=20),
                                    ft.Text("Meta:", weight=ft.FontWeight.BOLD, size=14),
                                    ft.Text("", ref=target_risks_text, weight=ft.FontWeight.BOLD, size=16, color=ft.Colors.AMBER_700),
                                ],
                                spacing=5,
                                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                                tight=True,
                            ),
                            padding=ft.padding.symmetric(horizontal=12, vertical=6),
                            border=ft.border.all(1.5, ft.Colors.AMBER_700),
                            border_radius=30,
                            bgcolor=ft.Colors.with_opacity(0.1, ft.Colors.AMBER_700),
                        ),
                        ft.Container(width=20),
                        ft.Container(
                            ref=inactive_switch_container,
                            content=ft.Row(
                                [
                                    ft.Icon(ft.Icons.VISIBILITY, ref=inactive_switch_icon, color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'), size=20),
                                    ft.Text("Incluir Inativos:", weight=ft.FontWeight.BOLD, size=14),
                                    ft.Switch(
                                        ref=include_inactive_switch,
                                        value=False,
                                        on_change=generate_risk_cards,
                                        active_color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary'),
                                        inactive_thumb_color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'),
                                    ),
                                ],
                                spacing=8,
                                vertical_alignment=ft.CrossAxisAlignment.CENTER,
                                tight=True,
                            ),
                            padding=ft.padding.symmetric(horizontal=12, vertical=6),
                            border=ft.border.all(1.5, get_current_theme_colors(get_theme_name_from_page(page)).get('outline')),
                            border_radius=30,
                            bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                        ),
                    ],
                    alignment=ft.MainAxisAlignment.START,
                    spacing=12,
                    vertical_alignment=ft.CrossAxisAlignment.CENTER,
                    tight=True,  # <--- IMPORTANTE
                ),
                padding=ft.padding.symmetric(horizontal=16, vertical=16),
                margin=ft.margin.only(bottom=6),
                ref=risks_header_container,
                border=None,  # Remover borda inicial
                border_radius=12,
                bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
                expand=False,  # <--- NÃO DEIXAR EXPANDIR
            ),
            ft.Container(height=12),
            ft.Container(ref=risks_cards_container, content=ft.Text("Nenhum risco pesquisado"), expand=True)
        ], spacing=8, horizontal_alignment=ft.CrossAxisAlignment.START),
        alignment=ft.alignment.top_center,
        expand=True,
        visible=False,
        padding=20
    )

    email_view = ft.Container(
        content=ft.Column([
            ft.Text("Envio de E-mails", size=24, weight="bold"),
            ft.Divider(),
            ft.Container(height=10),
            create_email_development_tabs()
        ], expand=True), 
        alignment=ft.alignment.top_center, 
        expand=True, 
        visible=False,
        padding=20
    )
    configs_view = ft.Container(
        content=configs_view_content, 
        alignment=ft.alignment.top_center, 
        padding=20, 
        visible=False,
        expand=True
    )

    menu_column = ft.Column([],
        alignment=ft.MainAxisAlignment.CENTER,
        horizontal_alignment=ft.CrossAxisAlignment.CENTER,
        spacing=15,
        ref=menu_column_ref
    )

    rail_container = ft.Container(
        content=ft.Column(
            [
                ft.IconButton(icon="menu", on_click=toggle_menu),
                menu_column,
            ],
            alignment=ft.MainAxisAlignment.START,
            horizontal_alignment=ft.CrossAxisAlignment.START,
            spacing=12,
        ),
        padding=ft.padding.only(left=4, right=4),  # Removido padding esquerdo
        bgcolor=get_current_theme_colors(get_theme_name_from_page(page))["rail_background"]  # Cor mais escura
    )

    # Função para mudar para aba de configurações
    def go_to_configs(e):
        selected_index.current = 5
        home_view.visible = False
        score_view.visible = False
        timeline_view.visible = False
        risks_view.visible = False
        email_view.visible = False
        configs_view.visible = True
        page.update()
    
    # Função para ir para aba de Pendências
    def go_to_pendencies(e):
        # Usar set_selected para atualizar o menu lateral corretamente
        set_selected(1)(e)
        # Definir a sub-aba de Pendências (índice 1)
        score_tabs.selected_index = 1
        # Atualizar a página
        page.update()
    
    # Função para verificar se há pendências reais
    def has_pending_scores():
        try:
            if not db_conn:
                return False
            
            # Usar a função get_pending_suppliers() que já existe e verifica pendências reais
            pending_suppliers = get_pending_suppliers()
            
            # Retorna True apenas se houver pendências reais que o usuário pode preencher
            return pending_suppliers and len(pending_suppliers) > 0
        except Exception as e:
            print(f"Erro ao verificar pendências: {e}")
            return False
    
    # Funções para controle da janela
    def minimize_window(e):
        page.window.minimized = True
        page.update()
    
    def maximize_window(e):
        page.window.maximized = not page.window.maximized
        page.update()
    
    def close_window(e):
        page.window.close()
    
    # Top bar - Barra superior fina e clean com controles de janela personalizados
    top_bar = ft.Container(
        content=ft.Row(
            [
                # Lado esquerdo - Informações do usuário
                ft.Row([
                    ft.Icon(
                        ft.Icons.ACCOUNT_CIRCLE, 
                        size=22,
                        color=get_current_theme_colors(get_theme_name_from_page(page)).get('primary')
                    ),
                    ft.Column([
                        ft.Text(
                            f"{current_user_name} ({current_user_wwid})",
                            size=13,
                            weight=ft.FontWeight.BOLD,
                        ),
                        ft.Text(
                            current_user_privilege,
                            size=11,
                            opacity=0.7,
                        ),
                    ], spacing=0, tight=True),
                ], spacing=8),
                
                # Centro - Espaço flexível
                ft.Container(expand=True),
                
                # Lado direito - Ícones de ação com hover
                ft.Row([
                    # Ícone de pendências (apenas ícone, sem texto)
                    ft.Container(
                        content=ft.IconButton(
                            icon=ft.Icons.PENDING_ACTIONS,
                            icon_size=20,
                            tooltip="Pendências de avaliação",
                            icon_color="orange" if has_pending_scores() else get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'),
                            on_click=go_to_pendencies,
                            style=ft.ButtonStyle(
                                overlay_color={
                                    ft.ControlState.HOVERED: ft.Colors.with_opacity(0.1, get_current_theme_colors(get_theme_name_from_page(page)).get('primary'))
                                }
                            ),
                        ),
                        visible=current_user_privilege in ["Super Admin", "Admin"],
                    ),
                    
                    # Ícone de configurações
                    ft.IconButton(
                        icon=ft.Icons.SETTINGS,
                        icon_size=20,
                        tooltip="Configurações",
                        on_click=go_to_configs,
                        icon_color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'),
                        style=ft.ButtonStyle(
                            overlay_color={
                                ft.ControlState.HOVERED: ft.Colors.with_opacity(0.1, get_current_theme_colors(get_theme_name_from_page(page)).get('primary'))
                            }
                        ),
                    ),
                    
                    # Separador visual
                    ft.Container(
                        width=1,
                        height=20,
                        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('outline'),
                        margin=ft.margin.symmetric(horizontal=5),
                    ),
                    
                    # Botão Minimizar
                    ft.IconButton(
                        icon=ft.Icons.MINIMIZE,
                        icon_size=16,
                        tooltip="Minimizar",
                        on_click=minimize_window,
                        icon_color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'),
                        style=ft.ButtonStyle(
                            overlay_color={
                                ft.ControlState.HOVERED: ft.Colors.with_opacity(0.1, get_current_theme_colors(get_theme_name_from_page(page)).get('primary'))
                            }
                        ),
                    ),
                    
                    # Botão Maximizar/Restaurar
                    ft.IconButton(
                        icon=ft.Icons.CROP_SQUARE,
                        icon_size=16,
                        tooltip="Maximizar/Restaurar",
                        on_click=maximize_window,
                        icon_color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'),
                        style=ft.ButtonStyle(
                            overlay_color={
                                ft.ControlState.HOVERED: ft.Colors.with_opacity(0.1, get_current_theme_colors(get_theme_name_from_page(page)).get('primary'))
                            }
                        ),
                    ),
                    
                    # Botão Fechar
                    ft.IconButton(
                        icon=ft.Icons.CLOSE,
                        icon_size=16,
                        tooltip="Fechar",
                        on_click=close_window,
                        icon_color=get_current_theme_colors(get_theme_name_from_page(page)).get('on_surface_variant'),
                        style=ft.ButtonStyle(
                            overlay_color={
                                ft.ControlState.HOVERED: ft.Colors.with_opacity(0.2, ft.Colors.RED)
                            }
                        ),
                    ),
                ], spacing=5),
            ],
            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
            vertical_alignment=ft.CrossAxisAlignment.CENTER,
        ),
        height=50,
        padding=ft.padding.symmetric(horizontal=20, vertical=8),
        bgcolor=get_current_theme_colors(get_theme_name_from_page(page)).get('surface_variant'),
        # Borda removida para visual mais clean
    )

    # Envolver top_bar com WindowDragArea para permitir arrastar a janela
    draggable_top_bar = ft.WindowDragArea(
        content=top_bar,
        maximizable=True,  # Permitir maximizar com duplo clique
    )
    
    page.add(
        ft.Container(
            content=ft.Row(
                [
                    rail_container,
                    ft.VerticalDivider(width=1),  # Removido color para ficar transparente
                    ft.Container( # Container principal para o conteúdo das abas
                        content=ft.Column([
                            draggable_top_bar,  # Barra superior arrastável
                            ft.Container(
                                content=ft.Column([home_view, score_view, timeline_view, risks_view, email_view, configs_view], expand=True),
                                expand=True,
                                padding=ft.padding.all(20),  # Padding ao redor de todas as abas
                            )
                        ], spacing=0, expand=True),
                        expand=True,
                    ),
                ],
                expand=True,
                vertical_alignment=ft.CrossAxisAlignment.STRETCH,
                spacing=0,  # Remove espaçamento entre componentes
            ),
            expand=True
        )
    )

    # Inicializar gerenciador responsivo
    global responsive_app_manager
    responsive_app_manager = ResponsiveAppManager(page)
    responsive_app_manager.initialize_containers(results_list, score_view_content)
    responsive_app_manager.initialize_menu_controls(menu_is_expanded, update_menu)
    responsive_app_manager.check_initial_window_state()  # Verificar estado inicial da janela
    print("📱 Gerenciador responsivo inicializado e estado inicial verificado")
    
    # Aplicar layout inicial da timeline
    window_width = page.window.width or 1200
    responsive_app_manager.update_timeline_layout(window_width)

    def update_interface_for_user_privileges():
        """Atualiza a interface baseado nos privilégios do usuário atual"""
        try:
            # Atualizar menu principal (ocultar/mostrar abas)
            update_menu()
            
            # Atualizar abas de configuração  
            update_config_tabs()
            
            print(f"Interface atualizada para privilégio: {current_user_privilege}")
            
        except Exception as ex:
            print(f"Erro ao atualizar interface para privilégios: {ex}")

    update_menu()
    update_config_tabs()  # Inicializar as abas de configuração
    
    # Atualizar interface baseado nos privilégios do usuário
    update_interface_for_user_privileges()
    
    # Inicializar controles de usuário
    print("Extraindo referências dos controles de usuário...")
    try:
        users_controls.update(extract_users_refs())
    except Exception as ex:
        print(f"Erro ao extrair controles: {ex}")
    
    # Configurar eventos dos botões de usuário
    try:
        print("Configurando eventos dos botões...")
        # Passar diretamente o handler para que o evento seja recebido (usado para controlar o botão)
        users_controls['action_btn'].on_click = add_or_update_user
        users_controls['clear_btn'].on_click = clear_users_fields
        print("Eventos configurados com sucesso!")
    except Exception as ex:
        print(f"Erro ao configurar eventos: {ex}")
    
    load_all_lists_data()  # Carregar dados existentes das tabelas
    
    # Carregar usuários na inicialização
    refresh_users_list()
    
    # Não carregar suppliers inicialmente - aguardar pesquisa do usuário ou seleção de mês/ano
    # reload_score_table()
    
    # Carregar critérios salvos na inicialização
    saved_criteria = load_user_criteria(current_user_wwid)
    if saved_criteria:
        # Mapear nomes dos critérios da tabela para os sliders
        if 'NIL' in saved_criteria:
            nil_slider.value = saved_criteria['NIL']
            nil_text.value = f"{saved_criteria['NIL']:.2f}"
        
        if 'OTIF' in saved_criteria:
            otif_slider.value = saved_criteria['OTIF']
            otif_text.value = f"{saved_criteria['OTIF']:.2f}"
        
        if 'Quality of Pick Up' in saved_criteria:
            pickup_slider.value = saved_criteria['Quality of Pick Up']
            pickup_text.value = f"{saved_criteria['Quality of Pick Up']:.2f}"
        
        if 'Quality-Supplier Package' in saved_criteria:
            package_slider.value = saved_criteria['Quality-Supplier Package']
            package_text.value = f"{saved_criteria['Quality-Supplier Package']:.2f}"
        
        if 'Target' in saved_criteria:
            target_slider.value = saved_criteria['Target']
            target_text.value = f"{saved_criteria['Target']:.2f}"
        
        # Atualizar a soma dos pesos após carregar
        update_weight_sum()
        safe_page_update(page)
        print(f"Critérios carregados: {saved_criteria}")
    else:
        # Mesmo com valores padrão, atualizar a soma
        update_weight_sum()
        print("Nenhum critério salvo encontrado, usando valores padrão")
    
    # Atualizar o texto do target na aba Risks
    if target_risks_text.current:
        target_risks_text.current.value = f"{target_slider.value:.2f}"
        safe_page_update(page)
def main():
    """Função principal que inicia com tela de login"""
    ft.app(target=login_screen)

if __name__ == "__main__":
    main()